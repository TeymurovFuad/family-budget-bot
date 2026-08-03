"""
excel_export.py — regenerate the Excel workbook from SQLite (Cycle S1 Phase 1).

Excel becomes a generated export: full regeneration from the template plus
all SQLite transaction rows. Reuses excel_schema.write_transaction_row for
layout/validation, but materializes Value (base) as a plain number instead
of the live VLOOKUP formula (SQLite already computed it).

Standalone callable — not wired into APScheduler yet (Phase 2).
"""

from logger import get_logger

import sqlite_ops
from sqlite_types import SyncDirection, SyncStatus
from excel_schema import (
    ListsSchema,
    MasterDataSchema,
    col_indices,
    ensure_monthly_summary_rows_from_masterdata,
    find_next_data_row,
    lists_currency_range,
    to_date,
    write_transaction_row,
    sync_dashboard_categories,
)
from cycle_dashboard import sync_cycle_dashboard_categories
from storage_backends import atomic_save

log = get_logger(__name__)


def _sqlite_row_to_excel_row(r: dict) -> dict:
    """Map a transactions-table dict to write_transaction_row's input dict."""
    return {
        "date":         to_date(r.get("date")),
        "year":         r.get("year"),
        "month":        r.get("month"),
        "value":        r.get("value"),
        "type":         r.get("type"),
        "category":     r.get("category"),
        "person":       r.get("person"),
        "description":  r.get("description"),
        "is_recurring": bool(r["is_recurring"]) if r.get("is_recurring") is not None else None,
        "is_done":      r.get("is_done"),
        "currency":     r.get("currency"),
    }


def generate_excel_from_sqlite(db_path, template_path, output_path) -> None:
    """
    Full regeneration: template workbook + every SQLite transaction row.
    Value (base) is written as the stored numeric value_base, not a formula.
    Saved crash-safely via atomic_save. Logs the run into sync_log.
    """
    from openpyxl import load_workbook

    conn = sqlite_ops.init_db(db_path)
    try:
        rows = sqlite_ops.list_transactions(conn)

        wb = load_workbook(template_path)
        ws = wb["MasterData"]
        lu_range = lists_currency_range(wb)
        idx = col_indices(ws, MasterDataSchema)
        vbase_col = idx.get("value_base", 12)

        r = find_next_data_row(ws)
        dm_col = None
        for row in rows:
            write_transaction_row(ws, r, _sqlite_row_to_excel_row(row), lu_range)
            # Override the VLOOKUP formula with the materialized value.
            ws.cell(r, vbase_col, row.get("value_base"))
            # Override the write-time timestamp write_transaction_row stamps
            # with the stored date_modified_utc, so re-importing this export
            # reproduces the same content_hash (idempotent round trip).
            if dm_col is None:
                # Recompute after the first write: write_transaction_row
                # creates the Date Modified header if the template lacks it.
                dm_col = col_indices(ws, MasterDataSchema).get("date_modified", 13)
            ws.cell(r, dm_col, row.get("date_modified_utc"))
            r += 1

        ensure_monthly_summary_rows_from_masterdata(wb)

        # Rebuild Dashboard and Cycle Dashboard category blocks so renamed or
        # newly added categories appear on those sheets after export.
        if "Lists" in wb.sheetnames:
            _idx_lists = col_indices(wb["Lists"], ListsSchema)
            _cat_col = _idx_lists.get("categories")
            _categories: list = []
            if _cat_col:
                for _r in range(2, wb["Lists"].max_row + 1):
                    _val = wb["Lists"].cell(_r, _cat_col).value
                    if _val is None or (isinstance(_val, str) and _val.startswith("←")):
                        break
                    _categories.append(_val)
            sync_dashboard_categories(wb, _categories)
        if "Cycle Dashboard" in wb.sheetnames:
            sync_cycle_dashboard_categories(wb)

        atomic_save(wb, output_path)
        sqlite_ops.log_sync(conn, SyncDirection.EXPORT, SyncStatus.OK,
                            f"{len(rows)} rows -> {output_path}")
        log.info("Exported %d SQLite transactions to %s", len(rows), output_path)
    except Exception as e:
        try:
            sqlite_ops.log_sync(conn, SyncDirection.EXPORT, SyncStatus.ERROR, str(e))
        except Exception:
            pass
        raise
    finally:
        conn.close()
