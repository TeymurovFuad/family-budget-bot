"""/bulk conversation — import transactions from photo, file, or pasted text."""

import asyncio
import json
import re
from datetime import datetime, date, timezone
from pathlib import Path

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from telegram.error import BadRequest
from telegram.ext import ContextTypes, ConversationHandler

import ai_parser
from ai_parser import parse_text, parse_image, _chunk_statement_text
from config import auth, auth_write, log
from data import load_dedup_evidence, load_reference_data
import settings
from excel_ops import async_append_batch
import merchant_map
from handlers.cycle import maybe_prompt_cycle_start
from models import Transaction
import statement_profiles as sp
from states import (
    BULK_RECEIVE, BULK_CONFIRM,
    BULK_STATEMENT, BULK_PROFILE_CONFIRM, BULK_PROFILE_NAME,
    BULK_PROFILE_FIX_COL, BULK_PROFILE_FIX_FIELD,
)
from validators import (
    clean_merchant_description,
    coerce_bool,
    make_dedup_key,
    make_loose_dedup_key,
    parse_amount,
    validate_parsed_row,
)

# ── Statement-profile helpers ─────────────────────────────────────────────────

_STATEMENT_EXTENSIONS = {".csv", ".xlsx", ".xls", ".txt"}

# All standard fields a user can assign a column to (plus "skip").
# "debit" and "credit" support split-column bank formats (one column per direction).
_MAPPABLE_FIELDS = ("date", "amount", "debit", "credit", "currency", "description", "time", "skip")


def _load_profiles() -> dict:
    """Load profiles from STATEMENT_PROFILES_DIR; returns fingerprint→profile dict."""
    return sp.load_profiles(settings.STATEMENT_PROFILES_DIR)


def _is_statement_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in _STATEMENT_EXTENSIONS


async def _cmd_bulk_profile_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    """List all saved profiles with inline Delete buttons."""
    profiles = await asyncio.to_thread(sp.list_profiles, settings.STATEMENT_PROFILES_DIR)
    if not profiles:
        await update.message.reply_text("No statement profiles saved yet.")
        return ConversationHandler.END

    lines = ["*Saved statement profiles:*"]
    buttons = []
    for p in profiles:
        name = p.get("name") or "?"
        fp_count = len(p.get("fingerprint") or [])
        sign = p.get("sign_convention") or "?"
        lines.append(f"• {name} ({fp_count} columns, {sign})")
        safe = sp.profile_safe_name(name)
        cb = f"profile_del:{safe}"
        if len(cb.encode()) <= 64:
            buttons.append([InlineKeyboardButton(f"❌ Delete {name}", callback_data=cb)])
        else:
            lines[-1] += f"  (use /bulk profile delete {name} — name too long for button)"

    await update.message.reply_text(
        "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons) if buttons else None,
    )
    return ConversationHandler.END


@auth_write
async def bulk_profile_list_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle inline-button callbacks for profile deletion:
      profile_del:<safe_name>         → show confirmation
      profile_del_confirm:<safe_name> → delete and confirm
      profile_del_cancel              → cancel
    """
    query = update.callback_query
    await query.answer()
    data = query.data or ""

    if data == "profile_del_cancel":
        await query.edit_message_text("Deletion cancelled.")
        return

    if data.startswith("profile_del_confirm:"):
        safe_name = data[len("profile_del_confirm:"):]
        profiles = await asyncio.to_thread(sp.list_profiles, settings.STATEMENT_PROFILES_DIR)
        target = next(
            (p for p in profiles if sp.profile_safe_name(p.get("name") or "") == safe_name),
            None,
        )
        name = target.get("name") if target else safe_name
        deleted = await asyncio.to_thread(sp.delete_profile, name, settings.STATEMENT_PROFILES_DIR)
        if deleted:
            await query.edit_message_text(f"✅ Profile '{name}' deleted.")
        else:
            await query.edit_message_text(
                f"Profile '{name}' not found — it may have already been deleted."
            )
        return

    if data.startswith("profile_del:"):
        safe_name = data[len("profile_del:"):]
        profiles = await asyncio.to_thread(sp.list_profiles, settings.STATEMENT_PROFILES_DIR)
        target = next(
            (p for p in profiles if sp.profile_safe_name(p.get("name") or "") == safe_name),
            None,
        )
        name = target.get("name") if target else safe_name
        confirm_cb = f"profile_del_confirm:{safe_name}"
        if len(confirm_cb.encode()) > 64:
            await query.edit_message_text("Profile name too long to delete via button.")
            return
        await query.edit_message_text(
            f"Delete profile '{name}'? This cannot be undone.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("Yes, delete", callback_data=confirm_cb),
                InlineKeyboardButton("Cancel", callback_data="profile_del_cancel"),
            ]]),
        )


def _read_statement_headers_and_sniff(
    file_bytes: bytes,
    filename: str,
) -> tuple[list[str], dict | None]:
    """
    Read the header row from a CSV/XLSX/TXT file.
    Returns (headers, provisional_profile_or_None).
    provisional_profile is only set for .txt files when a delimiter is detected;
    for CSV/XLSX it is None (profile comes from the registry).
    Returns ([], None) when the file cannot be sniffed (caller falls through).
    """
    import csv as _csv
    import io as _io
    ext = Path(filename).suffix.lower()

    if ext in {".xlsx", ".xls"}:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=True)
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()
            if first_row:
                return [str(c or "").strip() for c in first_row], None
        except Exception as exc:
            log.warning("_read_statement_headers XLSX error: %s", exc)
        return [], None

    # CSV or TXT — decode first.
    try:
        content = file_bytes.decode("utf-8", errors="replace")
    except Exception:
        content = file_bytes.decode("latin-1", errors="replace")

    if ext == ".txt":
        delim = sp.sniff_txt_delimiter(content)
        if not delim:
            return [], None  # caller falls through to AI free-form path
        reader = _csv.reader(_io.StringIO(content), delimiter=delim)
        headers = [h.strip() for h in next(reader, [])]
        if not headers:
            return [], None
        provisional = {
            "name": "",
            "delimiter": delim,
            "encoding": "utf-8",
            "header_row": 0,
            "fingerprint": headers,
            "column_map": {
                f: None for f in ("date", "amount", "debit", "credit", "currency", "description", "time")
            },
            "date_format": "%Y-%m-%d",
            "decimal_separator": ".",
            "sign_convention": "negative_expense",
        }
        return headers, provisional

    # Plain CSV — sniff delimiter.
    delim = sp.sniff_txt_delimiter(content, candidates=(";", ",", "\t")) or ","
    reader = _csv.reader(_io.StringIO(content), delimiter=delim)
    headers = [h.strip() for h in next(reader, [])]
    return headers, None


def _get_sample_rows(file_bytes: bytes, filename: str, delimiter: str = ",", n: int = 3) -> list[list]:
    """Return up to n data rows (after the header) as lists of strings."""
    import csv as _csv
    import io as _io
    ext = Path(filename).suffix.lower()
    if ext in {".xlsx", ".xls"}:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                rows.append([str(c) if c is not None else "" for c in row])
                if len(rows) >= n:
                    break
            wb.close()
            return rows
        except Exception:
            return []
    try:
        content = file_bytes.decode("utf-8", errors="replace")
    except Exception:
        content = file_bytes.decode("latin-1", errors="replace")
    reader = _csv.reader(_io.StringIO(content), delimiter=delimiter)
    next(reader, None)  # skip header
    rows = []
    for row in reader:
        rows.append(row)
        if len(rows) >= n:
            break
    return rows


def _format_profile_confirm_message(proposal: dict) -> str:
    """
    Render the profile-confirmation message in a structured layout:

      New statement format detected. My reading:
      ━━━━━━━━━━━━━━━━━━━━━━━
      Required fields:
        ✅ date      → 'Transaction date' (YYYY-MM-DD)
        ✅ amount    → split: 'Debits' (expense) + 'Credits' (income)
        ✅ currency  → 'Currency'

      Optional fields:
        ✅ description → 'Description'
        ➖ time        → not found (OK — used only for same-day dedup)

      Ignored columns: Account/Card Number, Balance, ...
    """
    col_map = proposal.get("column_map") or {}
    date_fmt = proposal.get("date_format") or "?"
    decimal_sep = proposal.get("decimal_separator") or "."
    sign = proposal.get("sign_convention") or "?"
    sep_word = "comma" if decimal_sep == "," else "dot"

    lines = ["New statement format detected. My reading:", "━━━━━━━━━━━━━━━━━━━━━━━"]

    # ── Required fields ───────────────────────────────────────────────────────
    lines.append("Required fields:")

    date_col = col_map.get("date")
    if date_col:
        lines.append(f"  ✅ date      → '{date_col}' ({date_fmt})")
    else:
        lines.append("  ❌ date      → not mapped (required)")

    if sign == sp.SIGN_DEBIT_CREDIT_SPLIT:
        debit_col = col_map.get("debit")
        credit_col = col_map.get("credit")
        if debit_col and credit_col:
            lines.append(
                f"  ✅ amount    → split: '{debit_col}' (expense) + '{credit_col}' (income)"
            )
        elif debit_col:
            lines.append(f"  ❌ amount    → debit '{debit_col}' mapped — credit column missing (required)")
        elif credit_col:
            lines.append(f"  ❌ amount    → credit '{credit_col}' mapped — debit column missing (required)")
        else:
            lines.append("  ❌ amount    → not mapped (required)")
    else:
        amount_col = col_map.get("amount")
        if amount_col:
            sign_word = "negative = expense" if sign == "negative_expense" else sign
            lines.append(f"  ✅ amount    → '{amount_col}' ({sep_word} decimal, {sign_word})")
        else:
            lines.append("  ❌ amount    → not mapped (required)")

    currency_col = col_map.get("currency")
    if currency_col:
        lines.append(f"  ✅ currency  → '{currency_col}'")
    else:
        lines.append("  ❌ currency  → not mapped (required)")

    # ── Optional fields ───────────────────────────────────────────────────────
    lines.append("")
    lines.append("Optional fields:")

    desc_col = col_map.get("description")
    if desc_col:
        lines.append(f"  ✅ description → '{desc_col}'")
    else:
        lines.append("  ➖ description → not found (OK — used for merchant memory)")

    time_col = col_map.get("time")
    if time_col:
        lines.append(f"  ✅ time        → '{time_col}'")
    else:
        lines.append("  ➖ time        → not found (OK — used only for same-day dedup)")

    # ── Ignored columns ───────────────────────────────────────────────────────
    mapped_cols = {v for v in col_map.values() if v}
    headers = proposal.get("fingerprint") or []
    ignored = [h for h in headers if h and h not in mapped_cols]
    if ignored:
        ignored_str = ", ".join(ignored)
        lines.append("")
        lines.append(f"Ignored columns: {ignored_str}")

    return "\n".join(lines)


def _profile_confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("Looks right", callback_data="profile_ok"),
        InlineKeyboardButton("Fix a column", callback_data="profile_fix"),
        InlineKeyboardButton("Cancel", callback_data="profile_cancel"),
    ]])


def _column_pick_keyboard(headers: list[str]) -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(h, callback_data=f"fix_col:{h}")]
        for h in headers
    ]
    buttons.append([InlineKeyboardButton("Cancel", callback_data="profile_cancel")])
    return InlineKeyboardMarkup(buttons)


_FIELD_LABELS = {
    "date": "date",
    "amount": "amount (single column)",
    "debit": "debit — expense column (split mode)",
    "credit": "credit — income column (split mode)",
    "currency": "currency",
    "description": "description",
    "time": "time",
    "skip": "skip (ignore this column)",
}


def _field_pick_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(_FIELD_LABELS.get(f, f), callback_data=f"fix_field:{f}")]
        for f in _MAPPABLE_FIELDS
    ]
    return InlineKeyboardMarkup(buttons)


async def _finish_profile_parse(
    update: Update,
    ctx: ContextTypes.DEFAULT_TYPE,
    file_bytes: bytes,
    filename: str,
    profile: dict,
    profile_name: str,
) -> int:
    """
    After a profile is confirmed (matched or newly created): parse the file,
    normalize rows, run merchant memory + validation, merge into the draft,
    run dedup, and send the preview. Returns the next conversation state.
    """
    uid = update.effective_user.id
    lists = ctx.user_data.get("lists") or load_reference_data()
    ctx.user_data["lists"] = lists

    loop = asyncio.get_running_loop()
    parsed = await loop.run_in_executor(
        None, lambda: sp.parse_statement(file_bytes, filename, profile)
    )

    if not parsed:
        await update.effective_message.reply_text("No transactions found in the statement.")
        return ConversationHandler.END

    parsed, corrections = _normalize_parsed_rows(parsed, lists)
    memory_notes = _apply_merchant_memory(parsed)
    parsed, validator_corrections = _validate_bulk_rows(parsed, lists)
    corrections += validator_corrections

    if memory_notes:
        shown = "\n".join(f"  • {n}" for n in memory_notes[:10])
        more = f"\n  … and {len(memory_notes) - 10} more" if len(memory_notes) > 10 else ""
        await update.effective_message.reply_text(
            f"🧠 {len(memory_notes)} row(s) categorized from merchant memory:\n{shown}{more}"
        )
    if corrections:
        shown = "\n".join(f"  • {c}" for c in corrections[:10])
        more = f"\n  … and {len(corrections) - 10} more" if len(corrections) > 10 else ""
        await update.effective_message.reply_text(
            f"🛡 Auto-corrected {len(corrections)} value(s):\n{shown}{more}"
        )

    parsed = _sort_bulk_rows(parsed)
    draft_rows, _ = _merge_bulk_draft(uid, parsed)
    ctx.user_data["bulk_parsed"] = draft_rows

    summary = _flag_master_duplicates(draft_rows)
    if summary["flagged"] or summary["identical_groups"] or summary["loose_matches"]:
        _save_bulk_draft(uid, draft_rows)
        for msg in _format_dedup_messages(summary):
            await update.effective_message.reply_text(msg)

    n_rows = len([r for r in parsed if r])
    await update.effective_message.reply_text(
        f"📄 Parsed with profile '{profile_name}' — {n_rows} row(s)."
    )
    await _send_bulk_preview_msg(update, draft_rows)
    return BULK_CONFIRM


async def _send_bulk_preview_msg(update: Update, parsed: list[dict]) -> None:
    """Send bulk preview — works from both message and callback_query update."""
    keyboard = ReplyKeyboardMarkup([["Save", "Cancel"]], one_time_keyboard=True, resize_keyboard=True)
    pages = _format_bulk_preview(parsed)
    msg_obj = update.message or (update.callback_query.message if update.callback_query else None)
    if not msg_obj:
        return
    for i, page in enumerate(pages):
        try:
            await msg_obj.reply_text(
                page,
                parse_mode="Markdown",
                reply_markup=keyboard if i == len(pages) - 1 else None,
            )
        except BadRequest:
            await msg_obj.reply_text(
                page,
                reply_markup=keyboard if i == len(pages) - 1 else None,
            )


# ── Profile-confirmation callback handler ────────────────────────────────────

@auth
async def bulk_profile_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Handle inline-keyboard callbacks during BULK_PROFILE_CONFIRM /
    BULK_PROFILE_FIX_COL / BULK_PROFILE_FIX_FIELD states.
    """
    query = update.callback_query
    await query.answer()
    data = query.data or ""

    if data == "profile_cancel":
        ctx.user_data.pop("_stmt_proposal", None)
        ctx.user_data.pop("_stmt_file_bytes", None)
        ctx.user_data.pop("_stmt_filename", None)
        ctx.user_data.pop("_stmt_headers", None)
        await query.edit_message_text("Statement import cancelled.")
        return ConversationHandler.END

    if data == "profile_ok":
        proposal = ctx.user_data.get("_stmt_proposal") or {}
        errors = sp.validate_profile_mapping(proposal)
        if errors:
            error_lines = "\n".join(f"  ❌ {e}" for e in errors)
            await query.edit_message_text(
                _format_profile_confirm_message(proposal)
                + f"\n\n⚠️ Cannot save yet:\n{error_lines}\n\nPlease fix the mapping first.",
                reply_markup=_profile_confirm_keyboard(),
            )
            return BULK_PROFILE_CONFIRM
        await query.edit_message_text(
            _format_profile_confirm_message(proposal) + "\n\nName this format? (e.g. MyBankA)"
        )
        return BULK_PROFILE_NAME

    if data == "profile_fix":
        headers = ctx.user_data.get("_stmt_headers") or []
        await query.edit_message_text(
            "Which column do you want to re-assign?",
            reply_markup=_column_pick_keyboard(headers),
        )
        return BULK_PROFILE_FIX_COL

    if data.startswith("fix_col:"):
        col = data[len("fix_col:"):]
        ctx.user_data["_stmt_fix_col"] = col
        await query.edit_message_text(
            f"Assign column '{col}' to which field?",
            reply_markup=_field_pick_keyboard(),
        )
        return BULK_PROFILE_FIX_FIELD

    if data.startswith("fix_field:"):
        field = data[len("fix_field:"):]
        col = ctx.user_data.get("_stmt_fix_col") or ""
        proposal = ctx.user_data.get("_stmt_proposal") or {}
        col_map = proposal.setdefault("column_map", {})
        # Clear old mapping for this column across all fields.
        for k in list(col_map.keys()):
            if col_map[k] == col:
                col_map[k] = None
        if field != "skip":
            col_map[field] = col
            # Enforce mutual exclusivity: amount vs debit/credit.
            if field == "amount":
                col_map.pop("debit", None)
                col_map.pop("credit", None)
                if proposal.get("sign_convention") == sp.SIGN_DEBIT_CREDIT_SPLIT:
                    proposal["sign_convention"] = "negative_expense"
            elif field in ("debit", "credit"):
                col_map.pop("amount", None)
                proposal["sign_convention"] = sp.SIGN_DEBIT_CREDIT_SPLIT
        ctx.user_data["_stmt_proposal"] = proposal
        await query.edit_message_text(
            _format_profile_confirm_message(proposal),
            reply_markup=_profile_confirm_keyboard(),
        )
        return BULK_PROFILE_CONFIRM

    # Unrecognised callback — ignore.
    return BULK_PROFILE_CONFIRM


@auth
async def bulk_profile_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Receive the profile name from the user, save profile, parse statement."""
    name = (update.message.text or "").strip()
    if not name:
        await update.message.reply_text("Please enter a name for this format.")
        return BULK_PROFILE_NAME

    proposal = ctx.user_data.get("_stmt_proposal") or {}
    file_bytes = ctx.user_data.get("_stmt_file_bytes") or b""
    filename = ctx.user_data.get("_stmt_filename") or "file.csv"

    # Final validation — catches edge cases (e.g. user edited mapping then named).
    errors = sp.validate_profile_mapping(proposal)
    if errors:
        error_lines = "\n".join(f"  ❌ {e}" for e in errors)
        await update.message.reply_text(
            f"⚠️ Cannot save — mapping is incomplete:\n{error_lines}"
        )
        await update.message.reply_text(
            _format_profile_confirm_message(proposal),
            reply_markup=_profile_confirm_keyboard(),
        )
        return BULK_PROFILE_CONFIRM

    profile = {**proposal, "name": name}
    # Fingerprint must always be present so load_profiles() can key on it.
    if not profile.get("fingerprint"):
        profile["fingerprint"] = ctx.user_data.get("_stmt_headers") or []
    # Determine delimiter for CSV: carry from provisional profile or sniff again.
    if not profile.get("delimiter"):
        ext = Path(filename).suffix.lower()
        if ext not in {".xlsx", ".xls"}:
            try:
                content = file_bytes.decode("utf-8", errors="replace")
            except Exception:
                content = file_bytes.decode("latin-1", errors="replace")
            profile["delimiter"] = sp.sniff_txt_delimiter(content, candidates=(";", ",", "\t")) or ","

    sp.save_profile(profile, settings.STATEMENT_PROFILES_DIR)
    log.info("User %s saved statement profile '%s'", update.effective_user.id, name)
    await update.message.reply_text(f"✅ Profile '{name}' saved — using it now.")

    # Clean up temp state.
    for key in ("_stmt_proposal", "_stmt_file_bytes", "_stmt_filename", "_stmt_headers", "_stmt_fix_col"):
        ctx.user_data.pop(key, None)

    return await _finish_profile_parse(update, ctx, file_bytes, filename, profile, name)



async def _announce_parse_plan(update: Update, text: str) -> None:
    """Tell the user up front when a large input will be parsed in chunks."""
    n_chunks = len(_chunk_statement_text(text))
    if n_chunks > 1:
        await update.message.reply_text(
            f"📄 That's a big statement — I'll parse it in {n_chunks} parts. "
            f"This can take a minute or two, hang on..."
        )
    else:
        await update.message.reply_text("🔍 Parsing transactions...")

def _bulk_draft_dir() -> Path:
    return settings.BULK_DRAFTS_DIR


def _user_draft_path(user_id: int) -> Path:
    return _bulk_draft_dir() / f"{user_id}.json"


def _load_bulk_drafts() -> dict[str, list[dict]]:
    draft_dir = _bulk_draft_dir()
    draft_dir.mkdir(parents=True, exist_ok=True)
    drafts: dict[str, list[dict]] = {}
    for path in draft_dir.glob("*.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, list):
                drafts[path.stem] = data
        except Exception:
            log.exception("Could not read bulk draft from %s", path)
    return drafts


def _load_user_draft(user_id: int) -> list[dict]:
    """Read one user's draft directly — avoids scanning every user's file."""
    path = _user_draft_path(user_id)
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        log.exception("Could not read bulk draft from %s", path)
        return []


def _save_bulk_draft(user_id: int, rows: list[dict]) -> None:
    path = _user_draft_path(user_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")


def _delete_bulk_draft(user_id: int) -> None:
    path = _user_draft_path(user_id)
    if path.exists():
        path.unlink(missing_ok=True)


def _sort_bulk_rows(parsed: list[dict]) -> list[dict]:
    def _date_key(item: dict) -> tuple[int, date]:
        value = str(item.get("date", "")).strip()
        try:
            return (0, date.fromisoformat(value))
        except Exception:
            return (1, date.max)

    return sorted([dict(item) for item in parsed], key=_date_key)


def _row_dedup_key(row: dict) -> str:
    """Strict dedup key for one draft row — same recipe as MasterData keys."""
    return make_dedup_key(
        row.get("date"), row.get("value"), row.get("currency", "PLN"), row.get("description"),
    )


def _row_loose_dedup_key(row: dict) -> str:
    """Loose dedup key (date|value|currency, no description) — advisory only."""
    return make_loose_dedup_key(row.get("date"), row.get("value"), row.get("currency", "PLN"))


def _fmt_row_numbers(nums: list[int]) -> str:
    """Render a list of 1-based row numbers as compact ranges: [4,5,6] -> '4-6'."""
    if not nums:
        return ""
    nums = sorted(nums)
    ranges = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
            continue
        ranges.append(f"{start}-{prev}" if prev != start else str(start))
        start = prev = n
    ranges.append(f"{start}-{prev}" if prev != start else str(start))
    return ", ".join(ranges)


def _fmt_short_date(date_iso: str) -> str:
    """'2024-05-12' -> '12 May' — matches the acceptance-criteria wording."""
    try:
        return date.fromisoformat(str(date_iso)[:10]).strftime("%-d %b")
    except Exception:
        try:
            # Windows strftime has no %-d.
            return date.fromisoformat(str(date_iso)[:10]).strftime("%#d %b")
        except Exception:
            return str(date_iso)


def _flag_master_duplicates(rows: list[dict]) -> dict:
    """
    Two-pass, count-aware scan of the draft against MasterData (dedup v2):

    Pass 1 (strict key: date|value|currency|cleaned-description) drives all
    automatic skip/keep behaviour and is COUNT-AWARE (multiset, not set) —
    if the batch has 3 identical rows and MasterData already has 2 in range,
    only the excess (2) is flagged; the rest (1) is treated as new. Rows
    flagged this way get row['dup'] = True unless overridden with `keep N`
    (row['dup_keep']). Within-batch identical rows with NO MasterData match
    are kept by default and annotated (row['identical_group']).

    Pass 2 (loose key: date|value|currency, no description) runs only on
    rows pass 1 left as new. It NEVER auto-skips — a match only sets
    row['loose_dup'] = True plus the matched entry's date/description, shown
    as an advisory in the preview. Reuses the same MasterData read as pass 1.

    Returns a summary dict used to build the user-facing messages:
        {
          "flagged": int,                 # rows skipped (dup, not dup_keep)
          "skip_groups": [ {...}, ... ],   # count-aware skip groups, g > 1
          "single_skips": [ {...}, ... ],  # count-aware skip groups, g == 1
          "identical_groups": [ [1-based row numbers], ... ],
          "loose_matches": [ {...}, ... ],
        }
    """
    summary = {
        "flagged": 0, "skip_groups": [], "single_skips": [],
        "identical_groups": [], "loose_matches": [],
    }
    if not rows:
        return summary

    for row in rows:
        row.pop("dup", None)
        row.pop("dup_evidence_date", None)
        row.pop("identical_group", None)
        row.pop("loose_dup", None)
        row.pop("loose_other_date", None)
        row.pop("loose_other_desc", None)

    row_dates = []
    for row in rows:
        try:
            row_dates.append(date.fromisoformat(str(row.get("date", "")).strip()[:10]))
        except ValueError:
            pass
    evidence = load_dedup_evidence(
        min(row_dates) if row_dates else None,
        max(row_dates) if row_dates else None,
    )
    strict_evidence = evidence.get("strict", {})
    loose_evidence = evidence.get("loose", {})

    # ── pass 1: strict, count-aware ──────────────────────────────────────────
    groups: dict[str, list[int]] = {}
    for idx, row in enumerate(rows):
        if row.get("dup_keep"):
            continue
        groups.setdefault(_row_dedup_key(row), []).append(idx)

    for key, idxs in groups.items():
        g = len(idxs)
        m = len(strict_evidence.get(key, []))
        skip_n = min(g, m)
        if skip_n <= 0:
            if g > 1:
                nums = [i + 1 for i in idxs]
                for i in idxs:
                    rows[i]["identical_group"] = nums
                summary["identical_groups"].append(nums)
            continue
        flagged_idxs = idxs[:skip_n]
        example_date = strict_evidence[key][0][0]
        for i in flagged_idxs:
            rows[i]["dup"] = True
            rows[i]["dup_evidence_date"] = example_date
        summary["flagged"] += skip_n
        entry = {
            "group_size": g, "master_count": m, "skip_n": skip_n,
            "flagged_rows": [i + 1 for i in flagged_idxs],
            "kept_rows": [i + 1 for i in idxs[skip_n:]],
            "example_date": example_date,
        }
        (summary["single_skips"] if g == 1 else summary["skip_groups"]).append(entry)

    # ── pass 2: loose, advisory only ─────────────────────────────────────────
    for idx, row in enumerate(rows):
        if row.get("dup") or row.get("dup_keep"):
            continue
        lkey = _row_loose_dedup_key(row)
        loose_hits = loose_evidence.get(lkey)
        if not loose_hits:
            continue
        other_date, other_desc = loose_hits[0]
        row["loose_dup"] = True
        row["loose_other_date"] = other_date
        row["loose_other_desc"] = other_desc
        summary["loose_matches"].append({
            "row": idx + 1,
            "value": row.get("value"), "currency": row.get("currency", "PLN"),
            "description": row.get("description", ""), "date": row.get("date", ""),
            "other_date": other_date, "other_desc": other_desc,
        })

    return summary


def _format_dedup_messages(summary: dict) -> list[str]:
    """
    Turn a _flag_master_duplicates() summary into the user-facing messages
    (BACKLOG "dedup v2 — agreed design"). Every automatic decision is
    reported with its reasoning and a one-command override — nothing this
    feature does is silent.
    """
    messages: list[str] = []

    for entry in summary["single_skips"]:
        n = entry["flagged_rows"][0]
        messages.append(
            f"↺ row {n}: matches an entry saved {_fmt_short_date(entry['example_date'])}. "
            f"Skipping. Reply `keep {n}` if this is a separate payment."
        )

    for entry in summary["skip_groups"]:
        g, m, skip_n = entry["group_size"], entry["master_count"], entry["skip_n"]
        saved = g - skip_n
        messages.append(
            f"↺ {g} identical rows found (rows {_fmt_row_numbers(entry['flagged_rows'] + entry['kept_rows'])}), "
            f"{m} already in your sheet → saving {saved}, skipping {skip_n} "
            f"(row(s) {_fmt_row_numbers(entry['flagged_rows'])}). "
            f"Reply `keep {_fmt_row_numbers(entry['flagged_rows'])}` or `keep all flagged` "
            f"if these are new payments."
        )

    for nums in summary["identical_groups"]:
        messages.append(
            f"rows {_fmt_row_numbers(nums)} are identical — keeping all {len(nums)}; "
            f"reply `drop N` if one is a scan error."
        )

    loose = summary["loose_matches"]
    if loose:
        lines = ["⚠️ Possible duplicates (matched on date+amount, but description differs):"]
        for m in loose[:10]:
            lines.append(
                f"row {m['row']}: {m['value']} {m['currency']} '{m['description']}' "
                f"({_fmt_short_date(m['date'])}) ↔ existing: {m['value']} {m['currency']} "
                f"'{m['other_desc']}' ({_fmt_short_date(m['other_date'])})."
            )
        more = f"\n… and {len(loose) - 10} more" if len(loose) > 10 else ""
        row_nums = [m["row"] for m in loose]
        lines.append(
            f"Saving them. Reply `drop {_fmt_row_numbers(row_nums)}` "
            f"if any of these are the same payment.{more}"
        )
        # Mass loose-match hint — bank likely reformatted descriptions between
        # exports; offer the one-command bulk override.
        total_new = sum(
            len(g) for g in summary["identical_groups"]
        ) + len(loose) + sum(e["skip_n"] for e in summary["skip_groups"] + summary["single_skips"])
        if len(loose) >= 3 and total_new and len(loose) >= total_new // 2:
            lines.append(
                "Most rows in this batch loose-matched an existing entry (likely a "
                "reformatted export). Reply `drop all flagged` to drop them all at once."
            )
        messages.append("\n".join(lines))

    return messages


def _merge_bulk_draft(user_id: int, parsed: list[dict]) -> tuple[list[dict], int]:
    """
    Merge new rows into the pending draft. Within-batch/within-draft repeats
    are KEPT by default (dedup v2 inverts PR #7's hard skip here — repetition
    inside one source is almost always real, e.g. several 2 PLN car-wash
    payments on the same day, or the same photo re-sent). They are annotated
    as an identical group in the preview instead (see _flag_master_duplicates),
    and MasterData-level dedup still applies once they're actually saved.
    Returns (merged draft, 0) — the second element is kept for API
    compatibility with callers that used to report a skip count.
    """
    previous = _load_user_draft(user_id)
    if not isinstance(previous, list):
        previous = []
    normalized_previous = []
    for item in previous:
        row = dict(item)
        row.setdefault("status", "pending")
        normalized_previous.append(row)
    fresh = [dict(item) for item in parsed]
    merged = _sort_bulk_rows(normalized_previous + fresh)
    for item in merged:
        item.setdefault("status", "pending")
    _save_bulk_draft(user_id, merged)
    log.info("User %s bulk draft updated: %d pending entries", user_id, len(merged))
    return merged, 0


# Maximum rows a pending draft may hold before new imports are refused.
_DRAFT_LIMIT_ENTRIES = 50


def _normalize_parsed_rows(parsed: list[dict], lists: dict) -> tuple[list[dict], list[str]]:
    """
    Enforce Lists reference data on AI output — the model drifts no matter what
    the prompt says (invents 'Shopping' next to 'Gifts & Shopping', puts
    transfer recipients into person). Auto-correct what's unambiguous, report
    every correction so the user sees them in the preview.
    """
    categories = lists.get("categories") or []
    cat_by_lower = {c.lower(): c for c in categories}
    persons = {str(p).strip() for p in (lists.get("persons") or [])}
    types = set(lists.get("txn_types") or ["Expense", "Income", "Savings"])
    corrections: list[str] = []

    for i, row in enumerate(parsed, 1):
        # Description: strip statement junk (masked PANs, BPID:, /OPT/ blocks,
        # country suffixes) so MasterData, dedup and merchant memory all see
        # the same clean merchant label.
        desc_raw = str(row.get("description") or "").strip()
        cleaned = clean_merchant_description(desc_raw)
        if desc_raw and cleaned != desc_raw:
            row["description"] = cleaned
            corrections.append(f"row {i}: description cleaned → '{cleaned}'")

        # Category: exact -> case-insensitive -> substring fuzzy -> Other
        cat = str(row.get("category") or "").strip()
        if cat and cat not in categories:
            fixed = cat_by_lower.get(cat.lower())
            if not fixed:
                fuzzy = [c for c in categories
                         if cat.lower() in c.lower() or c.lower() in cat.lower()]
                fixed = fuzzy[0] if len(fuzzy) == 1 else None
            row["category"] = fixed or "Other"
            corrections.append(f"row {i}: category '{cat}' → '{row['category']}'")

        # Person: must be a known household member — recipients go to description
        per = str(row.get("person") or "").strip()
        if per and per not in persons:
            desc = str(row.get("description") or "").strip()
            row["description"] = (f"{desc} — {per}" if desc else per)[:120]
            row["person"] = ""
            corrections.append(f"row {i}: person '{per}' moved to description")

        # Type: whitelist, default Expense
        typ = str(row.get("type") or "").strip()
        if typ and typ not in types:
            ci = next((t for t in types if t.lower() == typ.lower()), None)
            row["type"] = ci or "Expense"
            corrections.append(f"row {i}: type '{typ}' → '{row['type']}'")

    return parsed, corrections


def _apply_merchant_memory(parsed: list[dict]) -> list[str]:
    """
    Override AI categorization with remembered merchant defaults — the map is
    deterministic (seeded from history, corrected by the user), the model is
    not. Rows touched get row['mem'] = True (🧠 in the preview) and every
    override is reported so nothing changes silently.
    """
    mapping = merchant_map.load_merchant_map()
    if not mapping:
        return []
    notes: list[str] = []
    for i, row in enumerate(parsed, 1):
        entry = merchant_map.lookup(mapping, row.get("description"))
        if not entry:
            continue
        changed = []
        for field in ("category", "type"):
            remembered = str(entry.get(field) or "").strip()
            if remembered and remembered != str(row.get(field) or "").strip():
                changed.append(f"{field} '{row.get(field) or ''}' → '{remembered}'")
                row[field] = remembered
        if entry.get("person") and not str(row.get("person") or "").strip():
            row["person"] = entry["person"]
            changed.append(f"person → '{entry['person']}'")
        if entry.get("is_recurring") and not row.get("is_recurring"):
            row["is_recurring"] = True
            changed.append("is_recurring → yes")
        if changed:
            row["mem"] = True
            notes.append(f"row {i}: {', '.join(changed)} (merchant memory)")
    return notes


def _revalidate_bulk_row(row: dict, lists: dict, row_no: int) -> list[str]:
    """
    Run the shared validator on one draft row: normalize what's unambiguous,
    flag what isn't (row['invalid'] = reason, shown in the preview and
    excluded from save). Returns correction notes for the user.
    """
    notes: list[str] = []
    if not str(row.get("type") or "").strip():
        row["type"] = "Expense"
    if not str(row.get("category") or "").strip() and lists.get("categories"):
        row["category"] = "Other"
        notes.append(f"row {row_no}: empty category → 'Other'")

    ok, reason, normalized, corrections = validate_parsed_row(row, lists)
    if ok:
        row.pop("invalid", None)
        for f in ("value", "type", "category", "currency", "person"):
            row[f] = normalized[f]
        notes.extend(f"row {row_no}: {c}" for c in corrections)
    else:
        row["invalid"] = reason
    return notes


def _validate_bulk_rows(parsed: list[dict], lists: dict) -> tuple[list[dict], list[str]]:
    """Validate every draft row with the shared validator (all entry paths)."""
    corrections: list[str] = []
    for i, row in enumerate(parsed, 1):
        corrections.extend(_revalidate_bulk_row(row, lists, i))
    return parsed, corrections


def _draft_limit_reached(user_id: int) -> bool:
    previous = _load_user_draft(user_id)
    if not isinstance(previous, list):
        return False
    return len(previous) > _DRAFT_LIMIT_ENTRIES


# Telegram hard limit is 4096 chars; leave headroom for the header/footer.
_PREVIEW_MSG_LIMIT = 3500


def _md_escape(text: str) -> str:
    """Escape Markdown control chars in untrusted text (bank descriptions etc.)."""
    return re.sub(r"([_*`\[\]])", r"\\\1", str(text))


def _bulk_footer(parsed: list[dict]) -> str:
    """
    Contextual command footer — content adapts to what's actually flagged in
    THIS draft, so a user with no duplicates never reads a word about
    duplicates (BACKLOG "Contextual command footer — show only what applies").
    """
    lines = [
        "✏️ Reply with edits like `2 category=Transport`, `drop 3`, `keep 3`, `drop 4-6`.\n"
        "Send `save` to store them all, or `cancel` to stop."
    ]

    dup_rows = [i for i, t in enumerate(parsed, 1) if t.get("dup") and not t.get("dup_keep")]
    if dup_rows:
        example = _fmt_row_numbers(dup_rows)
        lines.append(
            f"↺ {len(dup_rows)} row(s) skipped as already imported — "
            f"`keep {dup_rows[0]}`, `keep {example}`, or `keep all flagged` to save them anyway."
        )

    loose_rows = [i for i, t in enumerate(parsed, 1) if t.get("loose_dup")]
    if loose_rows:
        lines.append(
            f"⚠️ {len(loose_rows)} possible duplicate(s) flagged above (saved by default) — "
            f"`drop {loose_rows[0]}` or `drop all flagged` to remove them."
        )

    return "\n".join(lines)


def _format_bulk_preview(parsed: list[dict]) -> list[str]:
    """
    Render the draft preview as a LIST of messages, each under Telegram's
    length limit. Row numbers are stable across all pages.
    """
    footer = "\n" + _bulk_footer(parsed)
    row_lines = []
    for i, t in enumerate(parsed, 1):
        person = t.get("person") or ""
        txn_type = t.get("type") or ""
        person_suffix = f" | person={_md_escape(person)}" if person else ""
        type_suffix = f" | type={_md_escape(txn_type)}" if txn_type else ""
        mem_suffix = " 🧠" if t.get("mem") else ""
        invalid = t.get("invalid") or ""
        invalid_suffix = f"\n   ⚠️ {_md_escape(invalid)} (won't be saved — edit it first)" if invalid else ""
        dropped_suffix = (
            f"\n   ❌ dropped — reply `keep {i}` to restore" if t.get("dropped") else ""
        )
        dup_suffix = (
            f"\n   ↺ matches an entry saved {_md_escape(_fmt_short_date(t.get('dup_evidence_date', '')))} "
            f"— already imported (skipped — reply `keep {i}` to save anyway)"
            if t.get("dup") and not t.get("dup_keep") else ""
        )
        loose_suffix = (
            f"\n   ⚠️ possible duplicate: matches {_md_escape(_fmt_short_date(t.get('loose_other_date', '')))} "
            f"'{_md_escape(t.get('loose_other_desc', ''))}' but description differs "
            f"— saving (reply `drop {i}` if it's the same payment)"
            if t.get("loose_dup") else ""
        )
        identical = t.get("identical_group") or []
        identical_suffix = (
            f"\n   ↔ rows {_fmt_row_numbers(identical)} are identical — keeping all "
            f"{len(identical)} (reply `drop {i}` if this one's a scan error)"
            if identical else ""
        )
        row_lines.append(
            f"{i}. {t.get('date', '')} | {t.get('value', '')} {t.get('currency', 'PLN')} | "
            f"{_md_escape(t.get('category', ''))} | {_md_escape(t.get('description', ''))}"
            f"{mem_suffix}{type_suffix}{person_suffix}{invalid_suffix}"
            f"{dup_suffix}{loose_suffix}{identical_suffix}{dropped_suffix}"
        )

    messages = []
    current = [f"Found *{len(parsed)}* transaction(s):\n"]
    current_len = len(current[0])
    for line in row_lines:
        if current_len + len(line) + 1 > _PREVIEW_MSG_LIMIT:
            messages.append("\n".join(current))
            current = []
            current_len = 0
        current.append(line)
        current_len += len(line) + 1
    if current:
        messages.append("\n".join(current))

    messages[-1] += "\n" + footer
    return messages


async def _send_bulk_preview(update: Update, parsed: list[dict]) -> None:
    """Send the preview as one or more messages; keyboard only on the last."""
    keyboard = ReplyKeyboardMarkup([["Save", "Cancel"]], one_time_keyboard=True, resize_keyboard=True)
    pages = _format_bulk_preview(parsed)
    for i, page in enumerate(pages):
        try:
            await update.message.reply_text(
                page,
                parse_mode="Markdown",
                reply_markup=keyboard if i == len(pages) - 1 else None,
            )
        except BadRequest:
            # Malformed entity despite escaping — resend as plain text.
            await update.message.reply_text(
                page,
                reply_markup=keyboard if i == len(pages) - 1 else None,
            )


_RANGE_TOKEN_RE = re.compile(r"^(\d+)(?:-(\d+))?$")
_ROW_COMMAND_RE = re.compile(r"^(drop|keep)\s+(.+)$", re.IGNORECASE)


def _parse_row_targets(args: str, n: int) -> list[int] | None:
    """
    Parse space-separated row targets ('4', '4 6', '4-6 9 12') into a sorted
    list of 0-based indices, bounds-checked against n rows. Returns None if
    nothing valid parses (caller reports "invalid").
    """
    idxs: set[int] = set()
    for tok in args.split():
        m = _RANGE_TOKEN_RE.match(tok)
        if not m:
            return None
        a = int(m.group(1))
        b = int(m.group(2)) if m.group(2) else a
        if a > b:
            a, b = b, a
        for v in range(a, b + 1):
            if 1 <= v <= n:
                idxs.add(v - 1)
    return sorted(idxs) if idxs else None


def _apply_row_command(text: str, parsed: list[dict]) -> tuple[bool, str, list[str]] | None:
    """
    Unified `drop` / `keep` row-command grammar (BACKLOG dedup v2): targets
    `N`, `N M`, `N-M`, `all`, `all flagged`. One parser for every preview
    state (dedup skip flags, loose-match advisory, manual pruning).

    `all flagged` SCOPES to the block it answers: `keep all flagged` acts
    only on rows skipped as MasterData duplicates (row['dup']); `drop all
    flagged` acts only on the loose-match advisory rows (row['loose_dup']).
    Plain `drop all` / `keep all` act on the whole batch.

    Returns None if `text` isn't a drop/keep command (caller falls through to
    the single-row `N field=value` grammar). Otherwise returns the same
    (save, reason, notes) contract as _apply_bulk_edit.
    """
    m = _ROW_COMMAND_RE.match(text.strip())
    if not m:
        return None
    verb = m.group(1).lower()
    args = m.group(2).strip().lower()
    n = len(parsed)

    if args == "all flagged":
        if verb == "keep":
            idxs = [i for i, r in enumerate(parsed) if r.get("dup") and not r.get("dup_keep")]
        else:
            idxs = [i for i, r in enumerate(parsed) if r.get("loose_dup")]
        if not idxs:
            return False, "invalid", []
    elif args == "all":
        idxs = list(range(n))
    else:
        idxs = _parse_row_targets(args, n)
        if idxs is None:
            return False, "invalid", []

    notes: list[str] = []
    for i in idxs:
        row = parsed[i]
        if verb == "keep":
            if row.get("dup") and not row.get("dup_keep"):
                row["dup_keep"] = True
                row.pop("dup", None)
                notes.append(f"row {i + 1}: will be saved even though it looked already imported")
            elif row.get("dropped"):
                row.pop("dropped", None)
                notes.append(f"row {i + 1}: restored")
        else:
            if not row.get("dropped"):
                row["dropped"] = True
                notes.append(f"row {i + 1}: dropped — won't be saved")

    return False, "edited", notes


def _apply_bulk_edit(
    message: str, parsed: list[dict], lists: dict | None = None
) -> tuple[bool, str, list[str]]:
    """Returns (save, reason, correction_notes) — notes are 🛡-reported to the user."""
    text = message.strip()
    if not text:
        return False, "", []

    # Accept both "save" and "/save" — users naturally type slash commands.
    normalized = text.lower().lstrip("/")
    if normalized in {"save", "yes"}:
        return True, "", []
    if normalized in {"cancel", "no"}:
        return False, "cancel", []

    row_command = _apply_row_command(text, parsed)
    if row_command is not None:
        return row_command

    # Backward-compatible single-row "N keep" (pre-dedup-v2 grammar) — same
    # effect as the new "keep N".
    legacy_keep_match = re.match(r"^(\d+)\s+keep$", normalized)
    if legacy_keep_match:
        idx = int(legacy_keep_match.group(1)) - 1
        if not (0 <= idx < len(parsed)):
            return False, "invalid", []
        if not parsed[idx].get("dup"):
            return False, "invalid", []
        parsed[idx]["dup_keep"] = True
        parsed[idx].pop("dup", None)
        return False, "edited", [f"row {idx + 1}: will be saved even though it looks already imported"]

    match = re.match(r"^(\d+)\s+(\w+)=(.+)$", text)
    if not match:
        return False, "invalid", []

    idx = int(match.group(1)) - 1
    field = match.group(2).strip().lower()
    value = match.group(3).strip()
    if not (0 <= idx < len(parsed)):
        return False, "invalid", []
    if field not in {"date", "value", "currency", "type", "category", "description",
                     "person", "is_recurring"}:
        return False, "invalid", []

    notes: list[str] = []
    if field == "value":
        try:
            value = parse_amount(value)
        except ValueError:
            return False, "invalid", []
        if value < 0:
            # Signed bank-export amount: negative means money out.
            parsed[idx]["type"] = "Expense"
            value = abs(value)
            notes.append(f"row {idx + 1}: negative amount → type 'Expense'")
    elif field == "is_recurring":
        try:
            value = coerce_bool(value)
        except ValueError:
            return False, "invalid", []

    parsed[idx][field] = value
    # Manual edits go through the same normalizer/validator as AI output —
    # a typo'd category must not slip past just because it was typed by hand.
    if lists:
        notes.extend(_revalidate_bulk_row(parsed[idx], lists, idx + 1))
    # A human correction to categorization is the strongest signal there is —
    # remember it so future imports of this merchant skip the AI's guess.
    if field in {"category", "type", "person", "is_recurring"} and not parsed[idx].get("invalid"):
        learned = merchant_map.learn_from_row(parsed[idx])
        if learned:
            notes.append(f"row {idx + 1}: remembered '{learned}' → "
                         f"{parsed[idx].get('category')} for future imports")
    return False, "edited", notes


async def bulk_timeout(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Conversation expired while the user was reviewing — tell them how to resume."""
    if update.effective_message:
        await update.effective_message.reply_text(
            "⏰ Bulk import session expired, but your draft is safe. "
            "Run /bulk and send any text to see it again, then `save` or `cancel`.",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardRemove(),
        )
    return ConversationHandler.END


@auth_write
async def cmd_bulk(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    # Handle /bulk help — show focused usage text without starting the flow.
    args = ctx.args or []
    if args and args[0].lower() == "help":
        await update.message.reply_text(
            "📎 */bulk* — Bulk import\n\n"
            "Upload a CSV, XLSX, TXT file or a photo\\. The AI maps columns and shows you the parsed rows\\.\n"
            "To fix a row, reply: `2 category=Transport`\\. Reply `save` to import or `cancel` to discard\\.\n"
            "Bank statement profiles from previous uploads are matched automatically\\.",
            parse_mode="MarkdownV2",
        )
        return ConversationHandler.END

    # Handle /bulk profile [list|delete <name>] — owner-only profile management.
    if args and args[0].lower() == "profile":
        subcommand = args[1].lower() if len(args) > 1 else "list"
        if subcommand == "list":
            return await _cmd_bulk_profile_list(update, ctx)
        elif subcommand == "delete":
            name = " ".join(args[2:]).strip()
            if not name:
                await update.message.reply_text(
                    "Usage: /bulk profile delete <name>\n"
                    "Use /bulk profile list to see all saved profiles."
                )
                return ConversationHandler.END
            deleted = await asyncio.to_thread(sp.delete_profile, name, settings.STATEMENT_PROFILES_DIR)
            if deleted:
                await update.message.reply_text(f"✅ Profile '{name}' deleted.")
            else:
                await update.message.reply_text(
                    f"No profile named '{name}' found. "
                    f"Use /bulk profile list to see all saved profiles."
                )
            return ConversationHandler.END
        else:
            await update.message.reply_text(
                "Usage:\n"
                "  /bulk profile list — list all saved profiles\n"
                "  /bulk profile delete <name> — delete a profile"
            )
            return ConversationHandler.END

    # Resume an unfinished draft directly — no need to re-upload anything.
    draft = _load_user_draft(update.effective_user.id)
    if isinstance(draft, list) and draft:
        summary = _flag_master_duplicates(draft)
        _save_bulk_draft(update.effective_user.id, draft)
        ctx.user_data["bulk_parsed"] = draft
        await update.message.reply_text(
            f"📋 You have an unfinished draft with {len(draft)} transaction(s). "
            f"Review it below — `save` to store, `cancel` to discard, "
            f"or edit rows first."
        )
        for msg in _format_dedup_messages(summary):
            await update.message.reply_text(msg)
        await _send_bulk_preview(update, draft)
        return BULK_CONFIRM

    await update.message.reply_text(
        "📎 Send me a photo, document, or paste your transaction text.\n"
        "I'll extract and preview all transactions before saving.",
        reply_markup=ReplyKeyboardRemove(),
    )
    return BULK_RECEIVE


@auth
async def bulk_receive(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    parsed = []
    lists  = load_reference_data()
    try:
        uid = update.effective_user.id
        src = 'unknown'
        loop = asyncio.get_running_loop()
        if update.message.photo:
            src = 'photo'
            photo       = update.message.photo[-1]
            file        = await photo.get_file()
            image_bytes = bytes(await file.download_as_bytearray())
            await update.message.reply_text("🔍 Analysing image...")
            parsed = await loop.run_in_executor(None, lambda: parse_image(image_bytes, lists))

        elif update.message.document:
            src = 'document'
            document = update.message.document
            filename = document.file_name or "file.txt"
            file = await document.get_file()
            file_bytes = bytes(await file.download_as_bytearray())

            # ── Statement-profile fast path ──────────────────────────────────
            if _is_statement_file(filename):
                headers, provisional = _read_statement_headers_and_sniff(file_bytes, filename)
                if headers:
                    # Try exact profile match.
                    profiles = _load_profiles()
                    matched = sp.match_profile(headers, profiles)
                    if matched:
                        profile_name = matched.get("name") or filename
                        await update.message.reply_text(f"📄 Matched profile '{profile_name}', parsing…")
                        ctx.user_data["lists"] = lists
                        return await _finish_profile_parse(
                            update, ctx, file_bytes, filename, matched, profile_name
                        )
                    # No match — propose via AI (one call).
                    await update.message.reply_text("🔍 New format — asking AI to map columns…")
                    sample_rows = _get_sample_rows(
                        file_bytes, filename,
                        delimiter=(provisional or {}).get("delimiter") or ","
                    )
                    proposal = await loop.run_in_executor(
                        None,
                        lambda: sp.propose_mapping(headers, sample_rows, ai_parser.get_provider()),
                    )
                    if proposal:
                        # Merge provisional settings (delimiter etc.) into the AI proposal.
                        if provisional:
                            for k in ("delimiter", "encoding", "header_row", "fingerprint"):
                                if k not in proposal:
                                    proposal[k] = provisional.get(k)
                        ctx.user_data["_stmt_proposal"] = proposal
                        ctx.user_data["_stmt_file_bytes"] = file_bytes
                        ctx.user_data["_stmt_filename"] = filename
                        ctx.user_data["_stmt_headers"] = headers
                        ctx.user_data["lists"] = lists
                        await update.message.reply_text(
                            _format_profile_confirm_message(proposal),
                            reply_markup=_profile_confirm_keyboard(),
                        )
                        return BULK_PROFILE_CONFIRM
                    # AI returned nothing — fall through to AI text path.
                    log.info("propose_mapping returned empty for %s; falling back to AI text", filename)

                # .txt with no consistent delimiter or empty headers — fall through.
                ext = Path(filename).suffix.lower()
                if ext in {".xlsx", ".xls"}:
                    await update.message.reply_text(
                        "Could not read headers from this XLSX file. "
                        "Please make sure it is a valid bank export."
                    )
                    return BULK_RECEIVE
                # CSV/TXT fall-through: treat as raw text.
                try:
                    text = file_bytes.decode("utf-8", errors="replace")
                except Exception:
                    text = file_bytes.decode("latin-1", errors="replace")
                await _announce_parse_plan(update, text)
                parsed = await loop.run_in_executor(None, lambda: parse_text(text, lists))
            else:
                # Not a recognised statement extension — existing plain-text path.
                mime_type = (document.mime_type or "").lower()
                if mime_type and not mime_type.startswith("text/") and "plain" not in mime_type:
                    await update.message.reply_text("Please upload a plain text file (.txt).")
                    return BULK_RECEIVE
                text = file_bytes.decode("utf-8", errors="replace")
                await _announce_parse_plan(update, text)
                parsed = await loop.run_in_executor(None, lambda: parse_text(text, lists))

        elif update.message.text:
            src = 'text'
            text = update.message.text.strip()
            if text.startswith("/"):
                await update.message.reply_text("Send the transaction text, not a command.")
                return BULK_RECEIVE
            await _announce_parse_plan(update, text)
            parsed = await loop.run_in_executor(None, lambda: parse_text(text, lists))

        else:
            await update.message.reply_text("Send a photo, document, or text.")
            return BULK_RECEIVE

    except Exception as e:
        log.exception("bulk_receive parse error for user %s (src=%s)", update.effective_user.id, src)
        await update.message.reply_text(f"❌ Parsing failed: {e}")
        return ConversationHandler.END

    if not parsed:
        await update.message.reply_text("No transactions found in the input.")
        return ConversationHandler.END

    ctx.user_data["lists"] = lists
    parsed, corrections = _normalize_parsed_rows(parsed, lists)
    memory_notes = _apply_merchant_memory(parsed)
    parsed, validator_corrections = _validate_bulk_rows(parsed, lists)
    corrections += validator_corrections
    if memory_notes:
        shown = "\n".join(f"  • {n}" for n in memory_notes[:10])
        more = f"\n  … and {len(memory_notes) - 10} more" if len(memory_notes) > 10 else ""
        await update.message.reply_text(
            f"🧠 {len(memory_notes)} row(s) categorized from merchant memory "
            f"(marked 🧠 in the preview):\n{shown}{more}"
        )
        log.info("User %s bulk merchant-memory: %d rows", uid, len(memory_notes))
    if corrections:
        shown = "\n".join(f"  • {c}" for c in corrections[:10])
        more = f"\n  … and {len(corrections) - 10} more" if len(corrections) > 10 else ""
        await update.message.reply_text(
            f"🛡 Auto-corrected {len(corrections)} value(s) the AI got wrong:\n{shown}{more}"
        )
        log.info("User %s bulk normalize: %d corrections", uid, len(corrections))

    if _draft_limit_reached(uid):
        # Do NOT merge the new rows — but never discard them silently either.
        await update.message.reply_text(
            f"⚠️ Your draft is full ({_DRAFT_LIMIT_ENTRIES}+ entries), so the {len(parsed)} "
            f"row(s) I just parsed were NOT added. Send `save` to store the existing draft "
            f"or `cancel` to discard it — then re-send this input.",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup([["Save", "Cancel"]], one_time_keyboard=True, resize_keyboard=True),
        )
        ctx.user_data["bulk_parsed"] = _load_user_draft(uid)
        return BULK_CONFIRM

    parsed = _sort_bulk_rows(parsed)
    pre_merge_len = len(_load_user_draft(uid))
    draft_rows, _ = _merge_bulk_draft(uid, parsed)
    ctx.user_data["bulk_parsed"] = draft_rows
    log.info("User %s bulk parse completed: %d items (src=%s)", update.effective_user.id, len(parsed), src)
    merged_in = pre_merge_len
    if merged_in > 0:
        await update.message.reply_text(
            f"ℹ️ {merged_in} row(s) from a previous draft were merged in. "
            f"The preview below is the full set that `save` will store.",
            parse_mode="Markdown",
        )

    # Two-pass, count-aware scan against MasterData — also picks up within-batch
    # repeats (kept by default) since it runs on the full merged draft.
    summary = _flag_master_duplicates(draft_rows)
    if summary["flagged"] or summary["identical_groups"] or summary["loose_matches"]:
        _save_bulk_draft(uid, draft_rows)
        for msg in _format_dedup_messages(summary):
            await update.message.reply_text(msg)

    # Preview the MERGED draft — exactly what `save` will write.
    await _send_bulk_preview(update, draft_rows)
    return BULK_CONFIRM


@auth
async def bulk_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    parsed = _sort_bulk_rows(ctx.user_data.get("bulk_parsed", []))
    text = update.message.text.strip()
    lists = ctx.user_data.get("lists") or {}
    action, reason, edit_notes = _apply_bulk_edit(text, parsed, lists)

    if reason == "cancel":
        _delete_bulk_draft(update.effective_user.id)
        log.info("User %s cancelled bulk import draft", update.effective_user.id)
        await update.message.reply_text(
            "Cancelled — draft discarded.", reply_markup=ReplyKeyboardRemove()
        )
        return ConversationHandler.END

    if reason == "edited":
        parsed = _sort_bulk_rows(parsed)
        ctx.user_data["bulk_parsed"] = parsed
        # Persist so a timeout/restart doesn't roll back to pre-edit values.
        _save_bulk_draft(update.effective_user.id, parsed)
        log.info("User %s edited bulk draft row via message: %s", update.effective_user.id, text)
        if edit_notes:
            # Same "report every silent correction" pattern as bulk_receive.
            shown = "\n".join(f"  • {n}" for n in edit_notes)
            await update.message.reply_text(f"🛡 Auto-corrected:\n{shown}")
            log.info("User %s bulk edit auto-corrections: %s",
                     update.effective_user.id, "; ".join(edit_notes))
        await _send_bulk_preview(update, parsed)
        return BULK_CONFIRM

    if reason == "invalid":
        await update.message.reply_text(
            "Please reply with edits like `2 category=Transport` or `1 description=Lunch`.",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup([["Save", "Cancel"]], one_time_keyboard=True, resize_keyboard=True),
        )
        return BULK_CONFIRM

    if not action:
        log.info("User %s sent non-action bulk message: %s", update.effective_user.id, text)
        await update.message.reply_text(
            "Please reply with edits, or send `save` to store them all.",
            reply_markup=ReplyKeyboardMarkup([["Save", "Cancel"]], one_time_keyboard=True, resize_keyboard=True),
        )
        return BULK_CONFIRM

    parsed = _sort_bulk_rows(parsed)
    ctx.user_data["bulk_parsed"] = parsed

    # Re-check against MasterData right before writing (the flags set at
    # preview time may be stale) — count-aware, so within-batch repeats that
    # aren't real MasterData duplicates are saved, never silently dropped.
    _flag_master_duplicates(parsed)
    skipped_dups: list[int] = []
    dropped_rows: list[int] = []

    transactions  = []
    errors        = []
    failed_items  = []  # raw rows that never made it into `transactions`, kept for retry

    for i, item in enumerate(parsed, 1):
        if item.get("dup") and not item.get("dup_keep"):
            skipped_dups.append(i)
            continue
        if item.get("dropped"):
            dropped_rows.append(i)
            continue
        try:
            if item.get("invalid"):
                raise ValueError(item["invalid"])
            try:
                is_recurring = coerce_bool(item.get("is_recurring", False))
            except ValueError:
                is_recurring = False
            txn_date = date.fromisoformat(item.get("date", str(datetime.now(timezone.utc).date())))
            transactions.append(Transaction(
                date=txn_date,
                value=float(item["value"]),
                currency=item.get("currency", "PLN").upper(),
                transaction_type=item.get("type", "Expense"),
                category=item.get("category", "Other"),
                person=item.get("person", ""),
                description=item.get("description", ""),
                is_recurring=is_recurring,
            ))
        except Exception as e:
            errors.append(f"Row {i}: {e}")
            failed_items.append(item)

    saved = 0
    write_failed = False
    if transactions:
        try:
            log.info("User %s saving bulk batch: %d transactions", update.effective_user.id, len(transactions))
            await async_append_batch(transactions)
            saved = len(transactions)
            log.info("User %s bulk batch saved: %d transactions", update.effective_user.id, saved)
        except Exception as e:
            log.exception("bulk_confirm batch write failed for user %s", update.effective_user.id)
            errors.append(f"Write failed: {e}")
            write_failed = True

    # Offer a cycle-boundary prompt if any saved transaction is a salary-income entry.
    if not write_failed and transactions:
        salary_txns = [
            t for t in transactions
            if t.transaction_type == "Income"
            and (t.category or "").strip().lower() == settings.SALARY_CATEGORY.strip().lower()
        ]
        if salary_txns:
            latest = max(salary_txns, key=lambda t: t.date)
            try:
                await maybe_prompt_cycle_start(update, latest)
            except Exception:
                pass  # cycle prompt is best-effort

    if write_failed:
        # The whole batch write failed — nothing was saved, so keep every row
        # (including the ones that parsed fine) in the draft for a clean retry.
        _save_bulk_draft(update.effective_user.id, parsed)
    elif failed_items:
        # Partial save: the valid rows are already in the workbook. Keep only
        # the rows that failed to construct so the user can fix/retry them
        # instead of losing them when the whole draft is wiped.
        _save_bulk_draft(update.effective_user.id, failed_items)
    else:
        _delete_bulk_draft(update.effective_user.id)

    if settings.STORAGE_BACKEND == "gcs" or settings.GCS_BUCKET_NAME:
        destination = f"gs://{settings.GCS_BUCKET_NAME}/{settings.GCS_OBJECT_NAME}"
    elif settings.STORAGE_BACKEND == "s3" or settings.S3_BUCKET_NAME:
        destination = f"s3://{settings.S3_BUCKET_NAME}/{settings.S3_OBJECT_NAME}"
    else:
        destination = str(settings.XLSX_PATH)
    msg = (
        f"✅ Saved {saved} of {len(parsed)} transaction(s) "
        f"to the MasterData sheet of:\n{destination}"
    )
    if skipped_dups and not write_failed:
        nums = ", ".join(f"#{n}" for n in skipped_dups[:10])
        more = f", … and {len(skipped_dups) - 10} more" if len(skipped_dups) > 10 else ""
        msg += f"\n\n↺ {len(skipped_dups)} row(s) skipped as already imported: {nums}{more}"
    if dropped_rows and not write_failed:
        nums = ", ".join(f"#{n}" for n in dropped_rows[:10])
        more = f", … and {len(dropped_rows) - 10} more" if len(dropped_rows) > 10 else ""
        msg += f"\n\n❌ {len(dropped_rows)} row(s) dropped as requested: {nums}{more}"
    if errors:
        msg += "\n\n⚠️ Errors:\n" + "\n".join(errors[:5])
    if failed_items and not write_failed:
        msg += (
            f"\n\n{len(failed_items)} row(s) could not be saved and are kept in your draft — "
            "reply with edits (e.g. `1 value=12.50`) and send `save` again to retry them."
        )
    await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END
