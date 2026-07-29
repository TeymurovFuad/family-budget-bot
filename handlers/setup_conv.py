"""
setup_conv.py — /setup onboarding conversation (owner only).

Guides a first-time user from an empty deployment to a working budget:
  1. create the workbook from the template (atomic),
  2. review / rename / add categories,
  3. set per-category monthly budgets,
  4. pick the main currency,
  5. commit everything and fetch live exchange rates.

Also entered via /start when no workbook exists yet.

All workbook writes happen in two atomic checkpoints, each inside a single
ExcelFileContext save:
  A. categories → Lists!C + both dashboards + category dropdown validation
     (when the user taps "Done with categories"),
  B. budgets + currency → Lists!D / Lists!H (at the final summary).
Cancelling mid-flow keeps whatever checkpoints already committed.
"""

import re

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
    CallbackQueryHandler, CommandHandler, ContextTypes, ConversationHandler,
    MessageHandler, filters,
)

from config import auth, auth_write, log, set_display_currency
from log_decorators import log_call
from states import (
    SETUP_WELCOME, SETUP_REVIEW, SETUP_RENAME, SETUP_ADD,
    SETUP_BUDGET, SETUP_CURRENCY, SETUP_SUMMARY, SETUP_REMOVE,
)

# ── Default categories (name, type) seeded into a fresh workbook ──────────────

DEFAULT_CATEGORIES: list[tuple[str, str]] = [
    ("Salary",        "Income"),
    ("Other Income",  "Income"),
    ("Housing",       "Expense"),
    ("Groceries",     "Expense"),
    ("Transport",     "Expense"),
    ("Utilities",     "Expense"),
    ("Health",        "Expense"),
    ("Dining Out",    "Expense"),
    ("Shopping",      "Expense"),
    ("Entertainment", "Expense"),
    ("Subscriptions", "Expense"),
    ("Travel",        "Expense"),
    ("Savings",       "Savings"),
    ("Other",         "Expense"),
]

# Used to re-derive types when /setup is re-run on an existing workbook.
CATEGORY_TYPE_HINTS: dict[str, str] = {name: typ for name, typ in DEFAULT_CATEGORIES}

CATEGORY_EMOJI: dict[str, str] = {
    "Salary": "💰", "Other Income": "💵", "Housing": "🏠", "Groceries": "🛒",
    "Transport": "🚗", "Utilities": "💡", "Health": "🏥", "Dining Out": "🍽️",
    "Shopping": "🛍️", "Entertainment": "🎬", "Subscriptions": "📱",
    "Travel": "✈️", "Savings": "💾", "Other": "📁",
}
FALLBACK_EMOJI = "📁"

CURRENCY_CHOICES = ["USD", "EUR", "RUB", "TRY", "CNY"]
_CCY_RE = re.compile(r"^[A-Za-z]{3}$")
_RENAME_BUTTONS_PER_ROW = 2
NO_LIMIT = 0  # budget value meaning "no limit"

CANCEL_TEXT = "⏹ Setup stopped. Changes so far are saved. Run /setup to continue."
CREATE_FAIL_TEXT = ("❌ Could not create the budget file. "
                    "Check XLSX_PATH is set and the path is writable.")

_SK = "setup"  # ctx.user_data key for the session dict


def _emoji(name: str) -> str:
    return CATEGORY_EMOJI.get(name, FALLBACK_EMOJI)


def _session(ctx) -> dict:
    return ctx.user_data.setdefault(_SK, {
        "categories": [], "budgets": {}, "budget_queue": [],
        "pending_rename": None, "pending_add": None, "currency": None,
    })


def _msg(update: Update):
    return update.message or (update.callback_query and update.callback_query.message)


async def _reply(update: Update, text: str, **kwargs) -> bool:
    """Reply via the effective message. Returns False if no message is available
    (e.g. an inaccessible callback message) — callers should end the conversation."""
    msg = _msg(update)
    if not msg:
        if update.callback_query:
            try:
                await update.callback_query.answer()
            except Exception:
                pass
        log.warning("/setup: no effective message to reply to")
        return False
    await msg.reply_text(text, **kwargs)
    return True


# ── Workbook state detection / IO ─────────────────────────────────────────────

def _workbook_exists() -> bool:
    import file_storage
    if file_storage._active_backend() != "local":
        return True  # remote backends: assume the object exists; load decides
    return file_storage.LOCAL_XLSX_PATH.exists()


def _load_existing_state() -> tuple[list[tuple[str, str]], dict[str, float]]:
    """Read categories (with persisted or hinted types) and budgets."""
    from file_storage import get_excel_path_for_reading, load_lists
    path = get_excel_path_for_reading()
    lists = load_lists(path)
    stored_types: dict[str, str] = {}
    try:
        from openpyxl import load_workbook
        from excel_schema import ListsSchema, col_indices
        wb = load_workbook(path, read_only=True)
        ws = wb["Lists"]
        idx = col_indices(ws, ListsSchema)
        cat_col, typ_col = idx.get("categories"), idx.get("category_type")
        if cat_col and typ_col:
            for row in ws.iter_rows(min_row=2, values_only=True):
                cat = row[cat_col - 1] if len(row) >= cat_col else None
                typ = row[typ_col - 1] if len(row) >= typ_col else None
                if cat is not None and typ:
                    stored_types[str(cat).strip()] = str(typ).strip()
        wb.close()
    except Exception:
        log.warning("/setup: could not read stored category types", exc_info=True)
    cats = [(str(c),
             stored_types.get(str(c)) or CATEGORY_TYPE_HINTS.get(str(c), "Expense"))
            for c in lists.get("categories", [])]
    return cats, dict(lists.get("budgets", {}))


def _extend_category_validation(wb, n_categories: int) -> None:
    """Make sure the MasterData Category dropdown covers all Lists rows."""
    if "MasterData" not in wb.sheetnames:
        return
    pattern = re.compile(r"(Lists!\$?C\$?2:\$?C\$?)(\d+)")
    for dv in wb["MasterData"].data_validations.dataValidation:
        f1 = str(dv.formula1 or "")
        m = pattern.search(f1)
        if m and int(m.group(2)) < n_categories + 1:
            dv.formula1 = pattern.sub(rf"\g<1>{n_categories + 1}", f1)


def _commit_categories(session: dict) -> None:
    """Checkpoint A: Lists!C + both dashboards + validation, one atomic save."""
    from openpyxl import load_workbook
    from cycle_dashboard import CYCLE_DASHBOARD_SHEET_NAME, sync_cycle_dashboard_categories
    from excel_schema import (
        ListsSchema, col_indices, header_of, sync_dashboard_categories,
    )
    from file_storage import ExcelFileContext, atomic_save

    names = [name for name, _typ in session["categories"]]
    types = {name: typ for name, typ in session["categories"]}
    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["Lists"]
        idx = col_indices(ws, ListsSchema)
        cat_col = idx["categories"]
        bud_col = idx.get("budget_base")
        typ_col = idx.get("category_type")
        if typ_col is None:  # older workbooks: add the column at the first free slot
            typ_col = ws.max_column + 1
            ws.cell(1, typ_col, header_of(ListsSchema, "category_type"))
        # preserve budgets of surviving categories, clear the rest
        old_budgets = {}
        for r in range(2, ws.max_row + 1):
            cat = ws.cell(r, cat_col).value
            if cat is not None and bud_col:
                old_budgets[str(cat).strip()] = ws.cell(r, bud_col).value
            ws.cell(r, cat_col).value = None
            ws.cell(r, typ_col).value = None
            if bud_col:
                ws.cell(r, bud_col).value = None
        for i, name in enumerate(names, start=2):
            ws.cell(i, cat_col, name)
            ws.cell(i, typ_col, types.get(name, "Expense"))
            if bud_col and old_budgets.get(name) is not None:
                ws.cell(i, bud_col, old_budgets[name])
        sync_dashboard_categories(wb, names)
        if CYCLE_DASHBOARD_SHEET_NAME in wb.sheetnames:
            sync_cycle_dashboard_categories(wb)
        _extend_category_validation(wb, len(names))
        atomic_save(wb, excel_path)
    log.info("/setup: committed %d categories", len(names))


def _commit_budgets_and_currency(session: dict) -> None:
    """Checkpoint B: budgets into Lists!D, currency into Lists!H, one save."""
    from openpyxl import load_workbook
    from excel_schema import ListsSchema, col_indices
    from file_storage import ExcelFileContext, atomic_save

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["Lists"]
        idx = col_indices(ws, ListsSchema)
        cat_col, bud_col = idx.get("categories"), idx.get("budget_base")
        if cat_col and bud_col:
            for r in range(2, ws.max_row + 1):
                cat = ws.cell(r, cat_col).value
                if cat is None:
                    break
                amount = session["budgets"].get(str(cat).strip())
                if amount is not None and amount > NO_LIMIT:
                    ws.cell(r, bud_col, round(float(amount), 2))
        ccy = session.get("currency")
        ccy_col, rate_col = idx.get("currency"), idx.get("rate_to_base")
        if ccy and ccy_col:
            existing, first_free = set(), None
            for r in range(2, ws.max_row + 2):
                val = ws.cell(r, ccy_col).value
                if val is None or str(val).startswith("←"):
                    first_free = first_free or r
                    if val is None:
                        break
                else:
                    existing.add(str(val).strip().upper())
            if ccy not in existing and first_free:
                ws.cell(first_free, ccy_col, ccy)
                if rate_col:
                    ws.cell(first_free, rate_col, 1.0)  # placeholder until rates refresh
        atomic_save(wb, excel_path)
    log.info("/setup: committed budgets (%d) and currency %s",
             len(session["budgets"]), session.get("currency"))


async def _refresh_rates_best_effort() -> dict[str, float]:
    """Fetch live rates from frankfurter and update Excel. Raises on failure."""
    import httpx
    from excel_ops import async_update_currency_rates
    urls = ["https://api.frankfurter.dev/v1/latest?from=PLN",
            "https://api.frankfurter.app/latest?from=PLN"]
    data = None
    async with httpx.AsyncClient(follow_redirects=True, timeout=10.0) as client:
        for url in urls:
            try:
                resp = await client.get(url)
                resp.raise_for_status()
                data = resp.json()
                break
            except Exception:
                continue
    if data is None:
        raise RuntimeError("no rate source reachable")
    live = {c.upper(): round(1 / r, 4) for c, r in data["rates"].items() if r > 0}
    live["PLN"] = 1.0
    await async_update_currency_rates(live)
    return live


# ── Message builders ──────────────────────────────────────────────────────────

def _review_view(session: dict) -> tuple[str, InlineKeyboardMarkup]:
    lines = ["📋 *Your categories:*"]
    for i, (name, typ) in enumerate(session["categories"], 1):
        lines.append(f"{i}. {_emoji(name)} {name} ({typ})")
    lines.append("\nRename or add, then confirm.")
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✏️ Rename a category", callback_data="setup:rename")],
        [InlineKeyboardButton("➕ Add a category", callback_data="setup:add")],
        [InlineKeyboardButton("🗑️ Remove a category", callback_data="setup:remove")],
        [InlineKeyboardButton("✅ Done with categories", callback_data="setup:done")],
        [InlineKeyboardButton("Cancel", callback_data="setup:cancel")],
    ])
    return "\n".join(lines), kb


def _rename_picker(session: dict) -> InlineKeyboardMarkup:
    buttons = [
        InlineKeyboardButton(f"{_emoji(name)} {name}", callback_data=f"setup:ren:{i}")
        for i, (name, _typ) in enumerate(session["categories"])
    ]
    rows = [buttons[i:i + _RENAME_BUTTONS_PER_ROW]
            for i in range(0, len(buttons), _RENAME_BUTTONS_PER_ROW)]
    rows.append([InlineKeyboardButton("Cancel", callback_data="setup:cancel")])
    return InlineKeyboardMarkup(rows)


def _remove_picker(session: dict) -> InlineKeyboardMarkup:
    buttons = [
        InlineKeyboardButton(f"{_emoji(name)} {name}", callback_data=f"setup:del:{i}")
        for i, (name, _typ) in enumerate(session["categories"])
    ]
    rows = [buttons[i:i + _RENAME_BUTTONS_PER_ROW]
            for i in range(0, len(buttons), _RENAME_BUTTONS_PER_ROW)]
    rows.append([InlineKeyboardButton("Cancel", callback_data="setup:cancel")])
    return InlineKeyboardMarkup(rows)


def _currency_keyboard() -> InlineKeyboardMarkup:
    codes = CURRENCY_CHOICES + ["Other"]
    buttons = [InlineKeyboardButton(c, callback_data=f"setup:ccy:{c}") for c in codes]
    rows = [buttons[i:i + 2] for i in range(0, len(buttons), 2)]
    rows.append([InlineKeyboardButton("Cancel", callback_data="setup:cancel")])
    return InlineKeyboardMarkup(rows)


def _summary_text(session: dict) -> str:
    lines = ["✅ *Setup complete.*",
             f"*Currency:* {session['currency']}",
             "*Categories:*"]
    for name, _typ in session["categories"]:
        amount = session["budgets"].get(name, NO_LIMIT)
        limit = f"{amount:g}" if amount and amount > NO_LIMIT else "no limit"
        lines.append(f"{_emoji(name)} {name} — {limit}")
    lines.append("\nFetching exchange rates now. "
                 "Use /add to log your first transaction.")
    return "\n".join(lines)


async def _show_review(update: Update, ctx) -> int:
    text, kb = _review_view(_session(ctx))
    if not await _reply(update, text, parse_mode="Markdown", reply_markup=kb):
        return ConversationHandler.END
    return SETUP_REVIEW


# ── Entry points ──────────────────────────────────────────────────────────────

async def _begin_fresh(update: Update, ctx) -> int:
    """Create the workbook if needed, seed defaults, show Step 1 + Step 2."""
    import file_storage
    if not _workbook_exists():
        try:
            file_storage.create_workbook_from_template(file_storage.LOCAL_XLSX_PATH)
        except Exception:
            log.exception("/setup: workbook creation failed")
            await _reply(update, CREATE_FAIL_TEXT)
            return ConversationHandler.END
    session = _session(ctx)
    session["categories"] = list(DEFAULT_CATEGORIES)
    session["budgets"] = {}
    if not await _reply(
        update,
        "🚀 *Setup started.* Review your categories below.",
        parse_mode="Markdown",
    ):
        return ConversationHandler.END
    return await _show_review(update, ctx)


@auth_write
@log_call()
async def cmd_setup(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    had_stale_session = _SK in ctx.user_data
    ctx.user_data.pop(_SK, None)
    if had_stale_session:
        await _reply(update,
                     "⚠️ Previous setup session was abandoned. "
                     "Any unsaved category edits are lost. Starting fresh.")
    if not _workbook_exists():
        return await _begin_fresh(update, ctx)

    from openpyxl import load_workbook
    from file_storage import get_excel_path_for_reading, lists_categories_populated
    try:
        wb = load_workbook(get_excel_path_for_reading(), read_only=True)
        populated = lists_categories_populated(wb)
        wb.close()
    except Exception:
        log.exception("/setup: could not read workbook")
        await _reply(update, CREATE_FAIL_TEXT)
        return ConversationHandler.END

    if not populated:
        return await _begin_fresh(update, ctx)

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("Continue", callback_data="setup:continue"),
        InlineKeyboardButton("Cancel", callback_data="setup:cancel"),
    ]])
    if not await _reply(
        update,
        "⚠️ Setup already ran. Running it again edits your existing "
        "categories and budgets. Continue?",
        reply_markup=kb,
    ):
        return ConversationHandler.END
    return SETUP_WELCOME


@auth
@log_call()
async def cmd_start_or_setup(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    """/start entry: fall through to the menu unless the workbook is missing.
    Only the owner (ALLOWED_USERS[0]) is routed into setup."""
    from config import ALLOWED_USERS
    if not _workbook_exists():
        if update.effective_user.id == ALLOWED_USERS[0]:
            ctx.user_data.pop(_SK, None)
            return await _begin_fresh(update, ctx)
        await _reply(update,
                     "No workbook found. The bot owner needs to run /setup first.")
        return ConversationHandler.END
    from handlers.menu import cmd_menu
    await cmd_menu(update, ctx)
    return ConversationHandler.END


# ── Step handlers ─────────────────────────────────────────────────────────────

@log_call()
async def setup_welcome_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "setup:cancel":
        return await setup_cancel(update, ctx)
    try:
        cats, budgets = _load_existing_state()
    except Exception:
        log.exception("/setup: could not load existing state")
        await _reply(update, "❌ Could not read the budget file. Try again.")
        return ConversationHandler.END
    session = _session(ctx)
    session["categories"] = cats or list(DEFAULT_CATEGORIES)
    session["budgets"] = budgets
    return await _show_review(update, ctx)


@log_call()
async def setup_review_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    session = _session(ctx)

    if query.data == "setup:cancel":
        return await setup_cancel(update, ctx)
    if query.data == "setup:rename":
        await query.message.reply_text(
            "Tap the category to rename:", reply_markup=_rename_picker(session))
        return SETUP_RENAME
    if query.data == "setup:add":
        session["pending_add"] = None
        await query.message.reply_text("Send the new category name:")
        return SETUP_ADD
    if query.data == "setup:remove":
        if len(session["categories"]) <= 1:
            await query.message.reply_text("You need at least 1 category.")
            return await _show_review(update, ctx)
        await query.message.reply_text(
            "Tap the category to remove:", reply_markup=_remove_picker(session))
        return SETUP_REMOVE
    if query.data == "setup:done":
        try:
            _commit_categories(session)
        except Exception:
            log.exception("/setup: category commit failed")
            await query.message.reply_text(CREATE_FAIL_TEXT)
            return ConversationHandler.END
        session["budget_queue"] = [
            name for name, typ in session["categories"] if typ == "Expense"]
        return await _next_budget_or_currency(update, ctx)
    return SETUP_REVIEW


async def _next_budget_or_currency(update: Update, ctx) -> int:
    session = _session(ctx)
    if session["budget_queue"]:
        cat = session["budget_queue"][0]
        if not await _reply(
            update,
            f"💸 Monthly budget for *{cat}*? Send a number. *0* = no limit.",
            parse_mode="Markdown",
        ):
            return ConversationHandler.END
        return SETUP_BUDGET
    if not await _reply(
        update, "💰 Pick your main currency:", reply_markup=_currency_keyboard()
    ):
        return ConversationHandler.END
    return SETUP_CURRENCY


@log_call()
async def setup_rename_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    session = _session(ctx)
    if query.data == "setup:cancel":
        session["pending_rename"] = None  # discard pending, keep committed work
        return await setup_cancel(update, ctx)
    try:
        idx = int(query.data.rsplit(":", 1)[1])
    except ValueError:
        return await _show_review(update, ctx)
    if not 0 <= idx < len(session["categories"]):
        return await _show_review(update, ctx)
    session["pending_rename"] = idx
    old = session["categories"][idx][0]
    await query.message.reply_text(
        f"Send the new name for *{old}*:", parse_mode="Markdown")
    return SETUP_RENAME


@log_call()
async def setup_rename_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    session = _session(ctx)
    idx = session.get("pending_rename")
    if idx is None:
        return SETUP_RENAME
    new = (update.message.text or "").strip()
    if not new:
        await update.message.reply_text("❌ Name cannot be empty. Send a name:")
        return SETUP_RENAME
    old, typ = session["categories"][idx]
    session["categories"][idx] = (new, typ)
    if old in session["budgets"]:
        session["budgets"][new] = session["budgets"].pop(old)
    session["pending_rename"] = None
    await update.message.reply_text(
        f"✅ *{old}* → *{new}*", parse_mode="Markdown")
    return await _show_review(update, ctx)


@log_call()
async def setup_remove_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    session = _session(ctx)
    if query.data == "setup:cancel":
        return await setup_cancel(update, ctx)
    try:
        idx = int(query.data.rsplit(":", 1)[1])
    except ValueError:
        return await _show_review(update, ctx)
    if not 0 <= idx < len(session["categories"]):
        return await _show_review(update, ctx)
    name, _typ = session["categories"].pop(idx)
    session["budgets"].pop(name, None)
    await query.message.reply_text(f"🗑️ Removed *{name}*.", parse_mode="Markdown")
    return await _show_review(update, ctx)


@log_call()
async def setup_add_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    session = _session(ctx)
    name = (update.message.text or "").strip()
    if not name:
        await update.message.reply_text("❌ Name cannot be empty. Send a name:")
        return SETUP_ADD
    session["pending_add"] = name
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(t, callback_data=f"setup:type:{t}")
        for t in ("Expense", "Income", "Savings")
    ]])
    await update.message.reply_text(
        f"What type is *{name}*?", parse_mode="Markdown", reply_markup=kb)
    return SETUP_ADD


@log_call()
async def setup_add_type_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    session = _session(ctx)
    if query.data == "setup:cancel":
        return await setup_cancel(update, ctx)
    typ = query.data.rsplit(":", 1)[1]
    name = session.get("pending_add")
    if not name:
        return await _show_review(update, ctx)
    session["categories"].append((name, typ))
    session["pending_add"] = None
    await query.message.reply_text(
        f"✅ Added *{name}* ({typ}).", parse_mode="Markdown")
    return await _show_review(update, ctx)


@log_call()
async def setup_budget_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    from validators import parse_amount
    session = _session(ctx)
    if not session["budget_queue"]:
        return await _next_budget_or_currency(update, ctx)
    cat = session["budget_queue"][0]
    try:
        amount, _ = parse_amount(update.message.text)
        if amount < 0:
            raise ValueError
    except (ValueError, TypeError):
        await update.message.reply_text(
            f"❌ Numbers only. Budget for *{cat}* (0 = no limit):",
            parse_mode="Markdown",
        )
        return SETUP_BUDGET
    session["budget_queue"].pop(0)
    if amount > NO_LIMIT:
        session["budgets"][cat] = amount
        note = f"✅ *{cat}*: {amount:g}"
    else:
        note = f"✅ *{cat}*: no limit"
    await update.message.reply_text(note, parse_mode="Markdown")
    return await _next_budget_or_currency(update, ctx)


async def _finish_with_currency(update: Update, ctx, ccy: str) -> int:
    session = _session(ctx)
    session["currency"] = ccy
    try:
        _commit_budgets_and_currency(session)
    except Exception:
        log.exception("/setup: final commit failed")
        await _reply(update, CREATE_FAIL_TEXT)
        return ConversationHandler.END
    set_display_currency(update.effective_user.id, ccy)

    kb = InlineKeyboardMarkup([[InlineKeyboardButton("Done", callback_data="setup:finish")]])
    if not await _reply(
        update, _summary_text(session), parse_mode="Markdown", reply_markup=kb
    ):
        return ConversationHandler.END

    try:
        live = await _refresh_rates_best_effort()
        if ccy not in live:
            await _reply(
                update,
                f"⚠️ No live rate found for {ccy} — its rate is set to 1.0. "
                "Update it in the Lists sheet or run /rates refresh later.",
            )
    except Exception as e:
        log.warning("/setup: rate refresh failed: %s", e)
        await _reply(
            update,
            "⚠️ Could not fetch live exchange rates. Run /rates refresh later.")
    return SETUP_SUMMARY


@log_call()
async def setup_currency_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "setup:cancel":
        return await setup_cancel(update, ctx)
    choice = query.data.rsplit(":", 1)[1]
    if choice == "Other":
        await query.message.reply_text("Send the 3-letter currency code (e.g. GBP):")
        return SETUP_CURRENCY
    return await _finish_with_currency(update, ctx, choice.upper())


@log_call()
async def setup_currency_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    code = (update.message.text or "").strip()
    if not _CCY_RE.fullmatch(code):
        await update.message.reply_text("❌ Not a valid 3-letter code. Try again:")
        return SETUP_CURRENCY
    return await _finish_with_currency(update, ctx, code.upper())


@log_call()
async def setup_summary_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    ctx.user_data.pop(_SK, None)
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass
    return ConversationHandler.END


@log_call()
async def setup_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    ctx.user_data.pop(_SK, None)
    await _reply(update, CANCEL_TEXT)
    return ConversationHandler.END


# ── Handler factory ───────────────────────────────────────────────────────────

def setup_conversation_handler() -> ConversationHandler:
    text_only = filters.TEXT & ~filters.COMMAND
    return ConversationHandler(
        entry_points=[
            CommandHandler("setup", cmd_setup),
            CommandHandler("start", cmd_start_or_setup),
        ],
        states={
            SETUP_WELCOME: [CallbackQueryHandler(setup_welcome_cb, pattern="^setup:")],
            SETUP_REVIEW:  [CallbackQueryHandler(setup_review_cb, pattern="^setup:")],
            SETUP_RENAME: [
                CallbackQueryHandler(setup_rename_cb, pattern="^setup:"),
                MessageHandler(text_only, setup_rename_text),
            ],
            SETUP_ADD: [
                CallbackQueryHandler(setup_add_type_cb, pattern="^setup:"),
                MessageHandler(text_only, setup_add_text),
            ],
            SETUP_REMOVE: [CallbackQueryHandler(setup_remove_cb, pattern="^setup:")],
            SETUP_BUDGET:  [MessageHandler(text_only, setup_budget_text)],
            SETUP_CURRENCY: [
                CallbackQueryHandler(setup_currency_cb, pattern="^setup:"),
                MessageHandler(text_only, setup_currency_text),
            ],
            SETUP_SUMMARY: [CallbackQueryHandler(setup_summary_cb, pattern="^setup:")],
        },
        fallbacks=[
            CommandHandler("cancel", setup_cancel),
            CommandHandler("setup", cmd_setup),
        ],
    )
