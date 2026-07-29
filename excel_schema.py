"""
excel_schema.py
===============
Column declarations for every Excel sheet the bot reads or writes.

Each schema class is a plain dataclass where every field carries the exact
header text of the corresponding Excel column via metadata={"excel_header": ...}.
This replaces all hardcoded column positions across the codebase.

Usage
-----
    from excel_schema import ListsSchema, MasterDataSchema, find_col, col_indices

    # Find one column
    ccy_col = find_col(ws, ListsSchema.currency)   # returns int | None

    # Build full index for a sheet
    idx = col_indices(ws, MasterDataSchema)         # {field_name: col_int}
    ws.cell(row, idx["year"], 2025)
"""

from dataclasses import dataclass, field, fields
from datetime import date
from typing import Any


# ── Value helpers ─────────────────────────────────────────────────────────────

def to_date(value) -> date | None:
    """Coerce a cell value (datetime, date, or ISO string) to a date, else None."""
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip()[:10])
    except ValueError:
        return None


# ── Field helper ──────────────────────────────────────────────────────────────

def col(header: str) -> Any:
    """Declare an Excel column by its exact header text (case-insensitive match)."""
    return field(default=None, metadata={"excel_header": header})


# ── Lookup helpers ────────────────────────────────────────────────────────────

def find_col(ws, header: str) -> int | None:
    """Return the 1-based column index whose row-1 value matches header (case-insensitive)."""
    needle = header.strip().lower()
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().lower() == needle:
            return c
    return None


def col_indices(ws, schema_cls) -> dict[str, int]:
    """
    Return {field_name: column_index} for every declared field in schema_cls
    whose header is found in ws row 1.  Missing columns are silently omitted.
    """
    result = {}
    for f in fields(schema_cls):
        header = f.metadata.get("excel_header")
        if header:
            c = find_col(ws, header)
            if c is not None:
                result[f.name] = c
    return result


def header_of(schema_cls, field_name: str) -> str:
    """Return the declared excel_header string for a given field name."""
    for f in fields(schema_cls):
        if f.name == field_name:
            return f.metadata.get("excel_header", field_name)
    raise KeyError(f"Field {field_name!r} not found in {schema_cls.__name__}")


def load_currency_rates_from_path(excel_path) -> dict[str, float]:
    """
    Read {currency_code: rate_to_base} from the Lists sheet.
    Uses ListsSchema to locate columns by header name — no positional assumptions.
    Returns {"PLN": 1.0} on any failure.
    """
    import re
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=True)
        ws = wb["Lists"]
        idx      = col_indices(ws, ListsSchema)
        ccy_col  = idx.get("currency")
        rate_col = idx.get("rate_to_base")
        if not ccy_col or not rate_col:
            return {"PLN": 1.0}
        rates: dict[str, float] = {}
        for row in range(2, ws.max_row + 1):
            ccy  = ws.cell(row, ccy_col).value
            rate = ws.cell(row, rate_col).value
            if ccy is None:
                break
            ccy_str = str(ccy).strip().upper()
            if re.match(r"^[A-Z]{3}$", ccy_str) and rate is not None:
                try:
                    rates[ccy_str] = float(rate)
                except (TypeError, ValueError):
                    pass
        return rates or {"PLN": 1.0}
    except Exception:
        return {"PLN": 1.0}


# ── Shared MasterData row writer ──────────────────────────────────────────────

_EXCEL_MAX_ROW = 1048576  # Excel's hard limit — truly open-ended for any realistic dataset


def lists_currency_range(wb) -> str:
    """VLOOKUP range for Currency→Rate on the Lists sheet (open-ended to Excel max row)."""
    from openpyxl.utils import get_column_letter
    idx      = col_indices(wb["Lists"], ListsSchema)
    ccy_col  = idx.get("currency",    8)
    rate_col = idx.get("rate_to_base", 9)
    return f"${get_column_letter(ccy_col)}$2:${get_column_letter(rate_col)}${_EXCEL_MAX_ROW}"


def find_next_data_row(ws) -> int:
    """
    Next writable MasterData row based on actual content (Date/Value columns).
    ws.max_row lies when empty rows carry styling or data validations.
    """
    idx = col_indices(ws, MasterDataSchema)
    value_col = idx.get("value", 4)
    date_col  = idx.get("date", 1)
    last_data_row = 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, value_col).value is not None or ws.cell(row, date_col).value is not None:
            last_data_row = row
    return last_data_row + 1


_VALIDATION_MARGIN_ROWS = 500


def extend_validation_ranges(ws, last_row: int, margin: int = _VALIDATION_MARGIN_ROWS) -> None:
    """
    Dropdown validations are static ranges (e.g. F2:F103) — appended rows fall
    outside them and show no list in Excel. Extend every single-column range
    on the sheet so it covers at least last_row + margin.
    """
    import re
    target = last_row + margin
    for dv in ws.data_validations.dataValidation:
        parts = []
        changed = False
        for rng in str(dv.sqref).split():
            m = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", rng)
            if m and m.group(1) == m.group(3) and int(m.group(4)) < target:
                rng = f"{m.group(1)}{m.group(2)}:{m.group(3)}{target}"
                changed = True
            parts.append(rng)
        if changed:
            dv.sqref = " ".join(parts)


def write_transaction_row(ws, r: int, row: dict, lu_range: str) -> None:
    """
    Write one transaction dict into MasterData row r.
    The single source of truth for column layout and the Value (base) formula —
    used by single append, batch append, and recovery-queue replay.
    """
    from datetime import datetime, timezone
    from openpyxl.utils import get_column_letter

    idx = col_indices(ws, MasterDataSchema)
    c = lambda field, fallback: idx.get(field, fallback)

    # Older live files predate the Date Modified column — writing values into a
    # headerless column looks like a stray column in Excel. Create the header.
    if "date_modified" not in idx:
        hdr_col = ws.max_column + 1 if ws.cell(1, 13).value not in (None, "") else 13
        ws.cell(1, hdr_col, header_of(MasterDataSchema, "date_modified"))
        idx["date_modified"] = hdr_col

    ws.cell(r, c("date",         1),  row.get("date"))
    ws.cell(r, c("year",         2),  row.get("year"))
    ws.cell(r, c("month",        3),  row.get("month"))
    ws.cell(r, c("value",        4),  row.get("value"))
    ws.cell(r, c("type",         5),  row.get("type"))
    ws.cell(r, c("category",     6),  row.get("category"))
    ws.cell(r, c("person",       7),  row.get("person"))
    # Formula-injection guard: descriptions come from untrusted sources (AI
    # output, bank statements). A leading = + - @ becomes a live Excel formula.
    desc = row.get("description")
    if isinstance(desc, str) and desc[:1] in ("=", "+", "-", "@"):
        desc = "'" + desc
    ws.cell(r, c("description",  8),  desc)
    ws.cell(r, c("is_recurring", 9),  row.get("is_recurring"))
    is_done = row.get("is_done")
    ws.cell(r, c("is_done",      10), True if is_done is None else bool(is_done))

    ccy_col = c("currency", 11)
    ws.cell(r, ccy_col, row.get("currency", "PLN"))

    vbase_col    = c("value_base", 12)
    value_letter = get_column_letter(c("value", 4))
    ccy_letter   = get_column_letter(ccy_col)
    ws.cell(r, vbase_col,
        f'=IF(OR({ccy_letter}{r}="",{ccy_letter}{r}="PLN"),'
        f'{value_letter}{r},'
        f'{value_letter}{r}*VLOOKUP({ccy_letter}{r},Lists!{lu_range},2,0))'
    )
    ws.cell(r, c("date_modified", 13), datetime.now(timezone.utc).replace(tzinfo=None))
    extend_validation_ranges(ws, r)


# ── Dashboard category block (shared by Dashboard and Cycle Dashboard) ───────

# Both dashboards render "Expenses by Category" as H11.. with columns
# H=Category, I=Budget, J=Actual, K=Variance, L=Var % and a TOTAL row below.
CATEGORY_FIRST_ROW = 11
_CAT_COL_CATEGORY = 8   # H
_CAT_COL_BUDGET   = 9   # I
_CAT_COL_ACTUAL   = 10  # J
_CAT_COL_VARIANCE = 11  # K
_CAT_COL_VAR_PCT  = 12  # L
CATEGORY_TOTAL_LABEL = "TOTAL"


def _dashboard_actual_formula(r: int) -> str:
    """Main Dashboard 'Actual' SUMIFS for the category in H{r} — filtered by
    the Year ($B$2) and optional Month ($D$2) selectors, in display currency."""
    return (
        f'=(IF($D$2="",SUMIFS(MasterData!$L$2:$L${_EXCEL_MAX_ROW},'
        f'MasterData!$F$2:$F${_EXCEL_MAX_ROW},H{r},'
        f'MasterData!$B$2:$B${_EXCEL_MAX_ROW},$B$2,'
        f'MasterData!$J$2:$J${_EXCEL_MAX_ROW},TRUE),'
        f'SUMIFS(MasterData!$L$2:$L${_EXCEL_MAX_ROW},'
        f'MasterData!$F$2:$F${_EXCEL_MAX_ROW},H{r},'
        f'MasterData!$B$2:$B${_EXCEL_MAX_ROW},$B$2,'
        f'MasterData!$C$2:$C${_EXCEL_MAX_ROW},$D$2,'
        f'MasterData!$J$2:$J${_EXCEL_MAX_ROW},TRUE)))/($N$2)'
    )


def write_category_sumif_row(ws, r: int, actual_formula: str) -> None:
    """Budget / Actual / Variance / Var % formulas for the category in H{r}.
    `actual_formula` supplies the dashboard-specific 'Actual' SUMIFS."""
    ws.cell(r, _CAT_COL_BUDGET,
            f"=IFERROR(VLOOKUP(H{r},Lists!$C$2:$D${_EXCEL_MAX_ROW},2,0),0)/$N$2")
    ws.cell(r, _CAT_COL_ACTUAL, actual_formula)
    ws.cell(r, _CAT_COL_VARIANCE, f"=I{r}-J{r}")
    ws.cell(r, _CAT_COL_VAR_PCT, f'=IF(I{r}=0,"",K{r}/I{r})')


def write_category_sumif_block(ws, categories: list[str], actual_formula_fn,
                               first_row: int = CATEGORY_FIRST_ROW) -> None:
    """
    Write the full H{first_row}..TOTAL category block.
    `actual_formula_fn(r)` returns the 'Actual' formula string for row r.
    """
    r = first_row
    for cat in categories:
        ws.cell(r, _CAT_COL_CATEGORY, cat)
        write_category_sumif_row(ws, r, actual_formula_fn(r))
        r += 1
    # TOTAL row
    ws.cell(r, _CAT_COL_CATEGORY, CATEGORY_TOTAL_LABEL)
    if r > first_row:
        ws.cell(r, _CAT_COL_BUDGET,   f"=SUM(I{first_row}:I{r - 1})")
        ws.cell(r, _CAT_COL_ACTUAL,   f"=SUM(J{first_row}:J{r - 1})")
        ws.cell(r, _CAT_COL_VARIANCE, f"=SUM(K{first_row}:K{r - 1})")
    else:
        ws.cell(r, _CAT_COL_BUDGET, 0)
        ws.cell(r, _CAT_COL_ACTUAL, 0)
        ws.cell(r, _CAT_COL_VARIANCE, 0)


def read_category_block(ws, first_row: int = CATEGORY_FIRST_ROW) -> list[str] | None:
    """
    Category names from H{first_row} down to (exclusive) the first TOTAL row.
    Returns None when no TOTAL row exists (malformed / absent block).
    """
    cats: list[str] = []
    for r in range(first_row, ws.max_row + 1):
        v = str(ws.cell(r, _CAT_COL_CATEGORY).value or "").strip()
        if v == CATEGORY_TOTAL_LABEL:
            return cats
        if v:
            cats.append(v)
    return None


def clear_category_block(ws, first_row: int = CATEGORY_FIRST_ROW) -> None:
    """Blank the existing H..L category block including its TOTAL row."""
    current = read_category_block(ws, first_row)
    old_total_row = first_row + (len(current) if current is not None else 0)
    for r in range(first_row, max(old_total_row, ws.max_row) + 1):
        for c in range(_CAT_COL_CATEGORY, _CAT_COL_VAR_PCT + 1):
            ws.cell(r, c).value = None


def sync_dashboard_categories(wb, categories: list[str]) -> int:
    """
    Rewrite the main Dashboard 'Expenses by Category' block so it lists
    exactly `categories`. Returns the number of category rows written
    (0 = already in sync or no Dashboard sheet).
    """
    if "Dashboard" not in wb.sheetnames:
        return 0
    ws = wb["Dashboard"]
    if read_category_block(ws) == categories:
        return 0
    clear_category_block(ws)
    write_category_sumif_block(ws, categories, _dashboard_actual_formula)
    return len(categories)


# ── Dashboard formula-bound repair ───────────────────────────────────────────

def repair_dashboard_bounds(wb) -> int:
    """
    Rewrite hardcoded row bounds in Dashboard formula cells to Excel's max row.

    Old templates contain ranges like MasterData!$L$2:$L$2000 and
    Lists!$C$2:$D$100 that silently stop counting when data grows past the cap.
    This replaces every such bound with $1048576 (Excel max row) in every
    formula cell across all sheets that reference MasterData or Lists.

    Returns the number of cells updated.
    """
    import re

    pattern = re.compile(
        r"((MasterData|Lists)!\$[A-Z]+\$2:\$[A-Z]+\$)\d+"
    )
    replacement = rf"\g<1>{_EXCEL_MAX_ROW}"

    updated = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                new_val = pattern.sub(replacement, cell.value)
                if new_val != cell.value:
                    cell.value = new_val
                    updated += 1
    return updated


# ── Category cascade rename ───────────────────────────────────────────────────

def rename_category_in_workbook(wb, old_name: str, new_name: str) -> dict[str, int]:
    """
    Rename a category everywhere inside an already-open workbook object:
      - MasterData Category column (all data rows)
      - Dashboard plain-value cells
      - Formula string literals in Dashboard + Monthly Summary

    Returns counts per area. Does NOT save — caller must atomic_save.
    Lists!C and bulk drafts are handled by the caller (_commit_categories
    and rename_category.py respectively).
    """
    counts: dict[str, int] = {"MasterData": 0, "Dashboard": 0, "Formulas": 0}

    # MasterData: Category column
    if "MasterData" in wb.sheetnames:
        ws = wb["MasterData"]
        cat_col = next(
            (c for c in range(1, ws.max_column + 1)
             if str(ws.cell(1, c).value or "").strip() == "Category"),
            None,
        )
        if cat_col:
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, cat_col).value or "").strip() == old_name:
                    ws.cell(r, cat_col, new_name)
                    counts["MasterData"] += 1

    # Dashboard: plain-value cells (budget table rows use plain strings)
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() == old_name:
                    ws.cell(r, c, new_name)
                    counts["Dashboard"] += 1

    # Formula literals in Dashboard + Monthly Summary
    old_lit, new_lit = f'"{old_name}"', f'"{new_name}"'
    for sheet_name in ("Dashboard", "Monthly Summary"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and old_lit in v:
                    cell.value = v.replace(old_lit, new_lit)
                    counts["Formulas"] += 1

    return counts


# ── Monthly Summary auto-population ──────────────────────────────────────────

_MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _ms_col_indices(ws) -> dict[str, int]:
    """Return {name: col_index} for Monthly Summary header row."""
    result: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        hdr = str(ws.cell(1, c).value or "").strip().lower()
        if hdr:
            result[hdr] = c
    return result


def _write_monthly_summary_row(ws, r: int, year: int, month: str,
                                md_col: dict[str, int]) -> None:
    """
    Write one Monthly Summary row using open-ended whole-column SUMIFS.
    md_col maps MasterData field names → column letters (e.g. {"year": "B"}).
    """
    from openpyxl.utils import get_column_letter
    ms = _ms_col_indices(ws)

    yr_col  = get_column_letter(ms.get("year",  1))
    mo_col  = get_column_letter(ms.get("month", 2))

    def sumifs(txn_type: str) -> str:
        vb  = md_col["value_base"]
        yr  = md_col["year"]
        mo  = md_col["month"]
        typ = md_col["type"]
        return (
            f"=IFERROR(SUMIFS(MasterData!${vb}:${vb},"
            f"MasterData!${yr}:${yr},${yr_col}{r},"
            f"MasterData!${mo}:${mo},${mo_col}{r},"
            f"MasterData!${typ}:${typ},{txn_type!r}),0)"
        )

    ws.cell(r, ms.get("year",  1), year)
    ws.cell(r, ms.get("month", 2), month)

    inc_col = ms.get("income",        3)
    exp_col = ms.get("expenses",      4)
    sav_col = ms.get("savings",       5)
    net_col = ms.get("net",           6)
    sr_col  = ms.get("savings rate %", 7)

    inc_letter = get_column_letter(inc_col)
    exp_letter = get_column_letter(exp_col)
    sav_letter = get_column_letter(sav_col)

    ws.cell(r, inc_col, sumifs("Income"))
    ws.cell(r, exp_col, sumifs("Expense"))
    ws.cell(r, sav_col, sumifs("Savings"))
    if net_col:
        ws.cell(r, net_col,
                f"=${inc_letter}{r}-${exp_letter}{r}-${sav_letter}{r}")
    if sr_col:
        ws.cell(r, sr_col,
                f"=IF(${inc_letter}{r}=0,0,${sav_letter}{r}/${inc_letter}{r})")


def ensure_monthly_summary_row(wb, year: int, month: str) -> bool:
    """
    Ensure Monthly Summary has a formula row for (year, month).
    Returns True if a new row was appended, False if it already existed.
    Does nothing if the Monthly Summary sheet is absent.
    """
    if "Monthly Summary" not in wb.sheetnames:
        return False

    ws_ms = wb["Monthly Summary"]
    ws_md = wb["MasterData"]

    ms = _ms_col_indices(ws_ms)
    yr_col = ms.get("year",  1)
    mo_col = ms.get("month", 2)

    for r in range(2, ws_ms.max_row + 1):
        if ws_ms.cell(r, yr_col).value == year and ws_ms.cell(r, mo_col).value == month:
            return False  # row already exists

    # Derive MasterData column letters from schema
    from openpyxl.utils import get_column_letter
    md_idx = col_indices(ws_md, MasterDataSchema)
    md_col = {
        k: get_column_letter(md_idx.get(k, fallback))
        for k, fallback in [
            ("value_base", 12), ("year", 2), ("month", 3), ("type", 5),
        ]
    }

    # Find insertion point: sorted by year then month order
    insert_at = None
    for r in range(2, ws_ms.max_row + 1):
        row_yr = ws_ms.cell(r, yr_col).value
        row_mo = ws_ms.cell(r, mo_col).value
        if row_yr is None:
            insert_at = r
            break
        try:
            row_yr_int = int(row_yr)
        except (TypeError, ValueError):
            continue
        if row_yr_int > year:
            insert_at = r
            break
        if row_yr_int == year:
            row_mo_idx = _MONTH_ORDER.index(str(row_mo)) if row_mo in _MONTH_ORDER else 99
            new_mo_idx = _MONTH_ORDER.index(month) if month in _MONTH_ORDER else 99
            if row_mo_idx > new_mo_idx:
                insert_at = r
                break

    if insert_at is not None:
        ws_ms.insert_rows(insert_at)
        r = insert_at
    else:
        r = ws_ms.max_row + 1

    _write_monthly_summary_row(ws_ms, r, year, month, md_col)
    return True


def ensure_monthly_summary_rows_from_masterdata(wb) -> int:
    """
    Scan all MasterData rows and ensure Monthly Summary has a formula row
    for every distinct (Year, Month) combination found.
    Returns the count of new rows added.
    Suitable for post-batch-import reconciliation.
    """
    if "Monthly Summary" not in wb.sheetnames:
        return 0

    ws_md = wb["MasterData"]
    md_idx = col_indices(ws_md, MasterDataSchema)
    yr_col  = md_idx.get("year",  2)
    mo_col  = md_idx.get("month", 3)

    seen: set[tuple[int, str]] = set()
    for r in range(2, ws_md.max_row + 1):
        yr = ws_md.cell(r, yr_col).value
        mo = ws_md.cell(r, mo_col).value
        if yr is None and mo is None:
            continue
        try:
            yr_int = int(yr)
        except (TypeError, ValueError):
            continue
        mo_str = str(mo).strip() if mo else ""
        if mo_str:
            seen.add((yr_int, mo_str))

    added = 0
    for year, month in sorted(seen, key=lambda ym: (ym[0], _MONTH_ORDER.index(ym[1]) if ym[1] in _MONTH_ORDER else 99)):
        if ensure_monthly_summary_row(wb, year, month):
            added += 1
    return added


# ── MasterData sheet ──────────────────────────────────────────────────────────

@dataclass
class MasterDataSchema:
    """Column declarations for the MasterData sheet."""
    date:          Any = col("Date")
    year:          Any = col("Year")
    month:         Any = col("Month")
    value:         Any = col("Value")
    type:          Any = col("Type")
    category:      Any = col("Category")
    person:        Any = col("Person")
    description:   Any = col("Description")
    is_recurring:  Any = col("IsRecurring")
    is_done:       Any = col("IsDone")
    currency:      Any = col("Currency")
    value_base:    Any = col("Value (base)")
    date_modified: Any = col("Date Modified (UTC)")


# ── Lists sheet ───────────────────────────────────────────────────────────────

@dataclass
class ListsSchema:
    """Column declarations for the Lists sheet."""
    months:      Any = col("Months")
    txn_types:   Any = col("TxnTypes")
    categories:  Any = col("Categories")
    category_type: Any = col("Category Type")
    budget_base: Any = col("Budget (base)")
    persons:     Any = col("Persons")
    years:       Any = col("Years")
    currency:    Any = col("Currency")
    rate_to_base: Any = col("Rate to base")
    goal_name:   Any = col("Goal Name")
    alloc_pct:   Any = col("Alloc %")
    goal_pln:    Any = col("Goal (PLN)")
    salary_keyword: Any = col("Salary Keywords")


# ── Cycles sheet ──────────────────────────────────────────────────────────────

@dataclass
class CyclesSchema:
    """Column declarations for the Cycles sheet (one row per budget cycle)."""
    start_date: Any = col("StartDate")
    label:      Any = col("Label")
