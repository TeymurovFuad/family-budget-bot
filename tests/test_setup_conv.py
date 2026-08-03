"""
tests/test_setup_conv.py — /setup onboarding conversation.

All Telegram objects are mocked; all rate fetches are mocked (no live API
calls). Excel writes go to the pytest tmp_path via the excel_path fixture.
"""

import asyncio
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import load_workbook
from telegram.ext import ConversationHandler

import file_storage
from handlers import setup_conv
from handlers.setup_conv import (
    DEFAULT_CATEGORIES, _session, cmd_setup, cmd_start_or_setup,
    setup_add_text, setup_add_type_cb, setup_budget_text, setup_cancel,
    setup_conversation_handler, setup_currency_cb, setup_currency_text,
    setup_remove_cb, setup_rename_cb, setup_rename_text, setup_review_cb,
    setup_summary_cb, setup_welcome_cb,
)
from states import (
    SETUP_WELCOME, SETUP_REVIEW, SETUP_RENAME, SETUP_ADD,
    SETUP_BUDGET, SETUP_CURRENCY, SETUP_SUMMARY, SETUP_REMOVE,
)

OWNER_ID = 123  # first (and only) id in ALLOWED_TELEGRAM_IDS (conftest)


def run(coro):
    return asyncio.run(coro)


def make_update(user_id=OWNER_ID, text=None, callback_data=None):
    upd = MagicMock()
    upd.effective_user.id = user_id
    upd.effective_user.first_name = "Tester"
    if callback_data is not None:
        upd.message = None
        q = MagicMock()
        q.data = callback_data
        q.answer = AsyncMock()
        q.message.reply_text = AsyncMock()
        q.edit_message_reply_markup = AsyncMock()
        upd.callback_query = q
    else:
        upd.callback_query = None
        upd.message.text = text
        upd.message.reply_text = AsyncMock()
    return upd


def make_ctx(session=None):
    ctx = MagicMock()
    ctx.user_data = {}
    if session is not None:
        ctx.user_data["setup"] = session
    ctx.args = []
    return ctx


def fresh_session():
    return {
        "categories": list(DEFAULT_CATEGORIES), "budgets": {},
        "budget_queue": [], "pending_rename": None, "pending_add": None,
        "currency": None, "renames": [],
    }


def replies(upd) -> str:
    mock = (upd.message or upd.callback_query.message).reply_text
    return "\n".join(str(c.args[0]) for c in mock.await_args_list)


# ── file_storage helpers ──────────────────────────────────────────────────────

class TestFileStorageHelpers:
    def test_create_workbook_from_template_atomic(self, tmp_path, monkeypatch):
        monkeypatch.setattr(file_storage, "TEMPLATE_PATH",
                            tmp_path / "no_template.xlsx")
        dest = tmp_path / "sub" / "book.xlsx"
        file_storage.create_workbook_from_template(dest)
        assert dest.exists()
        assert not dest.with_suffix(".tmp").exists()
        wb = load_workbook(dest)
        assert "MasterData" in wb.sheetnames and "Lists" in wb.sheetnames

    def test_create_workbook_cleans_tmp_on_failure(self, tmp_path, monkeypatch):
        def boom(path):
            raise OSError("disk full")
        monkeypatch.setattr(file_storage, "create_blank_excel", boom)
        dest = tmp_path / "book.xlsx"
        with pytest.raises(OSError):
            file_storage.create_workbook_from_template(dest)
        assert not dest.exists()
        assert not dest.with_suffix(".tmp").exists()

    def test_lists_categories_populated(self, excel_path):
        wb = load_workbook(excel_path)
        assert file_storage.lists_categories_populated(wb) is True
        ws = wb["Lists"]
        from excel_schema import ListsSchema, col_indices
        cat_col = col_indices(ws, ListsSchema)["categories"]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, cat_col).value = None
        assert file_storage.lists_categories_populated(wb) is False


# ── shared dashboard category block ───────────────────────────────────────────

class TestSharedCategoryBlock:
    def test_sync_dashboard_categories_rewrites_block(self, excel_path):
        from excel_schema import (
            CATEGORY_FIRST_ROW, read_category_block, sync_dashboard_categories,
        )
        wb = load_workbook(excel_path)
        cats = ["Food", "Rent"]
        n = sync_dashboard_categories(wb, cats)
        assert n == 2
        ws = wb["Dashboard"]
        assert read_category_block(ws) == cats
        # formulas written
        assert str(ws.cell(CATEGORY_FIRST_ROW, 9).value).startswith("=IFERROR(VLOOKUP(H11")
        assert "SUMIFS" in str(ws.cell(CATEGORY_FIRST_ROW, 10).value)
        total_row = CATEGORY_FIRST_ROW + 2
        assert ws.cell(total_row, 8).value == "TOTAL"
        # idempotent
        assert sync_dashboard_categories(wb, cats) == 0

    def test_cycle_dashboard_uses_shared_block(self, excel_path):
        from cycle_dashboard import ensure_cycle_dashboard, _read_category_list
        from excel_schema import sync_dashboard_categories
        wb = load_workbook(excel_path)
        sync_dashboard_categories(wb, ["Food", "Rent"])
        ws = ensure_cycle_dashboard(wb)
        assert _read_category_list(ws) == ["Food", "Rent"]
        assert "$N$3" in str(ws.cell(11, 10).value)  # cycle-bounded Actual


# ── conversation flow ────────────────────────────────────────────────────────

class TestSetupEntry:
    def test_fresh_setup_creates_file_and_shows_review(self, tmp_path, monkeypatch):
        import settings
        monkeypatch.setattr(file_storage, "TEMPLATE_PATH", tmp_path / "no_t.xlsx")
        path = tmp_path / "new.xlsx"
        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", path)
        # Simulate no SQLite DB yet (new equivalent of no Excel file)
        monkeypatch.setattr(settings, "SQLITE_DB_PATH", tmp_path / "nonexistent.db")
        upd, ctx = make_update(), make_ctx()
        state = run(cmd_setup(upd, ctx))
        assert state == SETUP_REVIEW
        out = replies(upd)
        assert "Setup started" in out and "Your categories" in out
        assert len(ctx.user_data["setup"]["categories"]) == len(DEFAULT_CATEGORIES)

    def test_setup_on_configured_workbook_asks_confirmation(self, excel_path):
        upd, ctx = make_update(), make_ctx()
        state = run(cmd_setup(upd, ctx))
        assert state == SETUP_WELCOME
        assert "Setup already ran" in replies(upd)

    def test_setup_create_failure_message(self, tmp_path, monkeypatch):
        import settings
        import sqlite_ops
        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", tmp_path / "x.xlsx")
        # Simulate no SQLite DB and make init_db fail
        monkeypatch.setattr(settings, "SQLITE_DB_PATH", tmp_path / "nonexistent.db")
        monkeypatch.setattr(sqlite_ops, "init_db", MagicMock(side_effect=OSError("nope")))
        upd, ctx = make_update(), make_ctx()
        state = run(cmd_setup(upd, ctx))
        assert state == ConversationHandler.END
        assert "Could not create the budget file" in replies(upd)

    def test_setup_is_owner_only(self):
        # test_handlers_full patches config.auth_write to a pass-through for
        # the whole session, so the gate can't be exercised at runtime here.
        # Assert the wiring instead: cmd_setup must carry @auth_write.
        import inspect
        import handlers.setup_conv as mod
        src = inspect.getsource(mod)
        assert "@auth_write\n@log_call()\nasync def cmd_setup" in src

    def test_stale_session_with_renames_warns_discarded(self, excel_path):
        stale = fresh_session()
        stale["renames"] = [("Groceries", "Food")]
        upd, ctx = make_update(), make_ctx(stale)
        run(cmd_setup(upd, ctx))
        out = replies(upd)
        assert "still open" in out

    def test_stale_session_no_renames_shows_restart(self, excel_path):
        stale = fresh_session()  # renames == []
        upd, ctx = make_update(), make_ctx(stale)
        run(cmd_setup(upd, ctx))
        out = replies(upd)
        assert "↩️ Restarting setup." in out
        assert "discarded" not in out

    def test_no_stale_session_no_warning(self, excel_path):
        upd, ctx = make_update(), make_ctx()  # empty user_data
        run(cmd_setup(upd, ctx))
        out = replies(upd)
        assert "still open" not in out
        assert "Restarting" not in out

    def test_start_with_file_falls_through_to_menu(self, excel_path):
        upd, ctx = make_update(), make_ctx()
        with patch("handlers.menu.cmd_menu", new_callable=AsyncMock) as menu:
            state = run(cmd_start_or_setup(upd, ctx))
        assert state == ConversationHandler.END
        menu.assert_awaited_once()

    def test_start_without_file_enters_setup(self, tmp_path, monkeypatch):
        import settings
        monkeypatch.setattr(file_storage, "TEMPLATE_PATH", tmp_path / "no_t.xlsx")
        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", tmp_path / "n.xlsx")
        # Simulate no SQLite DB yet
        monkeypatch.setattr(settings, "SQLITE_DB_PATH", tmp_path / "nonexistent.db")
        upd, ctx = make_update(), make_ctx()
        state = run(cmd_start_or_setup(upd, ctx))
        assert state == SETUP_REVIEW

    def test_welcome_continue_loads_existing_state(self, excel_path):
        upd, ctx = make_update(callback_data="setup:continue"), make_ctx()
        state = run(setup_welcome_cb(upd, ctx))
        assert state == SETUP_REVIEW
        cats = ctx.user_data["setup"]["categories"]
        assert ("Groceries", "Expense") in cats  # from fallback workbook Lists

    def test_welcome_cancel(self, excel_path):
        upd, ctx = make_update(callback_data="setup:cancel"), make_ctx()
        state = run(setup_welcome_cb(upd, ctx))
        assert state == ConversationHandler.END
        assert "Setup stopped" in replies(upd)


class TestCategorySteps:
    def test_rename_flow(self, excel_path):
        session = fresh_session()
        upd, ctx = make_update(callback_data="setup:rename"), make_ctx(session)
        assert run(setup_review_cb(upd, ctx)) == SETUP_RENAME

        upd2 = make_update(callback_data="setup:ren:3")  # Groceries
        assert run(setup_rename_cb(upd2, ctx)) == SETUP_RENAME
        assert session["pending_rename"] == 3

        upd3 = make_update(text="Food")
        assert run(setup_rename_text(upd3, ctx)) == SETUP_REVIEW
        assert session["categories"][3] == ("Food", "Expense")
        assert "✅ *Groceries* → *Food*" in replies(upd3)
        assert session["renames"] == [("Groceries", "Food")]

    def test_two_sequential_renames_accumulate(self, excel_path):
        session = fresh_session()
        ctx = make_ctx(session)

        # First rename: Groceries (index 3) → Food
        upd1 = make_update(callback_data="setup:ren:3")
        run(setup_rename_cb(upd1, ctx))
        upd2 = make_update(text="Food")
        run(setup_rename_text(upd2, ctx))

        # Second rename: Transport (index 4) → Commute
        upd3 = make_update(callback_data="setup:ren:4")
        run(setup_rename_cb(upd3, ctx))
        upd4 = make_update(text="Commute")
        run(setup_rename_text(upd4, ctx))

        assert ("Groceries", "Food") in session["renames"]
        assert ("Transport", "Commute") in session["renames"]
        assert len(session["renames"]) == 2

    def test_cancel_mid_rename_discards_pending(self, excel_path):
        session = fresh_session()
        session["pending_rename"] = 2
        upd, ctx = make_update(callback_data="setup:cancel"), make_ctx(session)
        assert run(setup_rename_cb(upd, ctx)) == ConversationHandler.END
        assert "setup" not in ctx.user_data

    def test_add_flow(self, excel_path):
        session = fresh_session()
        ctx = make_ctx(session)
        upd = make_update(text="Pets")
        assert run(setup_add_text(upd, ctx)) == SETUP_ADD
        assert "What type is *Pets*?" in replies(upd)

        upd2 = make_update(callback_data="setup:type:Expense")
        assert run(setup_add_type_cb(upd2, ctx)) == SETUP_REVIEW
        assert ("Pets", "Expense") in session["categories"]
        assert "✅ Added *Pets* (Expense)." in replies(upd2)

    def test_done_commits_categories_and_starts_budgets(self, excel_path):
        import storage_facade
        session = fresh_session()
        upd, ctx = make_update(callback_data="setup:done"), make_ctx(session)
        state = run(setup_review_cb(upd, ctx))
        assert state == SETUP_BUDGET
        assert session["budget_queue"][0] == "Housing"  # first Expense category
        # Categories are now committed to SQLite via storage_facade
        cats = storage_facade.load_reference_data()["categories"]
        expected = [n for n, _t in DEFAULT_CATEGORIES]
        assert cats == expected

    def test_done_with_no_expense_categories_skips_budget(self, excel_path):
        session = fresh_session()
        session["categories"] = [("Salary", "Income"), ("Savings", "Savings")]
        upd, ctx = make_update(callback_data="setup:done"), make_ctx(session)
        state = run(setup_review_cb(upd, ctx))
        assert state == SETUP_CURRENCY
        assert "Pick your main currency" in replies(upd)


class TestBudgetStep:
    def test_budget_amounts_and_no_limit(self, excel_path):
        session = fresh_session()
        session["budget_queue"] = ["Housing", "Groceries"]
        ctx = make_ctx(session)

        upd = make_update(text="2500")
        assert run(setup_budget_text(upd, ctx)) == SETUP_BUDGET
        assert session["budgets"]["Housing"] == 2500
        assert "✅ *Housing*: 2500" in replies(upd)

        upd2 = make_update(text="0")
        state = run(setup_budget_text(upd2, ctx))
        assert "✅ *Groceries*: no limit" in replies(upd2)
        assert "Groceries" not in session["budgets"]
        assert state == SETUP_CURRENCY  # queue drained → currency step

    def test_budget_rejects_non_numeric(self, excel_path):
        session = fresh_session()
        session["budget_queue"] = ["Housing"]
        ctx = make_ctx(session)
        upd = make_update(text="lots")
        assert run(setup_budget_text(upd, ctx)) == SETUP_BUDGET
        assert "❌ Numbers only" in replies(upd)
        assert session["budget_queue"] == ["Housing"]

    def test_budget_accepts_plain_integer(self, excel_path):
        session = fresh_session()
        session["budget_queue"] = ["Housing"]
        ctx = make_ctx(session)
        upd = make_update(text="100")
        state = run(setup_budget_text(upd, ctx))
        assert state == SETUP_CURRENCY  # queue drained
        assert session["budgets"]["Housing"] == 100
        assert "❌" not in replies(upd)


class TestCurrencyAndSummary:
    def _patched_rates(self):
        return patch.object(
            setup_conv, "_refresh_rates_best_effort",
            new=AsyncMock(return_value={"USD": 3.9, "EUR": 4.3, "PLN": 1.0}))

    def test_pick_currency_button_finishes(self, excel_path):
        session = fresh_session()
        session["budgets"] = {"Housing": 2500.0}
        upd, ctx = make_update(callback_data="setup:ccy:USD"), make_ctx(session)
        with self._patched_rates():
            state = run(setup_currency_cb(upd, ctx))
        assert state == SETUP_SUMMARY
        out = replies(upd)
        assert "Setup complete" in out and "*Currency:* USD" in out
        assert "Housing — 2500" in out and "no limit" in out
        # budgets + currency now persisted in SQLite via storage_facade
        import storage_facade
        assert storage_facade.load_budgets().get("Housing") == 2500.0
        assert "USD" in storage_facade.load_rates()

    def test_other_currency_valid_code(self, excel_path):
        session = fresh_session()
        ctx = make_ctx(session)
        upd = make_update(callback_data="setup:ccy:Other")
        with self._patched_rates():
            assert run(setup_currency_cb(upd, ctx)) == SETUP_CURRENCY
            assert "3-letter currency code" in replies(upd)
            upd2 = make_update(text="gbp")
            state = run(setup_currency_text(upd2, ctx))
        assert state == SETUP_SUMMARY
        assert session["currency"] == "GBP"

    def test_other_currency_invalid_code(self, excel_path):
        upd, ctx = make_update(text="EURO"), make_ctx(fresh_session())
        assert run(setup_currency_text(upd, ctx)) == SETUP_CURRENCY
        assert "Not a valid 3-letter code" in replies(upd)

    def test_unknown_live_rate_warns(self, excel_path):
        session = fresh_session()
        upd, ctx = make_update(callback_data="setup:ccy:TRY"), make_ctx(session)
        with patch.object(setup_conv, "_refresh_rates_best_effort",
                          new=AsyncMock(return_value={"PLN": 1.0})):
            state = run(setup_currency_cb(upd, ctx))
        assert state == SETUP_SUMMARY
        assert "No live rate found for TRY" in replies(upd)

    def test_rate_fetch_failure_warns(self, excel_path):
        session = fresh_session()
        upd, ctx = make_update(callback_data="setup:ccy:EUR"), make_ctx(session)
        with patch.object(setup_conv, "_refresh_rates_best_effort",
                          new=AsyncMock(side_effect=RuntimeError("offline"))):
            state = run(setup_currency_cb(upd, ctx))
        assert state == SETUP_SUMMARY
        assert "Could not fetch live exchange rates" in replies(upd)

    def test_summary_done_ends(self, excel_path):
        upd, ctx = make_update(callback_data="setup:finish"), make_ctx(fresh_session())
        assert run(setup_summary_cb(upd, ctx)) == ConversationHandler.END
        assert "setup" not in ctx.user_data


class TestHandlerFactory:
    def test_factory_shape(self):
        handler = setup_conversation_handler()
        cmds = set()
        for entry in handler.entry_points:
            cmds |= set(entry.commands)
        assert cmds == {"setup", "start"}
        assert set(handler.states) == {
            SETUP_WELCOME, SETUP_REVIEW, SETUP_RENAME, SETUP_ADD,
            SETUP_BUDGET, SETUP_CURRENCY, SETUP_SUMMARY, SETUP_REMOVE,
        }
        fallback_cmds = {cmd for h in handler.fallbacks for cmd in getattr(h, "commands", [])}
        assert fallback_cmds == {"cancel", "setup"}

    def test_cancel_fallback(self, excel_path):
        upd, ctx = make_update(text="/cancel"), make_ctx(fresh_session())
        assert run(setup_cancel(upd, ctx)) == ConversationHandler.END
        assert "Setup stopped" in replies(upd)


# ── remove-category flow ──────────────────────────────────────────────────────

def _make_remove_ctx(categories=None, budgets=None):
    """Return a minimal fake context with a pre-populated session."""
    ctx = MagicMock()
    ctx.user_data = {}
    s = _session(ctx)
    s["categories"] = list(categories or [("Groceries", "Expense"), ("Housing", "Expense")])
    s["budgets"] = dict(budgets or {"Groceries": 200.0, "Housing": 500.0})
    return ctx


def _make_remove_update(callback_data: str):
    """Return a fake Update whose callback_query carries callback_data."""
    query = MagicMock()
    query.answer = AsyncMock()
    query.data = callback_data
    query.message = MagicMock()
    query.message.reply_text = AsyncMock()
    update = MagicMock()
    update.callback_query = query
    update.message = query.message
    update.effective_user = MagicMock()
    return update


class TestRemoveCategoryFlow:
    def test_remove_shows_picker(self):
        """setup:remove with 2 categories transitions to SETUP_REMOVE."""
        ctx = _make_remove_ctx()
        upd = _make_remove_update("setup:remove")
        result = run(setup_review_cb(upd, ctx))
        assert result == SETUP_REMOVE
        upd.callback_query.message.reply_text.assert_awaited_once()
        call_kwargs = upd.callback_query.message.reply_text.call_args
        assert "remove" in call_kwargs[0][0].lower() or "tap" in call_kwargs[0][0].lower()

    def test_remove_guard_single_category(self):
        """setup:remove with only 1 category shows error and stays on SETUP_REVIEW."""
        ctx = _make_remove_ctx(categories=[("Groceries", "Expense")], budgets={})
        upd = _make_remove_update("setup:remove")
        result = run(setup_review_cb(upd, ctx))
        assert result == SETUP_REVIEW
        first_call = upd.callback_query.message.reply_text.call_args_list[0]
        assert "1 category" in first_call[0][0] or "at least" in first_call[0][0]

    def test_remove_deletes_category_and_budget(self):
        """setup:del:0 removes the first category and its budget, returns SETUP_REVIEW."""
        ctx = _make_remove_ctx(
            categories=[("Groceries", "Expense"), ("Housing", "Expense")],
            budgets={"Groceries": 200.0, "Housing": 500.0},
        )
        upd = _make_remove_update("setup:del:0")
        result = run(setup_remove_cb(upd, ctx))
        assert result == SETUP_REVIEW
        session = _session(ctx)
        names = [n for n, _ in session["categories"]]
        assert "Groceries" not in names
        assert "Housing" in names
        assert "Groceries" not in session["budgets"]
        assert session["budgets"].get("Housing") == 500.0
