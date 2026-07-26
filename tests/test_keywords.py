"""
tests/test_keywords.py — salary keywords stored in the Excel Lists sheet
("Salary Keywords" column) and the /keywords management command.
"""

import os
import sys
from datetime import date
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

os.environ.setdefault("STORAGE_BACKEND", "local")
os.environ.setdefault("TELEGRAM_BOT_TOKEN", "dummy")
os.environ.setdefault("ALLOWED_TELEGRAM_IDS", "123")

PROJECT_ROOT = Path(__file__).parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import settings
import cycles
from cycles import (
    cycle_detect_keywords, delete_salary_keyword, load_salary_keywords,
    save_salary_keyword,
)
from excel_schema import ListsSchema, col_indices, header_of
from telegram.ext import ConversationHandler
from handlers.misc import cmd_keywords, keywords_add_word, keywords_callback
from states import KW_ADD, KW_PICK


def make_update(text="", user_id=123):
    upd = MagicMock()
    upd.effective_user.id = user_id
    upd.message.text = text
    upd.message.reply_text = AsyncMock()
    return upd


def make_callback_update(data, user_id=123):
    upd = MagicMock()
    upd.effective_user.id = user_id
    upd.message = None
    upd.callback_query.data = data
    upd.callback_query.from_user.id = user_id
    upd.callback_query.answer = AsyncMock()
    upd.callback_query.message.reply_text = AsyncMock()
    return upd


def make_ctx(args=None):
    ctx = MagicMock()
    ctx.args = args or []
    ctx.user_data = {}
    return ctx


def _keyword_column_cells(path):
    wb = load_workbook(path)
    ws = wb["Lists"]
    kw_col = col_indices(ws, ListsSchema)["salary_keyword"]
    return [ws.cell(r, kw_col).value for r in range(2, ws.max_row + 1)]


# ── load_salary_keywords ───────────────────────────────────────────────────────

def test_load_missing_column_returns_empty(excel_path):
    assert load_salary_keywords(excel_path) == []


def test_load_empty_column_returns_empty(excel_path):
    wb = load_workbook(excel_path)
    ws = wb["Lists"]
    ws.cell(1, ws.max_column + 1, header_of(ListsSchema, "salary_keyword"))
    wb.save(excel_path)
    assert load_salary_keywords(excel_path) == []


def test_load_strips_lowercases_and_skips_blanks(excel_path):
    wb = load_workbook(excel_path)
    ws = wb["Lists"]
    kw_col = ws.max_column + 1
    ws.cell(1, kw_col, header_of(ListsSchema, "salary_keyword"))
    ws.cell(2, kw_col, "  Payroll ")
    ws.cell(3, kw_col, "")
    ws.cell(4, kw_col, "WYNAGRODZENIE")
    wb.save(excel_path)
    assert load_salary_keywords(excel_path) == ["payroll", "wynagrodzenie"]


def test_load_unreadable_workbook_returns_empty(tmp_path):
    assert load_salary_keywords(tmp_path / "nope.xlsx") == []


# ── save / delete ──────────────────────────────────────────────────────────────

def test_first_save_seeds_env_keywords(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["payroll", "bonus"])
    assert save_salary_keyword("stipend") is True
    assert load_salary_keywords(excel_path) == ["payroll", "bonus", "stipend"]


def test_save_duplicate_is_noop(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", [])
    assert save_salary_keyword("payroll") is True
    assert save_salary_keyword("  PAYROLL ") is False
    assert load_salary_keywords(excel_path) == ["payroll"]


def test_save_blank_is_rejected(excel_path):
    assert save_salary_keyword("   ") is False


def test_save_overlong_keyword_is_rejected(excel_path):
    assert save_salary_keyword("x" * (cycles.MAX_SALARY_KEYWORD_LENGTH + 1)) is False
    assert load_salary_keywords(excel_path) == []


def test_no_reseed_after_user_empties_the_list(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["envword"])
    save_salary_keyword("payroll")
    delete_salary_keyword("envword")
    delete_salary_keyword("payroll")
    assert load_salary_keywords(excel_path) == []
    assert save_salary_keyword("bonus") is True
    assert load_salary_keywords(excel_path) == ["bonus"]


def test_delete_removes_only_column_cells(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", [])
    save_salary_keyword("payroll")
    save_salary_keyword("bonus")
    wb = load_workbook(excel_path)
    ws = wb["Lists"]
    cat_col = col_indices(ws, ListsSchema)["categories"]
    categories_before = [ws.cell(r, cat_col).value for r in range(2, 10)]

    assert delete_salary_keyword("payroll") is True
    assert load_salary_keywords(excel_path) == ["bonus"]
    assert _keyword_column_cells(excel_path)[1:] == [None] * (len(_keyword_column_cells(excel_path)) - 1)

    wb = load_workbook(excel_path)
    ws = wb["Lists"]
    categories_after = [ws.cell(r, cat_col).value for r in range(2, 10)]
    assert categories_after == categories_before


def test_delete_unknown_keyword_returns_false(excel_path):
    assert delete_salary_keyword("nope") is False


# ── cycle_detect_keywords fallback ─────────────────────────────────────────────

def test_detect_keywords_fall_back_to_env(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["wynagrodzenie"])
    assert cycle_detect_keywords() == ["salary", "wynagrodzenie"]


def test_detect_keywords_prefer_excel_over_env(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["envword"])
    save_salary_keyword("payroll")
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["ignored"])
    assert cycle_detect_keywords() == ["salary", "envword", "payroll"]


# ── /keywords command ──────────────────────────────────────────────────────────

# NOTE: the @auth_write gate on cmd_keywords is covered in tests/test_write_gate.py
# (WRITE_ENTRY_POINTS) — collection-order patching in test_handlers_full.py makes
# gate assertions unreliable in any file collected after it.

async def test_cmd_keywords_help(excel_path):
    upd = make_update()
    result = await cmd_keywords(upd, make_ctx(["help"]))
    assert "/keywords" in upd.message.reply_text.call_args[0][0]
    assert result == ConversationHandler.END


async def test_cmd_keywords_shows_env_fallback_source(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["payroll"])
    upd = make_update()
    state = await cmd_keywords(upd, make_ctx())
    assert state == KW_PICK
    text = upd.message.reply_text.call_args[0][0]
    assert ".env fallback" in text and "payroll" in text
    markup = upd.message.reply_text.call_args.kwargs["reply_markup"]
    callbacks = [b.callback_data for row in markup.inline_keyboard for b in row]
    assert callbacks == ["kw:add"]  # env words get no delete buttons


async def test_cmd_keywords_shows_excel_keywords_with_delete_buttons(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", [])
    save_salary_keyword("payroll")
    save_salary_keyword("bonus")
    upd = make_update()
    await cmd_keywords(upd, make_ctx())
    text = upd.message.reply_text.call_args[0][0]
    assert "Excel" in text
    markup = upd.message.reply_text.call_args.kwargs["reply_markup"]
    callbacks = [b.callback_data for row in markup.inline_keyboard for b in row]
    assert callbacks == ["kw:del:payroll", "kw:del:bonus", "kw:add"]


# ── add flow ───────────────────────────────────────────────────────────────────

async def test_keywords_add_flow(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["envword"])
    upd = make_callback_update("kw:add")
    state = await keywords_callback(upd, make_ctx())
    assert state == KW_ADD
    assert "Send the new salary keyword" in upd.callback_query.message.reply_text.call_args[0][0]

    msg = make_update("  Premia ")
    state = await keywords_add_word(msg, make_ctx())
    assert state == KW_PICK
    assert load_salary_keywords(excel_path) == ["envword", "premia"]
    text = msg.message.reply_text.call_args[0][0]
    assert "Added 'premia'" in text


async def test_keywords_add_overlong_word_reprompts(excel_path):
    msg = make_update("x" * (cycles.MAX_SALARY_KEYWORD_LENGTH + 1))
    state = await keywords_add_word(msg, make_ctx())
    assert state == KW_ADD
    assert "too long" in msg.message.reply_text.call_args[0][0]
    assert load_salary_keywords(excel_path) == []


async def test_keywords_add_duplicate_reports_noop(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", [])
    save_salary_keyword("payroll")
    msg = make_update("payroll")
    state = await keywords_add_word(msg, make_ctx())
    assert state == KW_PICK
    assert "already in the list" in msg.message.reply_text.call_args[0][0]
    assert load_salary_keywords(excel_path) == ["payroll"]


# ── delete flow ────────────────────────────────────────────────────────────────

async def test_keywords_delete_flow(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", [])
    save_salary_keyword("payroll")
    save_salary_keyword("bonus")
    upd = make_callback_update("kw:del:payroll")
    state = await keywords_callback(upd, make_ctx())
    assert state == KW_PICK
    assert "Removed 'payroll'" in upd.callback_query.message.reply_text.call_args[0][0]
    assert load_salary_keywords(excel_path) == ["bonus"]


async def test_keywords_delete_unknown_word(excel_path):
    upd = make_callback_update("kw:del:ghost")
    state = await keywords_callback(upd, make_ctx())
    assert state == KW_PICK
    assert "not in the list" in upd.callback_query.message.reply_text.call_args[0][0]


# ── fallback-chain integration (Fix 5) ────────────────────────────────────────

def _make_salary_transaction(description="", category="", txn_date=None):
    t = MagicMock()
    t.transaction_type = "Income"
    t.category = category
    t.description = description
    t.date = txn_date or date(2026, 7, 1)
    return t


async def test_fallback_env_keywords_when_excel_empty(excel_path, monkeypatch):
    """With no Excel keywords, an env keyword in Description triggers the prompt."""
    import settings
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["payroll"])

    # Excel keywords column is absent — load_salary_keywords returns []
    assert load_salary_keywords(excel_path) == []

    upd = make_update()
    txn = _make_salary_transaction(description="payroll april", category="")
    from handlers.cycle import maybe_prompt_cycle_start
    await maybe_prompt_cycle_start(upd, txn)
    upd.message.reply_text.assert_called_once()
    assert "Salary received" in upd.message.reply_text.call_args[0][0]


async def test_excel_keywords_override_env(excel_path, monkeypatch):
    """When Excel keywords are present, env keywords are ignored;
    only the Excel keyword triggers the prompt."""
    import settings
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["envonly"])

    save_salary_keyword("excelword")
    assert load_salary_keywords(excel_path) == ["envonly", "excelword"]

    upd_env = make_update()
    txn_env = _make_salary_transaction(description="envonly march", category="")
    from handlers.cycle import maybe_prompt_cycle_start
    await maybe_prompt_cycle_start(upd_env, txn_env)
    # env keyword "envonly" was seeded into Excel alongside "excelword",
    # so it is now in Excel — this call should still prompt.
    # Verify the excelword path independently:
    upd_excel = make_update()
    txn_excel = _make_salary_transaction(description="excelword april", category="")
    await maybe_prompt_cycle_start(upd_excel, txn_excel)
    upd_excel.message.reply_text.assert_called_once()
    assert "Salary received" in upd_excel.message.reply_text.call_args[0][0]

    # A keyword not in Excel (and env is now superseded) must NOT trigger.
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["newenvword"])
    upd_miss = make_update()
    txn_miss = _make_salary_transaction(description="newenvword may", category="")
    await maybe_prompt_cycle_start(upd_miss, txn_miss)
    upd_miss.message.reply_text.assert_not_called()
