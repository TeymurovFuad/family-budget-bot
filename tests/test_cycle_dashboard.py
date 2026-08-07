"""
tests/test_cycle_dashboard.py — Cycle Dashboard sheet creation and the
Dashboard→Cycle Dashboard category sync. All workbooks are in-memory
openpyxl objects; no file I/O anywhere.
"""

import os
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

os.environ.setdefault("STORAGE_BACKEND", "local")
os.environ.setdefault("TELEGRAM_BOT_TOKEN", "dummy")
os.environ.setdefault("ALLOWED_TELEGRAM_IDS", "123")

PROJECT_ROOT = Path(__file__).parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from cycle_dashboard import (
    CYCLE_DASHBOARD_SHEET_NAME,
    ensure_cycle_dashboard,
    sync_cycle_dashboard_categories,
)

FIRST_CAT_ROW = 11  # H11


def make_dashboard(wb, categories):
    """Minimal Dashboard sheet: category names in H11.., TOTAL row after."""
    ws = wb.create_sheet("Dashboard")
    r = FIRST_CAT_ROW
    for cat in categories:
        ws.cell(r, 8, cat)
        r += 1
    ws.cell(r, 8, "TOTAL")
    return ws


def read_cd_categories(ws):
    cats = []
    for r in range(FIRST_CAT_ROW, ws.max_row + 1):
        v = str(ws.cell(r, 8).value or "").strip()
        if v == "TOTAL":
            return cats
        if v:
            cats.append(v)
    return None  # no TOTAL row found


def find_total_row(ws):
    for r in range(FIRST_CAT_ROW, ws.max_row + 1):
        if str(ws.cell(r, 8).value or "").strip() == "TOTAL":
            return r
    return None


# ── ensure_cycle_dashboard ─────────────────────────────────────────────────────

def test_ensure_creates_sheet():
    wb = Workbook()
    ws = ensure_cycle_dashboard(wb)

    assert CYCLE_DASHBOARD_SHEET_NAME in wb.sheetnames
    # No Cycles ledger → B2 blank
    assert not (ws["B2"].value or "")
    # N3 helper formula present
    assert "MATCH($B$2,Cycles!" in str(ws["N3"].value)
    # TOTAL row present (no Dashboard → 0 categories, TOTAL at H11)
    assert find_total_row(ws) == FIRST_CAT_ROW
    # Structure labels
    assert ws["A1"].value == "⚙ Cycle Filter"
    assert ws["A15"].value == "Cycle Days"
    assert ws["H10"].value == "Category"


def test_ensure_idempotent():
    wb = Workbook()
    make_dashboard(wb, ["Groceries", "Rent"])
    ws1 = ensure_cycle_dashboard(wb)
    marker = "user-edited"
    ws1["B2"] = marker

    ws2 = ensure_cycle_dashboard(wb)
    assert ws2 is ws1
    assert ws2["B2"].value == marker
    assert read_cd_categories(ws2) == ["Groceries", "Rent"]


def test_ensure_copies_dashboard_categories_and_seeds_cycle():
    wb = Workbook()
    make_dashboard(wb, ["Groceries", "Rent", "Fun"])
    cyc = wb.create_sheet("Cycles")
    cyc.cell(1, 1, "StartDate")
    cyc.cell(1, 2, "Label")
    from datetime import date
    cyc.cell(2, 1, date(2026, 6, 25)); cyc.cell(2, 2, "Jun 2026")
    cyc.cell(3, 1, date(2026, 7, 24)); cyc.cell(3, 2, "Jul 2026")

    ws = ensure_cycle_dashboard(wb)
    assert read_cd_categories(ws) == ["Groceries", "Rent", "Fun"]
    assert ws["B2"].value == "Jul 2026"  # latest cycle seeded
    # Per-row formulas reference the right row
    assert "VLOOKUP(H11," in ws["I11"].value
    assert "SUMIFS(MasterData!" in ws["J11"].value
    assert ws["K12"].value == "=I12-J12"
    # TOTAL sums
    total = find_total_row(ws)
    assert ws.cell(total, 9).value == f"=SUM(I11:I{total - 1})"


# ── sync_cycle_dashboard_categories ────────────────────────────────────────────

def test_sync_no_change():
    wb = Workbook()
    make_dashboard(wb, ["Groceries", "Rent"])
    ensure_cycle_dashboard(wb)
    assert sync_cycle_dashboard_categories(wb) == 0


def test_sync_adds_missing():
    wb = Workbook()
    make_dashboard(wb, ["Groceries", "Rent"])
    ensure_cycle_dashboard(wb)

    # Dashboard grows by one category
    dash = wb["Dashboard"]
    total = find_total_row(dash)
    dash.cell(total, 8, "Travel")
    dash.cell(total + 1, 8, "TOTAL")

    count = sync_cycle_dashboard_categories(wb)
    assert count == 3
    ws = wb[CYCLE_DASHBOARD_SHEET_NAME]
    assert read_cd_categories(ws) == ["Groceries", "Rent", "Travel"]
    # New row has formulas, TOTAL moved down
    assert "SUMIFS(MasterData!" in ws["J13"].value
    assert find_total_row(ws) == FIRST_CAT_ROW + 3


def test_sync_removes_extra():
    wb = Workbook()
    make_dashboard(wb, ["Groceries", "Rent", "Fun"])
    ensure_cycle_dashboard(wb)

    # Dashboard shrinks to one category
    dash = wb["Dashboard"]
    for r in range(FIRST_CAT_ROW, dash.max_row + 1):
        dash.cell(r, 8).value = None
    dash.cell(FIRST_CAT_ROW, 8, "Groceries")
    dash.cell(FIRST_CAT_ROW + 1, 8, "TOTAL")

    count = sync_cycle_dashboard_categories(wb)
    assert count == 1
    ws = wb[CYCLE_DASHBOARD_SHEET_NAME]
    assert read_cd_categories(ws) == ["Groceries"]
    assert find_total_row(ws) == FIRST_CAT_ROW + 1
    # No stale rows below the new TOTAL
    for r in range(FIRST_CAT_ROW + 2, ws.max_row + 1):
        assert ws.cell(r, 8).value in (None, "")


def test_sync_creates_if_missing():
    wb = Workbook()
    make_dashboard(wb, ["Groceries", "Rent"])
    assert CYCLE_DASHBOARD_SHEET_NAME not in wb.sheetnames

    count = sync_cycle_dashboard_categories(wb)
    assert count == 2
    assert CYCLE_DASHBOARD_SHEET_NAME in wb.sheetnames
    assert read_cd_categories(wb[CYCLE_DASHBOARD_SHEET_NAME]) == ["Groceries", "Rent"]
