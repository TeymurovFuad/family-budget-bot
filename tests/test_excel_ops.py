"""
test_excel_ops.py — tests for _do_append_transaction, delete_transaction_row,
and get_recent_transactions from excel_ops.py and file_storage.py.

All tests use real temp files via the excel_path fixture — no mocking of I/O.
"""

import datetime

import openpyxl
import pytest

from excel_ops import _do_append_transaction
from file_storage import (
    get_recent_transactions,
    delete_transaction_row,
    append_transactions_batch,
)
from models import Transaction


# ── _do_append_transaction ────────────────────────────────────────────────────


class TestDoAppendTransaction:

    def test_append_creates_data_row(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        # Header is row 1, first data row is row 2
        assert ws.max_row == 2

    def test_appended_value_matches_transaction(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        # Build header-to-column mapping
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        row = 2
        assert ws.cell(row, headers["Value"]).value == pytest.approx(150.50)
        assert ws.cell(row, headers["Category"]).value == "Groceries"
        assert ws.cell(row, headers["Person"]).value == "Alice"
        assert ws.cell(row, headers["Currency"]).value == "PLN"
        assert ws.cell(row, headers["Type"]).value == "Expense"
        assert ws.cell(row, headers["Description"]).value == "weekly shop"

    def test_appended_date_matches_transaction(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        # openpyxl reads dates as datetime.datetime when data_only=True
        cell_val = ws.cell(2, headers["Date"]).value
        if isinstance(cell_val, datetime.datetime):
            assert cell_val.date() == datetime.date(2024, 6, 15)
        else:
            assert cell_val == datetime.date(2024, 6, 15)

    def test_appended_year_matches_transaction(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, headers["Year"]).value == 2024

    def test_appended_month_matches_transaction(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, headers["Month"]).value == "Jun"

    def test_is_done_is_true(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, headers["IsDone"]).value is True

    def test_value_pln_column_contains_formula(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        # Load WITHOUT data_only to see the raw formula
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        formula = ws.cell(2, headers["Value (PLN)"]).value
        assert isinstance(formula, str)
        assert formula.startswith("=")

    def test_date_modified_column_contains_utc_datetime(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        value = ws.cell(2, headers["Date Modified (UTC)"]).value
        assert isinstance(value, datetime.datetime)

    def test_multiple_appends_produce_correct_row_count(self, excel_path):
        for i in range(1, 4):
            txn = Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 10),
                currency="PLN",
                transaction_type="Expense",
                category="Transport",
            )
            _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        # 1 header + 3 data rows
        assert ws.max_row == 4

    def test_eur_transaction_currency_stored(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 6, 15),
            value=80.0,
            currency="EUR",
            transaction_type="Expense",
            category="Dining Out",
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, headers["Currency"]).value == "EUR"


# ── delete_transaction_row ────────────────────────────────────────────────────


class TestDeleteTransactionRow:

    def test_delete_reduces_row_count(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        # Row 2 is the first data row
        delete_transaction_row(2)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        # Only header remains
        assert ws.max_row == 1

    def test_delete_middle_row_leaves_others_intact(self, excel_path):
        dates = [datetime.date(2024, 6, d) for d in (1, 2, 3)]
        categories = ["Groceries", "Transport", "Housing"]
        for d, c in zip(dates, categories):
            txn = Transaction(
                date=d, value=10.0, currency="PLN",
                transaction_type="Expense", category=c,
            )
            _do_append_transaction(txn)

        # Delete the second data row (Excel row 3 = Transport)
        delete_transaction_row(3)

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

        # Header + 2 remaining rows
        assert ws.max_row == 3
        # Row 2 should still be Groceries
        assert ws.cell(2, headers["Category"]).value == "Groceries"
        # Row 3 should now be Housing (Transport was deleted)
        assert ws.cell(3, headers["Category"]).value == "Housing"


# ── get_recent_transactions ───────────────────────────────────────────────────


class TestGetRecentTransactions:

    def test_returns_empty_list_on_blank_excel(self, excel_path):
        result = get_recent_transactions(excel_path)
        assert result == []

    def test_returns_appended_transactions(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        result = get_recent_transactions(excel_path)
        assert len(result) == 1

    def test_each_result_has_row_idx(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        result = get_recent_transactions(excel_path)
        assert "_row_idx" in result[0]

    def test_row_idx_is_correct(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        result = get_recent_transactions(excel_path)
        # First data row is Excel row 2
        assert result[0]["_row_idx"] == 2

    def test_returns_at_most_n_results(self, excel_path):
        for i in range(1, 8):
            txn = Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i),
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            )
            _do_append_transaction(txn)
        result = get_recent_transactions(excel_path, n=5)
        assert len(result) == 5

    def test_returns_last_n_rows(self, excel_path):
        for i in range(1, 8):
            txn = Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 10),
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            )
            _do_append_transaction(txn)
        result = get_recent_transactions(excel_path, n=3)
        # Last 3 transactions have values 50, 60, 70
        values = [r["Value"] for r in result]
        assert values == [50.0, 60.0, 70.0]

    def test_result_contains_correct_category(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        result = get_recent_transactions(excel_path)
        assert result[0]["Category"] == "Groceries"

    def test_result_contains_correct_value(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        result = get_recent_transactions(excel_path)
        assert result[0]["Value"] == pytest.approx(150.50)


# ── append_transactions_batch ─────────────────────────────────────────────────


class TestAppendTransactionsBatch:

    def test_batch_append_writes_all_rows(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 25),
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            )
            for i in range(1, 6)
        ]
        append_transactions_batch(transactions)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        # 1 header + 5 data rows
        assert ws.max_row == 6

    def test_batch_append_empty_list_does_nothing(self, excel_path):
        append_transactions_batch([])
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        assert ws.max_row == 1  # header only

    def test_batch_append_rows_have_formulas(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, 1),
                value=100.0,
                currency="PLN",
                transaction_type="Income",
            )
        ]
        append_transactions_batch(transactions)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        formula = ws.cell(2, headers["Value (PLN)"]).value
        assert isinstance(formula, str)
        assert formula.startswith("=")

    def test_batch_value_pln_formula_references_i2_j100(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, 1),
                value=80.0,
                currency="EUR",
                transaction_type="Expense",
                category="Travel",
            )
        ]
        append_transactions_batch(transactions)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        formula = ws.cell(2, headers["Value (PLN)"]).value
        assert "$H$2:$I$100" in formula

    def test_batch_date_modified_is_datetime_not_string(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, 1),
                value=55.0,
                currency="PLN",
                transaction_type="Income",
            )
        ]
        append_transactions_batch(transactions)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        val = ws.cell(2, headers["Date Modified (UTC)"]).value
        assert isinstance(val, datetime.datetime)


# ── replay_recovery_queue (regression: phantom-row bug + re-queue on failure) ─

class TestReplayRecoveryQueue:

    def test_replay_writes_at_next_data_row_not_max_row(self, excel_path):
        """max_row inflated by styled empty rows must not push replayed rows down."""
        from excel_ops import replay_recovery_queue
        from file_storage import append_to_recovery_queue

        # Inflate max_row with a styled-but-empty faraway cell
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        ws.cell(500, 1).fill = openpyxl.styles.PatternFill("solid", fgColor="FFFF00")
        wb.save(excel_path)

        append_to_recovery_queue({
            "date": datetime.date(2024, 6, 1), "year": 2024, "month": "Jun",
            "value": 42.0, "type": "Expense", "category": "Groceries",
            "description": "recovered", "person": "", "is_recurring": False,
            "currency": "PLN",
        })
        replay_recovery_queue()

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, headers["Value"]).value == 42.0  # row 2, not row 501

    def test_replay_uses_schema_vlookup_range(self, excel_path):
        from excel_ops import replay_recovery_queue
        from file_storage import append_to_recovery_queue

        append_to_recovery_queue({
            "date": datetime.date(2024, 6, 2), "year": 2024, "month": "Jun",
            "value": 10.0, "type": "Expense", "category": "Groceries",
            "description": "x", "person": "", "is_recurring": False,
            "currency": "EUR",
        })
        replay_recovery_queue()
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        formula = ws.cell(2, headers["Value (PLN)"]).value
        assert "$H$2:$I$100" in formula
        assert "$I$2:$J$100" not in formula

    def test_replay_empty_queue_is_noop(self, excel_path):
        from excel_ops import replay_recovery_queue
        replay_recovery_queue()  # must not raise

    def test_replay_deletes_queue_file_after_success(self, excel_path, monkeypatch, tmp_path):
        import file_storage
        from excel_ops import replay_recovery_queue
        from file_storage import append_to_recovery_queue

        queue_path = tmp_path / "recovery_queue.json"
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)
        import excel_ops
        monkeypatch.setattr(excel_ops, "RECOVERY_QUEUE_PATH", queue_path, raising=False)

        append_to_recovery_queue({
            "date": datetime.date(2024, 6, 1), "year": 2024, "month": "Jun",
            "value": 42.0, "type": "Expense", "category": "Groceries",
            "description": "recovered", "person": "", "is_recurring": False,
            "currency": "PLN",
        })
        assert queue_path.exists()
        replay_recovery_queue()
        assert not queue_path.exists()

    def test_corrupt_queue_file_is_quarantined_not_raised(self, excel_path, monkeypatch, tmp_path):
        import file_storage
        from file_storage import flush_recovery_queue

        queue_path = tmp_path / "recovery_queue.json"
        queue_path.write_text("{not valid json")
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        rows = flush_recovery_queue()  # must not raise

        assert rows == []
        assert not queue_path.exists()
        corrupt_path = queue_path.with_name(queue_path.name + ".corrupt")
        assert corrupt_path.exists()
        assert corrupt_path.read_text() == "{not valid json"

    def test_replay_startup_survives_corrupt_queue_file(self, excel_path, monkeypatch, tmp_path):
        """replay_recovery_queue() must not raise on startup when the queue file is corrupt."""
        import file_storage
        from excel_ops import replay_recovery_queue

        queue_path = tmp_path / "recovery_queue.json"
        queue_path.write_text("not json at all }{")
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        replay_recovery_queue()  # must not raise, corrupt file quarantined instead

        corrupt_path = queue_path.with_name(queue_path.name + ".corrupt")
        assert corrupt_path.exists()

    def test_failed_replay_leaves_rows_recoverable(self, excel_path, monkeypatch, tmp_path):
        """
        If the replay attempt itself fails (e.g. the workbook can't be opened),
        the pending rows must survive as a re-queued file, not be lost.
        """
        import file_storage
        import excel_ops
        from excel_ops import replay_recovery_queue
        from file_storage import append_to_recovery_queue

        queue_path = tmp_path / "recovery_queue.json"
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        append_to_recovery_queue({
            "date": datetime.date(2024, 6, 1), "year": 2024, "month": "Jun",
            "value": 42.0, "type": "Expense", "category": "Groceries",
            "description": "recovered", "person": "", "is_recurring": False,
            "currency": "PLN",
        })

        def _boom(*args, **kwargs):
            raise RuntimeError("simulated crash mid-replay")

        monkeypatch.setattr(excel_ops, "ExcelFileContext", _boom)

        replay_recovery_queue()  # must not raise; row should be re-queued

        assert queue_path.exists()
        import json
        requeued = json.loads(queue_path.read_text())
        assert len(requeued) == 1
        assert requeued[0]["description"] == "recovered"


# ── atomic_save ────────────────────────────────────────────────────────────────

class TestAtomicSave:

    def test_atomic_save_keeps_backup(self, excel_path):
        from file_storage import atomic_save
        wb = openpyxl.load_workbook(excel_path)
        original_bytes = excel_path.read_bytes()
        atomic_save(wb, excel_path)
        bak = excel_path.with_name(excel_path.name + ".bak")
        assert bak.exists()
        # backup is byte-identical to the pre-save file
        assert bak.read_bytes() == original_bytes

    def test_atomic_save_leaves_no_tmp_file(self, excel_path):
        from file_storage import atomic_save
        wb = openpyxl.load_workbook(excel_path)
        atomic_save(wb, excel_path)
        assert not excel_path.with_name(excel_path.name + ".tmp").exists()

    def test_single_append_uses_atomic_save(self, excel_path, sample_transaction):
        _do_append_transaction(sample_transaction)
        bak = excel_path.with_name(excel_path.name + ".bak")
        assert bak.exists()


class TestDateModifiedHeaderCreation:
    """Regression: live files predating Date Modified got a headerless column."""

    def test_writer_creates_missing_date_modified_header(self, excel_path, sample_transaction):
        # Simulate an old live file: remove the Date Modified header
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        for c in range(1, ws.max_column + 1):
            if ws.cell(1, c).value == "Date Modified (UTC)":
                ws.cell(1, c).value = None
        wb.save(excel_path)

        _do_append_transaction(sample_transaction)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Date Modified (UTC)" in headers
        col = headers.index("Date Modified (UTC)") + 1
        assert ws.cell(2, col).value is not None


class TestValidationRangeExtension:
    """Regression: appended rows beyond the static validation ranges lost dropdowns."""

    def test_append_extends_validation_ranges(self, excel_path, sample_transaction):
        from openpyxl.worksheet.datavalidation import DataValidation

        # Give MasterData a validation covering only rows 2-3
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        dv = DataValidation(type="list", formula1="Lists!$C$2:$C$18")
        dv.add("F2:F3")
        ws.add_data_validation(dv)
        wb.save(excel_path)

        _do_append_transaction(sample_transaction)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        sqrefs = [str(d.sqref) for d in ws.data_validations.dataValidation]
        assert any(s.startswith("F2:F") and int(s.split(":F")[1]) >= 500 for s in sqrefs), sqrefs


class TestFormulaInjectionGuard:
    def test_leading_equals_description_is_neutralized(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 6, 15), value=10.0, currency="PLN",
            transaction_type="Expense", category="Groceries",
            description='=HYPERLINK("http://evil","x")',
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        val = ws.cell(2, headers["Description"]).value
        assert not str(val).startswith("="), "description stored as live formula"
