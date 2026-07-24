"""
test_e2e.py — end-to-end tests exercising the complete write / read / delete
cycle against real temp Excel files.

No Telegram, no AI, no network. Pure data-layer verification.
"""

import datetime

import openpyxl
import pytest

import file_storage
from excel_ops import _do_append_transaction
from file_storage import (
    create_blank_excel,
    get_excel_path_for_reading,
    get_recent_transactions,
    delete_transaction_row,
    append_transactions_batch,
)
from models import Transaction


# ── helpers ───────────────────────────────────────────────────────────────────


def _headers(ws) -> dict[str, int]:
    """Return {header_name: column_index} for the first row of a sheet."""
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def _cell(ws, row: int, header: str, headers: dict) -> object:
    """Read a cell value by row number and header name."""
    return ws.cell(row, headers[header]).value


# ── test_e2e_full_transaction_flow ────────────────────────────────────────────


class TestE2EFullTransactionFlow:

    def test_blank_excel_has_correct_sheets(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert "MasterData" in wb.sheetnames
        assert "Lists" in wb.sheetnames
        assert "Dashboard" in wb.sheetnames

    def test_file_is_in_tmp_path(self, excel_path, tmp_path):
        # Sanity: the file is under tmp_path (auto-cleaned by pytest)
        assert str(excel_path).startswith(str(tmp_path))

    def test_three_transactions_write_and_verify_via_openpyxl(self, excel_path):
        """Write 3 known transactions, read back with openpyxl, verify every cell."""
        t1 = Transaction(
            date=datetime.date(2024, 6, 15),
            value=150.0,
            currency="PLN",
            transaction_type="Expense",
            category="Groceries",
            person="Alice",
            description="supermarket run",
        )
        t2 = Transaction(
            date=datetime.date(2024, 6, 16),
            value=80.0,
            currency="EUR",
            transaction_type="Expense",
            category="Dining Out",
            person="Bob",
            description="restaurant",
        )
        t3 = Transaction(
            date=datetime.date(2024, 6, 17),
            value=25.0,
            currency="PLN",
            transaction_type="Expense",
            category="Transport",
            person="",
            description="bus",
        )

        for txn in (t1, t2, t3):
            _do_append_transaction(txn)

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        h = _headers(ws)

        # 1 header + 3 data rows
        assert ws.max_row == 4

        # ── Row 2: Groceries ──
        val = _cell(ws, 2, "Date", h)
        if isinstance(val, datetime.datetime):
            val = val.date()
        assert val == datetime.date(2024, 6, 15)
        assert _cell(ws, 2, "Category", h) == "Groceries"
        assert _cell(ws, 2, "Value", h) == pytest.approx(150.0)
        assert _cell(ws, 2, "Person", h) == "Alice"
        assert _cell(ws, 2, "Currency", h) == "PLN"

        # ── Row 3: Dining Out (EUR) ──
        val3 = _cell(ws, 3, "Date", h)
        if isinstance(val3, datetime.datetime):
            val3 = val3.date()
        assert val3 == datetime.date(2024, 6, 16)
        assert _cell(ws, 3, "Category", h) == "Dining Out"
        assert _cell(ws, 3, "Value", h) == pytest.approx(80.0)
        assert _cell(ws, 3, "Currency", h) == "EUR"

        # ── Row 4: Transport ──
        assert _cell(ws, 4, "Category", h) == "Transport"
        assert _cell(ws, 4, "Value", h) == pytest.approx(25.0)

        # All rows should have IsDone = True
        for row in range(2, 5):
            assert _cell(ws, row, "IsDone", h) is True

    def test_value_base_formula_is_set_on_all_rows(self, excel_path):
        """Value (base) column must contain Excel formulas, not plain numbers."""
        for i, cat in enumerate(["Groceries", "Transport", "Housing"], 1):
            txn = Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 50),
                currency="PLN",
                transaction_type="Expense",
                category=cat,
            )
            _do_append_transaction(txn)

        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        h = _headers(ws)

        for row in range(2, 5):
            formula = _cell(ws, row, "Value (base)", h)
            assert isinstance(formula, str), f"Row {row}: expected formula string"
            assert formula.startswith("="), f"Row {row}: formula should start with ="

    def test_get_recent_transactions_reads_back_correctly(self, excel_path):
        """After writing, get_recent_transactions must return correct data."""
        txn = Transaction(
            date=datetime.date(2024, 6, 15),
            value=150.0,
            currency="PLN",
            transaction_type="Expense",
            category="Groceries",
            person="Alice",
        )
        _do_append_transaction(txn)

        results = get_recent_transactions(excel_path)
        assert len(results) == 1
        r = results[0]
        assert r["Value"] == pytest.approx(150.0)
        assert r["Category"] == "Groceries"
        assert r["Person"] == "Alice"
        assert r["Currency"] == "PLN"
        assert "_row_idx" in r

    def test_get_recent_transactions_returns_three_items(self, excel_path):
        for i in range(1, 4):
            txn = Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 10),
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            )
            _do_append_transaction(txn)
        results = get_recent_transactions(excel_path)
        assert len(results) == 3

    def test_delete_first_transaction_leaves_two(self, excel_path):
        """Write 3 rows, delete the first data row, verify 2 remain."""
        categories = ["Groceries", "Transport", "Housing"]
        for i, cat in enumerate(categories, 1):
            txn = Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 20),
                currency="PLN",
                transaction_type="Expense",
                category=cat,
            )
            _do_append_transaction(txn)

        # Get the row_idx of the first transaction (oldest = Groceries)
        all_txns = get_recent_transactions(excel_path, n=10)
        first = all_txns[0]  # Groceries
        assert first["Category"] == "Groceries"
        row_to_delete = first["_row_idx"]

        delete_transaction_row(row_to_delete)

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        # 1 header + 2 remaining data rows
        assert ws.max_row == 3

    def test_after_delete_correct_rows_remain(self, excel_path):
        """Delete middle row — first and last should survive."""
        categories = ["Groceries", "Transport", "Housing"]
        for i, cat in enumerate(categories, 1):
            txn = Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 10),
                currency="PLN",
                transaction_type="Expense",
                category=cat,
            )
            _do_append_transaction(txn)

        # Delete Transport (Excel row 3)
        delete_transaction_row(3)

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        h = _headers(ws)

        assert ws.max_row == 3  # header + 2
        assert _cell(ws, 2, "Category", h) == "Groceries"
        assert _cell(ws, 3, "Category", h) == "Housing"

    def test_row_idx_references_correct_excel_row(self, excel_path):
        """_row_idx returned by get_recent_transactions must point to the right row."""
        txn = Transaction(
            date=datetime.date(2024, 6, 15),
            value=99.0,
            currency="PLN",
            transaction_type="Income",
        )
        _do_append_transaction(txn)
        results = get_recent_transactions(excel_path)
        # There is one data row; its _row_idx should be 2 (row 1 is header)
        assert results[0]["_row_idx"] == 2


# ── test_e2e_blank_excel_auto_creation ───────────────────────────────────────


class TestE2EBlankExcelAutoCreation:

    def test_get_excel_path_for_reading_creates_file(self, tmp_path, monkeypatch):
        missing_path = tmp_path / "brand_new.xlsx"
        assert not missing_path.exists()

        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", missing_path)

        result = get_excel_path_for_reading()

        assert result.exists()
        assert result == missing_path

    def test_auto_created_file_has_masterdata(self, tmp_path, monkeypatch):
        missing_path = tmp_path / "brand_new.xlsx"
        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", missing_path)
        get_excel_path_for_reading()

        wb = openpyxl.load_workbook(missing_path)
        assert "MasterData" in wb.sheetnames

    def test_auto_created_file_has_lists(self, tmp_path, monkeypatch):
        missing_path = tmp_path / "brand_new.xlsx"
        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", missing_path)
        get_excel_path_for_reading()

        wb = openpyxl.load_workbook(missing_path)
        assert "Lists" in wb.sheetnames

    def test_auto_created_file_has_dashboard(self, tmp_path, monkeypatch):
        missing_path = tmp_path / "brand_new.xlsx"
        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", missing_path)
        get_excel_path_for_reading()

        wb = openpyxl.load_workbook(missing_path)
        assert "Dashboard" in wb.sheetnames


# ── test_e2e_batch_append ─────────────────────────────────────────────────────


class TestE2EBatchAppend:

    def test_batch_append_five_transactions(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 15),
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
                description=f"item {i}",
            )
            for i in range(1, 6)
        ]
        append_transactions_batch(transactions)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        assert ws.max_row == 6  # 1 header + 5 data rows

    def test_batch_append_data_is_correct(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 100),
                currency="PLN",
                transaction_type="Expense",
                category="Transport",
            )
            for i in range(1, 4)
        ]
        append_transactions_batch(transactions)

        results = get_recent_transactions(excel_path, n=10)
        assert len(results) == 3

        values = sorted(r["Value"] for r in results)
        assert values == pytest.approx([100.0, 200.0, 300.0])

    def test_batch_each_row_has_value_base_formula(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 50),
                currency="EUR",
                transaction_type="Expense",
                category="Travel",
            )
            for i in range(1, 4)
        ]
        append_transactions_batch(transactions)

        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        h = _headers(ws)

        for row in range(2, 5):
            formula = _cell(ws, row, "Value (base)", h)
            assert isinstance(formula, str)
            assert formula.startswith("=")

    def test_batch_append_each_row_has_row_idx(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, i),
                value=float(i * 10),
                currency="PLN",
                transaction_type="Income",
            )
            for i in range(1, 4)
        ]
        append_transactions_batch(transactions)

        results = get_recent_transactions(excel_path, n=10)
        for r in results:
            assert "_row_idx" in r
            assert isinstance(r["_row_idx"], int)
            assert r["_row_idx"] >= 2  # row 1 is header

    def test_batch_append_mixed_currencies(self, excel_path):
        transactions = [
            Transaction(
                date=datetime.date(2024, 6, 1),
                value=100.0,
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            ),
            Transaction(
                date=datetime.date(2024, 6, 2),
                value=50.0,
                currency="EUR",
                transaction_type="Expense",
                category="Dining Out",
            ),
            Transaction(
                date=datetime.date(2024, 6, 3),
                value=30.0,
                currency="USD",
                transaction_type="Expense",
                category="Entertainment",
            ),
        ]
        append_transactions_batch(transactions)

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        h = _headers(ws)

        assert _cell(ws, 2, "Currency", h) == "PLN"
        assert _cell(ws, 3, "Currency", h) == "EUR"
        assert _cell(ws, 4, "Currency", h) == "USD"
