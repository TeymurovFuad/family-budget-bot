"""
tests/test_dedup.py — statement dedup against MasterData + within-draft dedup
(dedup v2: count-aware matching, within-batch keep-by-default, two-pass
strict/loose scan, unified drop/keep grammar, contextual footer).

Covers BACKLOG.md "Follow-up PR: dedup v2 — agreed design" and
"Follow-up: dedup review notes (PR #7)".

asyncio_mode = auto (pytest.ini) — no @pytest.mark.asyncio needed.
"""

import json
import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from validators import make_dedup_key, make_loose_dedup_key, parse_amount

ALLOWED_UID = 123


# ── helpers ───────────────────────────────────────────────────────────────────

def make_update(text="hello", user_id=ALLOWED_UID):
    upd = MagicMock()
    upd.message.text = text
    upd.message.reply_text = AsyncMock()
    upd.effective_user.id = user_id
    upd.message.photo = None
    upd.message.document = None
    return upd


def make_ctx():
    ctx = MagicMock()
    ctx.user_data = {}
    ctx.bot = MagicMock()
    return ctx


SAMPLE_LISTS = {
    "txn_types": ["Expense", "Income", "Savings"],
    "categories": ["Groceries", "Transport", "Health"],
    "persons": ["Alice", "Bob"],
    "currencies": ["PLN", "USD", "EUR"],
}


def evidence(strict=None, loose=None) -> dict:
    return {"strict": strict or {}, "loose": loose or {}}


# ═══════════════════════════════════════════════════════════════════════════
# make_dedup_key / make_loose_dedup_key — key stability & normalization
# ═══════════════════════════════════════════════════════════════════════════

class TestMakeDedupKey:
    def test_same_inputs_produce_same_key(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        k2 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        assert k1 == k2

    def test_whitespace_differences_do_not_defeat_dedup(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries   shop")
        k2 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        assert k1 == k2

    def test_case_differences_do_not_defeat_dedup(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "PLN", "GROCERIES SHOP")
        k2 = make_dedup_key("2024-06-15", 50.0, "PLN", "groceries shop")
        assert k1 == k2

    def test_leading_trailing_whitespace_ignored(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "PLN", "  Groceries shop  ")
        k2 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        assert k1 == k2

    def test_different_value_changes_key(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        k2 = make_dedup_key("2024-06-15", 50.01, "PLN", "Groceries shop")
        assert k1 != k2

    def test_different_currency_changes_key(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        k2 = make_dedup_key("2024-06-15", 50.0, "USD", "Groceries shop")
        assert k1 != k2

    def test_different_date_changes_key(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        k2 = make_dedup_key("2024-06-16", 50.0, "PLN", "Groceries shop")
        assert k1 != k2

    def test_string_amount_matches_numeric_amount(self):
        k1 = make_dedup_key("2024-06-15", "50.00", "PLN", "Groceries shop")
        k2 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        assert k1 == k2

    def test_missing_currency_defaults_to_pln(self):
        k1 = make_dedup_key("2024-06-15", 50.0, "", "Groceries shop")
        k2 = make_dedup_key("2024-06-15", 50.0, "PLN", "Groceries shop")
        assert k1 == k2

    # ── locale-formatted value normalization fix (review notes) ─────────────
    def test_locale_thousands_comma_matches_plain_float(self):
        """'1,234.56'-style strings must route through parse_amount instead
        of falling back to a raw-string key that never matches Excel's
        float-derived key."""
        k1 = make_dedup_key("2024-06-15", "1,234.56", "PLN", "shop")
        k2 = make_dedup_key("2024-06-15", 1234.56, "PLN", "shop")
        assert k1 == k2

    def test_locale_european_comma_decimal_matches_plain_float(self):
        k1 = make_dedup_key("2024-06-15", "1 234,56", "PLN", "shop")
        k2 = make_dedup_key("2024-06-15", 1234.56, "PLN", "shop")
        assert k1 == k2

    def test_garbage_value_falls_back_to_raw_string_without_crashing(self):
        # Never raises — dedup must never block/crash an import.
        make_dedup_key("2024-06-15", "not-a-number-at-all-###", "PLN", "shop")


class TestMakeLooseDedupKey:
    def test_ignores_description(self):
        k1 = make_loose_dedup_key("2024-06-15", 50.0, "PLN")
        k2 = make_loose_dedup_key("2024-06-15", 50.0, "PLN")
        assert k1 == k2

    def test_differs_from_strict_key(self):
        loose = make_loose_dedup_key("2024-06-15", 50.0, "PLN")
        strict = make_dedup_key("2024-06-15", 50.0, "PLN", "shop")
        assert loose != strict

    def test_same_date_value_currency_same_loose_key_regardless_of_description(self):
        k1 = make_loose_dedup_key("2024-06-15", 45.98, "PLN")
        # Two different descriptions still produce the same loose key —
        # that's the whole point of the advisory pass.
        assert k1 == make_loose_dedup_key("2024-06-15", 45.98, "PLN")


# ═══════════════════════════════════════════════════════════════════════════
# load_dedup_evidence / load_dedup_keys — MasterData read, schema-driven
# ═══════════════════════════════════════════════════════════════════════════

class TestLoadDedupEvidence:
    def _seed_master_data(self, excel_path, rows):
        from openpyxl import load_workbook
        from excel_schema import find_next_data_row, lists_currency_range, write_transaction_row
        from file_storage import atomic_save

        wb = load_workbook(excel_path)
        ws = wb["MasterData"]
        lu_range = lists_currency_range(wb)
        for row in rows:
            r = find_next_data_row(ws)
            write_transaction_row(ws, r, row, lu_range)
        atomic_save(wb, excel_path)

    def test_finds_strict_key_for_row_in_range(self, excel_path):
        from data import load_dedup_evidence

        self._seed_master_data(excel_path, [
            {"date": date(2024, 6, 15), "value": 50.0, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "shop", "person": ""},
        ])
        result = load_dedup_evidence(date(2024, 6, 1), date(2024, 6, 30))
        expected = make_dedup_key("2024-06-15", 50.0, "PLN", "shop")
        assert expected in result["strict"]
        assert len(result["strict"][expected]) == 1

    def test_multiset_count_reflects_duplicate_master_rows(self, excel_path):
        """Count-aware matching needs the ACTUAL count, not just presence."""
        from data import load_dedup_evidence

        self._seed_master_data(excel_path, [
            {"date": date(2024, 6, 15), "value": 50.0, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "shop", "person": ""},
            {"date": date(2024, 6, 15), "value": 50.0, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "shop", "person": ""},
        ])
        result = load_dedup_evidence()
        key = make_dedup_key("2024-06-15", 50.0, "PLN", "shop")
        assert len(result["strict"][key]) == 2

    def test_loose_key_present_even_when_description_differs(self, excel_path):
        from data import load_dedup_evidence

        self._seed_master_data(excel_path, [
            {"date": date(2024, 6, 15), "value": 45.98, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "Zabka", "person": ""},
        ])
        result = load_dedup_evidence()
        loose_key = make_loose_dedup_key("2024-06-15", 45.98, "PLN")
        assert loose_key in result["loose"]
        assert result["loose"][loose_key][0][1] == "Zabka"

    def test_row_outside_range_excluded(self, excel_path):
        from data import load_dedup_evidence

        self._seed_master_data(excel_path, [
            {"date": date(2024, 6, 15), "value": 50.0, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "shop", "person": ""},
        ])
        result = load_dedup_evidence(date(2024, 7, 1), date(2024, 7, 31))
        expected = make_dedup_key("2024-06-15", 50.0, "PLN", "shop")
        assert expected not in result["strict"]

    def test_range_boundaries_are_inclusive(self, excel_path):
        from data import load_dedup_evidence

        self._seed_master_data(excel_path, [
            {"date": date(2024, 6, 1), "value": 10.0, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "first", "person": ""},
            {"date": date(2024, 6, 30), "value": 20.0, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "last", "person": ""},
        ])
        result = load_dedup_evidence(date(2024, 6, 1), date(2024, 6, 30))
        assert make_dedup_key("2024-06-01", 10.0, "PLN", "first") in result["strict"]
        assert make_dedup_key("2024-06-30", 20.0, "PLN", "last") in result["strict"]

    def test_read_failure_returns_empty_dicts_never_raises(self, monkeypatch):
        from data import load_dedup_evidence
        import file_storage

        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", Path("Z:/does/not/exist.xlsx"))
        result = load_dedup_evidence(date(2024, 1, 1), date(2024, 1, 31))
        assert result == {"strict": {}, "loose": {}}


class TestLoadDedupKeysBackwardCompat:
    def test_returns_set_of_strict_keys(self, excel_path):
        from data import load_dedup_keys
        from openpyxl import load_workbook
        from excel_schema import find_next_data_row, lists_currency_range, write_transaction_row
        from file_storage import atomic_save

        wb = load_workbook(excel_path)
        ws = wb["MasterData"]
        lu_range = lists_currency_range(wb)
        r = find_next_data_row(ws)
        write_transaction_row(ws, r, {
            "date": date(2024, 6, 15), "value": 50.0, "currency": "PLN",
            "type": "Expense", "category": "Groceries", "description": "shop", "person": "",
        }, lu_range)
        atomic_save(wb, excel_path)

        keys = load_dedup_keys()
        assert isinstance(keys, set)
        assert make_dedup_key("2024-06-15", 50.0, "PLN", "shop") in keys

    def test_read_failure_returns_empty_set(self, monkeypatch):
        from data import load_dedup_keys
        import file_storage

        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", Path("Z:/does/not/exist.xlsx"))
        assert load_dedup_keys(date(2024, 1, 1), date(2024, 1, 31)) == set()


# ═══════════════════════════════════════════════════════════════════════════
# _merge_bulk_draft — within-batch/within-draft repeats are KEPT by default
# ═══════════════════════════════════════════════════════════════════════════

class TestMergeBulkDraftKeepsRepeats:
    def test_same_photo_uploaded_twice_keeps_all_rows(self):
        """Dedup v2 inverts PR #7: repetition inside one source is almost
        always real, so it's kept and annotated, not hard-dropped here —
        MasterData-level count-aware dedup still applies once rows are
        actually flagged/saved."""
        from handlers.bulk_conv import _merge_bulk_draft

        with tempfile.TemporaryDirectory() as tmpdir:
            draft_dir = Path(tmpdir) / "bulk_drafts"
            draft_dir.mkdir()
            with patch("handlers.bulk_conv._bulk_draft_dir", return_value=draft_dir):
                rows = [
                    {"date": "2024-06-15", "value": 50, "currency": "PLN",
                     "category": "Groceries", "description": "shop", "person": ""},
                    {"date": "2024-06-16", "value": 20, "currency": "PLN",
                     "category": "Transport", "description": "train", "person": ""},
                ]
                merged, _ = _merge_bulk_draft(11111, rows)
                assert len(merged) == 2

                merged2, _ = _merge_bulk_draft(11111, [dict(r) for r in rows])
                assert len(merged2) == 4, "genuine repeats are kept by default in dedup v2"

    def test_within_batch_duplicates_are_kept(self):
        from handlers.bulk_conv import _merge_bulk_draft

        with tempfile.TemporaryDirectory() as tmpdir:
            draft_dir = Path(tmpdir) / "bulk_drafts"
            draft_dir.mkdir()
            with patch("handlers.bulk_conv._bulk_draft_dir", return_value=draft_dir):
                row = {"date": "2024-06-15", "value": 50, "currency": "PLN",
                       "category": "Groceries", "description": "shop", "person": ""}
                merged, _ = _merge_bulk_draft(22222, [dict(row), dict(row)])
                assert len(merged) == 2

    def test_distinct_rows_all_kept(self):
        from handlers.bulk_conv import _merge_bulk_draft

        with tempfile.TemporaryDirectory() as tmpdir:
            draft_dir = Path(tmpdir) / "bulk_drafts"
            draft_dir.mkdir()
            with patch("handlers.bulk_conv._bulk_draft_dir", return_value=draft_dir):
                rows = [
                    {"date": "2024-06-15", "value": 50, "currency": "PLN",
                     "category": "Groceries", "description": "shop A", "person": ""},
                    {"date": "2024-06-15", "value": 50, "currency": "PLN",
                     "category": "Groceries", "description": "shop B", "person": ""},
                ]
                merged, _ = _merge_bulk_draft(33333, rows)
                assert len(merged) == 2


# ═══════════════════════════════════════════════════════════════════════════
# _flag_master_duplicates — count-aware, two-pass (strict decides, loose advises)
# ═══════════════════════════════════════════════════════════════════════════

class TestFlagMasterDuplicatesCountAware:
    def test_single_occurrence_strict_match_flags_and_skips(self):
        from handlers.bulk_conv import _flag_master_duplicates

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "description": "shop", "person": ""}]
        key = make_dedup_key("2024-06-15", 50, "PLN", "shop")
        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(strict={key: [("2024-06-15", "shop")]})):
            summary = _flag_master_duplicates(rows)
        assert summary["flagged"] == 1
        assert rows[0]["dup"] is True
        assert len(summary["single_skips"]) == 1

    def test_non_matching_row_not_flagged(self):
        from handlers.bulk_conv import _flag_master_duplicates

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "description": "shop", "person": ""}]
        with patch("handlers.bulk_conv.load_dedup_evidence", return_value=evidence()):
            summary = _flag_master_duplicates(rows)
        assert summary["flagged"] == 0
        assert "dup" not in rows[0]

    def test_dup_keep_row_is_not_reflagged(self):
        from handlers.bulk_conv import _flag_master_duplicates

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "description": "shop", "person": "", "dup_keep": True}]
        key = make_dedup_key("2024-06-15", 50, "PLN", "shop")
        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(strict={key: [("2024-06-15", "shop")]})):
            summary = _flag_master_duplicates(rows)
        assert summary["flagged"] == 0
        assert "dup" not in rows[0]

    def test_count_aware_partial_match_saves_excess_skips_rest(self):
        """3 identical rows in the batch, MasterData has 2 -> save 1, skip 2."""
        from handlers.bulk_conv import _flag_master_duplicates

        row = {"date": "2024-06-15", "value": 2, "currency": "PLN",
               "description": "car wash", "person": ""}
        rows = [dict(row), dict(row), dict(row)]
        key = make_dedup_key("2024-06-15", 2, "PLN", "car wash")
        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(strict={key: [
                       ("2024-06-15", "car wash"), ("2024-06-15", "car wash"),
                   ]})):
            summary = _flag_master_duplicates(rows)
        assert summary["flagged"] == 2
        flagged_count = sum(1 for r in rows if r.get("dup"))
        kept_count = sum(1 for r in rows if not r.get("dup"))
        assert flagged_count == 2
        assert kept_count == 1
        assert len(summary["skip_groups"]) == 1
        assert summary["skip_groups"][0]["group_size"] == 3
        assert summary["skip_groups"][0]["master_count"] == 2
        assert summary["skip_groups"][0]["skip_n"] == 2

    def test_count_aware_batch_smaller_than_master_flags_all(self):
        """Batch has 2 identical rows, MasterData has 5 -> skip both."""
        from handlers.bulk_conv import _flag_master_duplicates

        row = {"date": "2024-06-15", "value": 2, "currency": "PLN",
               "description": "car wash", "person": ""}
        rows = [dict(row), dict(row)]
        key = make_dedup_key("2024-06-15", 2, "PLN", "car wash")
        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(strict={key: [("2024-06-15", "car wash")] * 5})):
            summary = _flag_master_duplicates(rows)
        assert summary["flagged"] == 2
        assert all(r.get("dup") for r in rows)

    def test_within_batch_identical_no_master_match_all_kept_and_annotated(self):
        """Within-batch identical rows are KEPT by default (inverts PR #7)."""
        from handlers.bulk_conv import _flag_master_duplicates

        row = {"date": "2024-06-15", "value": 2, "currency": "PLN",
               "description": "car wash", "person": ""}
        rows = [dict(row), dict(row), dict(row)]
        with patch("handlers.bulk_conv.load_dedup_evidence", return_value=evidence()):
            summary = _flag_master_duplicates(rows)
        assert summary["flagged"] == 0
        assert not any(r.get("dup") for r in rows)
        assert summary["identical_groups"] == [[1, 2, 3]]
        assert all(r["identical_group"] == [1, 2, 3] for r in rows)

    def test_loose_match_advisory_not_auto_skipped_shows_both_descriptions(self):
        """Pass 2 (loose) never auto-skips — it's saved by default and
        surfaced with BOTH descriptions side by side."""
        from handlers.bulk_conv import _flag_master_duplicates

        rows = [{"date": "2024-05-12", "value": 45.98, "currency": "PLN",
                 "description": "Zabka Warszawa 4211", "person": ""}]
        loose_key = make_loose_dedup_key("2024-05-12", 45.98, "PLN")
        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(loose={loose_key: [("2024-05-12", "Zabka")]})):
            summary = _flag_master_duplicates(rows)
        assert "dup" not in rows[0], "loose matches are never auto-skipped"
        assert rows[0]["loose_dup"] is True
        assert rows[0]["loose_other_desc"] == "Zabka"
        assert len(summary["loose_matches"]) == 1
        assert summary["loose_matches"][0]["description"] == "Zabka Warszawa 4211"
        assert summary["loose_matches"][0]["other_desc"] == "Zabka"

    def test_loose_pass_skipped_for_rows_already_strict_flagged(self):
        from handlers.bulk_conv import _flag_master_duplicates

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "description": "shop", "person": ""}]
        strict_key = make_dedup_key("2024-06-15", 50, "PLN", "shop")
        loose_key = make_loose_dedup_key("2024-06-15", 50, "PLN")
        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(
                       strict={strict_key: [("2024-06-15", "shop")]},
                       loose={loose_key: [("2024-06-15", "shop")]},
                   )):
            summary = _flag_master_duplicates(rows)
        assert rows[0]["dup"] is True
        assert "loose_dup" not in rows[0]
        assert summary["loose_matches"] == []

    def test_recompute_resets_stale_flags(self):
        """A row flagged dup on a previous scan but no longer matching (e.g.
        after an edit) must not stay flagged."""
        from handlers.bulk_conv import _flag_master_duplicates

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "description": "shop", "person": "", "dup": True,
                 "dup_evidence_date": "2024-06-15"}]
        with patch("handlers.bulk_conv.load_dedup_evidence", return_value=evidence()):
            _flag_master_duplicates(rows)
        assert "dup" not in rows[0]


# ═══════════════════════════════════════════════════════════════════════════
# Unified drop / keep row-command grammar
# ═══════════════════════════════════════════════════════════════════════════

class TestUnifiedRowCommandGrammar:
    def _rows(self, n=6):
        return [
            {"date": "2024-06-15", "value": 10 + i, "currency": "PLN",
             "description": f"row{i}", "person": ""}
            for i in range(1, n + 1)
        ]

    def test_keep_single_row_overrides_dup_flag(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(1)
        rows[0]["dup"] = True
        action, reason, notes = _apply_bulk_edit("keep 1", rows)
        assert action is False and reason == "edited"
        assert rows[0]["dup_keep"] is True
        assert "dup" not in rows[0]
        assert "row 1" in notes[0]

    def test_keep_rejected_when_out_of_range(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(1)
        action, reason, notes = _apply_bulk_edit("keep 5", rows)
        assert reason == "invalid"

    def test_legacy_n_keep_syntax_still_works(self):
        """Backward compatible with the pre-dedup-v2 `N keep` grammar."""
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(1)
        rows[0]["dup"] = True
        action, reason, notes = _apply_bulk_edit("1 keep", rows)
        assert reason == "edited"
        assert rows[0]["dup_keep"] is True

    def test_drop_single_row_marks_dropped(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(3)
        action, reason, notes = _apply_bulk_edit("drop 2", rows)
        assert reason == "edited"
        assert rows[1]["dropped"] is True
        assert not rows[0].get("dropped") and not rows[2].get("dropped")

    def test_keep_restores_dropped_row(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(3)
        rows[1]["dropped"] = True
        action, reason, notes = _apply_bulk_edit("keep 2", rows)
        assert reason == "edited"
        assert "dropped" not in rows[1]

    def test_drop_multi_space_separated(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(6)
        _apply_bulk_edit("drop 4 6", rows)
        dropped = [i + 1 for i, r in enumerate(rows) if r.get("dropped")]
        assert dropped == [4, 6]

    def test_drop_range_and_extra_singles(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(12)
        _apply_bulk_edit("drop 4-6 9 12", rows)
        dropped = sorted(i + 1 for i, r in enumerate(rows) if r.get("dropped"))
        assert dropped == [4, 5, 6, 9, 12]

    def test_keep_range(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(9)
        for r in rows:
            r["dup"] = True
        _apply_bulk_edit("keep 3 7-9", rows)
        kept = sorted(i + 1 for i, r in enumerate(rows) if r.get("dup_keep"))
        assert kept == [3, 7, 8, 9]
        # rows not targeted stay flagged as dup
        assert rows[0]["dup"] is True

    def test_drop_all_targets_whole_batch(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(4)
        _apply_bulk_edit("drop all", rows)
        assert all(r.get("dropped") for r in rows)

    def test_keep_all_targets_whole_batch(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(4)
        for r in rows:
            r["dup"] = True
        _apply_bulk_edit("keep all", rows)
        assert all(r.get("dup_keep") for r in rows)
        assert not any(r.get("dup") for r in rows)

    def test_keep_all_flagged_only_targets_dup_skipped_rows(self):
        """`keep all flagged` under the skip list acts on skipped rows only."""
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(4)
        rows[0]["dup"] = True
        rows[2]["loose_dup"] = True  # advisory, NOT a dup skip
        action, reason, notes = _apply_bulk_edit("keep all flagged", rows)
        assert rows[0]["dup_keep"] is True
        assert "dup_keep" not in rows[1]
        assert "dup_keep" not in rows[2], "loose-match rows are not 'flagged' for `keep`"
        assert "dup_keep" not in rows[3]

    def test_drop_all_flagged_only_targets_loose_advisory_rows(self):
        """`drop all flagged` under the advisory acts on advisory rows only."""
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(4)
        rows[0]["dup"] = True
        rows[2]["loose_dup"] = True
        action, reason, notes = _apply_bulk_edit("drop all flagged", rows)
        assert rows[2]["dropped"] is True
        assert not rows[0].get("dropped"), "dup-skipped rows are not 'flagged' for `drop`"
        assert not rows[1].get("dropped")
        assert not rows[3].get("dropped")

    def test_all_flagged_invalid_when_nothing_matches(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(3)
        action, reason, notes = _apply_bulk_edit("keep all flagged", rows)
        assert reason == "invalid"

    def test_stable_numbering_across_multiple_commands(self):
        """Row numbers never shift mid-draft; two sequential drop commands
        must both resolve against the ORIGINAL numbering."""
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(5)
        _apply_bulk_edit("drop 2", rows)
        _apply_bulk_edit("drop 4", rows)
        assert len(rows) == 5, "rows are never removed from the list, only flagged"
        dropped = [i + 1 for i, r in enumerate(rows) if r.get("dropped")]
        assert dropped == [2, 4]

    def test_field_edit_grammar_still_single_row(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(2)
        action, reason, notes = _apply_bulk_edit("2 description=Lunch", rows)
        assert reason == "edited"
        assert rows[1]["description"] == "Lunch"
        assert rows[0]["description"] == "row1"

    def test_unrecognized_command_falls_through_to_invalid(self):
        from handlers.bulk_conv import _apply_bulk_edit

        rows = self._rows(1)
        action, reason, notes = _apply_bulk_edit("banana", rows)
        assert reason == "invalid"


# ═══════════════════════════════════════════════════════════════════════════
# Preview rendering — evidence, annotations, contextual footer
# ═══════════════════════════════════════════════════════════════════════════

class TestPreviewAnnotationsAndFooter:
    def test_flagged_row_shows_evidence_and_override(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "category": "Groceries", "description": "shop", "person": "",
                 "dup": True, "dup_evidence_date": "2024-05-12"}]
        pages = _format_bulk_preview(rows)
        assert "already imported" in pages[0]
        assert "keep 1" in pages[0]

    def test_kept_row_does_not_show_marker(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "category": "Groceries", "description": "shop", "person": "",
                 "dup_keep": True}]
        pages = _format_bulk_preview(rows)
        assert "already imported" not in pages[0]

    def test_loose_match_shows_both_descriptions_hint(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [{"date": "2024-05-12", "value": 45.98, "currency": "PLN",
                 "category": "Groceries", "description": "Zabka Warszawa 4211", "person": "",
                 "loose_dup": True, "loose_other_date": "2024-05-12", "loose_other_desc": "Zabka"}]
        pages = _format_bulk_preview(rows)
        assert "possible duplicate" in pages[0]
        assert "drop 1" in pages[0]

    def test_identical_group_annotation_shown(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [{"date": "2024-06-15", "value": 2, "currency": "PLN",
                 "category": "Transport", "description": "car wash", "person": "",
                 "identical_group": [4, 5, 6]} for _ in range(3)]
        pages = _format_bulk_preview(rows)
        assert "identical" in pages[0]
        assert "keeping all 3" in pages[0]

    def test_dropped_row_shows_restore_hint(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "category": "Groceries", "description": "shop", "person": "",
                 "dropped": True}]
        pages = _format_bulk_preview(rows)
        assert "dropped" in pages[0]
        assert "keep 1" in pages[0]

    def test_footer_shows_nothing_flagged_when_clean(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "category": "Groceries", "description": "shop", "person": ""}]
        pages = _format_bulk_preview(rows)
        assert "skipped as already imported" not in pages[-1]
        assert "possible duplicate" not in pages[-1]
        assert "save" in pages[-1] and "cancel" in pages[-1]

    def test_footer_adds_dup_hint_only_when_dup_present(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [
            {"date": "2024-06-15", "value": 50, "currency": "PLN",
             "category": "Groceries", "description": "shop", "person": "", "dup": True},
            {"date": "2024-06-16", "value": 10, "currency": "PLN",
             "category": "Groceries", "description": "milk", "person": ""},
        ]
        pages = _format_bulk_preview(rows)
        assert "skipped as already imported" in pages[-1]
        assert "keep all flagged" in pages[-1]
        assert "possible duplicate" not in pages[-1]

    def test_footer_adds_loose_hint_only_when_loose_present(self):
        from handlers.bulk_conv import _format_bulk_preview

        rows = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                 "category": "Groceries", "description": "shop", "person": "", "loose_dup": True}]
        pages = _format_bulk_preview(rows)
        assert "possible duplicate" in pages[-1]
        assert "drop all flagged" in pages[-1]
        assert "skipped as already imported" not in pages[-1]


# ═══════════════════════════════════════════════════════════════════════════
# End-to-end: bulk_receive / bulk_confirm reporting
# ═══════════════════════════════════════════════════════════════════════════

class TestBulkReceiveDedupReporting:
    async def test_within_batch_repeat_reported_as_identical_group(self):
        from handlers.bulk_conv import bulk_receive
        import states

        upd = make_update("some text")
        ctx = make_ctx()
        with tempfile.TemporaryDirectory() as tmpdir:
            draft_dir = Path(tmpdir) / "bulk_drafts"
            draft_dir.mkdir()
            draft_path = draft_dir / f"{ALLOWED_UID}.json"
            existing_row = {"date": "2024-06-15", "value": 50, "currency": "PLN",
                             "category": "Groceries", "description": "shop", "person": "",
                             "status": "pending"}
            draft_path.write_text(json.dumps([existing_row]))

            with patch("handlers.bulk_conv._bulk_draft_dir", return_value=draft_dir), \
                 patch("handlers.bulk_conv.load_reference_data", return_value=SAMPLE_LISTS), \
                 patch("handlers.bulk_conv.parse_text", return_value=[dict(existing_row)]), \
                 patch("handlers.bulk_conv.load_dedup_evidence", return_value=evidence()):
                result = await bulk_receive(upd, ctx)

            assert result == states.BULK_CONFIRM
            all_texts = " ".join(
                c.args[0] for c in upd.message.reply_text.call_args_list if c.args
            )
            assert "identical" in all_texts
            assert "keeping all 2" in all_texts

    async def test_master_data_duplicate_reported_to_user(self):
        from handlers.bulk_conv import bulk_receive
        import states

        upd = make_update("some text")
        ctx = make_ctx()
        with tempfile.TemporaryDirectory() as tmpdir:
            draft_dir = Path(tmpdir) / "bulk_drafts"
            draft_dir.mkdir()
            parsed = [{"date": "2024-06-15", "value": 50, "currency": "PLN",
                       "category": "Groceries", "description": "shop", "person": ""}]
            key = make_dedup_key("2024-06-15", 50, "PLN", "shop")

            with patch("handlers.bulk_conv._bulk_draft_dir", return_value=draft_dir), \
                 patch("handlers.bulk_conv.load_reference_data", return_value=SAMPLE_LISTS), \
                 patch("handlers.bulk_conv.parse_text", return_value=parsed), \
                 patch("handlers.bulk_conv.load_dedup_evidence",
                       return_value=evidence(strict={key: [("2024-06-15", "shop")]})):
                result = await bulk_receive(upd, ctx)

            assert result == states.BULK_CONFIRM
            all_texts = " ".join(
                c.args[0] for c in upd.message.reply_text.call_args_list if c.args
            )
            assert "↺" in all_texts
            assert "matches an entry saved" in all_texts
            assert "keep 1" in all_texts

    async def test_loose_match_reported_as_advisory_not_skip(self):
        from handlers.bulk_conv import bulk_receive
        import states

        upd = make_update("some text")
        ctx = make_ctx()
        with tempfile.TemporaryDirectory() as tmpdir:
            draft_dir = Path(tmpdir) / "bulk_drafts"
            draft_dir.mkdir()
            parsed = [{"date": "2024-05-12", "value": 45.98, "currency": "PLN",
                       "category": "Groceries", "description": "Zabka Warszawa 4211", "person": ""}]
            loose_key = make_loose_dedup_key("2024-05-12", 45.98, "PLN")

            with patch("handlers.bulk_conv._bulk_draft_dir", return_value=draft_dir), \
                 patch("handlers.bulk_conv.load_reference_data", return_value=SAMPLE_LISTS), \
                 patch("handlers.bulk_conv.parse_text", return_value=parsed), \
                 patch("handlers.bulk_conv.load_dedup_evidence",
                       return_value=evidence(loose={loose_key: [("2024-05-12", "Zabka")]})):
                result = await bulk_receive(upd, ctx)

            assert result == states.BULK_CONFIRM
            all_texts = " ".join(
                c.args[0] for c in upd.message.reply_text.call_args_list if c.args
            )
            assert "Possible duplicates" in all_texts
            assert "Zabka" in all_texts
            assert "Saving them" in all_texts
            assert ctx.user_data["bulk_parsed"][0].get("dup") is not True


class TestBulkConfirmSkipsDuplicatesAtSave:
    async def test_master_data_dup_skipped_and_reported(self):
        from handlers.bulk_conv import bulk_confirm
        from telegram.ext import ConversationHandler

        ctx = make_ctx()
        ctx.user_data["bulk_parsed"] = [
            {"date": "2024-06-15", "value": 50, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "shop", "person": ""},
            {"date": "2024-06-16", "value": 20, "currency": "PLN",
             "type": "Expense", "category": "Transport", "description": "train", "person": ""},
        ]
        upd = make_update("Yes")
        key = make_dedup_key("2024-06-15", 50, "PLN", "shop")

        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(strict={key: [("2024-06-15", "shop")]})), \
             patch("handlers.bulk_conv.async_append_batch", AsyncMock()) as mock_batch:
            result = await bulk_confirm(upd, ctx)

        assert result == ConversationHandler.END
        saved_txns = mock_batch.call_args.args[0]
        assert len(saved_txns) == 1
        assert saved_txns[0].description == "train"

        sent = upd.message.reply_text.call_args.args[0]
        assert "Saved 1" in sent
        assert "↺" in sent
        assert "already imported" in sent
        assert "#1" in sent

    async def test_dup_keep_row_is_saved_despite_flag(self):
        from handlers.bulk_conv import bulk_confirm
        from telegram.ext import ConversationHandler

        ctx = make_ctx()
        ctx.user_data["bulk_parsed"] = [
            {"date": "2024-06-15", "value": 50, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "shop", "person": "",
             "dup_keep": True},
        ]
        upd = make_update("Yes")
        key = make_dedup_key("2024-06-15", 50, "PLN", "shop")

        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(strict={key: [("2024-06-15", "shop")]})), \
             patch("handlers.bulk_conv.async_append_batch", AsyncMock()) as mock_batch:
            result = await bulk_confirm(upd, ctx)

        assert result == ConversationHandler.END
        saved_txns = mock_batch.call_args.args[0]
        assert len(saved_txns) == 1

        sent = upd.message.reply_text.call_args.args[0]
        assert "Saved 1" in sent
        assert "↺" not in sent

    async def test_within_batch_dup_now_saved_both_by_default(self):
        """dedup v2: two genuinely identical rows in one batch with NO
        MasterData match are both saved (inverts the old hard skip)."""
        from handlers.bulk_conv import bulk_confirm
        from telegram.ext import ConversationHandler

        ctx = make_ctx()
        row = {"date": "2024-06-15", "value": 50, "currency": "PLN",
               "type": "Expense", "category": "Groceries", "description": "shop", "person": ""}
        ctx.user_data["bulk_parsed"] = [dict(row), dict(row)]
        upd = make_update("Yes")

        with patch("handlers.bulk_conv.load_dedup_evidence", return_value=evidence()), \
             patch("handlers.bulk_conv.async_append_batch", AsyncMock()) as mock_batch:
            result = await bulk_confirm(upd, ctx)

        assert result == ConversationHandler.END
        saved_txns = mock_batch.call_args.args[0]
        assert len(saved_txns) == 2

        sent = upd.message.reply_text.call_args.args[0]
        assert "Saved 2" in sent

    async def test_within_batch_dup_with_master_match_saves_only_excess(self):
        """3 identical rows in the batch, MasterData already has 2 -> save 1."""
        from handlers.bulk_conv import bulk_confirm
        from telegram.ext import ConversationHandler

        ctx = make_ctx()
        row = {"date": "2024-06-15", "value": 2, "currency": "PLN",
               "type": "Expense", "category": "Transport", "description": "car wash", "person": ""}
        ctx.user_data["bulk_parsed"] = [dict(row), dict(row), dict(row)]
        upd = make_update("Yes")
        key = make_dedup_key("2024-06-15", 2, "PLN", "car wash")

        with patch("handlers.bulk_conv.load_dedup_evidence",
                   return_value=evidence(strict={key: [
                       ("2024-06-15", "car wash"), ("2024-06-15", "car wash"),
                   ]})), \
             patch("handlers.bulk_conv.async_append_batch", AsyncMock()) as mock_batch:
            result = await bulk_confirm(upd, ctx)

        assert result == ConversationHandler.END
        saved_txns = mock_batch.call_args.args[0]
        assert len(saved_txns) == 1

        sent = upd.message.reply_text.call_args.args[0]
        assert "Saved 1" in sent
        assert "↺" in sent

    async def test_dropped_row_excluded_from_save_and_reported(self):
        from handlers.bulk_conv import bulk_confirm
        from telegram.ext import ConversationHandler

        ctx = make_ctx()
        ctx.user_data["bulk_parsed"] = [
            {"date": "2024-06-15", "value": 50, "currency": "PLN",
             "type": "Expense", "category": "Groceries", "description": "shop", "person": "",
             "dropped": True},
            {"date": "2024-06-16", "value": 20, "currency": "PLN",
             "type": "Expense", "category": "Transport", "description": "train", "person": ""},
        ]
        upd = make_update("Yes")

        with patch("handlers.bulk_conv.load_dedup_evidence", return_value=evidence()), \
             patch("handlers.bulk_conv.async_append_batch", AsyncMock()) as mock_batch:
            result = await bulk_confirm(upd, ctx)

        assert result == ConversationHandler.END
        saved_txns = mock_batch.call_args.args[0]
        assert len(saved_txns) == 1
        assert saved_txns[0].description == "train"

        sent = upd.message.reply_text.call_args.args[0]
        assert "dropped as requested" in sent
        assert "#1" in sent
