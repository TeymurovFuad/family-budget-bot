"""
test_e2e_full.py — exhaustive E2E tests for file_storage, data, scheduled_report,
and scheduled job functions against real temp Excel files.
"""

import asyncio
import datetime
import json

import openpyxl
import pandas as pd
import pytest
from unittest.mock import AsyncMock, MagicMock, patch

import file_storage
import scheduled
from file_storage import (
    create_blank_excel,
    get_recent_transactions,
    delete_transaction_row,
    append_transactions_batch,
)
from models import Transaction
from data import load_data
from scheduled_report import (
    load_currency_rates,
    load_transaction_data,
    build_weekly_report,
    build_monthly_summary,
    format_with_currency,
    previous_month_year_and_name,
)
from scheduled import send_daily_reminder


# ── helpers ───────────────────────────────────────────────────────────────────

def _make_txn(value=100.0, currency="PLN", txn_type="Expense",
              category="Groceries", date=None, description="test",
              is_recurring=False):
    return Transaction(
        date=date or datetime.date(2024, 6, 15),
        value=value,
        currency=currency,
        transaction_type=txn_type,
        category=category,
        description=description,
        is_recurring=is_recurring,
    )


# ── E2E: basic append and load_data ──────────────────────────────────────────

class TestLoadData:

    def test_expense_base_load_data_row_present(self, excel_path):
        append_transactions_batch([_make_txn(100.0, "PLN", "Expense", "Groceries", datetime.date(2024, 6, 15))])
        df = load_data()
        assert len(df) == 1
        assert df.iloc[0]["Type"] == "Expense"
        assert df.iloc[0]["Category"] == "Groceries"
        assert df.iloc[0]["Value"] == pytest.approx(100.0)
        assert int(df.iloc[0]["Year"]) == 2024
        assert df.iloc[0]["Month"] == "Jun"

    def test_expense_base_is_done_true(self, excel_path):
        append_transactions_batch([_make_txn()])
        df = load_data()
        assert bool(df.iloc[0]["IsDone"]) is True

    def test_expense_date_modified_is_datetime(self, excel_path):
        append_transactions_batch([_make_txn()])
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        col = headers["Date Modified (UTC)"]
        val = ws.cell(2, col).value
        assert isinstance(val, datetime.datetime)

    def test_income_transaction_empty_category_accepted(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 6, 15),
            value=5000.0,
            currency="PLN",
            transaction_type="Income",
            category="",
        )
        append_transactions_batch([txn])
        df = load_data()
        assert len(df) == 1
        assert df.iloc[0]["Type"] == "Income"

    def test_year_month_derived_from_date(self, excel_path):
        append_transactions_batch([_make_txn(date=datetime.date(2024, 3, 20))])
        df = load_data()
        assert int(df.iloc[0]["Year"]) == 2024
        assert df.iloc[0]["Month"] == "Mar"

    def test_load_data_base_currency_computes_base(self, excel_path):
        append_transactions_batch([_make_txn(200.0, "PLN")])
        df = load_data()
        # Formula is NaN when read by pandas; fallback: Value * 1.0 = 200.0
        assert df.iloc[0]["_base"] == pytest.approx(200.0)

    def test_load_data_eur_transaction_computes_base_via_rate(self, excel_path):
        append_transactions_batch([_make_txn(100.0, "EUR")])
        df = load_data()
        # EUR rate from blank excel = 4.28; fallback: 100 * 4.28 = 428.0
        assert df.iloc[0]["_base"] == pytest.approx(428.0)


# ── E2E: delete then query ────────────────────────────────────────────────────

class TestDeleteAndQuery:

    def test_delete_middle_then_get_recent_two_rows(self, excel_path):
        txns = [
            _make_txn(category="Groceries", date=datetime.date(2024, 6, 1)),
            _make_txn(category="Transport", date=datetime.date(2024, 6, 2)),
            _make_txn(category="Housing",   date=datetime.date(2024, 6, 3)),
        ]
        append_transactions_batch(txns)
        all_rows = get_recent_transactions(excel_path, n=10)
        # find Transport row
        transport_row = next(r for r in all_rows if r["Category"] == "Transport")
        delete_transaction_row(transport_row["_row_idx"])
        remaining = get_recent_transactions(excel_path, n=10)
        assert len(remaining) == 2
        categories = [r["Category"] for r in remaining]
        assert "Transport" not in categories

    def test_delete_middle_no_data_from_deleted(self, excel_path):
        txns = [
            _make_txn(category="Groceries", date=datetime.date(2024, 6, 1)),
            _make_txn(category="Transport", date=datetime.date(2024, 6, 2)),
            _make_txn(category="Housing",   date=datetime.date(2024, 6, 3)),
        ]
        append_transactions_batch(txns)
        all_rows = get_recent_transactions(excel_path, n=10)
        transport_row = next(r for r in all_rows if r["Category"] == "Transport")
        delete_transaction_row(transport_row["_row_idx"])
        df = load_data()
        assert len(df) == 2
        assert "Transport" not in df["Category"].values


# ── E2E: batch append ─────────────────────────────────────────────────────────

class TestBatchAppend:

    def test_batch_five_rows_load_data_count(self, excel_path):
        txns = [
            _make_txn(value=float(i * 10), date=datetime.date(2024, 6, i))
            for i in range(1, 6)
        ]
        append_transactions_batch(txns)
        df = load_data()
        assert len(df) == 5

    def test_batch_date_modified_all_datetime(self, excel_path):
        txns = [
            _make_txn(date=datetime.date(2024, 6, i))
            for i in range(1, 4)
        ]
        append_transactions_batch(txns)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        col = headers["Date Modified (UTC)"]
        for row in range(2, 5):
            val = ws.cell(row, col).value
            assert isinstance(val, datetime.datetime), f"Row {row}: expected datetime, got {type(val)}"

    def test_batch_all_values_correct(self, excel_path):
        txns = [
            _make_txn(value=50.0,  currency="PLN", date=datetime.date(2024, 6, 1)),
            _make_txn(value=100.0, currency="PLN", date=datetime.date(2024, 6, 2)),
            _make_txn(value=150.0, currency="PLN", date=datetime.date(2024, 6, 3)),
        ]
        append_transactions_batch(txns)
        df = load_data()
        pln_values = sorted(df["_base"].tolist())
        assert pln_values == pytest.approx([50.0, 100.0, 150.0])


# ── E2E: scheduled_report ─────────────────────────────────────────────────────

class TestScheduledReport:

    def test_load_currency_rates_returns_base(self, excel_path):
        rates = load_currency_rates()
        assert "PLN" in rates
        assert rates["PLN"] == pytest.approx(1.0)

    def test_load_currency_rates_returns_eur(self, excel_path):
        rates = load_currency_rates()
        assert "EUR" in rates

    def test_load_transaction_data_eur_rate_recomputed(self, excel_path):
        append_transactions_batch([_make_txn(50.0, "EUR")])
        df = load_transaction_data()
        assert not df.empty
        assert df.iloc[0]["amount_base"] == pytest.approx(50.0 * 4.28)

    def test_build_weekly_report_contains_projected(self, excel_path):
        today = datetime.datetime.now()
        txns = [
            Transaction(
                date=datetime.date(today.year, today.month, 1),
                value=500.0,
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            ),
            Transaction(
                date=datetime.date(today.year, today.month, 1),
                value=10000.0,
                currency="PLN",
                transaction_type="Income",
                category="Salary",
            ),
        ]
        append_transactions_batch(txns)
        df = load_transaction_data()
        rates = load_currency_rates()
        result = build_weekly_report(df, rates, "PLN")
        assert "Projected" in result

    def test_build_monthly_summary_savings_rate(self, excel_path):
        prev_year, prev_month = previous_month_year_and_name()
        # Map month name to number
        month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        month_num = month_names.index(prev_month) + 1
        txns = [
            Transaction(
                date=datetime.date(prev_year, month_num, 15),
                value=10000.0,
                currency="PLN",
                transaction_type="Income",
                category="Salary",
            ),
            Transaction(
                date=datetime.date(prev_year, month_num, 15),
                value=2000.0,
                currency="PLN",
                transaction_type="Savings",
                category="Bank Deposit",
            ),
            Transaction(
                date=datetime.date(prev_year, month_num, 15),
                value=8000.0,
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            ),
        ]
        append_transactions_batch(txns)
        df = load_transaction_data()
        rates = load_currency_rates()
        result = build_monthly_summary(df, rates, "PLN")
        # savings rate = 2000/10000 = 20%
        assert "20%" in result

    def test_build_monthly_summary_zero_income_no_divide_error(self, excel_path):
        prev_year, prev_month = previous_month_year_and_name()
        month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        month_num = month_names.index(prev_month) + 1
        txns = [
            Transaction(
                date=datetime.date(prev_year, month_num, 15),
                value=500.0,
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            ),
        ]
        append_transactions_batch(txns)
        df = load_transaction_data()
        rates = load_currency_rates()
        # Should not raise ZeroDivisionError
        result = build_monthly_summary(df, rates, "PLN")
        assert "0%" in result

    def test_build_monthly_summary_fixed_vs_variable(self, excel_path):
        prev_year, prev_month = previous_month_year_and_name()
        month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        month_num = month_names.index(prev_month) + 1
        txns = [
            Transaction(
                date=datetime.date(prev_year, month_num, 15),
                value=800.0,
                currency="PLN",
                transaction_type="Expense",
                category="Housing",
                is_recurring=True,
            ),
            Transaction(
                date=datetime.date(prev_year, month_num, 15),
                value=200.0,
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
                is_recurring=False,
            ),
        ]
        append_transactions_batch(txns)
        df = load_transaction_data()
        rates = load_currency_rates()
        # Should not crash; IsRecurring column must be readable
        result = build_monthly_summary(df, rates, "PLN")
        assert isinstance(result, str)


# ── E2E: scheduled.py send_daily_reminder ────────────────────────────────────

class TestSendDailyReminder:

    def test_daily_reminder_no_message_when_transaction_today(self, excel_path):
        today = datetime.datetime.now(datetime.timezone.utc).date()
        today_dt = pd.Timestamp(today)
        df = pd.DataFrame({
            "Date": [today_dt],
            "Value": [100.0],
            "_base": [100.0],
            "Type": ["Expense"],
            "Year": pd.array([today.year], dtype="Int64"),
            "Month": ["Jun"],
            "IsDone": [True],
            "Category": ["Groceries"],
        })
        mock_app = MagicMock()
        mock_app.bot.send_message = AsyncMock()

        with patch("scheduled.load_data", return_value=df), \
             patch("scheduled.ALLOWED_USERS", [123]):
            asyncio.run(send_daily_reminder(mock_app))

        mock_app.bot.send_message.assert_not_called()

    def test_daily_reminder_sends_when_no_transaction_today(self, excel_path):
        # df with old date only
        old_date = pd.Timestamp("2020-01-01")
        df = pd.DataFrame({
            "Date": [old_date],
            "Value": [100.0],
            "_base": [100.0],
            "Type": ["Expense"],
            "Year": pd.array([2020], dtype="Int64"),
            "Month": ["Jan"],
            "IsDone": [True],
            "Category": ["Groceries"],
        })
        mock_app = MagicMock()
        mock_app.bot.send_message = AsyncMock()

        with patch("scheduled.load_data", return_value=df), \
             patch("scheduled.ALLOWED_USERS", [123]):
            asyncio.run(send_daily_reminder(mock_app))

        mock_app.bot.send_message.assert_called()

    def test_daily_reminder_empty_df_sends_reminder(self, excel_path):
        df = pd.DataFrame({
            "Date": pd.Series([], dtype="datetime64[ns]"),
            "Value": pd.Series([], dtype=float),
            "_base": pd.Series([], dtype=float),
            "Type": pd.Series([], dtype=str),
            "Year": pd.array([], dtype="Int64"),
            "Month": pd.Series([], dtype=str),
            "IsDone": pd.Series([], dtype=bool),
            "Category": pd.Series([], dtype=str),
        })
        mock_app = MagicMock()
        mock_app.bot.send_message = AsyncMock()

        with patch("scheduled.load_data", return_value=df), \
             patch("scheduled.ALLOWED_USERS", [123]):
            asyncio.run(send_daily_reminder(mock_app))

        # Empty df → no transaction today → should call send_message
        mock_app.bot.send_message.assert_called()
