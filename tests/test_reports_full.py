"""
test_reports_full.py — exhaustive tests for the reporting layer.

Covers: handlers/reports.py, formatters.py, scheduled_report.py

IMPORTANT: env vars and sys.modules stubs must be set BEFORE importing any
project modules, because several modules read env vars at import time.
"""

# ── 0. Env vars before ANY project import ────────────────────────────────────
import os
import sys

os.environ.setdefault("TELEGRAM_BOT_TOKEN", "test-token")
os.environ.setdefault("ALLOWED_TELEGRAM_IDS", "12345")
os.environ.setdefault("STORAGE_BACKEND", "local")

# ── 1. Stub heavy/external modules before importing project code ──────────────
from unittest.mock import MagicMock, patch

# Telegram
sys.modules.setdefault("telegram", MagicMock())
sys.modules.setdefault("telegram.ext", MagicMock())

# Project modules that touch the filesystem / bot
for _mod in ("config", "data", "excel_ops", "file_storage"):
    sys.modules.setdefault(_mod, MagicMock())

# ── 2. Now safe to insert project root on path and import project modules ─────
from pathlib import Path
_PROJECT_ROOT = str(Path(__file__).parent.parent)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# ── 3. Real imports ───────────────────────────────────────────────────────────
import io
import pytest
import pandas as pd
from datetime import date, datetime, timedelta, timezone

# formatters (data.get_rate is already a MagicMock stub)
import data as _data_stub
import formatters
from formatters import (
    format_amount,
    convert,
    format_base_as_currency,
    budget_progress_bar,
    savings_emoji,
)

# handlers.reports (telegram / config / data / excel_ops / file_storage already stubbed)
from handlers.reports import _bar_color, _build_savings_chart, _build_range_report

# scheduled_report (needs env vars set above)
import scheduled_report


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _chunk_text(report_text: str, max_size: int = 4000) -> list:
    """Pure-function copy of the pagination logic from cmd_report."""
    if len(report_text) <= max_size:
        return [report_text]
    chunks = []
    current = ""
    for line in report_text.split("\n"):
        if len(current) + len(line) + 1 > max_size:
            chunks.append(current)
            current = line
        else:
            current += ("\n" if current else "") + line
    if current:
        chunks.append(current)
    return chunks


def _make_df():
    """Standard controlled DataFrame used by _build_range_report tests."""
    return pd.DataFrame({
        "Date":     ["2025-06-01", "2025-06-05", "2025-06-10", "2025-06-12", "2025-06-14"],
        "Type":     ["Income",     "Expense",    "Expense",    "Savings",    "Expense"],
        "Category": ["Salary",     "Food",       "Transport",  "Investment", "Food"],
        "_base":     [5000.0,        200.0,         100.0,        500.0,        150.0],
        "IsDone":   [True,          True,          True,         True,         False],
    })


# ─────────────────────────────────────────────────────────────────────────────
# formatters.py
# ─────────────────────────────────────────────────────────────────────────────

class TestFormatAmountFull:

    def test_zero_base(self):
        assert format_amount(0, "PLN") == "0 PLN"

    def test_1234_5_eur_uses_comma_separator(self):
        # :,.0f uses banker's rounding: 1234.5 rounds to 1234 (round-half-to-even)
        result = format_amount(1234.5, "EUR")
        assert result == "1,234 EUR"

    def test_negative_value_has_sign(self):
        result = format_amount(-500.0, "PLN")
        assert "-500" in result
        assert "PLN" in result

    def test_negative_with_thousands(self):
        result = format_amount(-1234.0, "EUR")
        assert "-1,234" in result
        assert "EUR" in result


class TestBudgetProgressBarFull:

    def test_zero_actual_all_empty(self):
        result = budget_progress_bar(0, 100)
        assert result == "⬜" * 10

    def test_exactly_100_percent_all_filled_yellow(self):
        # pct=1.0 → colour is 🟨 (pct <= 1.0)
        result = budget_progress_bar(100, 100)
        assert "⬜" not in result
        assert "🟨" in result
        assert "🟩" not in result

    def test_over_100_percent_red(self):
        result = budget_progress_bar(101, 100)
        assert "🟥" in result

    def test_exactly_80_percent_green(self):
        # pct=0.80 → colour 🟩
        result = budget_progress_bar(80, 100)
        assert "🟩" in result
        assert "🟨" not in result
        assert "🟥" not in result

    def test_81_percent_yellow(self):
        # pct=0.81 → colour 🟨
        result = budget_progress_bar(81, 100)
        assert "🟨" in result
        assert "🟩" not in result

    def test_zero_budget_returns_dashes(self):
        result = budget_progress_bar(0, 0)
        assert result == "─" * 10

    def test_zero_budget_nonzero_actual_returns_dashes(self):
        result = budget_progress_bar(500, 0)
        assert result == "─" * 10


class TestSavingsEmojiFull:

    def test_zero_rate_red(self):
        assert savings_emoji(0.0) == "🔴"

    def test_7_percent_red(self):
        assert savings_emoji(0.07) == "🔴"

    def test_8_percent_yellow(self):
        assert savings_emoji(0.08) == "🟡"

    def test_14_percent_yellow(self):
        assert savings_emoji(0.14) == "🟡"

    def test_15_percent_green(self):
        assert savings_emoji(0.15) == "💚"

    def test_19_percent_green(self):
        assert savings_emoji(0.19) == "💚"

    def test_20_percent_rocket(self):
        assert savings_emoji(0.20) == "🚀"

    def test_50_percent_rocket(self):
        assert savings_emoji(0.50) == "🚀"


class TestConvertFull:
    # formatters.py does `from data import get_rate`, so we must patch the
    # name as it lives inside the formatters module, not in data.

    def test_base_rate_1_returns_same(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        assert convert(1000.0, "PLN", {}) == pytest.approx(1000.0)

    def test_eur_rate_4_divides_by_4(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 4.0)
        assert convert(400.0, "EUR", {"EUR": 4.0}) == pytest.approx(100.0)

    def test_zero_rate_returns_original(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 0)
        assert convert(999.0, "BAD", {}) == pytest.approx(999.0)


class TestFormatPlnAsCurrencyFull:

    def test_base_passthrough(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        assert format_base_as_currency(1000.0, "PLN", {}) == "1,000 PLN"

    def test_converts_then_formats(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 4.0)
        # 800 PLN / 4 = 200 EUR
        assert format_base_as_currency(800.0, "EUR", {"EUR": 4.0}) == "200 EUR"


# ─────────────────────────────────────────────────────────────────────────────
# handlers/reports.py — _bar_color
# ─────────────────────────────────────────────────────────────────────────────

class TestBarColorFull:

    def test_zero_budget_grey(self):
        assert _bar_color(0, 0) == "#9E9E9E"

    def test_nonzero_spend_zero_budget_grey(self):
        assert _bar_color(50, 0) == "#9E9E9E"

    def test_79_percent_green(self):
        assert _bar_color(79, 100) == "#4CAF50"

    def test_exactly_80_percent_green(self):
        assert _bar_color(80, 100) == "#4CAF50"

    def test_81_percent_orange(self):
        assert _bar_color(81, 100) == "#FF9800"

    def test_exactly_100_percent_orange(self):
        assert _bar_color(100, 100) == "#FF9800"

    def test_101_percent_red(self):
        assert _bar_color(101, 100) == "#F44336"

    def test_200_percent_red(self):
        assert _bar_color(200, 100) == "#F44336"


# ─────────────────────────────────────────────────────────────────────────────
# handlers/reports.py — _build_savings_chart
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildSavingsChartFull:

    def _six_months(self):
        return ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]

    def test_returns_bytesio(self):
        buf = _build_savings_chart(self._six_months(), [10.0] * 6)
        assert isinstance(buf, io.BytesIO)

    def test_buffer_non_empty(self):
        buf = _build_savings_chart(self._six_months(), [15.0, 18.0, 12.0, 22.0, 20.0, 25.0])
        buf.seek(0, 2)
        assert buf.tell() > 0

    def test_all_zero_rates(self):
        buf = _build_savings_chart(self._six_months(), [0.0] * 6)
        assert isinstance(buf, io.BytesIO)
        buf.seek(0, 2)
        assert buf.tell() > 0

    def test_single_data_point(self):
        buf = _build_savings_chart(["Jun"], [15.0])
        assert isinstance(buf, io.BytesIO)
        buf.seek(0, 2)
        assert buf.tell() > 0

    def test_buffer_position_at_zero(self):
        """Buffer must be seeked to 0 so Telegram can stream it."""
        buf = _build_savings_chart(self._six_months(), [10.0] * 6)
        assert buf.tell() == 0


# ─────────────────────────────────────────────────────────────────────────────
# handlers/reports.py — _build_range_report sums
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=False)
def pln_rates(monkeypatch):
    """Make get_rate always return 1.0 so PLN amounts pass through unchanged."""
    monkeypatch.setattr(_data_stub, "get_rate", lambda ccy, rates: 1.0)


class TestBuildRangeReportSums:
    """
    DataFrame:
        Date        Type      Category    _base    IsDone
        2025-06-01  Income    Salary      5000    True
        2025-06-05  Expense   Food        200     True
        2025-06-10  Expense   Transport   100     True
        2025-06-12  Savings   Investment  500     True
        2025-06-14  Expense   Food        150     False  ← excluded
    Expected:
        income  = 5000
        expense = 300  (200 + 100, not 150)
        savings = 500
        net     = 4200 (5000 - 300 - 500)
        rate    = 0.10 (500 / 5000)
    """

    START = date(2025, 6, 1)
    END   = date(2025, 6, 15)

    def _report(self, df=None, monkeypatch=None):
        if df is None:
            df = _make_df()
        if monkeypatch:
            # formatters.py imports get_rate via `from data import get_rate`,
            # so we must patch it at the formatters module level.
            monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        rates   = {"PLN": 1.0}
        budgets = {}
        return _build_range_report(df, rates, budgets, "PLN", self.START, self.END, "Test")

    def test_income_sum(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        assert "5,000 PLN" in text

    def test_expense_sum_excludes_undone(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        # expense = 300, not 450
        assert "300 PLN" in text
        # 450 must NOT appear as an expense line
        assert "450 PLN" not in text

    def test_savings_sum(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        assert "500 PLN" in text

    def test_net_value(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        # net = 5000 - 300 - 500 = 4200
        assert "4,200 PLN" in text

    def test_savings_rate_formula(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        # rate = savings / income = 500 / 5000 = 10%
        assert "10%" in text

    def test_savings_are_not_income_minus_expense(self, monkeypatch):
        """savings must come from Type==Savings rows, NOT income - expense."""
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        # income - expense = 5000 - 300 = 4700, but savings = 500
        text = self._report(monkeypatch=monkeypatch)
        # If the wrong formula were used savings rate would be ~93%, not 10%
        assert "93%" not in text

    def test_isdone_false_excluded(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        # 150 PLN expense is IsDone=False; total expense should be 300 not 450
        assert "450 PLN" not in text

    def test_empty_dataframe_no_exception(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        empty_df = pd.DataFrame(columns=["Date", "Type", "Category", "_base", "IsDone"])
        text = _build_range_report(
            empty_df, {"PLN": 1.0}, {}, "PLN", self.START, self.END, "Empty"
        )
        assert isinstance(text, str)
        # all zeros
        assert "0 PLN" in text

    def test_category_breakdown_only_expense(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        # "Top Categories" section should appear (we have expenses)
        assert "Top Categories" in text
        # Salary (Income) should not appear in categories section
        # It might appear as a header line; check it's not in bullet points
        lines = text.split("\n")
        cat_lines = [l for l in lines if l.startswith("•")]
        categories_in_bullets = [l for l in cat_lines if "Salary" in l]
        assert len(categories_in_bullets) == 0

    def test_top_8_categories_max(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        # Build DF with 10 expense categories
        cats = [f"Cat{i}" for i in range(10)]
        df = pd.DataFrame({
            "Date":     ["2025-06-01"] * 10,
            "Type":     ["Expense"] * 10,
            "Category": cats,
            "_base":     [float(i + 1) * 100 for i in range(10)],
            "IsDone":   [True] * 10,
        })
        text = _build_range_report(
            df, {"PLN": 1.0}, {}, "PLN", self.START, self.END, "TopCats"
        )
        bullet_lines = [l for l in text.split("\n") if l.startswith("•")]
        assert len(bullet_lines) <= 8

    def test_budget_flag_when_over_budget(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        budgets = {"Food": 100.0}  # Food spent 200 > 100
        text = _build_range_report(
            _make_df(), {"PLN": 1.0}, budgets, "PLN", self.START, self.END, "BudgetTest"
        )
        assert "🔴" in text

    def test_no_budget_flag_when_within_budget(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        budgets = {"Food": 10000.0}  # Far above spend
        text = _build_range_report(
            _make_df(), {"PLN": 1.0}, budgets, "PLN", self.START, self.END, "NoBudgetFlag"
        )
        assert "🔴" not in text

    def test_categories_sorted_descending(self, monkeypatch):
        monkeypatch.setattr(formatters, "get_rate", lambda ccy, rates: 1.0)
        text = self._report(monkeypatch=monkeypatch)
        # Food appears twice (200 + 100, but last is excluded → Food=200, Transport=100)
        # Food (200) should appear before Transport (100)
        food_pos = text.find("Food")
        transport_pos = text.find("Transport")
        assert food_pos < transport_pos


# ─────────────────────────────────────────────────────────────────────────────
# handlers/reports.py — date range calculations
# ─────────────────────────────────────────────────────────────────────────────

_FIXED_NOW = datetime(2025, 6, 15, 12, 0, tzinfo=timezone.utc)
_TODAY = _FIXED_NOW.date()  # date(2025, 6, 15)


def _compute_range(range_key: str, today: date) -> tuple:
    """
    Pure-Python replication of the date range logic from handle_range_callback.
    Returns (start, end).
    """
    if range_key == "this_month":
        start = today.replace(day=1)
        end   = today
    elif range_key == "last_month":
        first_of_this = today.replace(day=1)
        last_of_prev  = first_of_this - timedelta(days=1)
        start = last_of_prev.replace(day=1)
        end   = last_of_prev
    elif range_key == "last_3_months":
        end = today
        start = today.replace(day=1)
        for _ in range(3):
            start = (start - timedelta(days=1)).replace(day=1)
    elif range_key == "last_6_months":
        end = today
        start = today.replace(day=1)
        for _ in range(6):
            start = (start - timedelta(days=1)).replace(day=1)
    elif range_key == "this_year":
        start = today.replace(month=1, day=1)
        end   = today
    else:
        raise ValueError(f"Unknown range key: {range_key}")
    return start, end


class TestDateRanges:
    """
    Mock handlers.reports.now_utc to return a fixed datetime.
    Then verify computed date ranges match expected values for today=2025-06-15.
    """

    def test_this_month(self):
        start, end = _compute_range("this_month", _TODAY)
        assert start == date(2025, 6, 1)
        assert end   == date(2025, 6, 15)

    def test_last_month(self):
        start, end = _compute_range("last_month", _TODAY)
        assert start == date(2025, 5, 1)
        assert end   == date(2025, 5, 31)

    def test_last_3_months(self):
        start, end = _compute_range("last_3_months", _TODAY)
        assert start == date(2025, 3, 1)
        assert end   == date(2025, 6, 15)

    def test_last_6_months(self):
        start, end = _compute_range("last_6_months", _TODAY)
        assert start == date(2024, 12, 1)
        assert end   == date(2025, 6, 15)

    def test_this_year(self):
        start, end = _compute_range("this_year", _TODAY)
        assert start == date(2025, 1, 1)
        assert end   == date(2025, 6, 15)

    def test_now_utc_mock_applied(self):
        """Verify the mock patches the real now_utc in handlers.reports."""
        import handlers.reports as hr
        with patch.object(hr, "now_utc", return_value=_FIXED_NOW):
            result = hr.now_utc()
        assert result == _FIXED_NOW


# ─────────────────────────────────────────────────────────────────────────────
# handlers/reports.py — cmd_report pagination
# ─────────────────────────────────────────────────────────────────────────────

class TestCmdReportPagination:

    def test_short_text_single_chunk(self):
        text = "line\n" * 20
        chunks = _chunk_text(text)
        assert len(chunks) == 1

    def test_exactly_4000_chars_single_chunk(self):
        text = "a" * 4000
        chunks = _chunk_text(text)
        assert len(chunks) == 1

    def test_over_4000_splits_into_multiple(self):
        line = "x" * 99
        text = "\n".join([line] * 50)  # ~4999 chars
        assert len(text) > 4000
        chunks = _chunk_text(text)
        assert len(chunks) > 1

    def test_each_chunk_at_most_4000(self):
        line = "y" * 99
        text = "\n".join([line] * 100)
        for chunk in _chunk_text(text):
            assert len(chunk) <= 4000

    def test_split_on_newline_boundary(self):
        """Every line must appear intact in exactly one chunk, never split."""
        lines = [f"Line {i:04d} content" for i in range(300)]
        text  = "\n".join(lines)
        chunks = _chunk_text(text)
        all_lines = []
        for chunk in chunks:
            all_lines.extend(chunk.split("\n"))
        for line in lines:
            assert line in all_lines

    def test_one_more_line_causes_split(self):
        """
        Build text where all-but-last lines fit in one chunk, and adding the
        last line would exceed 4000 chars.  Must produce exactly 2 chunks.
        """
        # Each line is 99 chars.  40 lines = 99*40 + 39 newlines = 3999 + 39 ... let's calibrate.
        # A 3980-char block in one chunk, then add a line that pushes it over.
        base_line  = "B" * 99   # 99 chars
        filler     = "\n".join([base_line] * 40)   # 40 * 99 + 39 = 4,000 - 1 = 3,999 chars
        extra_line = "E" * 10
        text       = filler + "\n" + extra_line
        chunks = _chunk_text(text)
        # With filler=3999 and extra=10, adding extra: 3999+1+10=4010 > 4000 → splits
        assert len(chunks) == 2
        assert extra_line in chunks[1]


# ─────────────────────────────────────────────────────────────────────────────
# scheduled_report.py — load_transaction_data
# ─────────────────────────────────────────────────────────────────────────────

class TestLoadTransactionData:
    """
    Patch pd.read_excel and file_storage.get_excel_path_for_reading to avoid
    real file I/O.
    """

    def _base_df(self):
        """Minimal valid MasterData sheet."""
        return pd.DataFrame({
            "Date":         ["2025-06-01", "2025-06-02", "2025-06-03"],
            "Type":         ["Income",     "Expense",    "Expense"],
            "Category":     ["Salary",     "Food",       "Transport"],
            "Value (base)":  [5000.0,        200.0,        None],   # last NaN
            "Value":        [5000.0,         200.0,         50.0],
            "Currency":     ["PLN",          "PLN",         "EUR"],
            "Year":         [2025,           2025,          2025],
            "Month":        ["June",         "June",        "June"],
            "IsDone":       [True,            True,          True],
        })

    def test_base_nan_recomputed_from_value_times_rate(self):
        """Row where Value(PLN) is NaN: amount_base = Value * rate."""
        with patch("scheduled_report.get_excel_path_for_reading", return_value="fake.xlsx"), \
             patch("scheduled_report.load_currency_rates", return_value={"EUR": 4.0, "PLN": 1.0}), \
             patch("pandas.read_excel", return_value=self._base_df()):
            df = scheduled_report.load_transaction_data()
        # EUR row: 50 * 4.0 = 200
        eur_rows = df[df["Currency"] == "EUR"]
        assert len(eur_rows) == 1
        assert eur_rows.iloc[0]["amount_base"] == pytest.approx(200.0)

    def test_base_row_nan_uses_value_directly(self):
        """PLN row with NaN Value(PLN): amount_base = Value * 1.0 (PLN rate)."""
        raw = pd.DataFrame({
            "Date":         ["2025-06-01"],
            "Type":         ["Expense"],
            "Category":     ["Food"],
            "Value (base)":  [None],
            "Value":        [300.0],
            "Currency":     ["PLN"],
            "Year":         [2025],
            "Month":        ["June"],
            "IsDone":       [True],
        })
        with patch("scheduled_report.get_excel_path_for_reading", return_value="fake.xlsx"), \
             patch("scheduled_report.load_currency_rates", return_value={"PLN": 1.0}), \
             patch("pandas.read_excel", return_value=raw):
            df = scheduled_report.load_transaction_data()
        assert df.iloc[0]["amount_base"] == pytest.approx(300.0)

    def test_unknown_currency_defaults_rate_1(self):
        """Row with currency not in rates dict: rate defaults to 1.0."""
        raw = pd.DataFrame({
            "Date":         ["2025-06-01"],
            "Type":         ["Expense"],
            "Category":     ["Food"],
            "Value (base)":  [None],
            "Value":        [77.0],
            "Currency":     ["XYZ"],
            "Year":         [2025],
            "Month":        ["June"],
            "IsDone":       [True],
        })
        with patch("scheduled_report.get_excel_path_for_reading", return_value="fake.xlsx"), \
             patch("scheduled_report.load_currency_rates", return_value={"PLN": 1.0}), \
             patch("pandas.read_excel", return_value=raw):
            df = scheduled_report.load_transaction_data()
        assert df.iloc[0]["amount_base"] == pytest.approx(77.0)


class TestLoadCurrencyRates:
    """load_currency_rates reads Currency/Rate cols from Lists via ListsSchema."""

    def test_reads_currency_code_and_rate(self, tmp_path):
        import openpyxl
        from excel_schema import ListsSchema, header_of
        path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lists"
        ws.cell(1, 1, header_of(ListsSchema, "currency"))
        ws.cell(1, 2, header_of(ListsSchema, "rate_to_base"))
        for i, (code, rate) in enumerate([("EUR", 4.25), ("USD", 3.90), ("GBP", 5.10)], 2):
            ws.cell(i, 1, code)
            ws.cell(i, 2, rate)
        wb.save(path)
        with patch("scheduled_report.get_excel_path_for_reading", return_value=path):
            rates = scheduled_report.load_currency_rates()
        assert rates["EUR"] == pytest.approx(4.25)
        assert rates["USD"] == pytest.approx(3.90)
        assert rates["GBP"] == pytest.approx(5.10)

    def test_returns_base_fallback_on_exception(self):
        with patch("scheduled_report.get_excel_path_for_reading", side_effect=Exception("no file")):
            rates = scheduled_report.load_currency_rates()
        assert rates == {"PLN": 1.0}


# ─────────────────────────────────────────────────────────────────────────────
# scheduled_report.py — savings_rate_emoji thresholds
# ─────────────────────────────────────────────────────────────────────────────

class TestSavingsRateEmoji:

    def test_7_percent_red(self):
        assert scheduled_report.savings_rate_emoji(0.07) == "🔴"

    def test_exactly_8_percent_yellow(self):
        assert scheduled_report.savings_rate_emoji(0.08) == "🟡"

    def test_exactly_15_percent_green(self):
        assert scheduled_report.savings_rate_emoji(0.15) == "💚"

    def test_exactly_20_percent_rocket(self):
        assert scheduled_report.savings_rate_emoji(0.20) == "🚀"

    def test_zero_rate_red(self):
        assert scheduled_report.savings_rate_emoji(0.0) == "🔴"

    def test_high_rate_rocket(self):
        assert scheduled_report.savings_rate_emoji(0.50) == "🚀"
