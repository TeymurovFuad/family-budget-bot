"""
test_file_storage.py — tests for file_storage.py helper functions.

All tests use the `excel_path` fixture which provides a fresh blank Excel
in a temp directory with file_storage.LOCAL_XLSX_PATH already monkeypatched.
"""

import json

import openpyxl
import pytest

import file_storage
from file_storage import (
    create_blank_excel,
    get_excel_path_for_reading,
    load_lists,
    load_budgets_from_excel,
    get_recent_transactions,
    load_user_prefs,
    save_user_prefs,
)


# ── create_blank_excel ────────────────────────────────────────────────────────


class TestCreateBlankExcel:

    def test_file_is_created(self, excel_path):
        assert excel_path.exists()

    def test_has_masterdata_sheet(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert "MasterData" in wb.sheetnames

    def test_has_lists_sheet(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert "Lists" in wb.sheetnames

    def test_has_dashboard_sheet(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert "Dashboard" in wb.sheetnames

    def test_masterdata_has_13_headers(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        headers = [ws.cell(1, c).value for c in range(1, 14)]
        assert len([h for h in headers if h is not None]) == 13

    def test_masterdata_header_names(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        expected = [
            "Date", "Year", "Month", "Value", "Type", "Category",
            "Person", "Description", "IsRecurring", "IsDone",
            "Currency", "Value (base)", "Date Modified (UTC)",
        ]
        actual = [ws.cell(1, c).value for c in range(1, 14)]
        assert actual == expected

    def test_lists_sheet_has_month_header_in_col_a(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(1, 1).value == "Months"

    def test_lists_sheet_has_category_header_in_col_c(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(1, 3).value == "Categories"

    def test_lists_sheet_has_currency_header_in_col_h(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(1, 8).value == "Currency"

    def test_lists_sheet_has_rate_header_in_col_i(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(1, 9).value == "Rate to base"

    def test_lists_sheet_has_budget_header_in_col_d(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(1, 4).value == "Budget (base)"

    def test_lists_sheet_col_g_is_empty(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(1, 7).value is None

    def test_lists_sheet_months_start_with_jan(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(2, 1).value == "Jan"

    def test_lists_sheet_months_end_with_dec(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(13, 1).value == "Dec"

    def test_lists_sheet_budget_column_is_blank_by_default(self, excel_path):
        # Budget (base) col is present but unpopulated — user fills in limits
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        for c in range(1, ws.max_column + 1):
            if "budget" in str(ws.cell(1, c).value or "").lower():
                for row in range(2, ws.max_row + 1):
                    assert ws.cell(row, c).value is None, f"Budget col row {row} should be blank"
                break

    def test_pln_rate_is_1(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 8).value == "PLN":
                assert ws.cell(row, 9).value == 1.0
                found = True
                break
        assert found, "PLN not found in Lists currencies"


# ── get_excel_path_for_reading ────────────────────────────────────────────────


class TestGetExcelPathForReading:

    def test_returns_existing_path_for_local_backend(self, excel_path):
        result = get_excel_path_for_reading()
        assert result == excel_path

    def test_auto_creates_file_if_missing(self, tmp_path, monkeypatch):
        missing_path = tmp_path / "auto_created.xlsx"
        monkeypatch.setattr(file_storage, "LOCAL_XLSX_PATH", missing_path)
        result = get_excel_path_for_reading()
        assert result.exists()
        wb = openpyxl.load_workbook(result)
        assert "MasterData" in wb.sheetnames


# ── load_lists ────────────────────────────────────────────────────────────────


class TestLoadLists:

    def test_returns_dict_with_required_keys(self, excel_path):
        result = load_lists(excel_path)
        assert {"months", "txn_types", "categories", "persons", "years", "budgets"}.issubset(result.keys())

    def test_months_has_12_entries(self, excel_path):
        result = load_lists(excel_path)
        assert len(result["months"]) == 12

    def test_months_starts_with_jan(self, excel_path):
        result = load_lists(excel_path)
        assert result["months"][0] == "Jan"

    def test_months_ends_with_dec(self, excel_path):
        result = load_lists(excel_path)
        assert result["months"][-1] == "Dec"

    def test_txn_types_contains_expense_income_savings(self, excel_path):
        result = load_lists(excel_path)
        assert "Expense" in result["txn_types"]
        assert "Income" in result["txn_types"]
        assert "Savings" in result["txn_types"]

    def test_categories_is_non_empty_list(self, excel_path):
        result = load_lists(excel_path)
        assert isinstance(result["categories"], list)
        assert len(result["categories"]) > 0

    def test_categories_contains_groceries(self, excel_path):
        result = load_lists(excel_path)
        assert "Groceries" in result["categories"]

    def test_persons_is_a_list(self, excel_path):
        result = load_lists(excel_path)
        assert isinstance(result["persons"], list)

    def test_years_is_non_empty_list(self, excel_path):
        result = load_lists(excel_path)
        assert len(result["years"]) > 0

    def test_years_are_integers(self, excel_path):
        result = load_lists(excel_path)
        for y in result["years"]:
            assert isinstance(y, int)

    def test_categories_contains_income_type_entry(self, excel_path):
        result = load_lists(excel_path)
        assert "Salary" in result["categories"], "Unified categories must include income-type entries"

    def test_categories_contains_expense_type_entry(self, excel_path):
        result = load_lists(excel_path)
        assert "Groceries" in result["categories"], "Unified categories must include expense-type entries"

    def test_categories_contains_savings_type_entry(self, excel_path):
        result = load_lists(excel_path)
        assert "Bank Deposit" in result["categories"], "Unified categories must include savings-type entries"


# ── load_budgets_from_excel ───────────────────────────────────────────────────


class TestLoadBudgetsFromExcel:

    def _find_budget_col(self, ws):
        for c in range(1, ws.max_column + 1):
            if "budget" in str(ws.cell(1, c).value or "").lower():
                return c
        return None

    def _find_category_col(self, ws):
        for c in range(1, ws.max_column + 1):
            if "categor" in str(ws.cell(1, c).value or "").lower():
                return c
        return None

    def test_returns_dict(self, excel_path):
        result = load_budgets_from_excel(excel_path)
        assert isinstance(result, dict)

    def test_blank_excel_has_no_budgets(self, excel_path):
        # Blank Excel has no budget amounts filled in — all categories have 0 or None
        result = load_budgets_from_excel(excel_path)
        assert isinstance(result, dict)
        assert len(result) == 0

    def test_returns_numeric_budget_written_as_float(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        cat_col = self._find_category_col(ws)
        bud_col = self._find_budget_col(ws)
        assert cat_col and bud_col, "Lists sheet must have Category and Budget columns"
        # Write a category with a limit in the next empty row
        next_row = ws.max_row + 1
        ws.cell(next_row, cat_col).value = "SpecialCat"
        ws.cell(next_row, bud_col).value = 1200.0
        wb.save(excel_path)
        result = load_budgets_from_excel(excel_path)
        assert "SpecialCat" in result
        assert result["SpecialCat"] == pytest.approx(1200.0)

    def test_skips_categories_with_zero_budget(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        cat_col = self._find_category_col(ws)
        bud_col = self._find_budget_col(ws)
        next_row = ws.max_row + 1
        ws.cell(next_row, cat_col).value = "ZeroCat"
        ws.cell(next_row, bud_col).value = 0
        wb.save(excel_path)
        result = load_budgets_from_excel(excel_path)
        assert "ZeroCat" not in result

    def test_skips_categories_with_no_budget(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        cat_col = self._find_category_col(ws)
        next_row = ws.max_row + 1
        ws.cell(next_row, cat_col).value = "NoBudgetCat"
        wb.save(excel_path)
        result = load_budgets_from_excel(excel_path)
        assert "NoBudgetCat" not in result

    def test_returns_empty_dict_on_missing_lists_sheet(self, tmp_path):
        path = tmp_path / "no_lists.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "MasterData"
        wb.save(path)
        result = load_budgets_from_excel(path)
        assert result == {}


# ── get_recent_transactions ───────────────────────────────────────────────────


class TestGetRecentTransactions:

    def test_empty_masterdata_returns_empty_list(self, excel_path):
        result = get_recent_transactions(excel_path)
        assert result == []

    def test_returns_list(self, excel_path):
        result = get_recent_transactions(excel_path)
        assert isinstance(result, list)


# ── user preferences ──────────────────────────────────────────────────────────


class TestUserPrefs:

    def test_load_user_prefs_returns_empty_dict_when_no_file(self, excel_path, monkeypatch, tmp_path):
        # USER_PREFS_PATH is patched to tmp_path/user_prefs.json by excel_path fixture
        # which doesn't exist yet — should return {}
        result = load_user_prefs()
        assert result == {}

    def test_save_and_load_user_prefs_roundtrip(self, excel_path, tmp_path):
        prefs = {"display_currency": "EUR", "user_id": 12345}
        save_user_prefs(prefs)
        loaded = load_user_prefs()
        assert loaded == prefs

    def test_save_user_prefs_creates_file(self, excel_path, tmp_path):
        prefs_path = file_storage.USER_PREFS_PATH
        assert not prefs_path.exists()  # Should not exist yet
        save_user_prefs({"key": "value"})
        assert prefs_path.exists()

    def test_save_user_prefs_overwrites_existing(self, excel_path):
        save_user_prefs({"v": 1})
        save_user_prefs({"v": 2})
        loaded = load_user_prefs()
        assert loaded["v"] == 2


def test_repair_template_keeps_validation_ranges(tmp_path, monkeypatch):
    """delete_rows during template cleanup must not collapse dropdown ranges."""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    import file_storage

    # Build a template-like file: header + 500 stale data rows + validation to row 578
    src = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MasterData"
    for i, h in enumerate(["Date", "Year", "Month", "Value", "Type", "Category",
                           "Person", "Description", "IsRecurring", "IsDone",
                           "Currency", "Value (base)", "Date Modified (UTC)"], 1):
        ws.cell(1, i, h)
    for r in range(2, 500):
        ws.cell(r, 4, 1.0)
    dv = DataValidation(type="list", formula1="Lists!$C$2:$C$18")
    dv.add("F2:F578")
    ws.add_data_validation(dv)
    wb.create_sheet("Lists")
    wb.save(src)

    file_storage._repair_template_workbook(src)

    wb = openpyxl.load_workbook(src)
    ws = wb["MasterData"]
    sqrefs = [str(d.sqref) for d in ws.data_validations.dataValidation]
    assert sqrefs, "validations were dropped entirely"
    end_row = int(sqrefs[0].split(":F")[1])
    assert end_row >= 500, f"validation range collapsed: {sqrefs}"


# ── update_transaction_field — multi-field dict updates ──────────────────────
# Year/Month sync on date edits is domain logic that lives in the edit
# handler (handlers/edit_conv.py); the storage layer just writes whatever
# fields it is given — atomically, in one save, when passed a dict.


class TestMultiFieldUpdate:
    def _seed_row(self, excel_path):
        from datetime import date
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        ws.cell(2, 1, date(2024, 5, 12))   # Date
        ws.cell(2, 2, 2024)                # Year
        ws.cell(2, 3, "May")               # Month
        ws.cell(2, 4, 45.0)                # Value
        ws.cell(2, 8, "shop")              # Description
        wb.save(excel_path)

    def test_dict_updates_all_fields_in_one_write(self, excel_path):
        from datetime import date
        self._seed_row(excel_path)
        file_storage.update_transaction_field(
            2, {"Date": date(2023, 12, 31), "Year": 2023, "Month": "Dec"}
        )
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        assert ws.cell(2, 1).value.strftime("%Y-%m-%d") == "2023-12-31"
        assert ws.cell(2, 2).value == 2023
        assert ws.cell(2, 3).value == "Dec"

    def test_date_only_edit_does_not_touch_year_month(self, excel_path):
        """Storage stays generic: writing Date alone must not rewrite Year/Month."""
        from datetime import date
        self._seed_row(excel_path)
        file_storage.update_transaction_field(2, "Date", date(2023, 12, 31))
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        assert ws.cell(2, 2).value == 2024
        assert ws.cell(2, 3).value == "May"

    def test_non_date_edit_leaves_year_month_untouched(self, excel_path):
        self._seed_row(excel_path)
        file_storage.update_transaction_field(2, "Value", 99.0)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        assert ws.cell(2, 2).value == 2024
        assert ws.cell(2, 3).value == "May"


# ── quarantine + .bak fixes ───────────────────────────────────────────────────


class TestFlushRecoveryQueueQuarantine:

    def test_quarantine_rename_failure_falls_back_to_delete(self, tmp_path, monkeypatch):
        """If the .corrupt rename fails, the corrupt file must be deleted so the
        next flush doesn't hit the same JSONDecodeError forever."""
        from unittest.mock import patch as _patch

        queue_path = tmp_path / "recovery_queue.json"
        queue_path.write_text("{not valid json", encoding="utf-8")
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        with _patch.object(type(queue_path), "replace", side_effect=PermissionError("locked")):
            rows = file_storage.flush_recovery_queue()

        assert rows == []
        assert not queue_path.exists(), "corrupt queue file must not survive the flush"

    def test_quarantine_rename_success_moves_file(self, tmp_path, monkeypatch):
        queue_path = tmp_path / "recovery_queue.json"
        queue_path.write_text("{not valid json", encoding="utf-8")
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        rows = file_storage.flush_recovery_queue()

        assert rows == []
        assert not queue_path.exists()
        assert (tmp_path / "recovery_queue.json.corrupt").exists()


class TestAtomicSaveBackupPolicy:

    def test_no_bak_written_for_temp_download_files(self, tmp_path):
        """Remote-backend temp downloads must not accumulate orphan .bak files."""
        from file_storage import atomic_save, _temp_files

        path = tmp_path / "download.xlsx"
        wb = openpyxl.Workbook()
        wb.save(path)  # pre-existing file so the .bak branch is reachable

        _temp_files.add(path)
        try:
            atomic_save(wb, path)
        finally:
            _temp_files.discard(path)

        assert path.exists()
        assert not (tmp_path / "download.xlsx.bak").exists()

    def test_bak_still_written_for_regular_files(self, tmp_path):
        from file_storage import atomic_save

        path = tmp_path / "workbook.xlsx"
        wb = openpyxl.Workbook()
        wb.save(path)
        atomic_save(wb, path)

        assert (tmp_path / "workbook.xlsx.bak").exists()


# ── backend selection: explicit STORAGE_BACKEND must win ─────────────────────


class TestActiveBackendOverride:

    def test_explicit_local_wins_over_stray_gcs_bucket(self, monkeypatch):
        """A leftover GCS_BUCKET_NAME must not flip an explicit local backend."""
        import settings
        monkeypatch.setattr(settings, "STORAGE_BACKEND_EXPLICIT", True)
        monkeypatch.setattr(file_storage, "STORAGE_BACKEND", "local")
        monkeypatch.setattr(file_storage, "GCS_BUCKET_NAME", "stray-bucket")
        assert file_storage._active_backend() == "local"

    def test_explicit_local_wins_over_stray_s3_bucket(self, monkeypatch):
        import settings
        monkeypatch.setattr(settings, "STORAGE_BACKEND_EXPLICIT", True)
        monkeypatch.setattr(file_storage, "STORAGE_BACKEND", "local")
        monkeypatch.setattr(file_storage, "S3_BUCKET_NAME", "stray-bucket")
        assert file_storage._active_backend() == "local"

    def test_bucket_name_still_selects_backend_when_not_explicit(self, monkeypatch):
        """Backward compat: with no explicit STORAGE_BACKEND, a bucket var selects the backend."""
        import settings
        monkeypatch.setattr(settings, "STORAGE_BACKEND_EXPLICIT", False)
        monkeypatch.setattr(file_storage, "STORAGE_BACKEND", "local")
        monkeypatch.setattr(file_storage, "GCS_BUCKET_NAME", "my-bucket")
        assert file_storage._active_backend() == "gcs"

    def test_default_is_local(self, monkeypatch):
        import settings
        monkeypatch.setattr(settings, "STORAGE_BACKEND_EXPLICIT", False)
        monkeypatch.setattr(file_storage, "STORAGE_BACKEND", "local")
        monkeypatch.setattr(file_storage, "GCS_BUCKET_NAME", "")
        monkeypatch.setattr(file_storage, "S3_BUCKET_NAME", "")
        assert file_storage._active_backend() == "local"


# ── recovery queue: append-only JSONL journal ─────────────────────────────────


class TestRecoveryQueueJsonl:

    def _patch_queue(self, tmp_path, monkeypatch):
        queue_path = tmp_path / "recovery_queue.json"
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)
        return queue_path

    def test_append_is_one_jsonl_line_per_call(self, tmp_path, monkeypatch):
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        file_storage.append_to_recovery_queue({"value": 1.0})
        file_storage.append_to_recovery_queue({"value": 2.0})
        lines = [l for l in queue_path.read_text(encoding="utf-8").splitlines() if l.strip()]
        assert len(lines) == 2
        first = json.loads(lines[0])
        assert first["op"] == "append"
        assert first["row"] == {"value": 1.0}

    def test_flush_returns_rows_from_jsonl(self, tmp_path, monkeypatch):
        self._patch_queue(tmp_path, monkeypatch)
        file_storage.append_to_recovery_queue({"value": 1.0})
        file_storage.append_to_recovery_queue({"value": 2.0})
        rows = file_storage.flush_recovery_queue()
        assert rows == [{"value": 1.0}, {"value": 2.0}]

    def test_flush_migrates_legacy_json_list(self, tmp_path, monkeypatch):
        """Old-format queue (whole-file JSON array) must still be readable."""
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        queue_path.write_text(json.dumps([{"value": 1.0}, {"value": 2.0}]), encoding="utf-8")
        rows = file_storage.flush_recovery_queue()
        assert rows == [{"value": 1.0}, {"value": 2.0}]

    def test_flush_reads_legacy_then_jsonl_appends_mixed(self, tmp_path, monkeypatch):
        """Appending to a file that still holds the legacy array must lose nothing."""
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        queue_path.write_text(json.dumps([{"value": 1.0}]) + "\n", encoding="utf-8")
        file_storage.append_to_recovery_queue({"value": 2.0})
        rows = file_storage.flush_recovery_queue()
        assert rows == [{"value": 1.0}, {"value": 2.0}]

    def test_flush_skips_unsupported_ops(self, tmp_path, monkeypatch):
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        queue_path.write_text(
            json.dumps({"op": "append", "row": {"value": 1.0}}) + "\n"
            + json.dumps({"op": "frobnicate", "row": {"value": 9.0}}) + "\n",
            encoding="utf-8",
        )
        rows = file_storage.flush_recovery_queue()
        assert rows == [{"value": 1.0}]

    def test_flush_does_not_delete_file(self, tmp_path, monkeypatch):
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        file_storage.append_to_recovery_queue({"value": 1.0})
        file_storage.flush_recovery_queue()
        assert queue_path.exists()

    def test_delete_removes_file(self, tmp_path, monkeypatch):
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        file_storage.append_to_recovery_queue({"value": 1.0})
        file_storage.delete_recovery_queue_file()
        assert not queue_path.exists()

    def test_requeue_rows_writes_readable_queue(self, tmp_path, monkeypatch):
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        file_storage.requeue_rows([{"value": 3.0}])
        assert queue_path.exists()
        assert file_storage.flush_recovery_queue() == [{"value": 3.0}]

    def test_requeue_empty_writes_nothing(self, tmp_path, monkeypatch):
        queue_path = self._patch_queue(tmp_path, monkeypatch)
        file_storage.requeue_rows([])
        assert not queue_path.exists()


# ── lost-update protection (remote backends) ──────────────────────────────────


class _FakeBlob:
    """Records download/upload calls and mimics GCS generation semantics."""

    def __init__(self, generation=7, fail_upload_with=None):
        self.generation = generation
        self.upload_kwargs = None
        self.uploaded_from = None
        self._fail_upload_with = fail_upload_with

    def download_to_filename(self, filename):
        openpyxl.Workbook().save(filename)

    def reload(self):
        pass

    def upload_from_filename(self, filename, **kwargs):
        if self._fail_upload_with is not None:
            raise self._fail_upload_with
        self.uploaded_from = filename
        self.upload_kwargs = kwargs


class _FakeGcsClient:
    def __init__(self, blob):
        self._blob = blob

    def bucket(self, name):
        return self

    def blob(self, name):
        return self._blob


class PreconditionFailed(Exception):
    """Stand-in for google.api_core.exceptions.PreconditionFailed (matched by name)."""


class TestLostUpdateProtection:

    def _gcs_env(self, monkeypatch, blob):
        import settings
        import storage_backends
        monkeypatch.setattr(settings, "STORAGE_BACKEND_EXPLICIT", True)
        monkeypatch.setattr(file_storage, "STORAGE_BACKEND", "gcs")
        monkeypatch.setattr(file_storage, "GCS_BUCKET_NAME", "bucket")
        monkeypatch.setattr(storage_backends, "_gcs_client",
                            lambda: _FakeGcsClient(blob))

    def test_upload_carries_generation_precondition(self, monkeypatch):
        """The generation captured at download must gate the upload."""
        blob = _FakeBlob(generation=7)
        self._gcs_env(monkeypatch, blob)

        with file_storage.ExcelFileContext() as path:
            assert path.exists()
        assert blob.upload_kwargs == {"if_generation_match": 7}

    def test_conflict_raises_concurrent_modification_error(self, monkeypatch):
        """A precondition failure means another writer won — never overwrite."""
        blob = _FakeBlob(generation=7, fail_upload_with=PreconditionFailed("409"))
        self._gcs_env(monkeypatch, blob)

        with pytest.raises(file_storage.ConcurrentModificationError):
            with file_storage.ExcelFileContext():
                pass

    def test_upload_without_generation_is_unconditional(self, monkeypatch, tmp_path):
        """If the generation could not be determined, upload proceeds (old behaviour)."""
        import storage_backends
        blob = _FakeBlob(generation=7)
        self._gcs_env(monkeypatch, blob)

        some_file = tmp_path / "wb.xlsx"
        openpyxl.Workbook().save(some_file)
        storage_backends._upload_from_local_file(some_file, generation=None)
        assert blob.upload_kwargs == {}


# ── audit log lines ───────────────────────────────────────────────────────────


class TestAuditLog:

    def test_audit_span_emits_ok_line(self, caplog):
        import logging
        from audit import audit_span

        with caplog.at_level(logging.INFO, logger="audit"):
            with audit_span("append", rows=1, user=123):
                pass
        lines = [r.getMessage() for r in caplog.records if "AUDIT save" in r.getMessage()]
        assert len(lines) == 1
        assert "user=123" in lines[0]
        assert "source=append" in lines[0]
        assert "rows=1" in lines[0]
        assert "outcome=ok" in lines[0]
        assert "duration_ms=" in lines[0]

    def test_audit_span_emits_error_line_and_reraises(self, caplog):
        import logging
        from audit import audit_span

        with caplog.at_level(logging.INFO, logger="audit"):
            with pytest.raises(ValueError):
                with audit_span("append_batch", rows=3):
                    raise ValueError("boom")
        lines = [r.getMessage() for r in caplog.records if "AUDIT save" in r.getMessage()]
        assert len(lines) == 1
        assert "outcome=error" in lines[0]
        assert "rows=3" in lines[0]
        assert "user=-" in lines[0]
