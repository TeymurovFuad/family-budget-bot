"""
test_wave2_group_a.py — Group A PLN-neutrality: storage/schema layer only.

Tests cover:
  1. data.load_data   — Currency column default uses settings.DISPLAY_CURRENCY
  2. scheduled_report — fillna uses DISPLAY_CURRENCY
  3. excel_schema.write_transaction_row — currency default uses settings.DISPLAY_CURRENCY
  4. excel_schema.write_transaction_row — Value formula contains DISPLAY_CURRENCY
  5. ListsSchema.goal_base exists; goal_pln does not
  6. migrate_pln_headers_to_base.RENAMES includes "Goal (PLN)" → "Goal"
"""

import datetime
import sys
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


# ── helpers ───────────────────────────────────────────────────────────────────

def _blank_masterdata_ws():
    from dataclasses import fields
    from excel_schema import MasterDataSchema, header_of
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MasterData"
    for i, f in enumerate(fields(MasterDataSchema), 1):
        ws.cell(1, i, header_of(MasterDataSchema, f.name))
    return ws


def _build_masterdata_excel(path, rows: list[dict]) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["MasterData"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for i, row in enumerate(rows, start=2):
        for col_name, val in row.items():
            col_idx = headers.get(col_name)
            if col_idx:
                ws.cell(i, col_idx).value = val
    wb.save(path)


# ── test 1: data.load_data Currency default uses settings.DISPLAY_CURRENCY ───

class TestLoadDataCurrencyDefault:

    def test_load_data_currency_default_uses_settings(self, excel_path, monkeypatch):
        import settings
        import data as data_mod

        monkeypatch.setattr(settings, "DISPLAY_CURRENCY", "EUR")
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        # Build a row with no Currency column
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        row_num = 2
        ws.cell(row_num, headers["Date"]).value = datetime.date(2024, 1, 1)
        ws.cell(row_num, headers["Year"]).value = 2024
        ws.cell(row_num, headers["Month"]).value = "Jan"
        ws.cell(row_num, headers["Value"]).value = 100.0
        ws.cell(row_num, headers["Type"]).value = "Expense"
        ws.cell(row_num, headers["Value (base)"]).value = 100.0
        ws.cell(row_num, headers["IsDone"]).value = True
        # Leave Currency blank (None)
        wb.save(excel_path)

        df = data_mod.load_data()
        assert len(df) == 1
        assert df.iloc[0]["Currency"] == "EUR"


# ── test 2: scheduled_report fillna uses DISPLAY_CURRENCY ────────────────────

class TestScheduledReportFillna:

    def test_scheduled_report_fillna_uses_display_currency(self, excel_path, monkeypatch):
        import scheduled_report as sr

        monkeypatch.setattr(sr, "DISPLAY_CURRENCY", "CHF")
        monkeypatch.setattr(sr, "get_excel_path_for_reading", lambda: excel_path)

        # Row with Currency = None
        _build_masterdata_excel(excel_path, [{
            "Date":       datetime.date(2024, 3, 1),
            "Year":       2024,
            "Month":      "Mar",
            "Value":      50.0,
            "Value (base)": 50.0,
            "Type":       "Expense",
            "IsDone":     True,
            # Currency intentionally absent → will become NaN → fillna
        }])

        df = sr.load_transaction_data()
        assert len(df) == 1
        assert df.iloc[0]["Currency"] == "CHF"


# ── test 3: write_transaction_row currency default uses settings ──────────────

class TestWriteTransactionRowCurrencyDefault:

    LU = "$H$2:$I$100"

    def test_write_transaction_row_currency_default_uses_settings(self, monkeypatch):
        import settings
        from excel_schema import write_transaction_row, col_indices, MasterDataSchema

        monkeypatch.setattr(settings, "DISPLAY_CURRENCY", "EUR")

        ws = _blank_masterdata_ws()
        row = {
            "date":  datetime.date(2024, 1, 1),
            "year":  2024,
            "month": "Jan",
            "value": 10.0,
            "type":  "Expense",
            # "currency" key intentionally omitted
        }
        write_transaction_row(ws, 2, row, self.LU)

        idx = col_indices(ws, MasterDataSchema)
        assert ws.cell(2, idx["currency"]).value == "EUR"


# ── test 4: Value formula uses DISPLAY_CURRENCY ───────────────────────────────

class TestValueBaseFormulaUsesDisplayCurrency:

    LU = "$H$2:$I$100"

    def test_value_base_formula_uses_display_currency(self, monkeypatch):
        import settings
        from excel_schema import write_transaction_row, col_indices, MasterDataSchema

        monkeypatch.setattr(settings, "DISPLAY_CURRENCY", "EUR")

        ws = _blank_masterdata_ws()
        row = {
            "date":     datetime.date(2024, 1, 1),
            "year":     2024,
            "month":    "Jan",
            "value":    10.0,
            "type":     "Expense",
            "currency": "EUR",
        }
        write_transaction_row(ws, 2, row, self.LU)

        idx = col_indices(ws, MasterDataSchema)
        formula = ws.cell(2, idx["value_base"]).value
        assert "EUR" in formula
        assert "PLN" not in formula


# ── test 5: goal_base exists; goal_pln does not ───────────────────────────────

class TestListsSchemaGoalField:

    def test_goal_base_field_exists_goal_pln_removed(self):
        from excel_schema import ListsSchema
        assert hasattr(ListsSchema, "goal_base")
        assert not hasattr(ListsSchema, "goal_pln")


# ── test 6: migrate script includes goal_pln rename ──────────────────────────

class TestMigrateScriptGoalRename:

    def test_migrate_script_includes_goal_rename(self):
        import importlib.util
        scripts_dir = str(PROJECT_ROOT / "scripts")
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)
        script_path = PROJECT_ROOT / "scripts" / "migrate_base_currency_headers.py"
        spec = importlib.util.spec_from_file_location("migrate_base_currency_headers", script_path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        renames = mod.RENAMES
        assert "Goal (PLN)" in renames
        assert renames["Goal (PLN)"] == "Goal"


# ── test 7: load_dedup_evidence null-currency fallback uses settings ──────────

class TestLoadDedupEvidenceNullCurrency:

    def test_load_dedup_evidence_null_currency_uses_settings(self, excel_path, monkeypatch):
        import settings
        import data as data_mod

        monkeypatch.setattr(settings, "DISPLAY_CURRENCY", "EUR")
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        # Write a row with a null Currency value
        _build_masterdata_excel(excel_path, [{
            "Date":         datetime.date(2024, 6, 1),
            "Year":         2024,
            "Month":        "Jun",
            "Value":        99.0,
            "Value (base)": 99.0,
            "Type":         "Expense",
            "IsDone":       True,
            # Currency intentionally absent → NaN → should fall back to settings.DISPLAY_CURRENCY
        }])

        evidence = data_mod.load_dedup_evidence()
        from validators import make_dedup_key, make_loose_dedup_key
        expected_strict = make_dedup_key("2024-06-01", 99.0, "EUR", "")
        expected_loose  = make_loose_dedup_key("2024-06-01", 99.0, "EUR")
        unexpected_strict = make_dedup_key("2024-06-01", 99.0, "PLN", "")
        unexpected_loose  = make_loose_dedup_key("2024-06-01", 99.0, "PLN")

        assert expected_strict in evidence["strict"], (
            f"Expected strict key for EUR not found. Keys: {list(evidence['strict'].keys())}"
        )
        assert expected_loose in evidence["loose"], (
            f"Expected loose key for EUR not found. Keys: {list(evidence['loose'].keys())}"
        )
        assert unexpected_strict not in evidence["strict"], "PLN strict key should not be present"
        assert unexpected_loose  not in evidence["loose"],  "PLN loose key should not be present"
