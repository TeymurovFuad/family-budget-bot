"""
tests/test_cycles.py — budget-cycles core: settings flag, Cycles sheet ledger,
/cycle handler, salary-triggered prompt, and cycle-scoped /summary.

No AI calls anywhere in this feature — nothing to mock on that front.
"""

import os
import sys
from datetime import date, timedelta
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import pandas as pd
import pytest

os.environ.setdefault("STORAGE_BACKEND", "local")
os.environ.setdefault("TELEGRAM_BOT_TOKEN", "dummy")
os.environ.setdefault("ALLOWED_TELEGRAM_IDS", "123")

PROJECT_ROOT = Path(__file__).parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import settings
import cycles
from cycles import (
    CYCLES_SHEET_NAME, cycle_label, cycle_totals, current_cycle_start,
    ensure_cycles_sheet, load_cycles, record_cycle_start, should_prompt_new_cycle,
)
from excel_schema import CyclesSchema, header_of
from handlers.cycle import cmd_cycle, handle_cycle_callback, maybe_prompt_cycle_start


def make_update(text="", user_id=123):
    upd = MagicMock()
    upd.effective_user.id = user_id
    upd.message.text = text
    upd.message.reply_text = AsyncMock()
    return upd


def make_callback_update(data, user_id=123):
    upd = MagicMock()
    upd.effective_user.id = user_id
    upd.message = None
    upd.callback_query.data = data
    upd.callback_query.from_user.id = user_id
    upd.callback_query.answer = AsyncMock()
    upd.callback_query.message.reply_text = AsyncMock()
    return upd


def make_ctx(args=None):
    ctx = MagicMock()
    ctx.args = args or []
    ctx.user_data = {}
    return ctx


def make_transaction(txn_type="Income", category="Salary", txn_date=None):
    t = MagicMock()
    t.transaction_type = txn_type
    t.category = category
    t.date = txn_date or date(2026, 7, 23)
    return t


# ── settings flag ──────────────────────────────────────────────────────────────

def test_budget_cycle_flag_off_by_default():
    assert settings.BUDGET_CYCLE is False
    assert settings.CYCLE_REPROMPT_MIN_AGE_DAYS == 20
    assert settings.SALARY_CATEGORY == "Salary"


# ── ledger ─────────────────────────────────────────────────────────────────────

def test_cycle_label_always_carries_year():
    assert cycle_label(date(2026, 8, 25)) == "Aug 2026"
    assert cycle_label(date(2025, 1, 2)) == "Jan 2025"


def test_ensure_cycles_sheet_creates_headers():
    from openpyxl import Workbook
    wb = Workbook()
    ws = ensure_cycles_sheet(wb)
    assert CYCLES_SHEET_NAME in wb.sheetnames
    assert ws.cell(1, 1).value == header_of(CyclesSchema, "start_date")
    assert ws.cell(1, 2).value == header_of(CyclesSchema, "label")
    assert ensure_cycles_sheet(wb) is ws


def test_record_and_load_cycles(excel_path):
    assert load_cycles() == []
    assert record_cycle_start(date(2026, 6, 25)) == "Jun 2026"
    assert record_cycle_start(date(2026, 7, 23)) == "Jul 2026"
    got = load_cycles()
    assert got == [(date(2026, 6, 25), "Jun 2026"), (date(2026, 7, 23), "Jul 2026")]


def test_record_duplicate_boundary_is_noop(excel_path):
    assert record_cycle_start(date(2026, 7, 23)) == "Jul 2026"
    assert record_cycle_start(date(2026, 7, 23)) is None
    assert len(load_cycles()) == 1


def test_current_cycle_start_picks_latest_past_boundary():
    ledger = [(date(2026, 5, 24), "May 2026"), (date(2026, 6, 25), "Jun 2026")]
    assert current_cycle_start(date(2026, 7, 1), ledger) == (date(2026, 6, 25), "Jun 2026")
    assert current_cycle_start(date(2026, 6, 1), ledger) == (date(2026, 5, 24), "May 2026")
    assert current_cycle_start(date(2026, 5, 1), ledger) is None
    assert current_cycle_start(date(2026, 7, 1), []) is None


def test_should_prompt_new_cycle_age_gate(excel_path):
    today = date(2026, 7, 23)
    assert should_prompt_new_cycle(today) is True  # no ledger yet
    record_cycle_start(today - timedelta(days=5))
    assert should_prompt_new_cycle(today) is False  # too young
    record_cycle_start(today - timedelta(days=settings.CYCLE_REPROMPT_MIN_AGE_DAYS))
    # latest boundary is the 5-day-old one, still too young
    assert should_prompt_new_cycle(today) is False


def test_should_prompt_new_cycle_old_cycle(excel_path):
    today = date(2026, 7, 23)
    record_cycle_start(today - timedelta(days=25))
    assert should_prompt_new_cycle(today) is True


# ── unaccounted math ───────────────────────────────────────────────────────────

def _cycle_df():
    return pd.DataFrame({
        "Date":     ["2026-06-25", "2026-06-26", "2026-07-01", "2026-07-02", "2026-06-01"],
        "Type":     ["Income",     "Income",     "Expense",    "Savings",    "Expense"],
        "Category": ["Salary",     "Freelance",  "Groceries",  "Bank Deposit", "Groceries"],
        "_base":     [6000.0,       900.0,        1500.0,       1000.0,       999.0],
        "IsDone":   [True,         True,         True,         True,         True],
    })


def test_cycle_totals_unaccounted_uses_salary_only():
    totals = cycle_totals(_cycle_df(), date(2026, 6, 25), date(2026, 7, 23))
    assert totals["income"] == 6900.0
    assert totals["salary"] == 6000.0
    assert totals["expense"] == 1500.0  # 999 row is before the cycle start
    assert totals["savings"] == 1000.0
    assert totals["unaccounted"] == 6000.0 - 1500.0 - 1000.0


def test_cycle_totals_negative_unaccounted_means_over_reported():
    df = _cycle_df()
    df.loc[df["Category"] == "Groceries", "_base"] = 7000.0
    totals = cycle_totals(df, date(2026, 6, 25), date(2026, 7, 23))
    assert totals["unaccounted"] < 0


# ── /cycle list and /cycle remove ──────────────────────────────────────────────

async def test_cmd_cycle_list_shows_all_boundaries(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 5, 24))
    record_cycle_start(date(2026, 6, 25))
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["list"]))
    text = upd.message.reply_text.call_args[0][0]
    assert "May 2026" in text and "2026-05-24 → 2026-06-24" in text
    assert "Jun 2026" in text and "2026-06-25 → today" in text


async def test_cmd_cycle_list_empty_ledger(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["list"]))
    assert "No cycle boundaries" in upd.message.reply_text.call_args[0][0]


async def test_cmd_cycle_remove_deletes_boundary(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 5, 24))
    record_cycle_start(date(2026, 6, 25))
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["remove", "2026-05-24"]))
    assert "Removed" in upd.message.reply_text.call_args[0][0]
    assert load_cycles() == [(date(2026, 6, 25), "Jun 2026")]


async def test_cmd_cycle_remove_unknown_date(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 6, 25))
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["remove", "2026-01-01"]))
    assert "No cycle boundary" in upd.message.reply_text.call_args[0][0]
    assert len(load_cycles()) == 1


async def test_cmd_cycle_remove_bad_args(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["remove"]))
    assert "Usage" in upd.message.reply_text.call_args[0][0]
    upd2 = make_update()
    await cmd_cycle(upd2, make_ctx(["remove", "not-a-date"]))
    assert "Could not parse" in upd2.message.reply_text.call_args[0][0]


def test_remove_cycle_start_roundtrip(excel_path):
    record_cycle_start(date(2026, 6, 25))
    assert cycles.remove_cycle_start(date(2026, 6, 25)) is True
    assert cycles.remove_cycle_start(date(2026, 6, 25)) is False
    assert load_cycles() == []
    assert record_cycle_start(date(2026, 6, 26)) == "Jun 2026"


# ── detect_cycle_candidates ────────────────────────────────────────────────────

def _detect_df():
    """Rows shaped like real bulk-imported salary data: category empty,
    'Salary' in Description."""
    return pd.DataFrame({
        "Date":        ["2024-07-01", "2024-08-01", "2024-08-01", "2024-07-01"],
        "Type":        ["Income",     "Income",     "Income",     "Expense"],
        "Category":    ["",           "",           "",           "Groceries"],
        "Description": ["Salary",     "Salary",     "Salary",     ""],
        "_base":        [12027.0,      11871.0,      11856.0,      2000.0],
        "IsDone":      [True,         True,         True,         True],
    })


def test_detect_matches_salary_in_description():
    results = cycles.detect_cycle_candidates(_detect_df(), existing_cycles=[])
    assert [r["date"] for r in results] == [date(2024, 7, 1), date(2024, 8, 1)]
    assert results[0]["unambiguous"] is True
    assert results[0]["amounts"] == [12027.0]
    assert results[1]["unambiguous"] is False
    assert results[1]["amounts"] == [11871.0, 11856.0]


def test_detect_skips_already_recorded_dates():
    results = cycles.detect_cycle_candidates(
        _detect_df(), existing_cycles=[(date(2024, 7, 1), "Jul 2024")]
    )
    assert [r["date"] for r in results] == [date(2024, 8, 1)]


def test_detect_matches_salary_in_category_without_description_column():
    df = _cycle_df()
    assert "Description" not in df.columns  # pins the guard this test covers
    results = cycles.detect_cycle_candidates(df, existing_cycles=[])
    assert [r["date"] for r in results] == [date(2026, 6, 25)]


def test_detect_contains_match_on_bank_transfer_titles():
    df = _detect_df()
    df["Description"] = [
        "WYNAGRODZENIE ZA LIPIEC ACME SP Z OO",
        "SALARY JUL 2024",
        "salary",
        "",
    ]
    results = cycles.detect_cycle_candidates(
        df, existing_cycles=[], extra_keywords=["wynagrodzenie"]
    )
    assert [r["date"] for r in results] == [date(2024, 7, 1), date(2024, 8, 1)]


def test_detect_extra_keywords_from_settings(monkeypatch):
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["wynagrodzenie"])
    df = _detect_df()
    df["Description"] = ["Wynagrodzenie", "x", "y", ""]
    results = cycles.detect_cycle_candidates(df, existing_cycles=[])
    assert [r["date"] for r in results] == [date(2024, 7, 1)]


def test_salary_mask_empty_keyword_matches_nothing(monkeypatch):
    monkeypatch.setattr(settings, "SALARY_CATEGORY", "")
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", [])
    assert not cycles.salary_mask(_detect_df()).any()


def test_salary_mask_empty_category_with_stored_keyword(monkeypatch):
    """With SALARY_CATEGORY="" and one stored keyword, rows containing that
    keyword in Description (when Category is blank) are masked; others are not.
    No empty-alternation regex explosion occurs."""
    monkeypatch.setattr(settings, "SALARY_CATEGORY", "")
    monkeypatch.setattr(settings, "CYCLE_DETECT_KEYWORDS", ["payroll"])
    monkeypatch.setattr(cycles, "load_salary_keywords", lambda excel_path=None: [])
    df = pd.DataFrame({
        "Date":        ["2026-01-10", "2026-01-15", "2026-01-20"],
        "Type":        ["Income",     "Income",     "Income"],
        "Category":    ["",           "",           "Freelance"],
        "Description": ["payroll jan", "bonus",     "payroll jan"],
        "_base":        [5000.0,        500.0,        3000.0],
        "IsDone":      [True,          True,          True],
    })
    mask = cycles.salary_mask(df)
    # row 0: category blank, description "payroll jan" -> matches
    # row 1: category blank, description "bonus"       -> no match
    # row 2: category "Freelance" (non-blank)          -> description ignored
    assert list(mask) == [True, False, False]


def test_salary_mask_description_ignored_when_category_present():
    df = _detect_df()
    df["Category"] = ["Freelance", "", "Salary Bonus", "Groceries"]
    df["Description"] = ["Salary", "SALARY JUL 2024", "x", ""]
    mask = cycles.salary_mask(df)
    assert list(mask) == [False, True, True, False]


def test_salary_mask_nan_category_falls_back_to_description():
    df = _detect_df()
    df["Category"] = [float("nan"), None, float("nan"), "Groceries"]
    df["Description"] = ["Salary", "SALARY JUL 2024", "x", ""]
    mask = cycles.salary_mask(df)
    assert list(mask) == [True, True, False, False]


def test_salary_keyword_not_matched_as_substring():
    df = _detect_df()
    df["Description"] = ["salaryman payment", "x", "y", ""]
    df["Category"] = ["", "", "", "Groceries"]
    assert not cycles.salary_mask(df).any()


# ── /cycle command ─────────────────────────────────────────────────────────────

async def test_cmd_cycle_flag_off_is_inert(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", False)
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["started"]))
    assert "disabled" in upd.message.reply_text.call_args[0][0]
    assert load_cycles() == []


async def test_cmd_cycle_started_with_date(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["started", "2026-07-01"]))
    assert load_cycles() == [(date(2026, 7, 1), "Jul 2026")]
    assert "✅" in upd.message.reply_text.call_args[0][0]


async def test_cmd_cycle_started_defaults_to_today(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["started"]))
    ledger = load_cycles()
    assert len(ledger) == 1


async def test_cmd_cycle_detect_warns_extra_keywords_not_saved(excel_path, monkeypatch):
    """Per-scan `/cycle detect <word>` keywords are not persisted — the user
    must be told to add them via /keywords."""
    import handlers.cycle as hc
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    monkeypatch.setattr(hc, "load_data", lambda: pd.DataFrame())
    monkeypatch.setattr(hc, "load_cycles", lambda: [])
    candidates = [{"date": date(2026, 6, 25), "amounts": [5000.0], "unambiguous": True}]
    monkeypatch.setattr(hc, "detect_cycle_candidates", lambda df, cyc, extra: candidates)

    upd = make_update()
    await cmd_cycle(upd, make_ctx(["detect", "bonuspay"]))

    texts = [c.args[0] for c in upd.message.reply_text.call_args_list]
    assert any("this scan only" in t and "/keywords" in t for t in texts), texts


async def test_cmd_cycle_detect_no_warning_without_extra_keywords(excel_path, monkeypatch):
    import handlers.cycle as hc
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    monkeypatch.setattr(hc, "load_data", lambda: pd.DataFrame())
    monkeypatch.setattr(hc, "load_cycles", lambda: [])
    candidates = [{"date": date(2026, 6, 25), "amounts": [5000.0], "unambiguous": True}]
    monkeypatch.setattr(hc, "detect_cycle_candidates", lambda df, cyc, extra: candidates)

    upd = make_update()
    await cmd_cycle(upd, make_ctx(["detect"]))

    texts = [c.args[0] for c in upd.message.reply_text.call_args_list]
    assert not any("not saved" in t for t in texts), texts


async def test_cmd_cycle_rejects_bad_and_future_dates(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["started", "not-a-date"]))
    assert "Could not parse" in upd.message.reply_text.call_args[0][0]
    await cmd_cycle(upd, make_ctx(["started", "2099-01-01"]))
    assert "future" in upd.message.reply_text.call_args[0][0]
    assert load_cycles() == []


async def test_cmd_cycle_bare_shows_status(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await cmd_cycle(upd, make_ctx())
    assert "No budget cycle recorded" in upd.message.reply_text.call_args[0][0]
    record_cycle_start(date(2026, 7, 1))
    await cmd_cycle(upd, make_ctx())
    assert "Jul 2026" in upd.message.reply_text.call_args[0][0]


async def test_cmd_cycle_duplicate_reports_noop(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 7, 1))
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["started", "2026-07-01"]))
    assert "already recorded" in upd.message.reply_text.call_args[0][0]


async def test_cmd_cycle_owner_only(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update(user_id=999)
    await cmd_cycle(upd, make_ctx(["started"]))
    assert "not authorized" in upd.message.reply_text.call_args[0][0]
    assert load_cycles() == []


# ── salary-triggered prompt ────────────────────────────────────────────────────

async def test_maybe_prompt_flag_off_no_prompt(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", False)
    upd = make_update()
    await maybe_prompt_cycle_start(upd, make_transaction())
    upd.message.reply_text.assert_not_called()


async def test_maybe_prompt_salary_income_prompts_with_wording(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await maybe_prompt_cycle_start(upd, make_transaction(txn_date=date(2026, 7, 23)))
    text = upd.message.reply_text.call_args[0][0]
    assert text.startswith("💰 Salary received. Start the new budget cycle from 23 Jul 2026?")
    assert "(yes / no / different date)" in text
    markup = upd.message.reply_text.call_args.kwargs["reply_markup"]
    callbacks = [b.callback_data for row in markup.inline_keyboard for b in row]
    assert callbacks == ["cycle:yes:2026-07-23", "cycle:no", "cycle:diff"]


async def test_maybe_prompt_ignores_non_salary(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    upd = make_update()
    await maybe_prompt_cycle_start(upd, make_transaction(txn_type="Expense"))
    await maybe_prompt_cycle_start(upd, make_transaction(category="Freelance"))
    upd.message.reply_text.assert_not_called()


async def test_maybe_prompt_young_cycle_stays_silent(excel_path, monkeypatch):
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    from datetime import datetime
    from config import TIMEZONE
    today = datetime.now(TIMEZONE).date()
    record_cycle_start(today - timedelta(days=3))
    upd = make_update()
    await maybe_prompt_cycle_start(upd, make_transaction(txn_date=today))
    upd.message.reply_text.assert_not_called()


# ── prompt callback ────────────────────────────────────────────────────────────

async def test_cycle_callback_yes_records_boundary(excel_path):
    upd = make_callback_update("cycle:yes:2026-07-23")
    await handle_cycle_callback(upd, make_ctx())
    assert load_cycles() == [(date(2026, 7, 23), "Jul 2026")]
    assert "Jul 2026" in upd.callback_query.message.reply_text.call_args[0][0]


async def test_cycle_callback_no_keeps_current_cycle(excel_path):
    upd = make_callback_update("cycle:no")
    await handle_cycle_callback(upd, make_ctx())
    assert load_cycles() == []
    assert "current cycle continues" in upd.callback_query.message.reply_text.call_args[0][0]


async def test_cycle_callback_diff_points_at_command(excel_path):
    upd = make_callback_update("cycle:diff")
    await handle_cycle_callback(upd, make_ctx())
    assert load_cycles() == []
    assert "/cycle started" in upd.callback_query.message.reply_text.call_args[0][0]


async def test_cycle_callback_owner_only(excel_path):
    upd = make_callback_update("cycle:yes:2026-07-23", user_id=999)
    await handle_cycle_callback(upd, make_ctx())
    assert load_cycles() == []


# ── cycle-scoped reports ───────────────────────────────────────────────────────

def _patch_report_data(monkeypatch, df):
    import handlers.reports as reports
    monkeypatch.setattr(reports, "load_data", lambda: df)
    monkeypatch.setattr(reports, "load_rates", lambda: {"PLN": 1.0})
    monkeypatch.setattr(reports, "load_budgets", lambda: {"Groceries": 2000.0})
    monkeypatch.setattr(
        reports, "load_reference_data",
        lambda: {"categories": ["Groceries", "Salary"]},
    )


async def test_summary_cycle_scoped_with_unaccounted(excel_path, monkeypatch):
    """'This cycle' quick button renders the cycle report with unaccounted."""
    from handlers.reports import handle_summary_callback
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 6, 25))
    df = _cycle_df()
    df["Year"] = 2026
    df["Month"] = "Jul"
    _patch_report_data(monkeypatch, df)
    upd = make_callback_update("sum:tc")
    await handle_summary_callback(upd, make_ctx())
    text = upd.callback_query.message.reply_text.call_args[0][0]
    assert "Cycle Jun 2026" in text
    assert "Unaccounted" in text
    assert "Salary received" in text
    assert "Projected month-end" not in text


async def test_summary_falls_back_to_calendar_without_boundary(excel_path, monkeypatch):
    """No boundary recorded → 'This month' quick button gives the calendar report."""
    from data import current_year_and_month
    from handlers.reports import handle_summary_callback
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    year, month = current_year_and_month()
    df = _cycle_df()
    df["Year"] = year
    df["Month"] = month
    _patch_report_data(monkeypatch, df)
    upd = make_callback_update("sum:tm")
    await handle_summary_callback(upd, make_ctx())
    text = upd.callback_query.message.reply_text.call_args[0][0]
    assert f"{month} {year} — Summary" in text
    assert "Unaccounted" not in text


async def test_summary_flag_off_is_calendar(excel_path, monkeypatch):
    """Flag off — bare /summary shows the picker without any cycle buttons."""
    from data import current_year_and_month
    from handlers.reports import cmd_summary, handle_summary_callback
    monkeypatch.setattr(settings, "BUDGET_CYCLE", False)
    record_cycle_start(date(2026, 6, 25))
    year, month = current_year_and_month()
    df = _cycle_df()
    df["Year"] = year
    df["Month"] = month
    _patch_report_data(monkeypatch, df)

    upd = make_update()
    await cmd_summary(upd, make_ctx())
    keyboard = upd.message.reply_text.call_args.kwargs["reply_markup"]
    labels = [b.text for row in keyboard.inline_keyboard for b in row]
    assert "This month" in labels and "Last month" in labels
    assert "This cycle" not in labels and "💰 Cycle" not in labels

    upd2 = make_callback_update("sum:tm")
    await handle_summary_callback(upd2, make_ctx())
    assert f"{month} {year} — Summary" in upd2.callback_query.message.reply_text.call_args[0][0]


async def test_budget_bars_cycle_scoped(excel_path, monkeypatch):
    from handlers.reports import cmd_budget
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 6, 25))
    df = _cycle_df()
    df["Year"] = 2026
    df["Month"] = "Jul"
    _patch_report_data(monkeypatch, df)
    upd = make_update()
    await cmd_budget(upd, make_ctx())
    text = upd.message.reply_text.call_args[0][0]
    assert "Cycle Jun 2026" in text
    # 999 PLN pre-cycle expense excluded: only the 1 500 in-cycle row counts
    assert "1,500" in text.replace(" ", ",")


# ── Wave 2 Group D: cycle_periods / before-cycles bucket ──────────────────────

def _periods_df():
    return pd.DataFrame({
        "Date":   ["2026-05-01", "2026-06-26", "2026-07-05"],
        "Type":   ["Expense",    "Expense",    "Expense"],
        "Category": ["Groceries", "Groceries", "Groceries"],
        "_base":   [100.0,        200.0,        300.0],
        "IsDone": [True, True, True],
    })


def test_cycle_periods_before_cycles_bucket():
    ledger = [(date(2026, 6, 25), "Jun 2026"), (date(2026, 7, 23), "Jul 2026")]
    periods = cycles.cycle_periods(_periods_df(), ledger, today=date(2026, 7, 30))
    assert periods[0] == (date(2026, 5, 1), date(2026, 6, 24), cycles.BEFORE_CYCLES_LABEL)
    assert periods[1] == (date(2026, 6, 25), date(2026, 7, 22), "Jun 2026")
    assert periods[2] == (date(2026, 7, 23), date(2026, 7, 30), "Jul 2026")


def test_cycle_periods_no_older_rows_no_bucket():
    ledger = [(date(2026, 4, 1), "Apr 2026")]
    periods = cycles.cycle_periods(_periods_df(), ledger, today=date(2026, 7, 30))
    assert [p[2] for p in periods] == ["Apr 2026"]


def test_cycle_periods_empty_ledger_returns_empty():
    assert cycles.cycle_periods(_periods_df(), [], today=date(2026, 7, 30)) == []


# ── Wave 2 Group D: detect_missing_boundaries ─────────────────────────────────

def test_detect_missing_boundaries_finds_gaps():
    ledger = [(date(2025, 7, 1), "Jul 2025"), (date(2025, 10, 2), "Oct 2025")]
    missing = cycles.detect_missing_boundaries(date(2025, 7, 1), date(2025, 10, 31), ledger)
    assert missing == [date(2025, 8, 1), date(2025, 9, 1)]


def test_detect_missing_boundaries_none_missing():
    ledger = [(date(2025, 7, 1), "Jul 2025"), (date(2025, 8, 1), "Aug 2025")]
    assert cycles.detect_missing_boundaries(date(2025, 7, 1), date(2025, 8, 31), ledger) == []


# ── Wave 2 Group D: fallback income candidates ────────────────────────────────

def _fallback_df():
    return pd.DataFrame({
        "Date":     ["2024-07-28", "2024-08-01", "2024-08-05", "2024-08-10", "2024-09-15"],
        "Type":     ["Income",     "Income",     "Income",     "Income",     "Income"],
        "Category": ["Refund",     "Transfer",   "Gift",       "Other",      "Other"],
        "Description": ["", "", "", "", ""],
        "_base":     [500.0,        9000.0,       50.0,         2000.0,       7000.0],
        "IsDone":   [True, True, True, True, True],
    })


def test_fallback_income_candidates_top3_in_window():
    got = cycles.fallback_income_candidates(_fallback_df(), date(2024, 8, 1), [])
    assert [c["date"] for c in got] == [date(2024, 8, 1), date(2024, 8, 10), date(2024, 7, 28)]
    assert got[0]["amounts"] == [9000.0]
    # 2024-09-15 is outside the +/-20-day window
    assert all(c["date"] != date(2024, 9, 15) for c in got)


def test_fallback_income_candidates_skips_recorded_and_empty_df():
    got = cycles.fallback_income_candidates(
        _fallback_df(), date(2024, 8, 1), [(date(2024, 8, 1), "Aug 2024")]
    )
    assert all(c["date"] != date(2024, 8, 1) for c in got)
    assert cycles.fallback_income_candidates(pd.DataFrame(), date(2024, 8, 1), []) == []


# ── Wave 2 Group D: detect review — none-this-month and multi-salary picker ───

def _detect_review_ctx(candidates):
    ctx = make_ctx()
    ctx.user_data["detect_candidates"] = [
        {"date_str": c[0], "amounts": c[1], "amounts_fmt": [f"{a:,.0f} PLN" for a in c[1]],
         "unambiguous": len(c[1]) == 1}
        for c in candidates
    ]
    return ctx


def _review_update():
    upd = make_callback_update("detect:review")
    upd.callback_query.edit_message_reply_markup = AsyncMock()
    upd.callback_query.edit_message_text = AsyncMock()
    return upd


async def test_detect_single_candidate_offers_none_this_month(excel_path):
    from handlers.cycle import handle_detect_callback
    ctx = _detect_review_ctx([("2026-06-25", [5000.0])])
    upd = _review_update()
    await handle_detect_callback(upd, ctx)
    markup = upd.callback_query.message.reply_text.call_args.kwargs["reply_markup"]
    callbacks = [b.callback_data for row in markup.inline_keyboard for b in row]
    assert "detect:pick:2026-06-25" in callbacks
    assert "detect:none:2026-06" in callbacks


async def test_detect_multi_candidates_one_button_each_largest_first(excel_path):
    from handlers.cycle import handle_detect_callback
    ctx = _detect_review_ctx([("2026-06-02", [3000.0]), ("2026-06-25", [8000.0])])
    upd = _review_update()
    await handle_detect_callback(upd, ctx)
    markup = upd.callback_query.message.reply_text.call_args.kwargs["reply_markup"]
    callbacks = [b.callback_data for row in markup.inline_keyboard for b in row]
    # largest amount first
    assert callbacks[0] == "detect:pick:2026-06-25"
    assert callbacks[1] == "detect:pick:2026-06-02"
    assert "detect:none:2026-06" in callbacks
    assert "detect:custom" in callbacks
    assert "detect:stop" in callbacks
    # both dates share the queue slot: one tap resolves the whole month
    assert ctx.user_data["detect_total"] == 1


async def test_detect_none_extends_previous_cycle(excel_path):
    from handlers.cycle import handle_detect_callback
    ctx = make_ctx()
    ctx.user_data["detect_queue"] = [[{"date_str": "2026-06-25", "amounts": [5000.0],
                                       "amounts_fmt": ["5,000 PLN"], "unambiguous": True}]]
    ctx.user_data["detect_total"] = 1
    upd = _review_update()
    upd.callback_query.data = "detect:none:2026-06"
    await handle_detect_callback(upd, ctx)
    text = upd.callback_query.edit_message_text.call_args[0][0]
    assert "No cycle in Jun 2026" in text
    assert "previous cycle" in text
    assert load_cycles() == []  # nothing recorded — "none" is a valid answer


async def test_detect_fallback_window_when_no_salary_rows(excel_path, monkeypatch):
    """No keyword hits -> largest income rows near the 1st offered as candidates."""
    import handlers.cycle as hc
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    monkeypatch.setattr(hc, "load_data", lambda: pd.DataFrame())
    monkeypatch.setattr(hc, "load_cycles", lambda: [])
    monkeypatch.setattr(hc, "detect_cycle_candidates", lambda df, cyc, extra: [])
    fb = [{"date": date(2026, 7, 1), "amounts": [9000.0], "unambiguous": True},
          {"date": date(2026, 7, 10), "amounts": [2000.0], "unambiguous": True}]
    monkeypatch.setattr(hc, "fallback_income_candidates", lambda df, anchor, cyc: fb)

    upd = make_update()
    await cmd_cycle(upd, make_ctx(["detect"]))

    texts = [c.args[0] for c in upd.message.reply_text.call_args_list]
    assert any("No salary" in t for t in texts), texts
    markup = upd.message.reply_text.call_args.kwargs["reply_markup"]
    callbacks = [b.callback_data for row in markup.inline_keyboard for b in row]
    assert "detect:pick:2026-07-01" in callbacks
    assert any(cb.startswith("detect:none:") for cb in callbacks)


async def test_detect_no_candidates_and_no_fallback_reports_nothing(excel_path, monkeypatch):
    import handlers.cycle as hc
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    monkeypatch.setattr(hc, "load_data", lambda: pd.DataFrame())
    monkeypatch.setattr(hc, "load_cycles", lambda: [])
    monkeypatch.setattr(hc, "detect_cycle_candidates", lambda df, cyc, extra: [])
    monkeypatch.setattr(hc, "fallback_income_candidates", lambda df, anchor, cyc: [])
    upd = make_update()
    await cmd_cycle(upd, make_ctx(["detect"]))
    texts = [c.args[0] for c in upd.message.reply_text.call_args_list]
    assert any("Nothing to backfill" in t for t in texts), texts


# ── Wave 2 Group D: reports — timezone, savings, before-cycles, backfill ─────

def test_current_cycle_bounds_uses_local_timezone(excel_path, monkeypatch):
    from datetime import datetime
    import handlers.reports as reports
    from settings import TIMEZONE
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 1, 1))
    got = reports._current_cycle_bounds()
    assert got is not None
    start, end, label = got
    assert end == datetime.now(TIMEZONE).date()


async def test_savings_cycle_aware_caption(excel_path, monkeypatch):
    import handlers.reports as reports
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 6, 25))
    _patch_report_data(monkeypatch, _cycle_df())
    upd = make_update()
    upd.message.reply_photo = AsyncMock()
    await reports.cmd_savings(upd, make_ctx())
    caption = upd.message.reply_photo.call_args.kwargs["caption"]
    assert "cycle" in caption and "month" not in caption


async def test_savings_calendar_when_flag_off(excel_path, monkeypatch):
    import handlers.reports as reports
    monkeypatch.setattr(settings, "BUDGET_CYCLE", False)
    df = _cycle_df()
    df["Year"] = 2026
    df["Month"] = "Jul"
    _patch_report_data(monkeypatch, df)
    upd = make_update()
    upd.message.reply_photo = AsyncMock()
    await reports.cmd_savings(upd, make_ctx())
    caption = upd.message.reply_photo.call_args.kwargs["caption"]
    assert "month" in caption


async def test_summary_entire_period_walk_includes_before_bucket(excel_path, monkeypatch):
    from datetime import datetime, timezone as _tz
    import handlers.reports as reports
    from handlers.reports import cmd_summary
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    monkeypatch.setattr(reports, "now_utc",
                        lambda: datetime(2026, 7, 27, 12, 0, tzinfo=_tz.utc))
    record_cycle_start(date(2026, 6, 25))
    record_cycle_start(date(2026, 7, 23))
    _patch_report_data(monkeypatch, _cycle_df())
    upd = make_update()
    await cmd_summary(upd, make_ctx(["all"]))
    texts = [c.args[0] for c in upd.message.reply_text.call_args_list]
    # 2026-06-01 expense predates the first boundary -> Before cycles bucket
    assert any("Before cycles" in t for t in texts), texts
    assert any("Cycle Jun 2026" in t for t in texts), texts
    # the before-cycles bucket has no salary anchor — no unaccounted math
    before = next(t for t in texts if "Before cycles" in t)
    assert "Unaccounted" not in before


async def test_summary_month_year_resolves_ledger_first(excel_path, monkeypatch):
    from handlers.reports import cmd_summary
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 6, 25))
    record_cycle_start(date(2026, 7, 23))
    _patch_report_data(monkeypatch, _cycle_df())
    upd = make_update()
    await cmd_summary(upd, make_ctx(["jun", "2026"]))
    text = upd.message.reply_text.call_args[0][0]
    assert "Cycle Jun 2026" in text  # ledger cycle, not calendar June


async def test_summary_month_year_calendar_when_no_ledger_label(excel_path, monkeypatch):
    from handlers.reports import cmd_summary
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 6, 25))
    df = _cycle_df()
    df["Year"] = 2026
    df["Month"] = "Mar"
    _patch_report_data(monkeypatch, df)
    upd = make_update()
    await cmd_summary(upd, make_ctx(["mar", "2026"]))
    text = upd.message.reply_text.call_args[0][0]
    assert "Mar 2026 — Summary" in text


async def test_summary_range_prompts_lazy_backfill(excel_path, monkeypatch):
    from handlers.reports import cmd_summary
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    record_cycle_start(date(2026, 5, 24))
    _patch_report_data(monkeypatch, _cycle_df())
    upd = make_update()
    ctx = make_ctx(["may", "2026", "-", "jul", "2026"])
    await cmd_summary(upd, ctx)
    text = upd.message.reply_text.call_args[0][0]
    assert "missing cycle boundaries" in text
    assert "Jun 2026" in text and "Jul 2026" in text
    markup = upd.message.reply_text.call_args.kwargs["reply_markup"]
    callbacks = [b.callback_data for row in markup.inline_keyboard for b in row]
    assert callbacks == ["sum:bf:yes", "sum:bf:skip"]
    assert ctx.user_data["sum_pending"]["kind"] == "range"


async def test_summary_backfill_skip_renders_pending(excel_path, monkeypatch):
    from handlers.reports import handle_summary_callback
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    _patch_report_data(monkeypatch, _cycle_df())
    upd = make_callback_update("sum:bf:skip")
    upd.callback_query.message.edit_reply_markup = AsyncMock()
    ctx = make_ctx()
    ctx.user_data["sum_pending"] = {
        "kind": "range", "start": date(2026, 6, 1), "end": date(2026, 7, 31),
        "label": "Jun 2026 - Jul 2026",
    }
    await handle_summary_callback(upd, ctx)
    text = upd.callback_query.message.reply_text.call_args[0][0]
    assert "Range Report" in text
    assert "sum_pending" not in ctx.user_data


async def test_summary_backfill_yes_points_to_detect(excel_path, monkeypatch):
    from handlers.reports import handle_summary_callback
    monkeypatch.setattr(settings, "BUDGET_CYCLE", True)
    _patch_report_data(monkeypatch, _cycle_df())
    upd = make_callback_update("sum:bf:yes")
    upd.callback_query.message.edit_text = AsyncMock()
    ctx = make_ctx()
    ctx.user_data["sum_pending"] = {"kind": "entire", "start": date(2026, 1, 1),
                                    "end": date(2026, 7, 31)}
    await handle_summary_callback(upd, ctx)
    text = upd.callback_query.message.edit_text.call_args[0][0]
    assert "/cycle detect" in text
    assert "sum_pending" not in ctx.user_data


def test_detect_missing_boundaries_empty_ledger_returns_empty():
    """Empty ledger = no boundaries to compare against = no gaps to fill."""
    assert cycles.detect_missing_boundaries(date(2025, 1, 1), date(2025, 3, 31), []) == []


def test_fallback_income_candidates_missing_base_column():
    df = _fallback_df().drop(columns=["_base"])
    assert cycles.fallback_income_candidates(df, date(2024, 8, 1), []) == []


def test_group_by_month_order_independent():
    from handlers.cycle import _group_by_month
    entries = [
        {"date_str": "2026-06-25"},
        {"date_str": "2026-05-24"},
        {"date_str": "2026-06-02"},
    ]
    groups = _group_by_month(entries)
    assert [[e["date_str"] for e in g] for g in groups] == [
        ["2026-05-24"], ["2026-06-02", "2026-06-25"],
    ]
