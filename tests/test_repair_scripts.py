"""
test_repair_scripts.py — regression tests for the one-off maintenance scripts
scripts/rename_category.py and scripts/fix_import_errors.py.

Covers:
- rename_category: formula string-literal rewriting on Dashboard
  (='"OldName"' criteria inside SUMIFS etc.) and pending bulk-draft JSON
  renaming in settings.BULK_DRAFTS_DIR.
- fix_import_errors: rule chaining — rule 3 (recipient moved into
  Description) must operate on the description already rewritten by rule 2
  (transfer-to-self -> Expense/Groceries), not the pre-rule-2 value.

The repair_guard is stubbed out (its own behavior is covered by
test_repair_guard.py); all file I/O happens under pytest tmp dirs.
"""
import json
import sys
from contextlib import contextmanager
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).parent.parent
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
for p in (str(PROJECT_ROOT), str(SCRIPTS_DIR)):
    if p not in sys.path:
        sys.path.insert(0, p)

import settings

import fix_import_errors
import rename_category


@contextmanager
def _no_guard():
    yield


def _make_workbook(path: Path) -> None:
    """Minimal workbook with the sheets/columns both scripts expect."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lists"
    ws["A1"] = "Categories"
    ws["A2"] = "OldName"
    ws["A3"] = "Groceries"
    ws["B1"] = "Persons"
    ws["B2"] = "Alice"

    md = wb.create_sheet("MasterData")
    for c, h in enumerate(["Date", "Type", "Category", "Value",
                           "Person", "Description"], start=1):
        md.cell(1, c, h)
    md.append(["2026-07-01", "Expense", "OldName", 10, "Alice", "weekly shop"])

    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "OldName"
    dash["B1"] = '=SUMIFS(MasterData!D:D,MasterData!C:C,"OldName")'

    wb.save(path)


class TestRenameCategory:
    def test_formula_string_literal_rewritten(self, tmp_path, monkeypatch):
        xlsx = tmp_path / "book.xlsx"
        _make_workbook(xlsx)
        monkeypatch.setattr(rename_category, "repair_guard", _no_guard)
        monkeypatch.setattr(settings, "BULK_DRAFTS_DIR", tmp_path / "no_drafts")
        monkeypatch.setattr(sys, "argv",
                            ["rename_category.py", "OldName", "NewName", str(xlsx)])

        rename_category.main()

        wb = openpyxl.load_workbook(xlsx, data_only=False)
        dash = wb["Dashboard"]
        assert dash["B1"].value == \
            '=SUMIFS(MasterData!D:D,MasterData!C:C,"NewName")'
        # Plain-value occurrences renamed too
        assert dash["A1"].value == "NewName"

    def test_bulk_draft_json_renamed(self, tmp_path, monkeypatch):
        xlsx = tmp_path / "book.xlsx"
        _make_workbook(xlsx)
        drafts = tmp_path / "bulk_drafts"
        drafts.mkdir(exist_ok=True)
        draft = drafts / "123.json"
        draft.write_text(json.dumps([
            {"category": "OldName", "value": 5},
            {"category": "Groceries", "value": 7},
        ]), encoding="utf-8")

        monkeypatch.setattr(rename_category, "repair_guard", _no_guard)
        monkeypatch.setattr(settings, "BULK_DRAFTS_DIR", drafts)
        monkeypatch.setattr(sys, "argv",
                            ["rename_category.py", "OldName", "NewName", str(xlsx)])

        rename_category.main()

        rows = json.loads(draft.read_text(encoding="utf-8"))
        assert rows[0]["category"] == "NewName"
        assert rows[1]["category"] == "Groceries"  # untouched


class TestFixImportErrorsRuleChaining:
    def test_rule3_sees_rule2_rewritten_description(self, tmp_path, monkeypatch):
        """A row hitting rule 2 (Savings transfer-to-self) AND rule 3
        (Person not a household member) must end with the rule-2 description
        plus the recipient appended — not the pre-rule-2 description."""
        xlsx = tmp_path / "book.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lists"
        ws["A1"] = "Persons"
        ws["A2"] = "Alice"
        md = wb.create_sheet("MasterData")
        for c, h in enumerate(["Date", "Type", "Category", "Value",
                               "Person", "Description"], start=1):
            md.cell(1, c, h)
        md.append(["2026-07-21", "Transfer", "Savings", 100,
                   "Stranger", "sent to self card"])
        wb.save(xlsx)

        monkeypatch.setattr(fix_import_errors, "repair_guard", _no_guard)
        monkeypatch.setattr(sys, "argv", ["fix_import_errors.py", str(xlsx)])

        fix_import_errors.main()

        out = openpyxl.load_workbook(xlsx)["MasterData"]
        assert out.cell(2, 2).value == "Expense"      # Type (rule 2)
        assert out.cell(2, 3).value == "Groceries"    # Category (rule 2)
        assert out.cell(2, 5).value is None           # Person cleared (rule 3)
        # Regression: description must chain rule 2 -> rule 3
        assert out.cell(2, 6).value == \
            "Transfer to own card for groceries — Stranger"
