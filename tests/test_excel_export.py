"""Tests for excel_export.generate_excel_from_sqlite — SQLite → workbook."""

import pytest
from openpyxl import load_workbook

import sqlite_ops
from excel_export import generate_excel_from_sqlite
from excel_schema import MasterDataSchema, col_indices


@pytest.fixture()
def template(excel_path):
    """The blank fixture workbook doubles as the export template."""
    return excel_path


@pytest.fixture()
def seeded_db(tmp_path):
    db = tmp_path / "export.db"
    conn = sqlite_ops.init_db(db)
    sqlite_ops.insert_transaction(conn, {
        "date": "2024-06-15", "year": 2024, "month": "Jun", "value": 150.5,
        "currency": "PLN", "value_base": 150.5, "rate_used": 1.0,
        "type": "Expense", "category": "Groceries", "person": "Alice",
        "description": "weekly shop", "is_recurring": 0, "is_done": 1,
    })
    sqlite_ops.insert_transaction(conn, {
        "date": "2024-07-01", "year": 2024, "month": "Jul", "value": 100.0,
        "currency": "USD", "value_base": 400.0, "rate_used": 4.0,
        "type": "Income", "category": "Salary", "person": "Bob",
        "description": "contract", "is_recurring": 0, "is_done": 1,
    })
    conn.close()
    return db


def test_export_row_count_and_plain_value_base(seeded_db, template, tmp_path):
    out = tmp_path / "export_out.xlsx"
    generate_excel_from_sqlite(seeded_db, template, out)

    wb = load_workbook(out)
    ws = wb["MasterData"]
    idx = col_indices(ws, MasterDataSchema)

    data_rows = [r for r in range(2, ws.max_row + 1)
                 if ws.cell(r, idx["value"]).value is not None]
    assert len(data_rows) == 2

    vbases = [ws.cell(r, idx["value_base"]).value for r in data_rows]
    for v in vbases:
        assert isinstance(v, (int, float)), f"value_base should be numeric, got {v!r}"
        assert not (isinstance(v, str) and v.startswith("=")), "must not be a formula"
    assert sorted(vbases) == [150.5, 400.0]

    cats = {ws.cell(r, idx["category"]).value for r in data_rows}
    assert cats == {"Groceries", "Salary"}


def test_export_logs_sync(seeded_db, template, tmp_path):
    generate_excel_from_sqlite(seeded_db, template, tmp_path / "out.xlsx")
    conn = sqlite_ops.init_db(seeded_db)
    try:
        row = conn.execute(
            "SELECT * FROM sync_log ORDER BY id DESC LIMIT 1").fetchone()
        assert row["direction"] == "export"
        assert row["status"] == "ok"
    finally:
        conn.close()


def test_reconcile_matches(seeded_db, template, tmp_path):
    from scripts.reconcile_sqlite_export import reconcile
    mismatches = reconcile(seeded_db, template, tmp_path / "reconcile.xlsx")
    assert mismatches == []


def test_export_calls_dashboard_resync(seeded_db, template, tmp_path, monkeypatch):
    """Both dashboard sync functions are called exactly once during export."""
    from unittest.mock import MagicMock
    from openpyxl import load_workbook
    import excel_export as _mod

    # Ensure the template workbook has a Cycle Dashboard sheet so the guard
    # inside generate_excel_from_sqlite lets the call through.
    wb_tmp = load_workbook(template)
    if "Cycle Dashboard" not in wb_tmp.sheetnames:
        wb_tmp.create_sheet("Cycle Dashboard")
        wb_tmp.save(template)

    mock_dash = MagicMock(return_value=0)
    mock_cycle = MagicMock(return_value=0)
    # Patch the names as they are bound in excel_export's module namespace.
    monkeypatch.setattr(_mod, "sync_dashboard_categories", mock_dash)
    monkeypatch.setattr(_mod, "sync_cycle_dashboard_categories", mock_cycle)

    out = tmp_path / "resync_out.xlsx"
    generate_excel_from_sqlite(seeded_db, template, out)

    mock_dash.assert_called_once()
    args, _ = mock_dash.call_args
    assert len(args[1]) > 0
    mock_cycle.assert_called_once()
