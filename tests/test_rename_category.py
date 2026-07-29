"""
tests/test_rename_category.py — unit tests for rename_category_in_workbook.

All workbooks are built in-memory with openpyxl — no disk I/O, no fixtures.
"""

import pytest
from openpyxl import Workbook

from excel_schema import rename_category_in_workbook


def _wb_with_masterdata(rows: list[str]) -> Workbook:
    """Return a workbook with a MasterData sheet whose Category column has rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MasterData"
    ws.cell(1, 1, "Date")
    ws.cell(1, 2, "Category")
    ws.cell(1, 3, "Amount")
    for i, cat in enumerate(rows, start=2):
        ws.cell(i, 2, cat)
    return wb


def _add_dashboard(wb: Workbook, plain_values: list[str], formula: str | None = None) -> None:
    ws = wb.create_sheet("Dashboard")
    for i, v in enumerate(plain_values, start=1):
        ws.cell(i, 1, v)
    if formula:
        ws.cell(len(plain_values) + 1, 2, formula)


class TestRenameInMasterData:
    def test_present_row_is_updated(self):
        wb = _wb_with_masterdata(["Groceries", "Housing", "Groceries"])
        counts = rename_category_in_workbook(wb, "Groceries", "Food")
        assert counts["MasterData"] == 2
        ws = wb["MasterData"]
        values = [ws.cell(r, 2).value for r in range(2, 5)]
        assert values == ["Food", "Housing", "Food"]

    def test_absent_name_leaves_workbook_unchanged(self):
        wb = _wb_with_masterdata(["Housing", "Transport"])
        counts = rename_category_in_workbook(wb, "Groceries", "Food")
        assert counts["MasterData"] == 0
        ws = wb["MasterData"]
        assert ws.cell(2, 2).value == "Housing"
        assert ws.cell(3, 2).value == "Transport"


class TestRenameFormulaLiterals:
    def test_formula_literal_replaced_in_dashboard(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        ws.cell(1, 1, '=SUMIFS(MasterData!$C:$C,MasterData!$B:$B,"Groceries")')
        counts = rename_category_in_workbook(wb, "Groceries", "Food")
        assert counts["Formulas"] == 1
        assert '"Food"' in ws.cell(1, 1).value
        assert '"Groceries"' not in ws.cell(1, 1).value

    def test_plain_dashboard_cell_renamed(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        ws.cell(1, 1, "Groceries")
        ws.cell(2, 1, "Housing")
        counts = rename_category_in_workbook(wb, "Groceries", "Food")
        assert counts["Dashboard"] == 1
        assert ws.cell(1, 1).value == "Food"
        assert ws.cell(2, 1).value == "Housing"


class TestMissingSheets:
    def test_no_masterdata_no_keyerror(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "SomeOtherSheet"
        counts = rename_category_in_workbook(wb, "Groceries", "Food")
        assert counts["MasterData"] == 0

    def test_no_dashboard_no_keyerror(self):
        wb = _wb_with_masterdata(["Groceries"])
        # no Dashboard sheet
        counts = rename_category_in_workbook(wb, "Groceries", "Food")
        assert counts["Dashboard"] == 0
        assert counts["Formulas"] == 0

    def test_empty_workbook_all_counts_zero(self):
        wb = Workbook()
        wb.active.title = "Empty"
        counts = rename_category_in_workbook(wb, "Groceries", "Food")
        assert counts == {"MasterData": 0, "Dashboard": 0, "Formulas": 0}
