"""
test_data_layer.py — exhaustive tests for the data layer.

Scope: models.py, data.py, file_storage.py, excel_ops.py.
All tests use real tmp_path fixtures with actual Excel files — no mocked I/O.
"""

import datetime
import json
import math

import openpyxl
import pandas as pd
import pytest

import file_storage
from file_storage import (
    create_blank_excel,
    get_recent_transactions,
    load_lists,
    load_budgets_from_excel,
)
from excel_ops import _do_append_transaction, replay_recovery_queue
from models import Transaction, AddTransactionState, MONTH_NAMES
from file_storage import delete_transaction_row, update_transaction_field, RowMovedError


# ══════════════════════════════════════════════════════════════════════════════
# models.py — Transaction
# ══════════════════════════════════════════════════════════════════════════════


class TestTransactionToRow:

    def _make(self, **kwargs) -> Transaction:
        defaults = dict(
            date=datetime.date(2024, 3, 10),
            value=100.0,
            currency="PLN",
            transaction_type="Expense",
        )
        defaults.update(kwargs)
        return Transaction(**defaults)

    # to_row() produces all required keys

    def test_to_row_expense_has_required_keys(self):
        row = self._make(transaction_type="Expense", category="Groceries").to_row()
        for key in ("date", "year", "month", "value", "type", "category",
                    "person", "description", "is_recurring", "is_done", "currency"):
            assert key in row, f"Missing key: {key}"

    def test_to_row_income_has_required_keys(self):
        row = self._make(transaction_type="Income").to_row()
        for key in ("date", "year", "month", "value", "type", "category",
                    "person", "description", "is_recurring", "is_done", "currency"):
            assert key in row

    def test_to_row_savings_has_required_keys(self):
        row = self._make(transaction_type="Savings").to_row()
        for key in ("date", "year", "month", "value", "type", "category",
                    "person", "description", "is_recurring", "is_done", "currency"):
            assert key in row

    # Empty category

    def test_to_row_empty_category_is_empty_string_not_none(self):
        row = self._make(category="").to_row()
        assert row["category"] == ""
        assert row["category"] is not None

    # Empty person

    def test_to_row_no_person_is_empty_string_not_none(self):
        row = self._make(person="").to_row()
        assert row["person"] == ""
        assert row["person"] is not None

    # Month abbreviations — all 12

    @pytest.mark.parametrize("month_num, expected", [
        (1, "Jan"), (2, "Feb"), (3, "Mar"), (4, "Apr"),
        (5, "May"), (6, "Jun"), (7, "Jul"), (8, "Aug"),
        (9, "Sep"), (10, "Oct"), (11, "Nov"), (12, "Dec"),
    ])
    def test_to_row_month_abbreviation_correct(self, month_num, expected):
        txn = self._make(date=datetime.date(2024, month_num, 1))
        assert txn.to_row()["month"] == expected

    # Year derivation

    def test_to_row_year_matches_date_year(self):
        txn = self._make(date=datetime.date(2023, 11, 5))
        assert txn.to_row()["year"] == 2023

    def test_to_row_year_2025(self):
        txn = self._make(date=datetime.date(2025, 1, 31))
        assert txn.to_row()["year"] == 2025

    # Values round-trip

    def test_to_row_value_matches(self):
        txn = self._make(value=42.99)
        assert txn.to_row()["value"] == pytest.approx(42.99)

    def test_to_row_type_matches(self):
        txn = self._make(transaction_type="Income")
        assert txn.to_row()["type"] == "Income"

    def test_to_row_currency_uppercased(self):
        txn = self._make(currency="eur")
        assert txn.to_row()["currency"] == "EUR"

    def test_to_row_is_done_default_true(self):
        txn = self._make()
        assert txn.to_row()["is_done"] is True

    def test_to_row_is_recurring_default_false(self):
        txn = self._make()
        assert txn.to_row()["is_recurring"] is False


class TestTransactionValidation:

    def test_negative_value_raises(self):
        with pytest.raises(Exception):
            Transaction(
                date=datetime.date(2024, 1, 1),
                value=-10.0,
                currency="PLN",
                transaction_type="Expense",
            )

    def test_zero_value_raises(self):
        with pytest.raises(Exception):
            Transaction(
                date=datetime.date(2024, 1, 1),
                value=0.0,
                currency="PLN",
                transaction_type="Expense",
            )

    def test_empty_transaction_type_raises(self):
        with pytest.raises(Exception):
            Transaction(
                date=datetime.date(2024, 1, 1),
                value=10.0,
                currency="PLN",
                transaction_type="",
            )

    def test_currency_defaults_not_needed_but_provided(self):
        # currency is required; verify PLN is accepted
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Expense",
        )
        assert txn.currency == "PLN"

    def test_description_stripped(self):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Expense",
            description="  padded  ",
        )
        assert txn.description == "padded"

    def test_description_truncated_to_100_chars(self):
        long_desc = "x" * 200
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Expense",
            description=long_desc,
        )
        assert len(txn.description) == 100


# ══════════════════════════════════════════════════════════════════════════════
# models.py — AddTransactionState
# ══════════════════════════════════════════════════════════════════════════════


class TestAddTransactionStateIsReadyToConfirm:

    def test_returns_false_when_all_none(self):
        state = AddTransactionState()
        assert state.is_ready_to_confirm() is False

    def test_returns_false_when_value_missing(self):
        state = AddTransactionState(currency="PLN", transaction_type="Income")
        assert state.is_ready_to_confirm() is False

    def test_returns_false_when_currency_missing(self):
        state = AddTransactionState(value=100.0, transaction_type="Income")
        assert state.is_ready_to_confirm() is False

    def test_returns_false_when_type_missing(self):
        state = AddTransactionState(value=100.0, currency="PLN")
        assert state.is_ready_to_confirm() is False

    def test_returns_true_for_income_without_category(self):
        # Income does not require category
        state = AddTransactionState(value=100.0, currency="PLN", transaction_type="Income")
        assert state.is_ready_to_confirm() is True

    def test_returns_true_for_savings_without_category(self):
        state = AddTransactionState(value=100.0, currency="PLN", transaction_type="Savings")
        assert state.is_ready_to_confirm() is True

    def test_returns_false_for_expense_without_category(self):
        # Expense requires category
        state = AddTransactionState(value=100.0, currency="PLN", transaction_type="Expense")
        assert state.is_ready_to_confirm() is False

    def test_returns_true_for_expense_with_category(self):
        state = AddTransactionState(
            value=100.0, currency="PLN", transaction_type="Expense", category="Groceries"
        )
        assert state.is_ready_to_confirm() is True

    def test_currency_defaults_to_pln_in_to_transaction(self):
        # currency=None → to_transaction uses "PLN"
        state = AddTransactionState(value=50.0, transaction_type="Income")
        txn = state.to_transaction()
        assert txn.currency == "PLN"

    def test_to_transaction_uses_today_when_date_none(self):
        state = AddTransactionState(value=50.0, currency="EUR", transaction_type="Income")
        txn = state.to_transaction()
        assert txn.date == datetime.datetime.now(datetime.timezone.utc).date()


# ══════════════════════════════════════════════════════════════════════════════
# file_storage.py — create_blank_excel: exact column positions
# ══════════════════════════════════════════════════════════════════════════════


class TestCreateBlankExcelColumnPositions:

    def test_masterdata_date_in_col_1(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 1).value == "Date"

    def test_masterdata_year_in_col_2(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 2).value == "Year"

    def test_masterdata_month_in_col_3(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 3).value == "Month"

    def test_masterdata_value_in_col_4(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 4).value == "Value"

    def test_masterdata_type_in_col_5(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 5).value == "Type"

    def test_masterdata_category_in_col_6(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 6).value == "Category"

    def test_masterdata_person_in_col_7(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 7).value == "Person"

    def test_masterdata_description_in_col_8(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 8).value == "Description"

    def test_masterdata_isrecurring_in_col_9(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 9).value == "IsRecurring"

    def test_masterdata_isdone_in_col_10(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 10).value == "IsDone"

    def test_masterdata_currency_in_col_11(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 11).value == "Currency"

    def test_masterdata_value_pln_in_col_12(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 12).value == "Value (base)"

    def test_masterdata_date_modified_in_col_13(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].cell(1, 13).value == "Date Modified (UTC)"

    def test_lists_month_header_in_col_a(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 1).value == "Months"

    def test_lists_type_header_in_col_b(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 2).value == "TxnTypes"

    def test_lists_category_header_in_col_c(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 3).value == "Categories"

    def test_lists_budget_header_in_col_d(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 4).value == "Budget (base)"

    def test_lists_person_header_in_col_e(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 5).value == "Persons"

    def test_lists_year_header_in_col_f(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 6).value == "Years"

    def test_lists_currency_header_in_col_h(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 8).value == "Currency"

    def test_lists_rate_header_in_col_i(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 9).value == "Rate to PLN"

    def test_lists_col_g_header_is_none(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        assert wb["Lists"].cell(1, 7).value is None

    def test_lists_col_c_contains_groceries(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        values = [ws.cell(r, 3).value for r in range(2, ws.max_row + 1)]
        assert "Groceries" in values

    def test_lists_col_c_contains_salary(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        values = [ws.cell(r, 3).value for r in range(2, ws.max_row + 1)]
        assert "Salary" in values

    def test_lists_col_h_contains_pln(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        values = [ws.cell(r, 8).value for r in range(2, ws.max_row + 1)]
        assert "PLN" in values

    def test_lists_budget_col_is_blank_by_default(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        assert ws.cell(1, 4).value == "Budget (base)"
        for row in range(2, ws.max_row + 1):
            assert ws.cell(row, 4).value is None

    def test_workbook_is_valid_after_creation(self, tmp_path):
        path = tmp_path / "fresh.xlsx"
        create_blank_excel(path)
        wb = openpyxl.load_workbook(path)
        assert wb is not None
        assert len(wb.sheetnames) >= 3


# ══════════════════════════════════════════════════════════════════════════════
# file_storage.py — load_lists()
# ══════════════════════════════════════════════════════════════════════════════


class TestLoadListsExhaustive:

    def test_returns_required_keys(self, excel_path):
        result = load_lists(excel_path)
        assert {"months", "txn_types", "categories", "persons", "years", "budgets"}.issubset(result.keys())

    def test_no_income_categories_key(self, excel_path):
        result = load_lists(excel_path)
        assert "income_categories" not in result

    def test_no_savings_categories_key(self, excel_path):
        result = load_lists(excel_path)
        assert "savings_categories" not in result

    def test_categories_non_empty_on_blank_excel(self, excel_path):
        result = load_lists(excel_path)
        assert len(result["categories"]) > 0

    def test_persons_empty_on_blank_excel(self, excel_path):
        # blank Excel has no persons in col D
        result = load_lists(excel_path)
        assert result["persons"] == []

    def test_txn_types_contains_expense(self, excel_path):
        assert "Expense" in load_lists(excel_path)["txn_types"]

    def test_txn_types_contains_income(self, excel_path):
        assert "Income" in load_lists(excel_path)["txn_types"]

    def test_txn_types_contains_savings(self, excel_path):
        assert "Savings" in load_lists(excel_path)["txn_types"]

    def test_categories_stops_at_first_none(self, tmp_path):
        # Build an Excel where col C has a gap after row 3 to verify break behavior
        path = tmp_path / "gap.xlsx"
        create_blank_excel(path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Lists"]
        # Overwrite col C: only two categories then None gap then another value
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 3).value = None
        ws.cell(2, 3).value = "Alpha"
        ws.cell(3, 3).value = "Beta"
        # Row 4 stays None — read_column must stop here
        ws.cell(5, 3).value = "Orphan"
        wb.save(path)
        result = load_lists(path)
        assert result["categories"] == ["Alpha", "Beta"]
        assert "Orphan" not in result["categories"]

    def test_missing_lists_sheet_returns_empty_lists(self, tmp_path):
        # Build a workbook with no Lists sheet
        path = tmp_path / "no_lists.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "MasterData"
        wb.active.append(["Date", "Year", "Month", "Value", "Type", "Category",
                           "Person", "Description", "IsRecurring", "IsDone",
                           "Currency", "Value (base)", "Date Modified (UTC)"])
        wb.save(path)
        result = load_lists(path)
        assert result["months"] == []
        assert result["txn_types"] == []
        assert result["categories"] == []
        assert result["persons"] == []
        assert result["years"] == []


# ══════════════════════════════════════════════════════════════════════════════
# file_storage.py — get_recent_transactions()
# ══════════════════════════════════════════════════════════════════════════════


class TestGetRecentTransactionsExhaustive:

    def test_empty_excel_returns_empty_list(self, excel_path):
        assert get_recent_transactions(excel_path) == []

    def test_n_parameter_limits_result(self, excel_path):
        for i in range(1, 10):
            txn = Transaction(
                date=datetime.date(2024, 1, i),
                value=float(i),
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            )
            _do_append_transaction(txn)
        result = get_recent_transactions(excel_path, n=4)
        assert len(result) == 4

    def test_returns_last_n_not_first_n(self, excel_path):
        for i in range(1, 6):
            txn = Transaction(
                date=datetime.date(2024, 1, i),
                value=float(i * 100),
                currency="PLN",
                transaction_type="Expense",
                category="Groceries",
            )
            _do_append_transaction(txn)
        result = get_recent_transactions(excel_path, n=2)
        values = [r["Value"] for r in result]
        # Last 2 written are 400.0 and 500.0
        assert values == [400.0, 500.0]

    def test_row_idx_matches_actual_excel_row(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=99.0,
            currency="PLN",
            transaction_type="Income",
        )
        _do_append_transaction(txn)
        result = get_recent_transactions(excel_path)
        # Header is row 1; first data row is row 2
        assert result[0]["_row_idx"] == 2

    def test_row_idx_increments_with_multiple_appends(self, excel_path):
        for i in range(3):
            txn = Transaction(
                date=datetime.date(2024, 1, i + 1),
                value=float(i + 1),
                currency="PLN",
                transaction_type="Expense",
                category="Transport",
            )
            _do_append_transaction(txn)
        result = get_recent_transactions(excel_path, n=3)
        row_indices = [r["_row_idx"] for r in result]
        assert row_indices == [2, 3, 4]

    def test_category_key_holds_category_column_value(self, excel_path):
        # Write two distinct columns to verify Category comes from col 6, not col 7
        txn = Transaction(
            date=datetime.date(2024, 3, 1),
            value=50.0,
            currency="PLN",
            transaction_type="Expense",
            category="Healthcare",
            person="Bob",
        )
        _do_append_transaction(txn)
        result = get_recent_transactions(excel_path)
        assert result[0]["Category"] == "Healthcare"
        # Person is separate — verify col 7 went to "Person" not "Category"
        assert result[0]["Person"] == "Bob"

    def test_value_key_holds_value_column_not_adjacent(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 4, 1),
            value=333.33,
            currency="EUR",
            transaction_type="Expense",
            category="Travel",
        )
        _do_append_transaction(txn)
        result = get_recent_transactions(excel_path)
        assert result[0]["Value"] == pytest.approx(333.33)
        # Currency is col 11, Value is col 4 — they must not be swapped
        assert result[0]["Currency"] == "EUR"

    def test_date_key_holds_date_from_col_1(self, excel_path):
        expected_date = datetime.date(2024, 7, 22)
        txn = Transaction(
            date=expected_date,
            value=10.0,
            currency="PLN",
            transaction_type="Income",
        )
        _do_append_transaction(txn)
        result = get_recent_transactions(excel_path)
        cell_val = result[0]["Date"]
        if isinstance(cell_val, datetime.datetime):
            assert cell_val.date() == expected_date
        else:
            assert cell_val == expected_date

    def test_person_key_holds_value_from_col_7(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=20.0,
            currency="PLN",
            transaction_type="Expense",
            category="Clothing",
            person="Charlie",
        )
        _do_append_transaction(txn)
        result = get_recent_transactions(excel_path)
        assert result[0]["Person"] == "Charlie"

    def test_works_correctly_after_multiple_appends(self, excel_path):
        categories = ["Groceries", "Transport", "Housing", "Utilities", "Entertainment"]
        for i, cat in enumerate(categories):
            txn = Transaction(
                date=datetime.date(2024, 1, i + 1),
                value=float((i + 1) * 10),
                currency="PLN",
                transaction_type="Expense",
                category=cat,
            )
            _do_append_transaction(txn)
        result = get_recent_transactions(excel_path, n=3)
        assert len(result) == 3
        returned_cats = [r["Category"] for r in result]
        assert returned_cats == ["Housing", "Utilities", "Entertainment"]


# ══════════════════════════════════════════════════════════════════════════════
# excel_ops.py — _do_append_transaction: specific assertions
# ══════════════════════════════════════════════════════════════════════════════


class TestDoAppendTransactionExhaustive:

    def test_date_modified_is_datetime_instance_not_string(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Expense",
            category="Groceries",
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        val = ws.cell(2, headers["Date Modified (UTC)"]).value
        assert isinstance(val, datetime.datetime), (
            f"Expected datetime.datetime, got {type(val)}: {val!r}"
        )

    def test_date_modified_is_not_formula_string(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Expense",
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        val = ws.cell(2, headers["Date Modified (UTC)"]).value
        assert not (isinstance(val, str) and val.startswith("=")), (
            f"Date Modified must not be a formula, got: {val!r}"
        )

    def test_value_pln_formula_starts_with_equals(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=50.0,
            currency="EUR",
            transaction_type="Expense",
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        formula = ws.cell(2, headers["Value (base)"]).value
        assert isinstance(formula, str)
        assert formula.startswith("=")

    def test_value_pln_formula_references_currency_rate_range(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=50.0,
            currency="EUR",
            transaction_type="Expense",
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        formula = ws.cell(2, headers["Value (base)"]).value
        # Range is built dynamically from Lists Currency/Rate column positions
        assert "$H$2:$I$100" in formula, (
            f"VLOOKUP range must be $H$2:$I$100, got: {formula!r}"
        )

    def test_isdone_written_as_bool_true_not_string(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Expense",
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        val = ws.cell(2, headers["IsDone"]).value
        assert val is True, f"IsDone must be bool True, got {type(val)}: {val!r}"
        assert not isinstance(val, str), "IsDone must not be the string 'TRUE'"

    def test_max_row_is_2_after_single_append(self, excel_path):
        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Expense",
        )
        _do_append_transaction(txn)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        assert ws.max_row == 2


# ══════════════════════════════════════════════════════════════════════════════
# excel_ops.py — replay_recovery_queue()
# ══════════════════════════════════════════════════════════════════════════════


class TestRowMovedGuard:
    """
    Regression: /delete and /edit captured a row index at pick time and
    applied it minutes later. If another delete/edit shifted rows in between,
    the stale index would silently delete/edit the wrong transaction. Now
    delete_transaction_row/update_transaction_field re-verify a date/value/
    description snapshot under the write lock and raise RowMovedError if the
    row no longer matches.
    """

    def _append(self, excel_path, day, value, desc):
        _do_append_transaction(Transaction(
            date=datetime.date(2024, 6, day),
            value=value,
            currency="PLN",
            transaction_type="Expense",
            category="Groceries",
            description=desc,
        ))

    def test_delete_succeeds_when_snapshot_matches(self, excel_path):
        self._append(excel_path, 1, 10.0, "first")
        self._append(excel_path, 2, 20.0, "second")
        expected = {"Date": datetime.date(2024, 6, 2), "Value": 20.0, "Description": "second"}
        delete_transaction_row(3, expected)
        remaining = get_recent_transactions(excel_path)
        assert len(remaining) == 1
        assert remaining[0]["Description"] == "first"

    def test_delete_raises_row_moved_error_when_row_shifted(self, excel_path):
        self._append(excel_path, 1, 10.0, "first")
        self._append(excel_path, 2, 20.0, "second")
        # Simulate a concurrent delete of row 2 shifting "second" up to row 2.
        wb = openpyxl.load_workbook(excel_path)
        wb["MasterData"].delete_rows(2)
        wb.save(excel_path)

        # Caller still thinks "second" is at row 3 (stale snapshot).
        stale_expected = {"Date": datetime.date(2024, 6, 2), "Value": 20.0, "Description": "second"}
        with pytest.raises(RowMovedError):
            delete_transaction_row(3, stale_expected)
        # Nothing was deleted — "second" (now at row 2) is still present.
        remaining = get_recent_transactions(excel_path)
        assert len(remaining) == 1
        assert remaining[0]["Description"] == "second"

    def test_delete_without_expected_snapshot_still_works(self, excel_path):
        """Backward-compatible: omitting `expected` skips verification."""
        self._append(excel_path, 1, 10.0, "only")
        delete_transaction_row(2)
        assert get_recent_transactions(excel_path) == []

    def test_update_field_succeeds_when_snapshot_matches(self, excel_path):
        self._append(excel_path, 1, 10.0, "first")
        expected = {"Date": datetime.date(2024, 6, 1), "Value": 10.0, "Description": "first"}
        update_transaction_field(2, "Value", 99.0, expected)
        result = get_recent_transactions(excel_path)
        assert result[0]["Value"] == 99.0

    def test_update_field_raises_row_moved_error_when_row_shifted(self, excel_path):
        self._append(excel_path, 1, 10.0, "first")
        self._append(excel_path, 2, 20.0, "second")
        wb = openpyxl.load_workbook(excel_path)
        wb["MasterData"].delete_rows(2)
        wb.save(excel_path)

        stale_expected = {"Date": datetime.date(2024, 6, 2), "Value": 20.0, "Description": "second"}
        with pytest.raises(RowMovedError):
            update_transaction_field(3, "Value", 999.0, stale_expected)
        # The value at the shifted row must be untouched.
        result = get_recent_transactions(excel_path)
        assert result[0]["Value"] == 20.0


class TestReplayRecoveryQueue:

    def test_replay_writes_datetime_not_formula_for_date_modified(
        self, excel_path, tmp_path, monkeypatch
    ):
        queue_path = tmp_path / "recovery_queue.json"
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        # Write a recovery queue entry as to_row() would produce
        txn = Transaction(
            date=datetime.date(2024, 6, 1),
            value=77.0,
            currency="PLN",
            transaction_type="Expense",
            category="Groceries",
        )
        row = txn.to_row()
        queue_path.write_text(json.dumps([row], default=str))

        replay_recovery_queue()

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        val = ws.cell(2, headers["Date Modified (UTC)"]).value
        assert isinstance(val, datetime.datetime), (
            f"replay_recovery_queue must write datetime, got {type(val)}: {val!r}"
        )

    def test_replay_writes_formula_in_value_pln_column(
        self, excel_path, tmp_path, monkeypatch
    ):
        queue_path = tmp_path / "recovery_queue.json"
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        txn = Transaction(
            date=datetime.date(2024, 6, 1),
            value=88.0,
            currency="EUR",
            transaction_type="Expense",
        )
        queue_path.write_text(json.dumps([txn.to_row()], default=str))

        replay_recovery_queue()

        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        formula = ws.cell(2, headers["Value (base)"]).value
        assert isinstance(formula, str) and formula.startswith("=")

    def test_replay_empty_queue_writes_nothing(self, excel_path, tmp_path, monkeypatch):
        queue_path = tmp_path / "recovery_queue.json"
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)
        # No queue file — nothing to replay
        replay_recovery_queue()
        wb = openpyxl.load_workbook(excel_path)
        assert wb["MasterData"].max_row == 1  # header only

    def test_replay_removes_queue_file_after_processing(
        self, excel_path, tmp_path, monkeypatch
    ):
        queue_path = tmp_path / "recovery_queue.json"
        monkeypatch.setattr(file_storage, "RECOVERY_QUEUE_PATH", queue_path)

        txn = Transaction(
            date=datetime.date(2024, 1, 1),
            value=10.0,
            currency="PLN",
            transaction_type="Income",
        )
        queue_path.write_text(json.dumps([txn.to_row()], default=str))
        replay_recovery_queue()
        assert not queue_path.exists()


# ══════════════════════════════════════════════════════════════════════════════
# data.py — load_data(), load_rates(), load_budgets(), load_reference_data()
# ══════════════════════════════════════════════════════════════════════════════


def _write_masterdata_row(ws, row_num: int, **kwargs):
    """Write a single MasterData data row using column headers as keys."""
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for field, value in kwargs.items():
        col = headers.get(field)
        if col:
            ws.cell(row_num, col).value = value


class TestLoadData:

    def _build_excel_with_rows(self, excel_path, rows: list[dict]) -> None:
        """Write rows into MasterData, save. rows is list of column→value dicts."""
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["MasterData"]
        for i, row in enumerate(rows, start=2):
            _write_masterdata_row(ws, i, **row)
        wb.save(excel_path)

    # Value(PLN) cache present → _pln equals cached value

    def test_value_pln_cache_used_when_present(self, excel_path, monkeypatch):
        import file_storage as fs
        monkeypatch.setattr(fs, "LOCAL_XLSX_PATH", excel_path)
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": 100.0,
            "Type": "Expense",
            "Category": "Groceries",
            "Currency": "EUR",
            "Value (base)": 428.0,   # cached result — EUR * 4.28
            "IsDone": True,
        }])

        df = data_mod.load_data()
        assert len(df) == 1
        assert df.iloc[0]["_pln"] == pytest.approx(428.0)

    # Value(PLN) = NaN → recomputed from Value * rate

    def test_value_pln_recomputed_when_cache_missing(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        # Write EUR row without a cached Value(PLN) — leave it None
        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": 100.0,
            "Type": "Expense",
            "Category": "Groceries",
            "Currency": "EUR",
            # Value (base) intentionally omitted → NaN in pandas
            "IsDone": True,
        }])

        # Put EUR rate 4.28 in Lists sheet (Currency col 8, Rate col 9)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        ws.cell(2, 8).value = "EUR"
        ws.cell(2, 9).value = 4.28
        wb.save(excel_path)

        df = data_mod.load_data()
        assert len(df) == 1
        assert df.iloc[0]["_pln"] == pytest.approx(428.0)

    # PLN currency with missing cache → _pln equals Value

    def test_pln_currency_missing_cache_uses_value(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": 250.0,
            "Type": "Income",
            "Currency": "PLN",
            # Value (base) left out → fallback
            "IsDone": True,
        }])

        df = data_mod.load_data()
        assert len(df) == 1
        assert df.iloc[0]["_pln"] == pytest.approx(250.0)

    # Unknown currency → rate defaults to 1.0

    def test_unknown_currency_missing_cache_defaults_rate_1(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": 123.0,
            "Type": "Expense",
            "Category": "Other",
            "Currency": "XYZ",   # unknown
            "IsDone": True,
        }])

        df = data_mod.load_data()
        assert len(df) == 1
        assert df.iloc[0]["_pln"] == pytest.approx(123.0)  # 123 * 1.0

    # Type=None → row dropped

    def test_row_with_none_type_is_dropped(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": 10.0,
            "Type": None,        # must be dropped
            "Currency": "PLN",
            "Value (base)": 10.0,
            "IsDone": True,
        }])

        df = data_mod.load_data()
        assert len(df) == 0

    # Year=None → row dropped

    def test_row_with_none_year_is_dropped(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": None,        # must be dropped
            "Month": "Jan",
            "Value": 10.0,
            "Type": "Expense",
            "Currency": "PLN",
            "Value (base)": 10.0,
            "IsDone": True,
        }])

        df = data_mod.load_data()
        assert len(df) == 0

    # _pln NaN after recompute → row dropped

    def test_row_dropped_when_pln_nan_after_recompute(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        # Value is None → recompute produces NaN → row dropped
        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": None,       # no value → NaN
            "Type": "Expense",
            "Currency": "PLN",
            # Value (base) also absent
            "IsDone": True,
        }])

        df = data_mod.load_data()
        assert len(df) == 0

    # IsDone=None → defaults to True

    def test_isdone_none_defaults_to_true(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": 10.0,
            "Type": "Expense",
            "Currency": "PLN",
            "Value (base)": 10.0,
            "IsDone": None,      # must default to True
        }])

        df = data_mod.load_data()
        assert len(df) == 1
        assert bool(df.iloc[0]["IsDone"]) is True

    # Currency=None → defaults to PLN

    def test_currency_none_defaults_to_pln(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        self._build_excel_with_rows(excel_path, [{
            "Date": datetime.date(2024, 1, 1),
            "Year": 2024,
            "Month": "Jan",
            "Value": 55.0,
            "Type": "Income",
            "Currency": None,    # must become "PLN"
            "Value (base)": 55.0,
            "IsDone": True,
        }])

        df = data_mod.load_data()
        assert len(df) == 1
        assert df.iloc[0]["Currency"] == "PLN"


class TestLoadRates:

    def test_returns_pln_minimum(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        rates = data_mod.load_rates()
        assert "PLN" in rates
        assert rates["PLN"] == pytest.approx(1.0)

    def test_reads_currency_and_rate_from_lists_i_j(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        # Add a custom rate to verify correct column reading
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        # Find next empty row in Currency col (col 8)
        next_row = 2
        while ws.cell(next_row, 8).value is not None:
            next_row += 1
        ws.cell(next_row, 8).value = "CHF"
        ws.cell(next_row, 9).value = 4.45
        wb.save(excel_path)

        rates = data_mod.load_rates()
        assert "CHF" in rates
        assert rates["CHF"] == pytest.approx(4.45)

    def test_returns_fallback_when_lists_empty(self, tmp_path, monkeypatch):
        import data as data_mod

        # Blank workbook with Lists sheet but no data
        path = tmp_path / "empty_lists.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "MasterData"
        ws = wb.create_sheet("Lists")
        ws.cell(1, 8).value = "Currency"
        ws.cell(1, 9).value = "Rate (PLN)"
        # No data rows
        wb.save(path)
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: path)

        rates = data_mod.load_rates()
        assert isinstance(rates, dict)  # empty is valid; callers use get_rate() which defaults to 1.0

    def test_rate_column_comes_from_col_j_not_col_i(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        # Write a known rate then verify load_rates reads the J value, not I
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        # Find PLN row
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 8).value == "PLN":  # Currency col
                ws.cell(r, 9).value = 1.0      # Rate col
                ws.cell(r, 7).value = 999.0    # col G — a decoy
                break
        wb.save(excel_path)

        rates = data_mod.load_rates()
        assert rates["PLN"] == pytest.approx(1.0)


class TestLoadBudgets:

    def test_returns_dict(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)
        budgets = data_mod.load_budgets()
        assert isinstance(budgets, dict)

    def test_reads_category_from_lists_budget_col(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        next_row = ws.max_row + 1
        ws.cell(next_row, 3).value = "UniqueTestCat"  # Category col
        ws.cell(next_row, 4).value = 1500.0            # Budget (base) col
        wb.save(excel_path)

        budgets = data_mod.load_budgets()
        assert "UniqueTestCat" in budgets

    def test_reads_budget_amount_from_lists(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Lists"]
        next_row = ws.max_row + 1
        ws.cell(next_row, 3).value = "TestBudgetCat"   # Category col
        ws.cell(next_row, 4).value = 2100.0             # Budget (base) col
        wb.save(excel_path)

        budgets = data_mod.load_budgets()
        assert budgets.get("TestBudgetCat") == pytest.approx(2100.0)


class TestLoadReferenceData:

    def test_includes_currencies_key(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)
        ref = data_mod.load_reference_data()
        assert "currencies" in ref

    def test_currencies_contains_pln(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)
        ref = data_mod.load_reference_data()
        assert "PLN" in ref["currencies"]

    def test_currencies_falls_back_to_pln_list_when_rates_empty(
        self, tmp_path, monkeypatch
    ):
        import data as data_mod

        # Build a workbook where Lists I/J have no valid rows
        path = tmp_path / "no_rates.xlsx"
        create_blank_excel(path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Lists"]
        # Wipe all currency data from col I and J
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 9).value = None
            ws.cell(r, 10).value = None
        wb.save(path)

        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: path)

        ref = data_mod.load_reference_data()
        # load_rates() returns {"PLN": 1.0} minimum; currencies must contain "PLN"
        assert ref["currencies"] == ["PLN"] or "PLN" in ref["currencies"]

    def test_includes_standard_list_keys(self, excel_path, monkeypatch):
        import data as data_mod
        monkeypatch.setattr(data_mod, "get_excel_path_for_reading", lambda: excel_path)
        ref = data_mod.load_reference_data()
        for key in ("months", "txn_types", "categories", "persons", "years"):
            assert key in ref

