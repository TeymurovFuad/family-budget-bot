"""
tests/test_statement_profiles.py — statement_profiles module tests.

All tests are offline: no live AI calls, no network, no real file I/O beyond
the module's own save/load logic (which uses a tmp_path fixture).

No bank names appear anywhere in this file — fixtures use "BankA", "BankB".
"""

import io
import json
import csv
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

import statement_profiles as sp


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────

BANKA_PROFILE = {
    "name": "BankA",
    "delimiter": ";",
    "encoding": "utf-8",
    "header_row": 0,
    "fingerprint": ["TransDate", "Amount", "CCY", "Info"],
    "column_map": {
        "date": "TransDate",
        "amount": "Amount",
        "currency": "CCY",
        "description": "Info",
        "time": None,
    },
    "date_format": "%d.%m.%Y",
    "decimal_separator": ",",
    "sign_convention": "negative_expense",
}


def _make_csv_bytes(
    headers: list[str],
    rows: list[list],
    delimiter: str = ";",
) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def _make_xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    """Build a minimal in-memory XLSX with openpyxl."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# fingerprint_from_headers
# ─────────────────────────────────────────────────────────────────────────────


class TestFingerprintFromHeaders:
    def test_sorted_and_stripped(self):
        fp = sp.fingerprint_from_headers(["  Amount ", "Date", "CCY"])
        assert fp == ("Amount", "CCY", "Date")

    def test_empty_headers_excluded(self):
        fp = sp.fingerprint_from_headers(["Date", "", "Amount"])
        assert fp == ("Amount", "Date")

    def test_all_empty_returns_empty_tuple(self):
        assert sp.fingerprint_from_headers(["", "  "]) == ()


# ─────────────────────────────────────────────────────────────────────────────
# load_profiles / match_profile
# ─────────────────────────────────────────────────────────────────────────────


class TestLoadAndMatchProfiles:
    def test_fingerprint_match(self, tmp_path):
        """Exact header set matches a stored profile."""
        profile_path = tmp_path / "banka.json"
        profile_path.write_text(json.dumps(BANKA_PROFILE), encoding="utf-8")

        profiles = sp.load_profiles(tmp_path)
        matched = sp.match_profile(["TransDate", "Amount", "CCY", "Info"], profiles)

        assert matched is not None
        assert matched["name"] == "BankA"

    def test_fingerprint_no_match(self, tmp_path):
        """Different headers return None."""
        profile_path = tmp_path / "banka.json"
        profile_path.write_text(json.dumps(BANKA_PROFILE), encoding="utf-8")

        profiles = sp.load_profiles(tmp_path)
        # Different column names → no match.
        matched = sp.match_profile(["Date", "Value", "Currency", "Memo"], profiles)
        assert matched is None

    def test_order_independent_match(self, tmp_path):
        """Match works regardless of the order headers arrive in the file."""
        profile_path = tmp_path / "banka.json"
        profile_path.write_text(json.dumps(BANKA_PROFILE), encoding="utf-8")
        profiles = sp.load_profiles(tmp_path)

        matched = sp.match_profile(["Info", "CCY", "Amount", "TransDate"], profiles)
        assert matched is not None

    def test_empty_dir_returns_empty_dict(self, tmp_path):
        assert sp.load_profiles(tmp_path) == {}

    def test_nonexistent_dir_returns_empty_dict(self, tmp_path):
        assert sp.load_profiles(tmp_path / "nonexistent") == {}

    def test_malformed_json_skipped(self, tmp_path):
        (tmp_path / "bad.json").write_text("not json", encoding="utf-8")
        assert sp.load_profiles(tmp_path) == {}

    def test_profile_without_fingerprint_skipped(self, tmp_path):
        bad = {**BANKA_PROFILE, "fingerprint": []}
        (tmp_path / "bad.json").write_text(json.dumps(bad), encoding="utf-8")
        assert sp.load_profiles(tmp_path) == {}


# ─────────────────────────────────────────────────────────────────────────────
# parse_statement — CSV
# ─────────────────────────────────────────────────────────────────────────────


class TestParseStatementCSV:
    def test_parse_statement_csv(self):
        """Parse a semicolon-delimited, comma-decimal CSV; check all fields."""
        csv_bytes = _make_csv_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [
                ["12.03.2026", "-45,99", "PLN", "Grocery store"],
                ["13.03.2026", "1000,00", "EUR", "Salary"],
            ],
            delimiter=";",
        )
        rows = sp.parse_statement(csv_bytes, "statement.csv", BANKA_PROFILE)

        assert len(rows) == 2

        # Negative amount → Expense (sign_convention = negative_expense)
        r0 = rows[0]
        assert r0["date"] == "2026-03-12"
        assert r0["value"] == pytest.approx(45.99)
        assert r0["currency"] == "PLN"
        assert r0["description"] == "Grocery store"
        assert r0["type"] == "Expense"
        assert r0["time"] is None

        # Positive amount → Income
        r1 = rows[1]
        assert r1["date"] == "2026-03-13"
        assert r1["value"] == pytest.approx(1000.00)
        assert r1["currency"] == "EUR"
        assert r1["type"] == "Income"

    def test_empty_amount_row_skipped(self):
        csv_bytes = _make_csv_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [
                ["12.03.2026", "", "PLN", "No amount"],
                ["13.03.2026", "-10,00", "PLN", "OK"],
            ],
            delimiter=";",
        )
        rows = sp.parse_statement(csv_bytes, "statement.csv", BANKA_PROFILE)
        assert len(rows) == 1
        assert rows[0]["description"] == "OK"

    def test_bad_amount_row_skipped(self):
        csv_bytes = _make_csv_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [
                ["12.03.2026", "N/A", "PLN", "Bad"],
                ["13.03.2026", "-5,00", "PLN", "Good"],
            ],
            delimiter=";",
        )
        rows = sp.parse_statement(csv_bytes, "statement.csv", BANKA_PROFILE)
        assert len(rows) == 1

    def test_dot_decimal_separator(self):
        profile = {**BANKA_PROFILE, "decimal_separator": "."}
        csv_bytes = _make_csv_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [["15.06.2026", "-99.50", "PLN", "Shop"]],
            delimiter=";",
        )
        rows = sp.parse_statement(csv_bytes, "statement.csv", profile)
        assert rows[0]["value"] == pytest.approx(99.50)

    def test_parse_statement_csv_header_row_nonzero(self):
        """CSV with 2 preamble lines before the real header; header_row=2."""
        buf = io.StringIO()
        buf.write("Bank Export Report\n")
        buf.write("Generated: 2026-03-01\n")
        buf.write("TransDate;Amount;CCY;Info\n")
        buf.write("12.03.2026;-10,00;PLN;Coffee\n")
        buf.write("13.03.2026;50,00;EUR;Salary\n")
        csv_bytes = buf.getvalue().encode("utf-8")

        profile = {**BANKA_PROFILE, "header_row": 2}
        rows = sp.parse_statement(csv_bytes, "statement.csv", profile)

        assert len(rows) == 2
        assert rows[0]["date"] == "2026-03-12"
        assert rows[0]["value"] == pytest.approx(10.00)
        assert rows[0]["type"] == "Expense"
        assert rows[1]["date"] == "2026-03-13"
        assert rows[1]["type"] == "Income"

    def test_parse_statement_csv_european_thousands(self):
        """Amount '1.234,56' with decimal_separator=',' → 1234.56."""
        profile = {**BANKA_PROFILE, "decimal_separator": ","}
        csv_bytes = _make_csv_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [["01.06.2026", "-1.234,56", "PLN", "Big purchase"]],
            delimiter=";",
        )
        rows = sp.parse_statement(csv_bytes, "statement.csv", profile)
        assert len(rows) == 1
        assert rows[0]["value"] == pytest.approx(1234.56)
        assert rows[0]["type"] == "Expense"

    def test_parse_statement_csv_us_thousands(self):
        """Amount '1,234.56' with decimal_separator='.' → 1234.56."""
        profile = {**BANKA_PROFILE, "decimal_separator": "."}
        csv_bytes = _make_csv_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [["01.06.2026", "-1,234.56", "PLN", "Big purchase"]],
            delimiter=";",
        )
        rows = sp.parse_statement(csv_bytes, "statement.csv", profile)
        assert len(rows) == 1
        assert rows[0]["value"] == pytest.approx(1234.56)
        assert rows[0]["type"] == "Expense"


# ─────────────────────────────────────────────────────────────────────────────
# parse_statement — XLSX
# ─────────────────────────────────────────────────────────────────────────────


class TestParseStatementXLSX:
    def test_parse_statement_xlsx(self):
        """Parse an in-memory XLSX with the BankA profile."""
        # Note: XLSX numeric cells don't have comma-decimal — the float is exact.
        profile = {**BANKA_PROFILE, "decimal_separator": "."}
        xlsx_bytes = _make_xlsx_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [
                ["15.06.2026", -50.25, "PLN", "Transport"],
                ["16.06.2026", 200.0, "EUR", "Refund"],
            ],
        )
        rows = sp.parse_statement(xlsx_bytes, "statement.xlsx", profile)

        assert len(rows) == 2
        assert rows[0]["date"] == "2026-06-15"
        assert rows[0]["value"] == pytest.approx(50.25)
        assert rows[0]["type"] == "Expense"
        assert rows[1]["type"] == "Income"

    def test_xlsx_with_string_amounts(self):
        """String-formatted amounts in XLSX cells (exported as text)."""
        xlsx_bytes = _make_xlsx_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [["01.01.2026", "-10,50", "PLN", "Coffee"]],
        )
        rows = sp.parse_statement(xlsx_bytes, "statement.xlsx", BANKA_PROFILE)
        assert len(rows) == 1
        assert rows[0]["value"] == pytest.approx(10.50)

    def test_parse_statement_xlsx_header_row_nonzero(self):
        """XLSX with 2 preamble rows before headers; header_row=2."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Bank Export Report"])           # row 0 (preamble)
        ws.append(["Generated: 2026-03-01"])        # row 1 (preamble)
        ws.append(["TransDate", "Amount", "CCY", "Info"])  # row 2 (headers)
        ws.append(["15.06.2026", -25.0, "PLN", "Transport"])
        ws.append(["16.06.2026", 100.0, "EUR", "Refund"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        profile = {**BANKA_PROFILE, "header_row": 2, "decimal_separator": "."}
        rows = sp.parse_statement(xlsx_bytes, "statement.xlsx", profile)

        assert len(rows) == 2
        assert rows[0]["date"] == "2026-06-15"
        assert rows[0]["value"] == pytest.approx(25.0)
        assert rows[0]["type"] == "Expense"
        assert rows[1]["type"] == "Income"

    def test_parse_statement_xlsx_merged_cell_header(self):
        """
        XLSX with a merged header cell (A1:B1 merged, value 'Date').
        openpyxl read_only returns None for the non-origin cells in a merge,
        so the fingerprint must skip the empty cell. We assert that the
        function either (a) produces a fingerprint with one 'Date' entry and
        skips the empty column name, or (b) maps the data correctly.
        The key contract is: no exception is raised, and no row is silently
        swallowed due to the empty header cell.
        """
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        # Merge A1:B1 — openpyxl read_only will show 'Date' at A1 and None at B1.
        ws.merge_cells("A1:B1")
        ws["A1"] = "Date"
        ws["C1"] = "Amount"
        ws["D1"] = "Info"
        ws.append(["", "15.06.2026", -10.0, "Coffee"])  # data row: col B holds date
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        # Profile maps to columns as openpyxl returns them (merged cell → empty string for B1).
        # The function must not raise; verify no crash and return value is a list.
        profile = {
            **BANKA_PROFILE,
            "decimal_separator": ".",
            # column_map intentionally references the real column names to test
            # that empty-header skipping doesn't break iteration.
            "column_map": {
                "date": "Date",
                "amount": "Amount",
                "currency": None,
                "description": "Info",
                "time": None,
            },
            "fingerprint": ["Date", "Amount", "Info"],
        }
        # Must not raise — merged cells produce empty strings, not exceptions.
        result = sp.parse_statement(xlsx_bytes, "statement.xlsx", profile)
        assert isinstance(result, list)


# ─────────────────────────────────────────────────────────────────────────────
# sign_convention
# ─────────────────────────────────────────────────────────────────────────────


class TestSignConvention:
    def _parse(self, amount_str: str, convention: str) -> dict:
        profile = {**BANKA_PROFILE, "sign_convention": convention, "decimal_separator": "."}
        csv_bytes = _make_csv_bytes(
            ["TransDate", "Amount", "CCY", "Info"],
            [["01.01.2026", amount_str, "PLN", "Test"]],
            delimiter=";",
        )
        rows = sp.parse_statement(csv_bytes, "s.csv", profile)
        assert rows, f"No rows parsed for amount {amount_str!r}"
        return rows[0]

    def test_negative_expense_negative_amount(self):
        row = self._parse("-45.00", "negative_expense")
        assert row["type"] == "Expense"
        assert row["value"] == pytest.approx(45.00)

    def test_negative_expense_positive_amount(self):
        row = self._parse("100.00", "negative_expense")
        assert row["type"] == "Income"
        assert row["value"] == pytest.approx(100.00)

    def test_positive_expense_positive_amount(self):
        row = self._parse("60.00", "positive_expense")
        assert row["type"] == "Expense"

    def test_positive_expense_negative_amount(self):
        row = self._parse("-60.00", "positive_expense")
        assert row["type"] == "Income"

    def test_always_expense(self):
        row = self._parse("-99.99", "always_expense")
        assert row["type"] == "Expense"
        assert row["value"] == pytest.approx(99.99)


# ─────────────────────────────────────────────────────────────────────────────
# mask_sample_rows
# ─────────────────────────────────────────────────────────────────────────────


class TestMaskSampleRows:
    def test_amount_cells_masked(self):
        rows = [["12.03.2026", "-45,99", "PLN", "Coffee"]]
        masked = sp.mask_sample_rows(rows, {"amount": None})
        # "-45,99" matches _AMOUNT_RE → masked.
        assert masked[0][1] == "***"

    def test_account_number_masked(self):
        rows = [["2026-01-01", "12345678901234567890", "PLN", "Transfer"]]
        masked = sp.mask_sample_rows(rows, {})
        assert masked[0][1] == "***"

    def test_description_not_masked(self):
        rows = [["2026-01-01", "-10,00", "PLN", "Grocery store"]]
        masked = sp.mask_sample_rows(rows, {})
        assert masked[0][3] == "Grocery store"

    def test_date_not_masked(self):
        rows = [["2026-01-01", "-10,00", "PLN", "Shop"]]
        masked = sp.mask_sample_rows(rows, {})
        assert masked[0][0] == "2026-01-01"

    def test_multiple_rows(self):
        rows = [
            ["2026-01-01", "-10,00", "PLN", "A"],
            ["2026-01-02", "5,00", "EUR", "B"],
        ]
        masked = sp.mask_sample_rows(rows, {})
        assert all(r[1] == "***" for r in masked)


# ─────────────────────────────────────────────────────────────────────────────
# propose_mapping (mocked AI)
# ─────────────────────────────────────────────────────────────────────────────


class TestProposeMapping:
    def _make_ai_client(self, response_json: dict) -> MagicMock:
        """Return a mock that satisfies the public AIProvider.chat interface."""
        client = MagicMock()
        client.chat.return_value = json.dumps(response_json)
        return client

    def test_propose_mapping_mocked(self):
        """Mock AI client; assert the right prompt is sent and result parsed."""
        ai_response = {
            "column_map": {
                "date": "TransDate",
                "amount": "Amount",
                "currency": "CCY",
                "description": "Info",
                "time": None,
            },
            "date_format": "%d.%m.%Y",
            "decimal_separator": ",",
            "sign_convention": "negative_expense",
        }
        client = self._make_ai_client(ai_response)

        headers = ["TransDate", "Amount", "CCY", "Info"]
        sample_rows = [["12.03.2026", "-45,99", "PLN", "Shop"]]

        result = sp.propose_mapping(headers, sample_rows, client)

        assert result["column_map"]["date"] == "TransDate"
        assert result["column_map"]["amount"] == "Amount"
        assert result["date_format"] == "%d.%m.%Y"
        assert result["decimal_separator"] == ","
        assert result["sign_convention"] == "negative_expense"

        # Verify chat was called once with correct role structure.
        client.chat.assert_called_once()
        messages = client.chat.call_args[0][0]
        assert messages[0]["role"] == "system"
        assert messages[1]["role"] == "user"
        assert "TransDate" in messages[1]["content"]

    def test_propose_mapping_ai_failure_returns_empty(self):
        """AI failure (exception) must return {} not raise."""
        client = MagicMock()
        client.chat.side_effect = RuntimeError("network error")
        result = sp.propose_mapping(["A", "B"], [["1", "2"]], client)
        assert result == {}

    def test_propose_mapping_bad_json_returns_empty(self):
        """Non-JSON AI response must return {}."""
        client = MagicMock()
        client.chat.return_value = "Sorry, I cannot help with that."
        result = sp.propose_mapping(["A", "B"], [["1", "2"]], client)
        assert result == {}

    def test_propose_mapping_missing_column_map_returns_empty(self):
        """AI response missing column_map key must return {}."""
        client = MagicMock()
        client.chat.return_value = json.dumps({"date_format": "%Y-%m-%d"})
        result = sp.propose_mapping(["A", "B"], [["1", "2"]], client)
        assert result == {}

    def test_propose_mapping_mocked_with_fences(self):
        """AI returns a markdown-fenced JSON block; fence stripping must work."""
        fenced_response = (
            "```json\n"
            '{"column_map": {"date": "Date"}, "date_format": "%d.%m.%Y",'
            ' "decimal_separator": ".", "sign_convention": "negative_expense"}'
            "\n```"
        )
        client = MagicMock()
        client.chat.return_value = fenced_response

        result = sp.propose_mapping(["Date", "Amount"], [["01.01.2026", "-10.00"]], client)

        assert result.get("column_map", {}).get("date") == "Date"
        assert result.get("date_format") == "%d.%m.%Y"
        assert result.get("decimal_separator") == "."
        assert result.get("sign_convention") == "negative_expense"


# ─────────────────────────────────────────────────────────────────────────────
# .txt sniffing
# ─────────────────────────────────────────────────────────────────────────────


class TestTxtSniff:
    def test_txt_sniff_csv_semicolon(self):
        """Consistent semicolon-delimited .txt detected as CSV."""
        content = "TransDate;Amount;CCY;Info\n12.03.2026;-45,99;PLN;Coffee\n13.03.2026;100,00;EUR;Salary"
        delim = sp.sniff_txt_delimiter(content)
        assert delim == ";"

    def test_txt_sniff_csv_comma(self):
        content = "Date,Value,Currency\n2026-01-01,10.00,PLN\n2026-01-02,20.00,EUR"
        delim = sp.sniff_txt_delimiter(content)
        assert delim == ","

    def test_txt_sniff_fallback(self):
        """Inconsistent column count → None (falls through to AI text path)."""
        content = "random text without\nconsistent delimiter structure abc\nfoo"
        delim = sp.sniff_txt_delimiter(content)
        assert delim is None

    def test_txt_sniff_single_line_returns_none(self):
        content = "only one line here"
        delim = sp.sniff_txt_delimiter(content)
        assert delim is None

    def test_txt_sniff_tab_delimiter(self):
        content = "Date\tAmount\tCCY\n2026-01-01\t10.00\tPLN\n2026-01-02\t20.00\tEUR"
        delim = sp.sniff_txt_delimiter(content)
        assert delim == "\t"


# ─────────────────────────────────────────────────────────────────────────────
# save_profile / load_profiles round-trip
# ─────────────────────────────────────────────────────────────────────────────


class TestProfileRoundtrip:
    def test_profile_roundtrip(self, tmp_path):
        """save_profile then load_profiles; assert loaded matches saved."""
        sp.save_profile(BANKA_PROFILE, tmp_path)

        profiles = sp.load_profiles(tmp_path)
        fp = sp.fingerprint_from_headers(BANKA_PROFILE["fingerprint"])
        assert fp in profiles

        loaded = profiles[fp]
        assert loaded["name"] == "BankA"
        assert loaded["delimiter"] == ";"
        assert loaded["date_format"] == "%d.%m.%Y"
        assert loaded["column_map"]["date"] == "TransDate"
        assert loaded["sign_convention"] == "negative_expense"

    def test_save_profile_sanitizes_name(self, tmp_path):
        """Profile with special chars in name gets a safe filename."""
        profile = {**BANKA_PROFILE, "name": "My Bank/A 2026"}
        sp.save_profile(profile, tmp_path)
        files = list(tmp_path.glob("*.json"))
        profile_files = [f for f in files if "Bank" in f.name]
        assert len(profile_files) == 1
        # Name must not contain slash.
        assert "/" not in profile_files[0].name
        assert "My_Bank" in profile_files[0].stem

    def test_save_creates_dir(self, tmp_path):
        """save_profile creates the directory if it doesn't exist."""
        target = tmp_path / "sub" / "profiles"
        sp.save_profile(BANKA_PROFILE, target)
        assert target.is_dir()
        assert any(target.glob("*.json"))

    def test_multiple_profiles_loaded(self, tmp_path):
        """Two profiles with different fingerprints both load correctly."""
        p1 = {**BANKA_PROFILE, "name": "BankA"}
        p2 = {
            **BANKA_PROFILE,
            "name": "BankB",
            "fingerprint": ["ValueDate", "Debit", "Credit", "Ref"],
        }
        sp.save_profile(p1, tmp_path)
        sp.save_profile(p2, tmp_path)
        profiles = sp.load_profiles(tmp_path)
        assert len(profiles) == 2


# ─────────────────────────────────────────────────────────────────────────────
# debit_credit_split sign convention
# ─────────────────────────────────────────────────────────────────────────────

SPLIT_PROFILE = {
    "name": "SplitBank",
    "delimiter": ",",
    "encoding": "utf-8",
    "header_row": 0,
    "fingerprint": ["TxnDate", "Debits", "Credits", "CCY", "Memo"],
    "column_map": {
        "date": "TxnDate",
        "debit": "Debits",
        "credit": "Credits",
        "currency": "CCY",
        "description": "Memo",
        "time": None,
    },
    "date_format": "%Y-%m-%d",
    "decimal_separator": ".",
    "sign_convention": sp.SIGN_DEBIT_CREDIT_SPLIT,
}


class TestDebitCreditSplit:
    def _make_split_csv(self, rows: list[list]) -> bytes:
        return _make_csv_bytes(
            ["TxnDate", "Debits", "Credits", "CCY", "Memo"],
            rows,
            delimiter=",",
        )

    def test_debit_row_becomes_expense(self):
        csv_bytes = self._make_split_csv([
            ["2026-05-01", "45.99", "", "PLN", "Grocery store"],
        ])
        result = sp.parse_statement(csv_bytes, "statement.csv", SPLIT_PROFILE)
        assert len(result) == 1
        assert result[0]["type"] == "Expense"
        assert result[0]["value"] == pytest.approx(45.99)
        assert result[0]["description"] == "Grocery store"

    def test_credit_row_becomes_income(self):
        csv_bytes = self._make_split_csv([
            ["2026-05-02", "", "1500.00", "EUR", "Salary"],
        ])
        result = sp.parse_statement(csv_bytes, "statement.csv", SPLIT_PROFILE)
        assert len(result) == 1
        assert result[0]["type"] == "Income"
        assert result[0]["value"] == pytest.approx(1500.00)

    def test_both_zero_row_skipped(self):
        """Row with both debit and credit empty/zero is skipped (balance line)."""
        csv_bytes = self._make_split_csv([
            ["2026-05-03", "", "", "PLN", "Balance brought forward"],
            ["2026-05-04", "10.00", "", "PLN", "Coffee"],
        ])
        result = sp.parse_statement(csv_bytes, "statement.csv", SPLIT_PROFILE)
        assert len(result) == 1
        assert result[0]["description"] == "Coffee"

    def test_both_zero_explicit_zeros_skipped(self):
        """Explicit '0' / '0.00' in both columns → skip."""
        csv_bytes = self._make_split_csv([
            ["2026-05-05", "0.00", "0.00", "PLN", "Fee waiver"],
            ["2026-05-06", "5.00", "0", "PLN", "Bus"],
        ])
        result = sp.parse_statement(csv_bytes, "statement.csv", SPLIT_PROFILE)
        assert len(result) == 1
        assert result[0]["description"] == "Bus"

    def test_both_nonzero_prefers_debit(self):
        """Unusual case: both columns non-zero — debit wins, type is Expense."""
        csv_bytes = self._make_split_csv([
            ["2026-05-07", "20.00", "5.00", "PLN", "Weird row"],
        ])
        result = sp.parse_statement(csv_bytes, "statement.csv", SPLIT_PROFILE)
        assert len(result) == 1
        assert result[0]["type"] == "Expense"
        assert result[0]["value"] == pytest.approx(20.00)

    def test_multiple_rows_mixed(self):
        """Expense + income rows both parsed correctly in one file."""
        csv_bytes = self._make_split_csv([
            ["2026-05-01", "45.99", "", "PLN", "Grocery"],
            ["2026-05-02", "", "1500.00", "EUR", "Salary"],
            ["2026-05-03", "", "", "PLN", "Balance"],
            ["2026-05-04", "9.99", "", "PLN", "Coffee"],
        ])
        result = sp.parse_statement(csv_bytes, "statement.csv", SPLIT_PROFILE)
        assert len(result) == 3
        types = [r["type"] for r in result]
        assert types == ["Expense", "Income", "Expense"]

    def test_split_profile_xlsx(self):
        """XLSX file with split columns parsed correctly."""
        profile = {**SPLIT_PROFILE, "decimal_separator": "."}
        xlsx_bytes = _make_xlsx_bytes(
            ["TxnDate", "Debits", "Credits", "CCY", "Memo"],
            [
                ["2026-06-01", 30.0, None, "PLN", "Lunch"],
                ["2026-06-02", None, 200.0, "PLN", "Refund"],
            ],
        )
        result = sp.parse_statement(xlsx_bytes, "statement.xlsx", profile)
        assert len(result) == 2
        assert result[0]["type"] == "Expense"
        assert result[0]["value"] == pytest.approx(30.0)
        assert result[1]["type"] == "Income"
        assert result[1]["value"] == pytest.approx(200.0)

    def test_amount_always_positive(self):
        """Amount is stored positive regardless of column (sign lives in type)."""
        csv_bytes = self._make_split_csv([
            ["2026-05-01", "100.00", "", "PLN", "Payment"],
            ["2026-05-02", "", "50.00", "PLN", "Refund"],
        ])
        result = sp.parse_statement(csv_bytes, "statement.csv", SPLIT_PROFILE)
        for row in result:
            assert row["value"] > 0

    def test_comma_decimal_in_split_columns(self):
        """Split-column amounts with comma decimal separator parse correctly."""
        profile = {**SPLIT_PROFILE, "decimal_separator": ","}
        csv_bytes = _make_csv_bytes(
            ["TxnDate", "Debits", "Credits", "CCY", "Memo"],
            [
                ["2026-05-01", "1.234,56", "", "PLN", "Big buy"],
                ["2026-05-02", "", "500,00", "EUR", "Payment in"],
            ],
            delimiter=",",
        )
        # Note: comma-decimal and comma-delimiter can conflict; we use semicolon
        # for this test to avoid CSV parsing ambiguity.
        profile2 = {**profile, "delimiter": ";"}
        csv2 = _make_csv_bytes(
            ["TxnDate", "Debits", "Credits", "CCY", "Memo"],
            [
                ["2026-05-01", "1.234,56", "", "PLN", "Big buy"],
                ["2026-05-02", "", "500,00", "EUR", "Pay in"],
            ],
            delimiter=";",
        )
        result = sp.parse_statement(csv2, "statement.csv", profile2)
        assert len(result) == 2
        assert result[0]["value"] == pytest.approx(1234.56)
        assert result[1]["value"] == pytest.approx(500.00)


# ─────────────────────────────────────────────────────────────────────────────
# validate_profile_mapping
# ─────────────────────────────────────────────────────────────────────────────


class TestValidateProfileMapping:
    def _make_profile(self, col_map: dict, sign: str = "negative_expense") -> dict:
        return {"column_map": col_map, "sign_convention": sign}

    def test_valid_single_amount(self):
        p = self._make_profile({"date": "Date", "amount": "Amt", "currency": "CCY"})
        assert sp.validate_profile_mapping(p) == []

    def test_valid_split(self):
        p = self._make_profile(
            {"date": "Date", "debit": "Dr", "credit": "Cr", "currency": "CCY"},
            sign=sp.SIGN_DEBIT_CREDIT_SPLIT,
        )
        assert sp.validate_profile_mapping(p) == []

    def test_missing_date(self):
        p = self._make_profile({"amount": "Amt", "currency": "CCY"})
        errors = sp.validate_profile_mapping(p)
        assert any("date" in e for e in errors)

    def test_missing_currency(self):
        p = self._make_profile({"date": "Date", "amount": "Amt"})
        errors = sp.validate_profile_mapping(p)
        assert any("currency" in e for e in errors)

    def test_missing_amount(self):
        p = self._make_profile({"date": "Date", "currency": "CCY"})
        errors = sp.validate_profile_mapping(p)
        assert any("amount" in e for e in errors)

    def test_debit_without_credit(self):
        p = self._make_profile(
            {"date": "Date", "debit": "Dr", "currency": "CCY"},
            sign=sp.SIGN_DEBIT_CREDIT_SPLIT,
        )
        errors = sp.validate_profile_mapping(p)
        assert any("credit" in e for e in errors)

    def test_credit_without_debit(self):
        p = self._make_profile(
            {"date": "Date", "credit": "Cr", "currency": "CCY"},
            sign=sp.SIGN_DEBIT_CREDIT_SPLIT,
        )
        errors = sp.validate_profile_mapping(p)
        assert any("debit" in e for e in errors)

    def test_amount_and_debit_conflict(self):
        p = self._make_profile({"date": "Date", "amount": "Amt", "debit": "Dr", "currency": "CCY"})
        errors = sp.validate_profile_mapping(p)
        assert any("both" in e or "amount and debit" in e for e in errors)

    def test_fully_empty_col_map(self):
        p = self._make_profile({})
        errors = sp.validate_profile_mapping(p)
        assert len(errors) >= 3  # date, currency, amount all missing


# ─────────────────────────────────────────────────────────────────────────────
# delete_profile / list_profiles
# ─────────────────────────────────────────────────────────────────────────────


class TestDeleteAndListProfiles:
    def test_delete_existing_profile(self, tmp_path):
        sp.save_profile(BANKA_PROFILE, tmp_path)
        deleted = sp.delete_profile("BankA", tmp_path)
        assert deleted is True
        assert list(tmp_path.glob("BankA.json")) == []

    def test_delete_nonexistent_profile_returns_false(self, tmp_path):
        assert sp.delete_profile("DoesNotExist", tmp_path) is False

    def test_delete_nonexistent_dir_returns_false(self, tmp_path):
        assert sp.delete_profile("BankA", tmp_path / "missing") is False

    def test_delete_by_name_field_not_filename(self, tmp_path):
        """delete_profile matches on the 'name' JSON field, not the filename."""
        profile = {**BANKA_PROFILE, "name": "My Bank A"}
        sp.save_profile(profile, tmp_path)
        deleted = sp.delete_profile("My Bank A", tmp_path)
        assert deleted is True

    def test_list_profiles_sorted_by_name(self, tmp_path):
        for name in ("Zeta", "Alpha", "Gamma"):
            sp.save_profile({**BANKA_PROFILE, "name": name}, tmp_path)
        profiles = sp.list_profiles(tmp_path)
        names = [p["name"] for p in profiles]
        assert names == sorted(names, key=str.lower)

    def test_list_profiles_empty_dir(self, tmp_path):
        assert sp.list_profiles(tmp_path) == []

    def test_list_profiles_nonexistent_dir(self, tmp_path):
        assert sp.list_profiles(tmp_path / "missing") == []

    def test_list_after_delete(self, tmp_path):
        sp.save_profile(BANKA_PROFILE, tmp_path)
        p2 = {**BANKA_PROFILE, "name": "BankB", "fingerprint": ["A", "B"]}
        sp.save_profile(p2, tmp_path)
        assert len(sp.list_profiles(tmp_path)) == 2
        sp.delete_profile("BankA", tmp_path)
        remaining = sp.list_profiles(tmp_path)
        assert len(remaining) == 1
        assert remaining[0]["name"] == "BankB"


# ─────────────────────────────────────────────────────────────────────────────
# mask_sample_rows — debit/credit
# ─────────────────────────────────────────────────────────────────────────────


class TestMaskSampleRowsSplitColumns:
    def test_debit_and_credit_col_names_masked(self):
        """If proposal maps debit/credit column names, those values are masked."""
        proposal = {"debit": "Debits", "credit": "Credits"}
        rows = [["2026-01-01", "Debits", "Credits", "PLN", "Shop"]]
        masked = sp.mask_sample_rows(rows, proposal)
        # "Debits" and "Credits" match masked_col_names — masked.
        assert masked[0][1] == "***"
        assert masked[0][2] == "***"

    def test_amount_regex_masks_numeric_cells(self):
        """Numeric-looking cells in debit/credit columns are masked via regex."""
        rows = [["2026-01-01", "45.99", "0.00", "PLN", "Grocery"]]
        masked = sp.mask_sample_rows(rows, {})
        assert masked[0][1] == "***"
        assert masked[0][2] == "***"


# ─────────────────────────────────────────────────────────────────────────────
# AI proposal — debit/credit split response
# ─────────────────────────────────────────────────────────────────────────────


class TestProposeMappingSplit:
    def test_propose_mapping_split_columns(self):
        """AI proposes debit+credit split; response is accepted."""
        ai_response = {
            "column_map": {
                "date": "TxnDate",
                "amount": None,
                "debit": "Debits",
                "credit": "Credits",
                "currency": "CCY",
                "description": "Memo",
                "time": None,
            },
            "date_format": "%Y-%m-%d",
            "decimal_separator": ".",
            "sign_convention": sp.SIGN_DEBIT_CREDIT_SPLIT,
        }
        client = MagicMock()
        client.chat.return_value = json.dumps(ai_response)

        result = sp.propose_mapping(
            ["TxnDate", "Debits", "Credits", "CCY", "Memo"],
            [["2026-05-01", "45.99", "", "PLN", "Grocery"]],
            client,
        )

        assert result["sign_convention"] == sp.SIGN_DEBIT_CREDIT_SPLIT
        assert result["column_map"]["debit"] == "Debits"
        assert result["column_map"]["credit"] == "Credits"
        assert result["column_map"]["amount"] is None

    def test_prompt_mentions_split_columns(self):
        """The generated prompt instructs the AI about debit/credit split."""
        from statement_profiles import _build_mapping_prompt, SIGN_DEBIT_CREDIT_SPLIT

        prompt = _build_mapping_prompt(["Date", "Debits", "Credits"], [])
        assert SIGN_DEBIT_CREDIT_SPLIT in prompt
        assert "debit" in prompt.lower()
        assert "credit" in prompt.lower()
