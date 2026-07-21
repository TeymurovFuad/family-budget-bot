"""
test_validators.py — shared validation layer for all entry paths
(BACKLOG "Follow-up PR: data validation").
"""

import datetime
from datetime import date

import pytest
from openpyxl import Workbook

from excel_schema import MasterDataSchema, col_indices, header_of, write_transaction_row
from models import Transaction
from validators import (
    MAX_PAST_DAYS,
    coerce_bool,
    parse_amount,
    validate_parsed_row,
)

LISTS = {
    "txn_types": ["Expense", "Income", "Savings"],
    "categories": ["Groceries", "Transport", "Savings", "Salary", "Other"],
    "persons": ["Alice", "Bob"],
    "currencies": ["PLN", "EUR", "USD"],
}

TODAY = date(2026, 7, 22)


def _row(**overrides):
    base = {
        "date": "2026-07-20",
        "value": 50,
        "currency": "PLN",
        "type": "Expense",
        "category": "Groceries",
        "description": "shop",
        "person": "",
    }
    base.update(overrides)
    return base


# ═══════════════════════════════════════════════════════════════════════════════
# parse_amount — one shared normalizer for every entry path
# ═══════════════════════════════════════════════════════════════════════════════

class TestParseAmount:
    @pytest.mark.parametrize("raw,expected", [
        ("45.00", 45.0),
        ("-45.00", -45.0),
        ("1 234,56", 1234.56),
        ("1.234,56", 1234.56),
        ("1,234.56", 1234.56),
        ("1 234,56", 1234.56),       # non-breaking space thousands
        ("12,5", 12.5),
        ("1,234,567", 1234567.0),          # repeated separator = thousands
        ("1.234.567", 1234567.0),
        ("50 PLN", 50.0),                   # currency suffix ignored
        ("zł 89", 89.0),
        (150.5, 150.5),                     # already numeric
        ("10.567", 10.57),                  # rounded to 2 decimals
    ])
    def test_parses(self, raw, expected):
        assert parse_amount(raw) == expected

    @pytest.mark.parametrize("raw", ["", "abc", None, "-", ",", "PLN"])
    def test_rejects_non_numbers(self, raw):
        with pytest.raises(ValueError):
            parse_amount(raw)


class TestCoerceBool:
    @pytest.mark.parametrize("raw,expected", [
        ("yes", True), ("No", False), ("TRUE", True), ("false", False),
        ("y", True), ("n", False), ("1", True), ("0", False),
        (True, True), (False, False),
    ])
    def test_coerces(self, raw, expected):
        assert coerce_bool(raw) is expected

    def test_rejects_garbage(self):
        with pytest.raises(ValueError):
            coerce_bool("maybe")


# ═══════════════════════════════════════════════════════════════════════════════
# validate_parsed_row — shared validator
# ═══════════════════════════════════════════════════════════════════════════════

class TestValidateParsedRow:
    def test_valid_row_normalized(self):
        ok, reason, norm, corr = validate_parsed_row(
            _row(type="expense", category="groceries", currency="pln", person="alice"),
            LISTS, today=TODAY,
        )
        assert ok
        assert norm["type"] == "Expense"
        assert norm["category"] == "Groceries"
        assert norm["currency"] == "PLN"
        assert norm["person"] == "Alice"
        assert norm["date"] == date(2026, 7, 20)
        assert corr == []

    def test_typo_category_rejected(self):
        ok, reason, norm, corr = validate_parsed_row(
            _row(category="Grocries"), LISTS, today=TODAY)
        assert not ok
        assert "Grocries" in reason

    def test_unknown_person_rejected(self):
        ok, reason, *_ = validate_parsed_row(_row(person="Carol"), LISTS, today=TODAY)
        assert not ok and "Carol" in reason

    def test_unknown_currency_rejected(self):
        ok, reason, *_ = validate_parsed_row(_row(currency="GBP"), LISTS, today=TODAY)
        assert not ok and "GBP" in reason

    def test_zero_value_rejected(self):
        ok, reason, *_ = validate_parsed_row(_row(value=0), LISTS, today=TODAY)
        assert not ok

    def test_string_value_normalized(self):
        ok, _, norm, _ = validate_parsed_row(_row(value="1.234,56"), LISTS, today=TODAY)
        assert ok and norm["value"] == 1234.56

    def test_negative_value_flips_type_to_expense(self):
        ok, _, norm, corr = validate_parsed_row(
            _row(value=-45.0, type="Income", category="Groceries"), LISTS, today=TODAY)
        assert ok
        assert norm["value"] == 45.0
        assert norm["type"] == "Expense"
        assert any("negative" in c for c in corr)

    # Type↔Category coherence (observed live: 2000 PLN transfer-to-self as Expense)
    def test_savings_category_forces_savings_type(self):
        ok, _, norm, corr = validate_parsed_row(
            _row(type="Expense", category="Savings"), LISTS, today=TODAY)
        assert ok
        assert norm["type"] == "Savings"
        assert any("Savings" in c for c in corr)

    def test_salary_category_forces_income_type(self):
        ok, _, norm, corr = validate_parsed_row(
            _row(type="Expense", category="Salary"), LISTS, today=TODAY)
        assert ok and norm["type"] == "Income"

    def test_coherent_row_gets_no_correction(self):
        ok, _, norm, corr = validate_parsed_row(
            _row(type="Savings", category="Savings"), LISTS, today=TODAY)
        assert ok and corr == []

    # Date sanity — aligned with /add's future/90-day check
    def test_future_date_rejected(self):
        ok, reason, *_ = validate_parsed_row(
            _row(date="2026-07-23"), LISTS, today=TODAY)
        assert not ok and "future" in reason.lower()

    def test_old_date_rejected_with_max_past_days(self):
        ok, reason, *_ = validate_parsed_row(
            _row(date="2026-01-01"), LISTS, max_past_days=MAX_PAST_DAYS, today=TODAY)
        assert not ok and "/add" in reason

    def test_old_date_allowed_without_max_past_days(self):
        ok, *_ = validate_parsed_row(_row(date="2024-01-01"), LISTS, today=TODAY)
        assert ok

    def test_invalid_date_rejected(self):
        ok, reason, *_ = validate_parsed_row(_row(date="2026-13-01"), LISTS, today=TODAY)
        assert not ok and "YYYY-MM-DD" in reason

    def test_empty_row_rejected(self):
        ok, *_ = validate_parsed_row({}, LISTS)
        assert not ok


# ═══════════════════════════════════════════════════════════════════════════════
# Transaction model — value rounded to 2 decimals
# ═══════════════════════════════════════════════════════════════════════════════

class TestTransactionValueRounding:
    def test_value_rounded_to_two_decimals(self):
        t = Transaction(date=date(2026, 7, 1), value=10.567, currency="PLN",
                        transaction_type="Expense")
        assert t.value == 10.57

    def test_exact_value_unchanged(self):
        t = Transaction(date=date(2026, 7, 1), value=10.5, currency="PLN",
                        transaction_type="Expense")
        assert t.value == 10.5


# ═══════════════════════════════════════════════════════════════════════════════
# write_transaction_row — honors is_done instead of hardcoding True
# ═══════════════════════════════════════════════════════════════════════════════

def _blank_masterdata_ws():
    wb = Workbook()
    ws = wb.active
    ws.title = "MasterData"
    from dataclasses import fields
    for i, f in enumerate(fields(MasterDataSchema), 1):
        ws.cell(1, i, header_of(MasterDataSchema, f.name))
    return ws


class TestWriteTransactionRowIsDone:
    LU = "$H$2:$I$100"

    def _write(self, row):
        ws = _blank_masterdata_ws()
        write_transaction_row(ws, 2, row, self.LU)
        idx = col_indices(ws, MasterDataSchema)
        return ws.cell(2, idx["is_done"]).value

    def _row(self, **overrides):
        base = Transaction(date=date(2026, 7, 1), value=10, currency="PLN",
                           transaction_type="Expense").to_row()
        base.update(overrides)
        return base

    def test_is_done_false_written(self):
        assert self._write(self._row(is_done=False)) is False

    def test_is_done_true_written(self):
        assert self._write(self._row(is_done=True)) is True

    def test_missing_is_done_defaults_true(self):
        row = self._row()
        row.pop("is_done")
        assert self._write(row) is True

    def test_none_is_done_defaults_true(self):
        assert self._write(self._row(is_done=None)) is True


# ═══════════════════════════════════════════════════════════════════════════════
# Bulk integration — validator on parse + manual edits, is_recurring editable
# ═══════════════════════════════════════════════════════════════════════════════

class TestValidateBulkRows:
    def test_bad_value_row_flagged(self):
        from handlers.bulk_conv import _validate_bulk_rows
        rows = [_row(value="not-a-number")]
        rows, corr = _validate_bulk_rows(rows, LISTS)
        assert "invalid" in rows[0]

    def test_valid_row_not_flagged_and_normalized(self):
        from handlers.bulk_conv import _validate_bulk_rows
        rows = [_row(category="groceries")]
        rows, corr = _validate_bulk_rows(rows, LISTS)
        assert "invalid" not in rows[0]
        assert rows[0]["category"] == "Groceries"

    def test_empty_category_defaults_to_other(self):
        from handlers.bulk_conv import _validate_bulk_rows
        rows = [_row(category="")]
        rows, corr = _validate_bulk_rows(rows, LISTS)
        assert rows[0]["category"] == "Other"
        assert any("Other" in c for c in corr)

    def test_savings_coherence_corrected(self):
        from handlers.bulk_conv import _validate_bulk_rows
        rows = [_row(type="Expense", category="Savings")]
        rows, corr = _validate_bulk_rows(rows, LISTS)
        assert rows[0]["type"] == "Savings"

    def test_flagged_row_shown_in_preview(self):
        from handlers.bulk_conv import _format_bulk_preview
        rows = [_row(), dict(_row(), invalid="Transaction value must be a positive number.")]
        pages = _format_bulk_preview(rows)
        assert "⚠️" in pages[0]
        assert "won't be saved" in pages[0]


class TestApplyBulkEditValidation:
    def test_is_recurring_yes_coerced(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [_row()]
        action, reason = _apply_bulk_edit("1 is_recurring=yes", rows)
        assert reason == "edited" and rows[0]["is_recurring"] is True

    def test_is_recurring_false_coerced(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [_row(is_recurring=True)]
        action, reason = _apply_bulk_edit("1 is_recurring=false", rows)
        assert reason == "edited" and rows[0]["is_recurring"] is False

    def test_is_recurring_garbage_rejected(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [_row()]
        action, reason = _apply_bulk_edit("1 is_recurring=maybe", rows)
        assert reason == "invalid"

    def test_value_locale_format_accepted(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [_row()]
        action, reason = _apply_bulk_edit("1 value=1.234,56", rows)
        assert reason == "edited" and rows[0]["value"] == 1234.56

    def test_negative_value_maps_to_expense(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [_row(type="Income")]
        action, reason = _apply_bulk_edit("1 value=-45.00", rows)
        assert reason == "edited"
        assert rows[0]["value"] == 45.0
        assert rows[0]["type"] == "Expense"

    def test_typo_category_edit_flagged_when_lists_known(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [_row()]
        action, reason = _apply_bulk_edit("1 category=Grocries", rows, LISTS)
        assert reason == "edited"
        assert "Grocries" in rows[0]["invalid"]

    def test_valid_edit_clears_invalid_flag(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [dict(_row(), invalid="Unknown category 'Grocries'.", category="Grocries")]
        action, reason = _apply_bulk_edit("1 category=groceries", rows, LISTS)
        assert reason == "edited"
        assert "invalid" not in rows[0]
        assert rows[0]["category"] == "Groceries"

    def test_edit_without_lists_skips_membership_check(self):
        from handlers.bulk_conv import _apply_bulk_edit
        rows = [_row()]
        action, reason = _apply_bulk_edit("1 category=Whatever", rows)
        assert reason == "edited" and rows[0]["category"] == "Whatever"
