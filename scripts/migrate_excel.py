"""
migrate_excel.py
================
Migration script -- keeps the Lists sheet in sync with the current schema.

Current target schema (v2 -- unified categories):
  col A = Month
  col B = Type
  col C = Category  (single unified list for Expense + Income + Savings)
  col D = Person
  col E = Year
  col F = (empty -- no longer used)
  col G = (empty -- no longer used)
  col I = Currency code
  col J = Rate (PLN)

MasterData col F dropdown points at Lists col C (plain list, no INDIRECT).

Safe to run multiple times -- idempotent.
"""

from pathlib import Path


def fix_currency_dropdown(wb, num_currencies: int) -> None:
    """Fix MasterData col K (Currency) validation to point at Lists!$I$2:$I$N."""
    from openpyxl.worksheet.datavalidation import DataValidation

    if "MasterData" not in wb.sheetnames:
        return

    ws_md = wb["MasterData"]
    ws_md.data_validations.dataValidation = [
        dv for dv in ws_md.data_validations.dataValidation
        if "K" not in str(dv.sqref)
    ]
    dv = DataValidation(
        type="list",
        formula1=f"Lists!$I$2:$I${1 + num_currencies}",
        allow_blank=True,
        showErrorMessage=False,
    )
    dv.sqref = "K2:K10000"
    ws_md.add_data_validation(dv)


def add_plain_category_dropdown(wb, num_categories: int) -> None:
    """
    Replace whatever col-F validation MasterData has with a plain list
    pointing at Lists!$C$2:$C${1 + num_categories}.

    Idempotent -- removes any existing F-column validation first.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    if "MasterData" not in wb.sheetnames:
        return

    ws_md = wb["MasterData"]
    ws_md.data_validations.dataValidation = [
        dv for dv in ws_md.data_validations.dataValidation
        if "F" not in str(dv.sqref)
    ]
    dv = DataValidation(
        type="list",
        formula1=f"Lists!$C$2:$C${1 + num_categories}",
        allow_blank=True,
        showErrorMessage=False,
    )
    dv.sqref = "F2:F10000"
    ws_md.add_data_validation(dv)


def _remove_named_ranges(wb, names) -> None:
    """Remove named ranges if they exist (ignore missing)."""
    for name in names:
        if name in wb.defined_names:
            del wb.defined_names[name]


EXCEL_PATH = Path(__file__).parent.parent / "data" / "Expenses_Improved.xlsx"


def main():
    if not EXCEL_PATH.exists():
        print(
            f"File not found: {EXCEL_PATH}\n"
            "Please place your Expenses_Improved.xlsx in the data/ directory first."
        )
        return

    from openpyxl import load_workbook

    print(f"Opening: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)

    if "Lists" not in wb.sheetnames:
        print("ERROR: No 'Lists' sheet found in the workbook.")
        return

    ws = wb["Lists"]
    changes = []

    # -- Idempotency check ----------------------------------------------------
    col_f_header = ws.cell(1, 6).value

    if col_f_header not in ("Income Category", "Savings Category"):
        # Already migrated (col F is empty/None) -- just reapply validation
        print("Col F header is not 'Income Category' -- looks already migrated.")
        print("Reapplying plain list validation on MasterData col F...")
        num_cats = 0
        for row in range(2, ws.max_row + 20):
            val = ws.cell(row, 3).value
            if val is None:
                break
            num_cats += 1
        add_plain_category_dropdown(wb, num_cats)
        _remove_named_ranges(wb, ["Expense", "Income", "Savings"])
        wb.save(EXCEL_PATH)
        print("Done.")
        return

    # -- Step 1: Collect all three category columns ---------------------------
    import re as _re
    _currency_pat = _re.compile(r'^[A-Z]{2,4}$')

    def _read_col(col_idx: int) -> list:
        vals = []
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row, col_idx).value
            if v is None:
                break
            s = str(v).strip()
            if s.startswith("←"):  # left-arrow annotation label
                break
            vals.append(s)
        return vals

    expense_cats = _read_col(3)   # col C
    income_cats  = _read_col(6)   # col F
    savings_cats = _read_col(7)   # col G

    def _safe(lst):
        return [v.encode("ascii", "replace").decode() for v in lst]
    print(f"  Col C (expense): {_safe(expense_cats)}")
    print(f"  Col F (income):  {_safe(income_cats)}")
    print(f"  Col G (savings): {_safe(savings_cats)}")

    # -- Step 2: Merge, deduplicate, preserve order ---------------------------
    # Known category exceptions — short uppercase strings that are NOT currencies
    _KNOWN_CATEGORIES = {"ZUS"}
    seen = set()
    unified = []
    for cat in expense_cats + income_cats + savings_cats:
        # Skip raw ISO currency codes (2-4 uppercase letters), but allow known categories
        if cat not in _KNOWN_CATEGORIES and _currency_pat.match(cat):
            print(f"  Skipping likely currency code: {cat!r}")
            continue
        if cat not in seen:
            seen.add(cat)
            unified.append(cat)

    # Move "Other" to end if present
    if "Other" in unified:
        unified.remove("Other")
        unified.append("Other")

    print(f"  Unified list ({len(unified)} items): {unified}")

    # -- Step 3: Write unified list to col C, update header -------------------
    ws.cell(1, 3).value = "Category"
    for row in range(2, ws.max_row + 10):
        ws.cell(row, 3).value = None
    for i, val in enumerate(unified, 2):
        ws.cell(i, 3).value = val
    changes.append(f"Wrote {len(unified)} unified categories to col C, header -> 'Category'.")

    # -- Step 4: Clear col F and col G (header + data) ------------------------
    for row in range(1, ws.max_row + 10):
        ws.cell(row, 6).value = None
        ws.cell(row, 7).value = None
    changes.append("Cleared col F and col G (Income Category / Savings Category).")

    # -- Step 5: Remove old named ranges --------------------------------------
    _remove_named_ranges(wb, ["Expense", "Income", "Savings"])
    changes.append("Removed named ranges Expense / Income / Savings (if present).")

    # -- Step 6: Apply plain list validation on MasterData col F --------------
    add_plain_category_dropdown(wb, len(unified))
    changes.append(
        f"Set MasterData col F dropdown -> Lists!$C$2:$C${1 + len(unified)} (plain list)."
    )

    # -- Step 7: Fix currency dropdown on MasterData col K --------------------
    num_currencies = sum(
        1 for row in range(2, ws.max_row + 1)
        if ws.cell(row, 9).value and str(ws.cell(row, 9).value).strip()
    )
    fix_currency_dropdown(wb, num_currencies)
    changes.append(
        f"Set MasterData col K dropdown -> Lists!$I$2:$I${1 + num_currencies} (currency list)."
    )

    # -- Save -----------------------------------------------------------------
    wb.save(EXCEL_PATH)
    print("\nMigration complete. Summary of changes:")
    for c in changes:
        print(f"  - {c}")
    print(f"\nFile saved: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
