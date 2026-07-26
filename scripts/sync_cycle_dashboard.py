"""
sync_cycle_dashboard.py
=======================
Keep the Cycle Dashboard category table in step with the Dashboard.

Reads the Dashboard's H11..TOTAL category list and rebuilds the Cycle
Dashboard block when they differ; creates the whole Cycle Dashboard sheet
if it's absent. Safe to run any number of times — no change → no save.

Usage (on the bot machine):
  python scripts/sync_cycle_dashboard.py
  python scripts/sync_cycle_dashboard.py --dry-run
  python scripts/sync_cycle_dashboard.py --path /path/to/Expenses_Improved.xlsx
"""

import sys
import argparse
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


def main(dry_run: bool) -> None:
    from openpyxl import load_workbook

    from file_storage import ExcelFileContext, atomic_save
    from cycle_dashboard import (
        CYCLE_DASHBOARD_SHEET_NAME,
        sync_cycle_dashboard_categories,
    )

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        existed = CYCLE_DASHBOARD_SHEET_NAME in wb.sheetnames
        count = sync_cycle_dashboard_categories(wb)

        if not existed:
            print(f"{CYCLE_DASHBOARD_SHEET_NAME} sheet was missing — created "
                  f"with {count} category row(s).")
        elif count:
            print(f"Category block rebuilt: {count} row(s) written.")
        else:
            print("Already in sync — nothing to do.")

        if dry_run:
            print("Dry run — no changes saved.")
        elif count or not existed:
            atomic_save(wb, excel_path)
            print(f"Saved: {excel_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Sync Cycle Dashboard categories with the Dashboard.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Report what would change without saving.")
    parser.add_argument("--path", type=Path, default=None,
                        help="Path to Excel file (default: XLSX_PATH env / settings)")
    args = parser.parse_args()

    if args.path is not None:
        import os
        os.environ["XLSX_PATH"] = str(args.path)

    main(args.dry_run)
