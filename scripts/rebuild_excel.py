"""
rebuild_excel.py
================
Creates a fresh, correctly structured Excel workbook and migrates all
data from the existing file into it.

What it does:
  1.  Reads every row from MasterData (old file)
  2.  Reads Lists data: categories, persons, currencies+rates, years
  3.  Creates a new blank workbook via create_blank_excel()
  4.  Writes Lists data into the new file
  5.  Writes all MasterData rows into the new file
  6.  Applies MasterData styling: navy headers, freeze B2, row CF (Expense/Income/Savings)
  7.  Applies Lists styling: blue headers, alternating rows, F-H gap fill
  8.  Applies Dashboard styling: freeze A4, filter row fill, F2 currency validation
  9.  Applies Monthly Summary arrow/delta columns and totals row
  10. Rebuilds the Guide sheet
  11. Verifies row count matches — aborts if not
  12. Saves as data/Expenses_Improved.xlsx (replaces old file)

Safe to run multiple times — always reads from the live file and writes
a fresh copy. A backup is saved before overwriting.

Usage:
  python scripts/rebuild_excel.py
  python scripts/rebuild_excel.py --source path/to/other.xlsx
"""

import sys
import shutil
import argparse
from datetime import datetime, timezone
from pathlib import Path

# Make sure project root is on the path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from file_storage import create_blank_excel, TEMPLATE_PATH

SOURCE_DEFAULT = ROOT / "data" / "Expenses_Improved.xlsx"
DEST           = ROOT / "data" / "Expenses_Improved.xlsx"


# ── helpers ───────────────────────────────────────────────────────────────────

def read_column(ws, col_idx: int) -> list:
    """Read non-None values from a column starting at row 2."""
    values = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, col_idx).value
        if v is None:
            break
        if isinstance(v, str) and v.startswith("←"):
            break
        values.append(v)
    return values


def find_col(ws, header: str) -> int | None:
    """Return 1-based column index by header name (case-insensitive)."""
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h and str(h).strip().lower() == header.lower():
            return c
    return None


def rebuild_guide(ws):
    """Write a clean, formatted Guide sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment

    TITLE_FONT = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
    TITLE_FILL = PatternFill('solid', fgColor='1F3864')
    H2_FONT    = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    H2_FILL    = PatternFill('solid', fgColor='2E75B6')
    BODY_FONT  = Font(name='Calibri', size=11)
    GREEN_FONT = Font(name='Calibri', bold=True, size=11, color='1F3864')
    LABEL_FONT = Font(name='Calibri', bold=True, size=11, color='375623')
    CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    INDENT1    = Alignment(indent=1, vertical='center')
    WRAP       = Alignment(wrap_text=True, vertical='center', indent=1)

    def title_row(row, text):
        ws.row_dimensions[row].height = 32
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row, 1, text)
        c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = CENTER

    def h2_row(row, text):
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row, 1, text)
        c.font = H2_FONT; c.fill = H2_FILL; c.alignment = INDENT1

    def body_row(row, left, right='', green_left=False):
        ws.row_dimensions[row].height = 18
        c1 = ws.cell(row, 1, left)
        c1.font = GREEN_FONT if green_left else LABEL_FONT
        c1.alignment = INDENT1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c2 = ws.cell(row, 2, right)
        c2.font = BODY_FONT; c2.alignment = WRAP

    def blank(row, h=7):
        ws.row_dimensions[row].height = h

    r = 1
    title_row(r, '📖  Budget Bot — Excel Guide'); r += 1
    blank(r); r += 1

    h2_row(r, '📋  Lists Sheet — Column Layout'); r += 1
    for col_ltr, header, desc in [
        ('A', 'Month',      'Jan, Feb … Dec — do not edit'),
        ('B', 'Type',       'Expense | Income | Savings — do not edit'),
        ('C', 'Category',   'Shared by all types. Add / remove rows freely.'),
        ('D', 'Person',     'Family members. Add rows freely.'),
        ('E', 'Year',       'Years shown in dropdowns. Add future years here.'),
        ('F–H', '(empty)', 'Reserved — do not use'),
        ('I', 'Currency',   '3-letter ISO code e.g. EUR, USD, GBP'),
        ('J', 'Rate (PLN)', 'How many PLN = 1 unit of that currency'),
    ]:
        body_row(r, f'{col_ltr}  {header}', desc, green_left=True); r += 1
    blank(r); r += 1

    h2_row(r, '➕  How to Add Reference Data'); r += 1
    for left, right in [
        ('New category', 'Add a row in column C of the Lists sheet. Bot picks it up immediately — no restart.'),
        ('New person',   'Add a row in column D. No restart needed.'),
        ('New currency', 'Add code in col I and PLN rate in col J. Use /rates to refresh from internet.'),
        ('New year',     'Add the year number in column E. No restart needed.'),
    ]:
        body_row(r, left, right); r += 1
    blank(r); r += 1

    h2_row(r, '📊  MasterData Sheet — Column Layout'); r += 1
    for col_ltr, header, desc in [
        ('A', 'Date',            'Transaction date — set by bot'),
        ('B', 'Year',            'Auto-filled from date'),
        ('C', 'Month',           'Auto-filled from date'),
        ('D', 'Value',           'Amount in original currency'),
        ('E', 'Type',            'Expense | Income | Savings'),
        ('F', 'Category',        'Dropdown from Lists col C'),
        ('G', 'Person',          'Dropdown from Lists col D'),
        ('H', 'Description',     'Free-text note (optional)'),
        ('I', 'IsRecurring',     'TRUE / FALSE'),
        ('J', 'IsDone',          'TRUE = confirmed, FALSE = pending'),
        ('K', 'Currency',        'Dropdown from Lists col I'),
        ('L', 'Value (PLN)',     'Formula: Value × exchange rate'),
        ('M', 'Date Modified',   'Auto-set by bot (UTC)'),
    ]:
        body_row(r, f'{col_ltr}  {header}', desc, green_left=True); r += 1
    blank(r); r += 1

    h2_row(r, '🤖  Bot Restart Rules'); r += 1
    body_row(r, 'No restart needed', 'Any change to the Lists sheet (categories, persons, currencies, years)'); r += 1
    body_row(r, 'Restart required',  'Changes to .py files, .env variables, or bot configuration'); r += 1
    blank(r); r += 1

    h2_row(r, '💬  Bot Commands'); r += 1
    for cmd, desc in [
        ('/add',     'Log a new transaction step-by-step'),
        ('/summary', 'Income, expenses and net for the current month'),
        ('/report',  'Full transaction list for this month'),
        ('/chart',   'Bar chart of expenses by category'),
        ('/savings', '6-month savings trend line chart'),
        ('/budget',  'Budget vs actual spend by category'),
        ('/top',     'Top 5 expenses this month'),
        ('/rates',   'Show current exchange rates + refresh button'),
        ('/range',   'Report for a selected date range (presets or custom)'),
        ('/menu',    'Open the main menu with all buttons'),
    ]:
        body_row(r, cmd, desc, green_left=True); r += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 60
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 10
    ws.sheet_view.showGridLines = False


# ── styling helpers ───────────────────────────────────────────────────────────

def _fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)

def _font(bold=False, color="FF000000", size=11, name="Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)

def _side(style="thin") -> Side:
    return Side(style=style)

def _border(bottom=None) -> Border:
    return Border(bottom=bottom)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def style_masterdata(ws, last_row: int) -> None:
    """Apply all MasterData styling: headers, widths, freeze, row CF."""
    HDR_FILL = _fill("FF1F3864")
    HDR_FONT = _font(bold=True, color="FFFFFFFF", size=11)
    HDR_BORDER = _border(bottom=_side("medium"))

    col_widths = [12, 6, 9, 10, 10, 20, 14, 30, 12, 8, 10, 12, 20]

    for c_idx in range(1, 14):
        cell = ws.cell(1, c_idx)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.alignment = _center()
        cell.border = HDR_BORDER
        if c_idx <= len(col_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = col_widths[c_idx - 1]

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = True

    if last_row < 2:
        return

    rng = f"A2:M{last_row}"

    # Expense → salmon
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['$E2="Expense"'],
        fill=_fill("FFFFC7CE"),
        stopIfTrue=False,
    ))
    # Income → light green
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['$E2="Income"'],
        fill=_fill("FFC6EFCE"),
        stopIfTrue=False,
    ))
    # Savings → light blue
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['$E2="Savings"'],
        fill=_fill("FFDAE8FC"),
        stopIfTrue=False,
    ))


def style_lists(ws, num_categories: int, num_persons: int,
                num_years: int, num_currencies: int) -> None:
    """Apply Lists sheet styling: blue headers, alternating rows, F-H gap fill."""
    HDR_FILL = _fill("FF2E75B6")
    HDR_FONT = _font(bold=True, color="FFFFFFFF")
    GAP_FILL = _fill("FFDEEBF7")
    ALT_FILL = _fill("FFF2F8FF")

    col_widths = [10, 14, 18, 14, 8, 4, 4, 4, 10, 12]
    for c_idx in range(1, 11):
        cell = ws.cell(1, c_idx)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = _center()
        if c_idx <= len(col_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = col_widths[c_idx - 1]

    # F-H gap columns: light blue fill for all data rows
    max_data = max(num_categories, num_persons, num_years, num_currencies, 12)
    for r in range(2, max_data + 5):
        for c in [6, 7, 8]:
            ws.cell(r, c).fill = GAP_FILL

    # Alternating fill for data columns
    for r in range(2, max_data + 5):
        if r % 2 == 0:
            for c in [1, 2, 3, 4, 5, 9, 10]:
                ws.cell(r, c).fill = ALT_FILL

    ws.sheet_view.showGridLines = False


def style_dashboard(ws) -> None:
    """Apply Dashboard styling: freeze panes, filter row fill."""
    ws.freeze_panes = "A4"
    FILTER_FILL = _fill("FFD6E4F0")
    for c in range(1, 12):
        ws.cell(2, c).fill = FILTER_FILL


def _ensure_monthly_summary(wb) -> None:
    """Add arrow/delta columns to Monthly Summary if present."""
    if "Monthly Summary" not in wb.sheetnames:
        return

    ws = wb["Monthly Summary"]

    # Check if col H header is already set (arrow columns already exist)
    if ws.cell(1, 8).value and "Net" in str(ws.cell(1, 8).value):
        return

    # Collect existing data to compute MoM deltas
    months_data = []
    for r in range(2, ws.max_row + 1):
        month_name = ws.cell(r, 1).value
        income     = ws.cell(r, 2).value
        expenses   = ws.cell(r, 3).value
        savings    = ws.cell(r, 4).value
        if not month_name:
            break
        try:
            inc = float(income or 0)
            exp = float(expenses or 0)
            sav = float(savings or 0)
        except (TypeError, ValueError):
            inc = exp = sav = 0.0
        net = inc - exp - sav
        months_data.append((r, month_name, inc, exp, sav, net))

    if not months_data:
        return

    # Write headers for extra cols
    HDRF = _fill("FF1F3864")
    HDRT = _font(bold=True, color="FFFFFFFF")
    for c_idx, hdr in [(8, "Net ↕"), (9, "Exp Δ MoM"), (10, "Cum. Net")]:
        cell = ws.cell(1, c_idx, hdr)
        cell.fill      = HDRF
        cell.font      = HDRT
        cell.alignment = _center()
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 12

    GREEN = _fill("FFC6EFCE")
    RED   = _fill("FFFFC7CE")
    cumulative = 0.0
    prev_exp   = None

    for i, (r, month_name, inc, exp, sav, net) in enumerate(months_data):
        cumulative += net

        # Net arrow
        arrow = "▲" if net >= 0 else "▼"
        net_cell = ws.cell(r, 8, f"{arrow} {net:,.0f}")
        net_cell.fill      = GREEN if net >= 0 else RED
        net_cell.alignment = _center()

        # Expense MoM delta
        if prev_exp is not None:
            delta = exp - prev_exp
            arrow_e = "▲" if delta > 0 else ("▼" if delta < 0 else "─")
            exp_cell = ws.cell(r, 9, f"{arrow_e} {abs(delta):,.0f}")
            exp_cell.fill      = RED if delta > 0 else GREEN
            exp_cell.alignment = _center()
        else:
            ws.cell(r, 9, "─").alignment = _center()
        prev_exp = exp

        # Cumulative net
        cum_cell = ws.cell(r, 10, f"{cumulative:,.0f}")
        cum_cell.fill      = GREEN if cumulative >= 0 else RED
        cum_cell.alignment = _center()

    # Totals row
    if months_data:
        tr = months_data[-1][0] + 1
        tot_inc = sum(d[2] for d in months_data)
        tot_exp = sum(d[3] for d in months_data)
        tot_sav = sum(d[4] for d in months_data)
        tot_net = sum(d[5] for d in months_data)

        TOT_FILL = _fill("FF1F3864")
        TOT_FONT = _font(bold=True, color="FFFFFFFF")
        for c_idx, val in [(1, "TOTAL"), (2, tot_inc), (3, tot_exp),
                           (4, tot_sav), (8, tot_net)]:
            cell = ws.cell(tr, c_idx, val)
            cell.fill      = TOT_FILL
            cell.font      = TOT_FONT
            cell.alignment = _center()


# ── main ──────────────────────────────────────────────────────────────────────

def main(source: Path, dest: Path):
    print(f"Source: {source}")
    print(f"Dest:   {dest}")

    if not source.exists():
        print(f"ERROR: source file not found: {source}")
        sys.exit(1)

    # ── 1. Read old data ──────────────────────────────────────────────────────
    print("\n── Reading source file...")
    old_wb = openpyxl.load_workbook(source, data_only=True)

    # Lists
    old_li = old_wb["Lists"]
    categories = read_column(old_li, find_col(old_li, "Category") or 3)
    persons    = read_column(old_li, find_col(old_li, "Persons") or find_col(old_li, "Person") or 4)
    years      = read_column(old_li, find_col(old_li, "Years") or find_col(old_li, "Year") or 5)

    # Currencies: find by header name
    cur_col  = find_col(old_li, "Currency")
    rate_col = next(
        (c for c in range(1, old_li.max_column + 1)
         if old_li.cell(1, c).value and
            "rate" in str(old_li.cell(1, c).value).lower() and
            "pln" in str(old_li.cell(1, c).value).lower()),
        None
    )
    currencies = []
    if cur_col and rate_col:
        for row in range(2, old_li.max_row + 1):
            code = old_li.cell(row, cur_col).value
            rate = old_li.cell(row, rate_col).value
            if code and rate is not None:
                currencies.append((str(code).strip().upper(), float(rate)))

    print(f"  Categories : {len(categories)}")
    print(f"  Persons    : {len(persons)}")
    print(f"  Years      : {years}")
    print(f"  Currencies : {[c[0] for c in currencies]}")

    # MasterData
    old_md = old_wb["MasterData"]
    md_headers = [old_md.cell(1, c).value for c in range(1, old_md.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(md_headers) if h}

    rows = []
    for r in range(2, old_md.max_row + 1):
        row_vals = {h: old_md.cell(r, idx).value for h, idx in col.items()}
        # Skip completely empty rows
        if all(v is None for v in row_vals.values()):
            continue
        rows.append(row_vals)

    print(f"  MasterData rows: {len(rows)}")

    # Dashboard budgets (col H=Category, col I=Budget)
    budgets = {}
    if "Dashboard" in old_wb.sheetnames:
        old_db = old_wb["Dashboard"]
        for r in range(2, old_db.max_row + 1):
            cat = old_db.cell(r, 8).value
            bud = old_db.cell(r, 9).value
            if cat and bud is not None:
                budgets[str(cat)] = bud

    print(f"  Budget entries: {len(budgets)}")

    # ── 2. Backup old file ────────────────────────────────────────────────────
    if dest.exists():
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup  = dest.parent / f"Expenses_Improved_backup_{ts}.xlsx"
        shutil.copy2(dest, backup)
        print(f"\n── Backup saved: {backup.name}")

    # ── 3. Start from template (preserves Dashboard formulas, Monthly Summary, styling) ──
    print("\n── Loading template...")
    template = TEMPLATE_PATH if TEMPLATE_PATH.exists() else None
    if template:
        print(f"   Using template: {template}")
        new_wb = openpyxl.load_workbook(template, data_only=False)
    else:
        print("   Template not found — falling back to create_blank_excel()")
        tmp = dest.parent / "_rebuild_tmp.xlsx"
        create_blank_excel(tmp)
        new_wb = openpyxl.load_workbook(tmp)
        tmp.unlink()

    # ── 4. Write Lists data ───────────────────────────────────────────────────
    new_li = new_wb["Lists"]

    # Clear existing placeholder data in cols C/D/E/I/J (keep A=months, B=types)
    for col_idx in [3, 4, 5, 9, 10]:
        for r in range(2, 200):
            new_li.cell(r, col_idx).value = None

    for i, val in enumerate(categories, 2):
        new_li.cell(i, 3, val)

    for i, val in enumerate(persons, 2):
        new_li.cell(i, 4, val)

    for i, val in enumerate(years, 2):
        new_li.cell(i, 5, val)

    for i, (code, rate) in enumerate(currencies, 2):
        new_li.cell(i, 9, code)
        new_li.cell(i, 10, rate)

    print(f"  Lists written.")

    # ── 5. Fix validations ────────────────────────────────────────────────────
    new_md = new_wb["MasterData"]

    # Remove all existing validations then re-add correct ones
    new_md.data_validations.dataValidation.clear()

    def add_dv(sqref, formula):
        dv = DataValidation(type="list", formula1=formula,
                            allow_blank=True, showErrorMessage=False)
        dv.sqref = sqref
        new_md.add_data_validation(dv)

    add_dv("E2:E10000", "Lists!$B$2:$B$4")                             # Type
    add_dv("F2:F10000", f"Lists!$C$2:$C${1 + len(categories)}")        # Category
    add_dv("G2:G10000", f"Lists!$D$2:$D${max(1, len(persons)) + 1}")   # Person
    add_dv("B2:B10000", f"Lists!$E$2:$E${1 + len(years)}")             # Year
    add_dv("K2:K10000", f"Lists!$I$2:$I${1 + len(currencies)}")        # Currency

    # ── 6. Write MasterData rows ──────────────────────────────────────────────
    print(f"\n── Writing {len(rows)} MasterData rows...")

    # Expected column order in new file (matches create_blank_excel header)
    NEW_HEADERS = [
        "Date", "Year", "Month", "Value", "Type", "Category",
        "Person", "Description", "IsRecurring", "IsDone",
        "Currency", "Value (PLN)", "Date Modified (UTC)",
    ]

    # Rates dict for recomputing Value(PLN) where missing
    rates_dict = {code: rate for code, rate in currencies}

    for r_idx, row in enumerate(rows, 2):
        date_val     = row.get("Date")
        year_val     = row.get("Year")
        month_val    = row.get("Month")
        value_val    = row.get("Value")
        type_val     = row.get("Type")
        cat_val      = row.get("Category") or ""
        person_val   = row.get("Person") or ""
        desc_val     = row.get("Description") or ""
        recur_val    = row.get("IsRecurring")
        done_val     = row.get("IsDone")
        ccy_val      = row.get("Currency") or "PLN"
        pln_val      = row.get("Value (PLN)")
        mod_val      = row.get("Date Modified (UTC)")

        # Recompute Value(PLN) if missing
        if pln_val is None and value_val is not None:
            rate = rates_dict.get(str(ccy_val).upper(), 1.0)
            pln_val = round(float(value_val) * rate, 4)

        new_md.cell(r_idx,  1, date_val)
        new_md.cell(r_idx,  2, year_val)
        new_md.cell(r_idx,  3, month_val)
        new_md.cell(r_idx,  4, value_val)
        new_md.cell(r_idx,  5, type_val)
        new_md.cell(r_idx,  6, cat_val)
        new_md.cell(r_idx,  7, person_val)
        new_md.cell(r_idx,  8, desc_val)
        new_md.cell(r_idx,  9, recur_val)
        new_md.cell(r_idx, 10, done_val)
        new_md.cell(r_idx, 11, ccy_val)
        new_md.cell(r_idx, 12, pln_val)
        new_md.cell(r_idx, 13, mod_val)

    # ── 7. Dashboard budgets ──────────────────────────────────────────────────
    if "Dashboard" in new_wb.sheetnames:
        new_db = new_wb["Dashboard"]
        # Clear placeholder budget rows
        for r in range(2, 200):
            if new_db.cell(r, 8).value or new_db.cell(r, 9).value:
                new_db.cell(r, 8).value = None
                new_db.cell(r, 9).value = None

        for i, cat in enumerate(categories, 2):
            new_db.cell(i, 8, cat)
            new_db.cell(i, 9, budgets.get(cat, 0))

        print(f"  Dashboard budgets written.")

    # ── 8. Apply MasterData styling ───────────────────────────────────────────
    style_masterdata(new_md, len(rows) + 1)
    print("  MasterData styling applied.")

    # ── 9. Apply Lists styling ────────────────────────────────────────────────
    style_lists(new_li, len(categories), len(persons), len(years), len(currencies))
    print("  Lists styling applied.")

    # ── 10. Apply Dashboard styling ───────────────────────────────────────────
    if "Dashboard" in new_wb.sheetnames:
        style_dashboard(new_wb["Dashboard"])
        # Fix F2 validation: point to Lists!$I$2:$I$N (Display Currency dropdown)
        ws_db = new_wb["Dashboard"]
        ws_db.data_validations.dataValidation = [
            dv for dv in ws_db.data_validations.dataValidation
            if "F2" not in str(dv.sqref)
        ]
        from openpyxl.worksheet.datavalidation import DataValidation as DV2
        dv_ccy = DV2(type="list",
                     formula1=f"Lists!$I$2:$I${1 + len(currencies)}",
                     allow_blank=True, showErrorMessage=False)
        dv_ccy.sqref = "F2"
        ws_db.add_data_validation(dv_ccy)
        print("  Dashboard styling + F2 validation applied.")

    # ── 11. Apply Monthly Summary extra columns ───────────────────────────────
    _ensure_monthly_summary(new_wb)
    print("  Monthly Summary arrow columns applied (if sheet exists).")

    # ── 12. Rebuild Guide sheet ───────────────────────────────────────────────
    if "📖 Guide" in new_wb.sheetnames:
        del new_wb["📖 Guide"]
    ws_guide = new_wb.create_sheet("📖 Guide", 0)
    rebuild_guide(ws_guide)
    print("  Guide sheet rebuilt.")

    # ── 13. Verify ────────────────────────────────────────────────────────────
    written = sum(
        1 for r in range(2, new_md.max_row + 1)
        if any(new_md.cell(r, c).value is not None for c in range(1, 14))
    )
    if written != len(rows):
        print(f"\nERROR: wrote {written} rows but expected {len(rows)} — aborting, not saving.")
        sys.exit(1)

    print(f"\n── Verification: {written}/{len(rows)} rows ✓")

    # ── 14. Save ──────────────────────────────────────────────────────────────
    new_wb.save(dest)
    print(f"\n✅  Done. Saved: {dest}")
    print(f"    Rows:       {written}")
    print(f"    Categories: {len(categories)}")
    print(f"    Persons:    {len(persons)}")
    print(f"    Currencies: {len(currencies)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Rebuild Excel from scratch, migrating all data.")
    parser.add_argument("--source", type=Path, default=SOURCE_DEFAULT,
                        help="Path to source Excel (default: data/Expenses_Improved.xlsx)")
    args = parser.parse_args()
    main(args.source, DEST)
