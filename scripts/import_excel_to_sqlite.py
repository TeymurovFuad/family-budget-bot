"""
import_excel_to_sqlite.py — one-time Excel → SQLite migration (Cycle S1 Phase 1).

Reads MasterData + Lists via the existing readers (data.load_data,
excel_schema.load_currency_rates_from_path, file_storage.load_lists),
computes value_base/rate_used via sqlite_ops.compute_value_base, and inserts
into the SQLite store. Safe to re-run: the content_hash UNIQUE constraint
skips rows already imported (skip count is reported).

Usage:
    python scripts/import_excel_to_sqlite.py [--db PATH] [--dry-run]
"""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd

import settings
import sqlite_ops
from excel_schema import load_currency_rates_from_path
from sqlite_types import SyncDirection, SyncStatus, TransactionRow, TransactionSource


def _clean(v):
    """NaN/NaT → None; pandas scalars → python scalars."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _strip_injection_guard(desc):
    """
    Undo excel_schema.write_transaction_row's formula-injection guard: a
    description starting with = + - @ is written to Excel with a leading '.
    Strip it on import so the round trip (export → re-import) does not change
    the description — and therefore the content_hash — of such rows.
    """
    if isinstance(desc, str) and desc[:1] == "'" and desc[1:2] in ("=", "+", "-", "@"):
        return desc[1:]
    return desc


def df_row_to_txn(row, rates: dict[str, float], conn=None) -> TransactionRow:
    """Map one data.load_data() DataFrame row to a TransactionRow."""
    value = float(row["Value"])
    currency = str(_clean(row.get("Currency")) or settings.DISPLAY_CURRENCY).strip().upper()
    value_base, rate_used = sqlite_ops.compute_value_base(
        value, currency, rates, settings.DISPLAY_CURRENCY, conn=conn)

    date = _clean(row.get("Date"))
    if hasattr(date, "date"):
        date = date.date()
    date_modified = _clean(row.get("Date Modified (UTC)"))

    txn = TransactionRow(
        date=date.isoformat() if date is not None else None,
        year=int(row["Year"]),
        month=_clean(row.get("Month")),
        value=value,
        currency=currency,
        value_base=value_base,
        rate_used=rate_used,
        type=_clean(row.get("Type")),
        category=_clean(row.get("Category")),
        person=_clean(row.get("Person")),
        description=_strip_injection_guard(_clean(row.get("Description"))),
        is_recurring=bool(_clean(row.get("IsRecurring")) or False),
        is_done=bool(row.get("IsDone", True)),
        date_modified_utc=str(date_modified) if date_modified is not None else None,
        source=TransactionSource.EXCEL_IMPORT,
    )
    # Hash over the raw source values (including a possibly-None
    # date_modified_utc) so re-running the import is idempotent —
    # insert_transaction must not stamp a fresh timestamp into the hash.
    txn.content_hash = sqlite_ops.content_hash_for_row(
        txn.date, txn.value, txn.currency, txn.type,
        txn.category, txn.person, txn.description, txn.date_modified_utc)
    return txn


def _load_excel_data(excel_path):
    """Read MasterData from Excel directly (bypasses SQLite — used only here)."""
    df = pd.read_excel(excel_path, sheet_name="MasterData")
    df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
    if "Value (base)" in df.columns:
        df["_base"] = pd.to_numeric(df["Value (base)"], errors="coerce")
    else:
        df["_base"] = pd.to_numeric(df["Value"], errors="coerce")
    if "Currency" not in df.columns:
        df["Currency"] = settings.DISPLAY_CURRENCY
    df["Currency"] = df["Currency"].fillna(settings.DISPLAY_CURRENCY)
    missing = df["_base"].isna() & df["Value"].notna()
    if missing.any():
        rates = load_currency_rates_from_path(excel_path)
        df.loc[missing, "_base"] = df.loc[missing].apply(
            lambda r: r["Value"] * rates.get(str(r["Currency"]).strip().upper(), 1.0),
            axis=1,
        )
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["_base", "Type", "Year", "Month"])
    df["IsDone"] = df["IsDone"].fillna(True).astype(bool)
    return df


def run_import(db_path=None, dry_run: bool = False) -> dict:
    """Import the current workbook into SQLite. Returns counters."""
    from file_storage import get_excel_path_for_reading, load_lists

    excel_path = get_excel_path_for_reading()
    df = _load_excel_data(excel_path)
    rates = load_currency_rates_from_path(excel_path)
    lists = load_lists(excel_path)

    db_path = db_path or settings.SQLITE_DB_PATH
    stats = {"total": len(df), "inserted": 0, "skipped": 0}

    if dry_run:
        # Report against the current DB state without writing.
        conn = sqlite_ops.init_db(db_path)
        try:
            existing = {r["content_hash"] for r in conn.execute(
                f"SELECT content_hash FROM {sqlite_ops.TABLE_TRANSACTIONS}")}
        finally:
            conn.close()
        for _, row in df.iterrows():
            txn = df_row_to_txn(row, rates)
            stats["skipped" if txn.content_hash in existing else "inserted"] += 1
        return stats

    conn = sqlite_ops.init_db(db_path)
    try:
        count_sql = f"SELECT COUNT(*) FROM {sqlite_ops.TABLE_TRANSACTIONS}"
        before = conn.execute(count_sql).fetchone()[0]
        for _, row in df.iterrows():
            sqlite_ops.insert_transaction(conn, df_row_to_txn(row, rates, conn=conn))
        after = conn.execute(count_sql).fetchone()[0]
        stats["inserted"] = after - before
        stats["skipped"] = stats["total"] - stats["inserted"]

        for ccy, rate in rates.items():
            sqlite_ops.upsert_rate(conn, ccy, rate)
        budgets = lists.get("budgets", {})
        for cat in lists.get("categories", []):
            sqlite_ops.upsert_category(conn, str(cat), budgets.get(str(cat)))
        for person in lists.get("persons", []):
            sqlite_ops.upsert_person(conn, str(person))

        # ── Cycles ────────────────────────────────────────────────────────────
        wb = None
        try:
            import openpyxl as _openpyxl
            wb = _openpyxl.load_workbook(excel_path, data_only=True)
            if "Cycles" in wb.sheetnames:
                ws_cycles = wb["Cycles"]
                for row in ws_cycles.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    cell_date, cell_label = row[0], row[1] if len(row) > 1 else None
                    if cell_label is None:
                        continue
                    if hasattr(cell_date, "date"):
                        date_str = cell_date.date().isoformat()
                    elif hasattr(cell_date, "isoformat"):
                        date_str = cell_date.isoformat()[:10]
                    else:
                        try:
                            from datetime import date as _dt
                            date_str = str(cell_date)[:10]
                            _dt.fromisoformat(date_str)
                        except (ValueError, TypeError):
                            continue
                    sqlite_ops.upsert_cycle(conn, date_str, str(cell_label))
        except Exception as _e:
            import logging as _logging
            _logging.getLogger(__name__).warning("Could not import Cycles sheet: %s", _e)

        # ── Salary keywords ────────────────────────────────────────────────────
        try:
            if wb is not None and "Lists" in wb.sheetnames:
                ws_lists = wb["Lists"]
                headers = [str(c.value).strip() if c.value is not None else "" for c in ws_lists[1]]
                try:
                    kw_col = headers.index("Salary Keywords")
                    for row in ws_lists.iter_rows(min_row=2, values_only=True):
                        if len(row) <= kw_col or row[kw_col] is None:
                            continue
                        kw = str(row[kw_col]).strip()
                        if kw:
                            sqlite_ops.insert_salary_keyword(conn, kw)
                except ValueError:
                    pass  # column not found — skip
        except Exception as _e:
            import logging as _logging
            _logging.getLogger(__name__).warning("Could not import salary keywords: %s", _e)

        # ── Category types ─────────────────────────────────────────────────────
        try:
            if wb is not None and "Lists" in wb.sheetnames:
                ws_lists = wb["Lists"]
                headers = [str(c.value).strip() if c.value is not None else "" for c in ws_lists[1]]
                try:
                    cat_col = headers.index("Category")
                    type_col = headers.index("Type")
                    for row in ws_lists.iter_rows(min_row=2, values_only=True):
                        cat_val = row[cat_col] if len(row) > cat_col else None
                        type_val = row[type_col] if len(row) > type_col else None
                        if cat_val and type_val:
                            sqlite_ops.upsert_category(conn, str(cat_val), category_type=str(type_val))
                except ValueError:
                    pass  # columns not found — skip
        except Exception as _e:
            import logging as _logging
            _logging.getLogger(__name__).warning("Could not import category types: %s", _e)

        sqlite_ops.log_sync(conn, SyncDirection.IMPORT, SyncStatus.OK,
                            f"inserted={stats['inserted']} skipped={stats['skipped']}")
    finally:
        conn.close()
    return stats


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Import Excel workbook into SQLite")
    parser.add_argument("--db", default=None, help="SQLite path (default: settings.SQLITE_DB_PATH)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Report what would be imported without writing")
    args = parser.parse_args(argv)

    stats = run_import(db_path=args.db, dry_run=args.dry_run)
    mode = "DRY RUN — would import" if args.dry_run else "Imported"
    print(f"{mode}: {stats['inserted']} new, {stats['skipped']} skipped "
          f"(already present), {stats['total']} total rows read")
    return 0


if __name__ == "__main__":
    sys.exit(main())




