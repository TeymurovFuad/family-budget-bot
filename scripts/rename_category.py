"""
rename_category.py — atomically rename a category everywhere in a workbook:
Lists Categories column, all MasterData rows, the Dashboard budget table,
category names quoted inside formulas (SUMIFS criteria etc.) on Dashboard and
Monthly Summary, and any pending bulk drafts in data/bulk_drafts/*.json.

Usage:  python scripts/rename_category.py "Old Name" "New Name" [path-to-xlsx]
"""
import json
import os
import shutil
import sys

import openpyxl

# Scripts share the bot's configuration — .env is the single source of truth.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import settings
from _repair_guard import repair_guard
from file_storage import atomic_save

sys.stdout.reconfigure(encoding="utf-8")


def main():
    if len(sys.argv) < 3:
        print('Usage: python scripts/rename_category.py "Old Name" "New Name" [xlsx]')
        sys.exit(1)

    old_name, new_name = sys.argv[1], sys.argv[2]
    path = sys.argv[3] if len(sys.argv) > 3 else str(settings.XLSX_PATH)

    with repair_guard():
        shutil.copy2(path, path + ".bak")
        print(f"Backup: {path}.bak")

        from excel_schema import rename_category_in_workbook
        wb = openpyxl.load_workbook(path, data_only=False)
        counts = {"Lists": 0, "Bulk drafts": 0}

        # Lists: Categories column
        ws = wb["Lists"]
        cat_col = next((c for c in range(1, ws.max_column + 1)
                        if str(ws.cell(1, c).value or "").strip() == "Categories"), None)
        if cat_col:
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, cat_col).value or "").strip() == old_name:
                    ws.cell(r, cat_col, new_name)
                    counts["Lists"] += 1

        # MasterData + Dashboard values + formula literals (shared with /setup)
        wb_counts = rename_category_in_workbook(wb, old_name, new_name)
        counts.update(wb_counts)

        # Merchant map: stored category suggestions must follow the rename too.
        import merchant_map
        counts["Merchant map"] = merchant_map.rename_category(old_name, new_name)

        # Pending bulk drafts: parsed-but-unsaved rows keep category by name.
        drafts_dir = settings.BULK_DRAFTS_DIR
        if drafts_dir.is_dir():
            for draft_path in sorted(drafts_dir.glob("*.json")):
                try:
                    rows = json.loads(draft_path.read_text(encoding="utf-8"))
                except Exception as e:
                    print(f"Skipping unreadable draft {draft_path.name}: {e}")
                    continue
                if not isinstance(rows, list):
                    continue
                changed = 0
                for row in rows:
                    if isinstance(row, dict) and str(row.get("category", "")).strip() == old_name:
                        row["category"] = new_name
                        changed += 1
                if changed:
                    draft_path.write_text(
                        json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8"
                    )
                    counts["Bulk drafts"] += changed
                    print(f"Draft {draft_path.name}: {changed} row(s) renamed")

        atomic_save(wb, path)
        for sheet, n in counts.items():
            print(f"{sheet}: {n} cell(s) renamed")
        print(f"Done: '{old_name}' -> '{new_name}'")


if __name__ == "__main__":
    main()
