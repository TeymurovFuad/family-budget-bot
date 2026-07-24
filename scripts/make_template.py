"""
make_template.py
================
Strips all personal data from Expenses_Improved.xlsx and saves a clean
template to data/Expenses_Template.xlsx that can be committed to the repo.

What is removed / replaced:
  - MasterData       : all data rows (header kept)
  - Monthly Summary  : all data rows (header kept)
  - Lists col D      : personal names → ["Person 1", "Person 2"]
  - Dashboard row 2  : filter state reset (Year=current, Month="", Currency=PLN)
  - Dashboard budgets: PLN amounts zeroed (formula structure kept: =0/$N$2)

What is preserved:
  - All sheet structure, formulas, conditional formatting, data validations
  - Lists: months, types, categories, years, currencies, rates
  - Dashboard: all SUMIFS display formulas, layout, styling
  - Guide sheet: untouched

Usage:
  python scripts/make_template.py
  python scripts/make_template.py --source path/to/other.xlsx
"""

import sys
import re
import argparse
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

SOURCE_DEFAULT = ROOT / "data" / "Expenses_Improved.xlsx"
DEST_DEFAULT   = ROOT / "data" / "Expenses_Template.xlsx"

GENERIC_PERSONS = ["Person 1", "Person 2"]


def _clear_rows(ws, from_row: int) -> None:
    """Set all cells to None from from_row to max_row."""
    for r in range(from_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None


def scrub_masterdata(ws) -> None:
    _clear_rows(ws, 2)
    print(f"  MasterData: all data rows cleared.")


def scrub_monthly_summary(ws) -> None:
    _clear_rows(ws, 2)
    print(f"  Monthly Summary: all data rows cleared.")


def scrub_persons(ws) -> None:
    """Replace personal names in Lists col D with generic placeholders."""
    # Find Person column
    person_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().lower() in ("person", "persons"):
            person_col = c
            break
    if person_col is None:
        print("  Lists: Person column not found — skipping.")
        return

    # Clear all existing person values
    for r in range(2, ws.max_row + 1):
        ws.cell(r, person_col).value = None

    # Write generic placeholders
    for i, name in enumerate(GENERIC_PERSONS, 2):
        ws.cell(i, person_col, name)

    print(f"  Lists: persons replaced with {GENERIC_PERSONS}.")


def scrub_dashboard(ws) -> None:
    """Reset filter state and zero out budget PLN amounts."""
    cur_year = datetime.now(timezone.utc).year

    # Row 2 = filter area: Year | value | Month | value | Display | currency
    # Reset: Year → current year, Month → "" (all months), Currency → PLN
    for c in range(1, ws.max_column + 1):
        label = str(ws.cell(2, c).value or "").strip()
        if label == "Year":
            ws.cell(2, c + 1, cur_year)
        elif label == "Month":
            ws.cell(2, c + 1, "")
        elif label == "Display":
            ws.cell(2, c + 1, "PLN")

    # Budget column (col I) — replace =NNNN/$N$2 formulas with =0/$N$2
    budget_col = None
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or "")
        if "budget" in v.lower():
            budget_col = c
            break
    # Also check row 10 (sub-header)
    if budget_col is None:
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(10, c).value or "")
            if "budget" in v.lower():
                budget_col = c
                break

    if budget_col:
        zeroed = 0
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, budget_col)
            val = str(cell.value or "")
            # Replace =NNNN/... or =NNNN.N/... with =0/...
            if re.match(r"^=[\d.]+/", val):
                cell.value = re.sub(r"^=[\d.]+/", "=0/", val)
                zeroed += 1

        print(f"  Dashboard: filter state reset, {zeroed} budget amounts zeroed.")
    else:
        print("  Dashboard: filter state reset. Budget column not found — amounts unchanged.")


def main(source: Path, dest: Path) -> None:
    if not source.exists():
        print(f"ERROR: source not found: {source}")
        sys.exit(1)

    print(f"Source: {source}")
    print(f"Dest:   {dest}\n")

    wb = openpyxl.load_workbook(source, data_only=False)
    print(f"Sheets: {wb.sheetnames}\n")

    if "MasterData" in wb.sheetnames:
        scrub_masterdata(wb["MasterData"])

    if "Monthly Summary" in wb.sheetnames:
        scrub_monthly_summary(wb["Monthly Summary"])

    if "Lists" in wb.sheetnames:
        scrub_persons(wb["Lists"])

    if "Dashboard" in wb.sheetnames:
        scrub_dashboard(wb["Dashboard"])

    from cycles import CYCLES_SHEET_NAME, ensure_cycles_sheet
    if CYCLES_SHEET_NAME in wb.sheetnames:
        ws = wb[CYCLES_SHEET_NAME]
        _clear_rows(ws, 2)
        print("  Cycles: ledger rows cleared.")
    else:
        ensure_cycles_sheet(wb)
        print("  Cycles: sheet created (empty ledger).")

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    print(f"\n✅  Template saved: {dest}")
    print("    All formulas, styling, validations, and structure preserved.")
    print("    Commit with: git add -f data/Expenses_Template.xlsx")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Strip personal data, produce repo template.")
    parser.add_argument("--source", type=Path, default=SOURCE_DEFAULT)
    parser.add_argument("--dest",   type=Path, default=DEST_DEFAULT)
    args = parser.parse_args()
    main(args.source, args.dest)
