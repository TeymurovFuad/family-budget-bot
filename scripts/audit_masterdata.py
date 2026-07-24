"""Audit MasterData rows against Lists reference data. Usage: python scripts/audit_masterdata.py <xlsx>"""
import sys
from collections import Counter

import openpyxl

# Scripts share the bot's configuration — .env is the single source of truth.
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import settings

sys.stdout.reconfigure(encoding="utf-8")

path = sys.argv[1] if len(sys.argv) > 1 else str(settings.XLSX_PATH)
wb = openpyxl.load_workbook(path, data_only=False)
ws_li, ws_md = wb["Lists"], wb["MasterData"]

def read_col(ws, header):
    col = next((c for c in range(1, ws.max_column + 1)
                if str(ws.cell(1, c).value or "").strip().lower() == header.lower()), None)
    if col is None:
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            break
        out.append(str(v).strip())
    return out

categories = set(read_col(ws_li, "Categories"))
persons    = set(read_col(ws_li, "Persons"))
types      = set(read_col(ws_li, "TxnTypes"))
currencies = set(c for c in read_col(ws_li, "Currency") if len(c) == 3 and c.isalpha())
print(f"Lists: {len(categories)} categories, persons={sorted(persons)}, types={sorted(types)}, currencies={sorted(currencies)}")

hdr = {str(ws_md.cell(1, c).value): c for c in range(1, ws_md.max_column + 1)}
issues = Counter()
detail = {}

def flag(kind, row, msg):
    issues[kind] += 1
    detail.setdefault(kind, [])
    if len(detail[kind]) < 8:
        detail[kind].append(f"  row {row}: {msg}")

n_rows = 0
seen_keys = Counter()
for r in range(2, ws_md.max_row + 1):
    val = ws_md.cell(r, hdr["Value"]).value
    dat = ws_md.cell(r, hdr["Date"]).value
    if val is None and dat is None:
        continue
    n_rows += 1
    typ  = str(ws_md.cell(r, hdr["Type"]).value or "")
    cat  = str(ws_md.cell(r, hdr["Category"]).value or "")
    per  = ws_md.cell(r, hdr["Person"]).value
    ccy  = str(ws_md.cell(r, hdr["Currency"]).value or "")
    desc = str(ws_md.cell(r, hdr["Description"]).value or "")
    vpln = ws_md.cell(r, hdr.get("Value (base)", 12)).value

    if cat and cat not in categories:
        flag("unknown_category", r, f"category={cat!r} ({typ}, {val}, {desc[:40]!r})")
    if per and str(per).strip() and str(per).strip() not in persons:
        flag("unknown_person", r, f"person={per!r} ({desc[:40]!r})")
    if typ and typ not in types:
        flag("unknown_type", r, f"type={typ!r}")
    if ccy and ccy not in currencies:
        flag("unknown_currency", r, f"currency={ccy!r}")
    if typ == "Expense" and cat in {"Salary", "Income"}:
        flag("type_category_conflict", r, f"{typ}/{cat} ({val}, {desc[:40]!r})")
    if typ != "Savings" and cat == "Savings":
        flag("type_category_conflict", r, f"{typ}/{cat} ({val}, {desc[:40]!r})")
    if isinstance(dat, str):
        flag("date_is_text", r, f"date={dat!r}")
    if isinstance(val, str):
        flag("value_is_text", r, f"value={val!r}")
    if isinstance(vpln, str) and vpln.startswith("=") and "$H$2:$I$" not in vpln:
        flag("bad_vlookup_range", r, vpln[:70])
    if vpln is None or (isinstance(vpln, str) and not vpln.startswith("=")):
        flag("missing_vpln_formula", r, f"vpln={vpln!r}")
    key = (str(dat)[:10], str(val), ccy, desc.lower()[:40])
    seen_keys[key] += 1

dups = {k: c for k, c in seen_keys.items() if c > 1}
print(f"\nAudited {n_rows} data rows in MasterData\n")
if not issues and not dups:
    print("No issues found.")
for kind, count in issues.most_common():
    print(f"❌ {kind}: {count}")
    for line in detail[kind]:
        print(line)
    if count > len(detail[kind]):
        print(f"  ... and {count - len(detail[kind])} more")
    print()
if dups:
    print(f"⚠️ duplicate keys (date|value|ccy|desc): {len(dups)}")
    for k, c in list(dups.items())[:10]:
        print(f"  {c}x {k}")
