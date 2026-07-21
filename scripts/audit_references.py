import sys, openpyxl
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('data/Expenses_Improved.xlsx', data_only=False)

# ── 1. Lists actual layout ────────────────────────────────────────────────────
ws_li = wb['Lists']
lists_layout = {}
for c in range(1, ws_li.max_column + 1):
    header = ws_li.cell(1, c).value
    letter = get_column_letter(c)
    data = [ws_li.cell(r, c).value for r in range(2, 20) if ws_li.cell(r, c).value is not None]
    lists_layout[letter] = (header, data[:6])

print("=== LISTS LAYOUT ===")
for letter, (header, data) in lists_layout.items():
    print(f"  {letter}: {header} -> {data}")

# ── 2. All data validations in every sheet ────────────────────────────────────
print("\n=== DATA VALIDATIONS ===")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    dvs = list(ws.data_validations.dataValidation)
    if dvs:
        print(f"\n{sheet}:")
        for dv in dvs:
            f = dv.formula1 or ''
            # Determine what column the formula references
            note = ''
            if 'Lists!$' in f:
                import re
                m = re.search(r'Lists!\$([A-Z]+)\$', f)
                if m:
                    col_letter = m.group(1)
                    header, _ = lists_layout.get(col_letter, ('??? MISSING', []))
                    note = f'  <- Lists col {col_letter} = "{header}"'
            print(f"  {dv.sqref}: {f}{note}")

# ── 3. All formulas referencing Lists in every sheet ──────────────────────────
print("\n=== FORMULAS REFERENCING LISTS ===")
import re
for sheet in wb.sheetnames:
    ws = wb[sheet]
    found = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and 'Lists!' in val:
                cols_ref = re.findall(r'Lists!\$([A-Z]+)\$', val)
                notes = []
                for cl in set(cols_ref):
                    header, _ = lists_layout.get(cl, ('??? MISSING', []))
                    notes.append(f'{cl}="{header}"')
                found.append(f"  {get_column_letter(c)}{r}: {val}  <- {', '.join(notes)}")
    if found:
        print(f"\n{sheet} ({len(found)} cells):")
        # deduplicate by formula content
        seen = set()
        for line in found:
            key = re.sub(r'\d+', 'N', line)
            if key not in seen:
                seen.add(key)
                print(line)
        if len(found) > len(seen):
            print(f"  ... ({len(found)} total cells, showing unique patterns)")
