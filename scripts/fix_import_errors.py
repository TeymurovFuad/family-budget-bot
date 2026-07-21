"""
fix_import_errors.py — one-time repair of the 2026-07-21 bulk import.

Fixes in MasterData:
  1. Categories not in Lists remapped (see category_remap below)
  2. Category 'Savings' used as category on transfer-to-self ->
     Expense / Groceries (money sent to own card for grocery spending)
  3. Person values that are transfer recipients (not household members) ->
     moved into Description, Person cleared.

Usage:  python scripts/fix_import_errors.py [path-to-xlsx]
Default path: data/Expenses_Improved.xlsx (or XLSX_PATH env var).
A .bak copy is written next to the file before any change.
"""
import os
import shutil
import sys

import openpyxl

# Scripts share the bot's configuration — .env is the single source of truth.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import settings
from _repair_guard import repair_guard
from file_storage import atomic_save

sys.stdout.reconfigure(encoding="utf-8")


def read_persons(wb):
    ws_li = wb["Lists"]
    col = next(c for c in range(1, ws_li.max_column + 1)
               if str(ws_li.cell(1, c).value or "").strip() == "Persons")
    out = set()
    for r in range(2, ws_li.max_row + 1):
        v = ws_li.cell(r, col).value
        if v is None:
            break
        out.add(str(v).strip())
    return out


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else str(settings.XLSX_PATH)

    with repair_guard():
        shutil.copy2(path, path + ".bak")
        print(f"Backup written: {path}.bak")

        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb["MasterData"]
        hdr = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}

        persons = read_persons(wb)
        fixed_shopping = fixed_savings = fixed_persons = 0

        for r in range(2, ws.max_row + 1):
            if ws.cell(r, hdr["Value"]).value is None and ws.cell(r, hdr["Date"]).value is None:
                continue
            cat  = str(ws.cell(r, hdr["Category"]).value or "")
            typ  = str(ws.cell(r, hdr["Type"]).value or "")
            per  = ws.cell(r, hdr["Person"]).value
            desc = str(ws.cell(r, hdr["Description"]).value or "")

            # NOTE: category 'Gifts & Shopping' was renamed to 'Shopping'
            # (scripts/rename_category.py) — the remap now points at the new name.
            category_remap = {
                "Gifts & Shopping": "Shopping",
                "Home Improvement": "Housing",
            }
            if cat in category_remap:
                ws.cell(r, hdr["Category"], category_remap[cat])
                fixed_shopping += 1

            if cat == "Savings" and "self" in desc.lower():
                ws.cell(r, hdr["Type"], "Expense")
                ws.cell(r, hdr["Category"], "Groceries")
                ws.cell(r, hdr["Description"], "Transfer to own card for groceries")
                fixed_savings += 1

            if per and str(per).strip() and str(per).strip() not in persons:
                recipient = str(per).strip()
                new_desc = f"{desc} — {recipient}" if desc else recipient
                ws.cell(r, hdr["Description"], new_desc[:120])
                ws.cell(r, hdr["Person"]).value = None
                fixed_persons += 1

        atomic_save(wb, path)
        print(f"Fixed: {fixed_shopping}x category remapped (Shopping/Home Improvement)")
        print(f"Fixed: {fixed_savings}x transfer-to-self -> Expense/Groceries")
        print(f"Fixed: {fixed_persons}x recipient moved from Person to Description")
        print("Saved.")


if __name__ == "__main__":
    main()
