import argparse
import io
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
import settings

if isinstance(sys.stdout, io.TextIOWrapper):
    sys.stdout.reconfigure(encoding="utf-8")

def main(path: Path) -> None:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws_db = wb['Dashboard']
    ws_md = wb['MasterData']

    # Print MasterData headers so we know what each column letter means
    print("=== MasterData column letters ===")
    for c in range(1, ws_md.max_column + 1):
        print(f"  {get_column_letter(c)}: {ws_md.cell(1, c).value}")

    # Print all Dashboard formulas (non-empty, non-trivial)
    print("\n=== Dashboard formulas ===")
    for r in range(1, ws_db.max_row + 1):
        for c in range(1, ws_db.max_column + 1):
            val = ws_db.cell(r, c).value
            if isinstance(val, str) and val.startswith('='):
                print(f"  {get_column_letter(c)}{r}: {val}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Audit Dashboard formulas.")
    parser.add_argument(
        "--path",
        type=Path,
        default=settings.XLSX_PATH,
        help="Path to workbook (default: XLSX_PATH env / settings)",
    )
    args = parser.parse_args()
    main(args.path)
