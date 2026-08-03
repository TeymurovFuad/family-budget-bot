"""
workbook_template.py — creating and repairing workbooks from the repo template.

Split out of file_storage.py (which remains the public facade). This module
owns everything about how a FRESH workbook comes into existence: copying the
template, repairing it, and the minimal fallback builder.

TEMPLATE_PATH is read lazily through file_storage so tests that monkeypatch
`file_storage.TEMPLATE_PATH` keep working unchanged.
"""

from pathlib import Path

import settings
from logger import get_logger
from excel_schema import ListsSchema, MasterDataSchema, find_col, header_of
from storage_backends import _replace_with_retry

log = get_logger(__name__)

# Last MasterData row covered by the category dropdown validation (row count).
_VALIDATION_LAST_ROW = 10000


def _template_path() -> Path:
    import file_storage
    return file_storage.TEMPLATE_PATH


def _repair_template_workbook(path: Path) -> None:
    """
    Minimal template repair for freshly copied workbooks.

    The template is canonical — headers and layout are already correct and
    resolved by excel_schema at runtime. Only ensure the Date Modified column
    exists, and clear any leftover data/placeholder rows.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path)
    changed = False

    if "MasterData" in wb.sheetnames:
        ws = wb["MasterData"]
        if find_col(ws, header_of(MasterDataSchema, "date_modified")) is None:
            ws.cell(1, ws.max_column + 1).value = header_of(MasterDataSchema, "date_modified")
            changed = True
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
            # delete_rows shrinks dropdown validation ranges along with the
            # rows — restore them so a fresh workbook has working dropdowns.
            from excel_schema import extend_validation_ranges
            extend_validation_ranges(ws, 1)
            changed = True

    if "Lists" in wb.sheetnames:
        ws = wb["Lists"]
        persons_col = find_col(ws, header_of(ListsSchema, "persons"))
        if persons_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, persons_col).value is not None:
                    ws.cell(row, persons_col).value = None
                    changed = True

    _CYCLES_SHEET = "Cycles"
    if _CYCLES_SHEET not in wb.sheetnames:
        _ws_cy = wb.create_sheet(_CYCLES_SHEET)
        _ws_cy.cell(1, 1, "StartDate")
        _ws_cy.cell(1, 2, "Label")
        changed = True

    from excel_schema import repair_dashboard_bounds
    n = repair_dashboard_bounds(wb)
    if n:
        log.info("_repair_template_workbook: updated %d formula cell(s) with open-ended bounds", n)
        changed = True

    if changed:
        wb.save(path)


def create_blank_excel(path: Path) -> None:
    """
    Copy the repo template to path.

    The template (data/Expenses_Template.xlsx) preserves the full sheet
    structure, formulas, styling, and data validations of the production
    workbook — but contains no personal data.  Using it as the base means
    the Dashboard SUMIFS, Monthly Summary layout, and all conditional
    formatting are available immediately without needing to be rebuilt
    from scratch in Python.

    Falls back to a minimal hand-built workbook if the template is missing
    (e.g. fresh clone before the template has been committed).
    """
    import shutil
    from datetime import datetime, timezone

    template_path = _template_path()
    log.info("Creating blank Excel workbook at %s (template: %s)", path, template_path)
    if template_path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, path)
        _repair_template_workbook(path)
        log.info("Created Excel workbook from template at %s", path)
        return

    # ── Fallback: minimal workbook (no formulas / styling) ────────────────────
    log.warning(
        "Template not found at %s — creating minimal fallback workbook. "
        "Run scripts/make_template.py to generate the template.",
        template_path,
    )
    from openpyxl import Workbook

    wb = Workbook()

    ws_md = wb.active
    ws_md.title = "MasterData"
    ws_md.append([
        "Date", "Year", "Month", "Value", "Type", "Category",
        "Person", "Description", "IsRecurring", "IsDone",
        "Currency", "Value (base)", "Date Modified (UTC)",
    ])

    ws_li = wb.create_sheet("Lists")
    _li_headers = [
        (1, header_of(ListsSchema, "months")),
        (2, header_of(ListsSchema, "txn_types")),
        (3, header_of(ListsSchema, "categories")),
        (4, header_of(ListsSchema, "budget_base")),
        (5, header_of(ListsSchema, "persons")),
        (6, header_of(ListsSchema, "years")),
        (8, header_of(ListsSchema, "currency")),
        (9, header_of(ListsSchema, "rate_to_base")),
    ]
    for _c, _h in _li_headers:
        ws_li.cell(1, _c, _h)

    months     = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    txn_types  = ["Expense","Income","Savings"]
    categories = [
        "Groceries", "Transport", "Housing", "Utilities", "Healthcare",
        "Entertainment", "Travel", "Insurance", "Education", "Salary",
        "Freelance", "Rental", "Bonus", "Bank Deposit", "Investment",
        "Emergency Fund", "Other",
    ]
    cur_year   = datetime.now(timezone.utc).year
    years      = [cur_year - 1, cur_year, cur_year + 1, cur_year + 2]
    currencies = [(settings.DISPLAY_CURRENCY, 1.0), ("EUR", 4.28), ("USD", 3.95), ("GBP", 5.05), ("CHF", 4.45)]

    for i, v in enumerate(months,     2): ws_li.cell(i, 1, v)
    for i, v in enumerate(txn_types,  2): ws_li.cell(i, 2, v)
    for i, v in enumerate(categories, 2): ws_li.cell(i, 3, v)
    # col 4 = Budget (base) — left blank; user fills in per-category limits
    for i, v in enumerate(years,      2): ws_li.cell(i, 6, v)
    for i, (code, rate) in enumerate(currencies, 2):
        ws_li.cell(i, 8, code)   # Currency
        ws_li.cell(i, 9, rate)   # Rate to base

    ws_db = wb.create_sheet("Dashboard")

    ws_ms = wb.create_sheet("Monthly Summary")
    for c_idx, hdr in enumerate(
        ["Year", "Month", "Income", "Expenses", "Savings", "Net", "Savings Rate %"], 1
    ):
        ws_ms.cell(1, c_idx, hdr)

    _ws_cy = wb.create_sheet("Cycles")
    _ws_cy.cell(1, 1, "StartDate")
    _ws_cy.cell(1, 2, "Label")

    try:
        from openpyxl.worksheet.datavalidation import DataValidation as _DV
        dv = _DV(type="list", formula1=f"Lists!$C$2:$C${1+len(categories)}", allow_blank=True)
        dv.sqref = f"F2:F{_VALIDATION_LAST_ROW}"
        ws_md.add_data_validation(dv)
    except Exception as _e:
        log.warning("Could not add category dropdown: %s", _e)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("Created fallback Excel workbook at %s", path)


def create_workbook_from_template(dest: Path) -> None:
    """
    Create a fresh workbook at `dest` atomically: the template is materialised
    to a sibling .tmp file first, then os.replace moves it into place, so a
    crash mid-creation never leaves a half-written workbook at `dest`.
    """
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_name(dest.stem + "._setup_tmp.xlsx")
    try:
        # Late-bound through the facade so tests/callers that monkeypatch
        # file_storage.create_blank_excel keep working.
        import file_storage
        file_storage.create_blank_excel(tmp)
        _replace_with_retry(tmp, dest)
    finally:
        tmp.unlink(missing_ok=True)
    log.info("Created workbook from template at %s", dest)


def lists_categories_populated(wb) -> bool:
    """True if the Lists sheet has at least one category in the Categories column."""
    if "Lists" not in wb.sheetnames:
        return False
    ws = wb["Lists"]
    cat_col = find_col(ws, header_of(ListsSchema, "categories"))
    if cat_col is None:
        return False
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, cat_col).value
        if val is not None and str(val).strip():
            return True
    return False
