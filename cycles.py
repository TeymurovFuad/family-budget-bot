"""
cycles.py — budget-cycle ledger (Cycles sheet) and cycle-scoped aggregation.

A cycle boundary is a RECORDED EVENT (salary confirmation or /cycle started),
never a date formula. Boundaries are written once and never recomputed.
Everything here is inert unless settings.BUDGET_CYCLE is on — callers gate on
the flag; these helpers just read/write the ledger.
"""

import asyncio
import re
from datetime import date, timedelta

import pandas as pd

import settings
from logger import get_logger
from excel_schema import CyclesSchema, ListsSchema, col_indices, header_of, to_date
from file_storage import (
    ExcelFileContext,
    _excel_write_lock,
    atomic_save,
    get_excel_path_for_reading,
)

log = get_logger(__name__)

CYCLES_SHEET_NAME = "Cycles"
LISTS_SHEET_NAME = "Lists"

# Implicit bucket for transactions older than the first recorded boundary.
# It has no salary anchor, so unaccounted math never applies to it.
BEFORE_CYCLES_LABEL = "Before cycles"


def cycle_label(start: date) -> str:
    """Ledger label for a cycle — always carries the year, e.g. 'Aug 2026'."""
    return start.strftime("%b %Y")


def _dedup_cycle_label(start: date, existing_dates) -> str:
    """
    Ledger label for a boundary, unique within its calendar month: the first
    boundary in a month keeps the plain label ('Jul 2026'); further ones get
    an index suffix ('Jul 2026 #2').
    """
    same_month = sum(
        1 for d in existing_dates
        if d != start and (d.year, d.month) == (start.year, start.month)
    )
    base = cycle_label(start)
    return base if same_month == 0 else f"{base} #{same_month + 1}"


def ensure_cycles_sheet(wb):
    """Return the Cycles worksheet, creating it with headers if missing."""
    if CYCLES_SHEET_NAME in wb.sheetnames:
        return wb[CYCLES_SHEET_NAME]
    ws = wb.create_sheet(CYCLES_SHEET_NAME)
    ws.cell(1, 1, header_of(CyclesSchema, "start_date"))
    ws.cell(1, 2, header_of(CyclesSchema, "label"))
    log.info("Created %s sheet in workbook", CYCLES_SHEET_NAME)
    return ws


def load_cycles() -> list[tuple[date, str]]:
    """
    Read the cycle ledger, sorted by start date ascending.
    Returns [] when the sheet is missing or unreadable — callers fall back to
    calendar behaviour.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(get_excel_path_for_reading(), data_only=True)
        if CYCLES_SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[CYCLES_SHEET_NAME]
        idx = col_indices(ws, CyclesSchema)
        start_col = idx.get("start_date")
        label_col = idx.get("label")
        if not start_col:
            return []
        cycles: list[tuple[date, str]] = []
        for row in range(2, ws.max_row + 1):
            start = to_date(ws.cell(row, start_col).value)
            if start is None:
                continue
            raw_label = ws.cell(row, label_col).value if label_col else None
            label = str(raw_label).strip() if raw_label else cycle_label(start)
            cycles.append((start, label))
        cycles.sort(key=lambda c: c[0])
        return cycles
    except Exception as e:
        log.warning("Could not load cycle ledger: %s", e)
        return []


def record_cycle_start(start: date) -> str | None:
    """
    Append one boundary row to the Cycles sheet.
    Returns the recorded cycle label, or None (no write) if that start date
    is already recorded — boundaries are written once, never recomputed.
    """
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = ensure_cycles_sheet(wb)
        from cycle_dashboard import ensure_cycle_dashboard
        ensure_cycle_dashboard(wb)
        idx = col_indices(ws, CyclesSchema)
        start_col = idx["start_date"]
        label_col = idx["label"]
        next_row = 2
        existing_dates: set[date] = set()
        for row in range(2, ws.max_row + 1):
            existing = to_date(ws.cell(row, start_col).value)
            if existing is None:
                continue
            if existing == start:
                return None
            existing_dates.add(existing)
            next_row = row + 1
        label = _dedup_cycle_label(start, existing_dates)
        ws.cell(next_row, start_col, start)
        ws.cell(next_row, label_col, label)
        atomic_save(wb, excel_path)
        log.info("Recorded cycle boundary %s (%s)", start, label)
        return label


async def async_record_cycle_start(start: date) -> str | None:
    """Async wrapper — returns the recorded label, or None if already recorded."""
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        return await loop.run_in_executor(None, record_cycle_start, start)


def remove_cycle_start(start: date) -> bool:
    """
    Delete one boundary row from the Cycles sheet.
    Returns False when that start date is not in the ledger.
    """
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        if CYCLES_SHEET_NAME not in wb.sheetnames:
            return False
        ws = wb[CYCLES_SHEET_NAME]
        idx = col_indices(ws, CyclesSchema)
        start_col = idx.get("start_date")
        if not start_col:
            return False
        for row in range(2, ws.max_row + 1):
            if to_date(ws.cell(row, start_col).value) == start:
                ws.delete_rows(row)
                atomic_save(wb, excel_path)
                log.info("Removed cycle boundary %s", start)
                return True
        return False


async def async_remove_cycle_start(start: date) -> bool:
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        return await loop.run_in_executor(None, remove_cycle_start, start)


def current_cycle_start(today: date, cycles: list[tuple[date, str]] | None = None) -> tuple[date, str] | None:
    """Latest recorded boundary on or before today, or None (→ calendar fallback)."""
    if cycles is None:
        cycles = load_cycles()
    past = [c for c in cycles if c[0] <= today]
    return past[-1] if past else None


def should_prompt_new_cycle(today: date) -> bool:
    """
    True when a Salary income should trigger the new-cycle prompt: either no
    cycle exists yet, or the current one is at least
    CYCLE_REPROMPT_MIN_AGE_DAYS old. Younger cycle → income inside the cycle,
    silently counted.
    """
    current = current_cycle_start(today)
    if current is None:
        return True
    return (today - current[0]).days >= settings.CYCLE_REPROMPT_MIN_AGE_DAYS


def load_salary_keywords(excel_path=None) -> list[str]:
    """
    Salary keywords stored in the Lists sheet "Salary Keywords" column —
    stripped, lowercased, deduplicated, blanks skipped. Returns [] when the
    column is missing/empty or the workbook is unreadable.
    """
    from openpyxl import load_workbook

    try:
        path = excel_path if excel_path is not None else get_excel_path_for_reading()
        wb = load_workbook(path, data_only=True)
        if LISTS_SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[LISTS_SHEET_NAME]
        kw_col = col_indices(ws, ListsSchema).get("salary_keyword")
        if not kw_col:
            return []
        return _keyword_column_words(ws, kw_col)
    except Exception as e:
        log.warning("Could not load salary keywords from Lists sheet: %s", e)
        return []


def _keyword_column_words(ws, kw_col) -> list[str]:
    words: list[str] = []
    for row in range(2, ws.max_row + 1):
        w = str(ws.cell(row, kw_col).value or "").strip().lower()
        if w and w not in words:
            words.append(w)
    return words


def _rewrite_keyword_column(ws, kw_col, words: list[str]) -> None:
    for i, w in enumerate(words, start=2):
        ws.cell(i, kw_col, w)
    for row in range(2 + len(words), ws.max_row + 1):
        ws.cell(row, kw_col).value = None


MAX_SALARY_KEYWORD_BYTES = 57  # Telegram callback_data limit is 64 bytes; "kw:del:" takes 7


def save_salary_keyword(keyword: str) -> bool:
    """
    Append one keyword to the Lists sheet "Salary Keywords" column. The very
    first write ever (column header not yet present) creates the column and
    seeds it with the current .env keywords so they are not silently dropped;
    a list the user has emptied via /keywords stays empty.
    Returns False (no write) when the keyword is already stored or invalid.
    """
    from openpyxl import load_workbook

    keyword = str(keyword or "").strip().lower()
    if not keyword or len(keyword.encode("utf-8")) > MAX_SALARY_KEYWORD_BYTES:
        return False
    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb[LISTS_SHEET_NAME]
        kw_col = col_indices(ws, ListsSchema).get("salary_keyword")
        first_write = kw_col is None
        if first_write:
            kw_col = ws.max_column + 1
            ws.cell(1, kw_col, header_of(ListsSchema, "salary_keyword"))
        words = _keyword_column_words(ws, kw_col)
        seeded = False
        if first_write:
            for w in settings.CYCLE_DETECT_KEYWORDS:
                w = str(w or "").strip().lower()
                if w and w not in words:
                    words.append(w)
            seeded = bool(words)
        added = keyword not in words
        if added:
            words.append(keyword)
        if not added and not seeded:
            return False
        _rewrite_keyword_column(ws, kw_col, words)
        atomic_save(wb, excel_path)
        log.info("Saved salary keyword %r (%d stored)", keyword, len(words))
        return added


async def async_save_salary_keyword(keyword: str) -> bool:
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        return await loop.run_in_executor(None, save_salary_keyword, keyword)


def delete_salary_keyword(keyword: str) -> bool:
    """
    Remove one keyword from the Lists sheet "Salary Keywords" column.
    Only that column's cells are shifted — Lists rows are shared with other
    reference columns, so whole-row deletion is never used here.
    Returns False when the keyword is not stored.
    """
    from openpyxl import load_workbook

    keyword = str(keyword or "").strip().lower()
    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        if LISTS_SHEET_NAME not in wb.sheetnames:
            return False
        ws = wb[LISTS_SHEET_NAME]
        kw_col = col_indices(ws, ListsSchema).get("salary_keyword")
        if not kw_col:
            return False
        words = _keyword_column_words(ws, kw_col)
        if keyword not in words:
            return False
        words = [w for w in words if w != keyword]
        _rewrite_keyword_column(ws, kw_col, words)
        atomic_save(wb, excel_path)
        log.info("Deleted salary keyword %r (%d remain)", keyword, len(words))
        return True


async def async_delete_salary_keyword(keyword: str) -> bool:
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        return await loop.run_in_executor(None, delete_salary_keyword, keyword)


def cycle_detect_keywords(extra: list[str] | None = None) -> list[str]:
    """SALARY_CATEGORY plus the stored salary keywords (Excel, falling back to
    .env CYCLE_DETECT_KEYWORDS) plus any ad-hoc extras, lowercased,
    deduplicated, blanks dropped."""
    stored = load_salary_keywords()
    if not stored:
        log.debug("No salary keywords in Excel — falling back to .env CYCLE_DETECT_KEYWORDS")
        stored = settings.CYCLE_DETECT_KEYWORDS
    words = [settings.SALARY_CATEGORY, *stored, *(extra or [])]
    seen: list[str] = []
    for w in words:
        w = str(w or "").strip().lower()
        if w and w not in seen:
            seen.append(w)
    return seen


def salary_mask(df: pd.DataFrame, extra_keywords: list[str] | None = None) -> pd.Series:
    """
    Boolean mask for salary rows: Income type AND a salary keyword in Category
    (word-boundary contains), or in Description when Category is blank.
    Description matters because bulk-imported salary rows carry the bank's
    transfer title (e.g. 'WYNAGRODZENIE ZA LIPIEC') with an empty category;
    a categorised row ('Freelance' + 'Salary' description) is not a salary.
    """
    keywords = cycle_detect_keywords(extra_keywords)
    if not keywords:
        return pd.Series(False, index=df.index)
    pattern = r"\b(?:" + "|".join(re.escape(k) for k in keywords) + r")\b"
    category = df["Category"].fillna("").astype(str).str.strip()
    matches = category.str.contains(pattern, case=False, regex=True)
    if "Description" in df.columns:
        matches |= (category == "") & df["Description"].astype(str).str.contains(
            pattern, case=False, regex=True
        )
    return (df["Type"] == "Income") & matches


def detect_cycle_candidates(
    df: pd.DataFrame,
    existing_cycles: list[tuple[date, str]] | None = None,
    extra_keywords: list[str] | None = None,
) -> list[dict]:
    """
    Scan transaction history for salary arrivals not yet recorded as cycle
    boundaries. Returns one entry per unique salary date, oldest first.

    Each entry: {date, amounts, unambiguous}
      - date: the transaction date (the cycle boundary if confirmed)
      - amounts: list of salary amounts on that date (usually one)
      - unambiguous: True when exactly one salary row on that date
    """
    if existing_cycles is None:
        existing_cycles = load_cycles()
    existing_starts = {c[0] for c in existing_cycles}

    df = df.copy()
    df["_date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date

    salary_rows = df[
        df["_date"].notna()
        & df["IsDone"].astype(bool)
        & salary_mask(df, extra_keywords)
        & (df["_date"] <= date.today())
    ].copy()

    if salary_rows.empty:
        return []

    results: list[dict] = []
    for salary_date, group in salary_rows.groupby("_date"):
        if salary_date in existing_starts:
            continue
        amounts = sorted(
            [round(float(r["_base"]), 2) for _, r in group.iterrows()],
            reverse=True,
        )
        results.append(
            {
                "date": salary_date,
                "amounts": amounts,
                "unambiguous": len(amounts) == 1,
            }
        )

    return results


def record_cycle_starts_batch(starts: list[date]) -> int:
    """
    Open the workbook ONCE and write all boundary rows. Returns the count
    actually written (skips dates that are already present).
    """
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = ensure_cycles_sheet(wb)
        from cycle_dashboard import ensure_cycle_dashboard
        ensure_cycle_dashboard(wb)
        idx = col_indices(ws, CyclesSchema)
        start_col = idx["start_date"]
        label_col = idx["label"]

        existing: set[date] = set()
        next_row = 2
        for row in range(2, ws.max_row + 1):
            existing_date = to_date(ws.cell(row, start_col).value)
            if existing_date is None:
                continue
            existing.add(existing_date)
            next_row = row + 1

        count = 0
        for start in starts:
            if start in existing:
                continue
            label = _dedup_cycle_label(start, existing)
            ws.cell(next_row, start_col, start)
            ws.cell(next_row, label_col, label)
            existing.add(start)
            next_row += 1
            count += 1
            log.info("Batch-recorded cycle boundary %s (%s)", start, label)

        if count:
            atomic_save(wb, excel_path)
        return count


def cycle_totals(df: pd.DataFrame, start: date, end: date) -> dict:
    """
    Aggregate MasterData over [start, end] (inclusive; end is today for the
    open-ended current cycle). All sums use the _base column.

    unaccounted = salary received − tracked expenses − tracked savings;
    negative means over-reported.
    """
    dates = pd.to_datetime(df["Date"], errors="coerce")
    sub = df[
        dates.notna()
        & (dates.dt.date >= start)
        & (dates.dt.date <= end)
        & df["IsDone"]
    ]
    income  = sub[sub["Type"] == "Income"]["_base"].sum()
    expense = sub[sub["Type"] == "Expense"]["_base"].sum()
    savings = sub[sub["Type"] == "Savings"]["_base"].sum()
    salary = sub[salary_mask(sub)]["_base"].sum()
    return {
        "sub": sub,
        "income": income,
        "expense": expense,
        "savings": savings,
        "salary": salary,
        "unaccounted": salary - expense - savings,
    }


def cycle_periods(
    df: pd.DataFrame,
    cycles: list[tuple[date, str]] | None = None,
    today: date | None = None,
) -> list[tuple[date, date, str]]:
    """
    Every cycle as (start, end, label), oldest first. Each cycle ends the day
    before the next begins; the newest ends today. When transactions exist
    that are older than the first recorded boundary, an implicit
    "Before cycles" bucket is prepended covering [earliest txn, first-1].
    Returns [] when the ledger is empty.
    """
    if cycles is None:
        cycles = load_cycles()
    if today is None:
        today = date.today()
    if not cycles:
        return []
    periods: list[tuple[date, date, str]] = []
    first_start = cycles[0][0]
    dates = pd.to_datetime(df["Date"], errors="coerce").dt.date.dropna()
    older = dates[dates < first_start]
    if not older.empty:
        periods.append((older.min(), first_start - timedelta(days=1), BEFORE_CYCLES_LABEL))
    for i, (start, label) in enumerate(cycles):
        end = cycles[i + 1][0] - timedelta(days=1) if i + 1 < len(cycles) else today
        periods.append((start, end, label))
    return periods


def detect_missing_boundaries(
    start: date,
    end: date,
    cycles: list[tuple[date, str]] | None = None,
) -> list[date]:
    """
    First-of-month markers for calendar months inside [start, end] that have
    no recorded cycle boundary. Used by reports to offer a lazy backfill
    before rendering a period with gaps.
    """
    if cycles is None:
        cycles = load_cycles()
    if not cycles:
        # No ledger at all → no boundaries to compare against, hence no gaps
        # to backfill (an empty ledger means cycles are effectively unused).
        return []
    covered = {(c[0].year, c[0].month) for c in cycles}
    missing: list[date] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        if (y, m) not in covered:
            missing.append(date(y, m, 1))
        m += 1
        if m > 12:
            y, m = y + 1, 1
    return missing


def fallback_income_candidates(
    df: pd.DataFrame,
    anchor: date,
    existing_cycles: list[tuple[date, str]] | None = None,
    window_days: int = 20,
    limit: int = 3,
) -> list[dict]:
    """
    When keyword detection finds nothing: the largest `limit` Income rows in
    the ±window_days window around `anchor` (usually the 1st of the target
    month), largest first. Catches salaries filed under non-salary categories.
    Each entry: {date, amounts, unambiguous} — same shape as
    detect_cycle_candidates() entries.
    """
    if existing_cycles is None:
        existing_cycles = load_cycles()
    existing_starts = {c[0] for c in existing_cycles}

    df = df.copy()
    if df.empty or "Date" not in df.columns or "_base" not in df.columns:
        return []
    df["_date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    lo, hi = anchor - timedelta(days=window_days), anchor + timedelta(days=window_days)
    rows = df[
        df["_date"].notna()
        & df["IsDone"].astype(bool)
        & (df["Type"] == "Income")
        & (df["_date"] >= lo)
        & (df["_date"] <= hi)
        & (df["_date"] <= date.today())
    ]
    rows = rows[~rows["_date"].isin(existing_starts)]
    if rows.empty:
        return []
    rows = rows.sort_values("_base", ascending=False).head(limit)
    return [
        {
            "date": r["_date"],
            "amounts": [round(float(r["_base"]), 2)],
            "unambiguous": True,
        }
        for _, r in rows.iterrows()
    ]
