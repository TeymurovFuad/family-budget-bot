"""
cycle_dashboard.py — Cycle Dashboard sheet: cycle-scoped mirror of Dashboard.

The sheet lets the user pick a recorded cycle (B2 dropdown fed by the Cycles
ledger) and see the same summary block and per-category budget table as the
Dashboard, but bounded by [cycle start, next cycle start) instead of a
calendar month. All formulas live in Excel — the bot only writes structure.

No function here saves the workbook — callers hold the write lock and run
atomic_save themselves.
"""

import settings
from logger import get_logger
from excel_schema import (
    CATEGORY_FIRST_ROW, CyclesSchema, clear_category_block, col_indices,
    read_category_block, to_date, write_category_sumif_block,
)

log = get_logger(__name__)

CYCLE_DASHBOARD_SHEET_NAME = "Cycle Dashboard"
DASHBOARD_SHEET_NAME = "Dashboard"

_MAX_ROW = 1048576

# Cycle boundary helpers (hidden column N):
#   N3 = selected cycle start date, N4 = next cycle start (exclusive end)
_RATE_FORMULA = f"=IFERROR(VLOOKUP($F$2,Lists!$H$2:$I${_MAX_ROW},2,0),1)"
_N3_FORMULA = (
    f'=IFERROR(INDEX(Cycles!$A$2:$A${_MAX_ROW},'
    f'MATCH($B$2,Cycles!$B$2:$B${_MAX_ROW},0)),"")'
)
_N4_FORMULA = (
    f'=IF($N$3="","",IF(MINIFS(Cycles!$A$2:$A${_MAX_ROW},'
    f'Cycles!$A$2:$A${_MAX_ROW},">"&$N$3)=0,TODAY()+1,'
    f'MINIFS(Cycles!$A$2:$A${_MAX_ROW},'
    f'Cycles!$A$2:$A${_MAX_ROW},">"&$N$3)))'
)

_DATE_FORMAT = "yyyy-mm-dd"

_CATEGORY_HEADERS = ["Category", "Budget", "Actual", "Variance", "Var %"]
_CATEGORY_FIRST_ROW = CATEGORY_FIRST_ROW  # H11 — same as Dashboard


def _sumifs_type(txn_type: str) -> str:
    """Summary-block SUMIFS: one Type, IsDone, dates in [N3, N4)."""
    return (
        f'=IF($N$3="",0,SUMIFS(MasterData!$L$2:$L${_MAX_ROW},'
        f'MasterData!$E$2:$E${_MAX_ROW},"{txn_type}",'
        f'MasterData!$A$2:$A${_MAX_ROW},">="&$N$3,'
        f'MasterData!$A$2:$A${_MAX_ROW},"<"&$N$4,'
        f'MasterData!$J$2:$J${_MAX_ROW},TRUE)/$N$2)'
    )


def _sumifs_salary(salary_category: str) -> str:
    """Salary SUMIFS: Income type AND Category = salary_category."""
    return (
        f'=IF($N$3="",0,SUMIFS(MasterData!$L$2:$L${_MAX_ROW},'
        f'MasterData!$E$2:$E${_MAX_ROW},"Income",'
        f'MasterData!$F$2:$F${_MAX_ROW},"{salary_category}",'
        f'MasterData!$A$2:$A${_MAX_ROW},">="&$N$3,'
        f'MasterData!$A$2:$A${_MAX_ROW},"<"&$N$4,'
        f'MasterData!$J$2:$J${_MAX_ROW},TRUE)/$N$2)'
    )


def _cycle_actual_formula(r: int) -> str:
    """Cycle-bounded 'Actual' SUMIFS for the category in H{r}."""
    return (
        f'=IF($N$3="",0,SUMIFS(MasterData!$L$2:$L${_MAX_ROW},'
        f'MasterData!$F$2:$F${_MAX_ROW},H{r},'
        f'MasterData!$A$2:$A${_MAX_ROW},">="&$N$3,'
        f'MasterData!$A$2:$A${_MAX_ROW},"<"&$N$4,'
        f'MasterData!$J$2:$J${_MAX_ROW},TRUE)/$N$2)'
    )


def _read_category_list(ws) -> list[str] | None:
    """Categories from H11 down to (exclusive) the first "TOTAL" row."""
    return read_category_block(ws)


def _dashboard_categories(wb) -> list[str]:
    if DASHBOARD_SHEET_NAME not in wb.sheetnames:
        return []
    return _read_category_list(wb[DASHBOARD_SHEET_NAME]) or []


def _latest_cycle_label(wb) -> str:
    """Latest cycle label read straight from the open workbook's Cycles sheet."""
    from cycles import CYCLES_SHEET_NAME, cycle_label

    if CYCLES_SHEET_NAME not in wb.sheetnames:
        return ""
    ws = wb[CYCLES_SHEET_NAME]
    idx = col_indices(ws, CyclesSchema)
    start_col = idx.get("start_date")
    label_col = idx.get("label")
    if not start_col:
        return ""
    latest = None
    for row in range(2, ws.max_row + 1):
        start = to_date(ws.cell(row, start_col).value)
        if start is None:
            continue
        raw_label = ws.cell(row, label_col).value if label_col else None
        label = str(raw_label).strip() if raw_label else cycle_label(start)
        if latest is None or start > latest[0]:
            latest = (start, label)
    return latest[1] if latest else ""


def _write_category_block(ws, categories: list[str]) -> None:
    """Write H11..TOTAL rows (headers in row 10 are written by ensure)."""
    write_category_sumif_block(ws, categories, _cycle_actual_formula)


def ensure_cycle_dashboard(wb, salary_category=None):
    """
    Return the Cycle Dashboard worksheet, creating it per spec if missing.
    Idempotent: an existing sheet is returned unchanged. No save here —
    the caller holds the write lock.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    if CYCLE_DASHBOARD_SHEET_NAME in wb.sheetnames:
        return wb[CYCLE_DASHBOARD_SHEET_NAME]

    if salary_category is None:
        salary_category = settings.SALARY_CATEGORY

    ws = wb.create_sheet(CYCLE_DASHBOARD_SHEET_NAME)

    # Filter area
    ws.cell(1, 1, "⚙ Cycle Filter")
    ws.cell(2, 1, "Cycle")
    ws.cell(2, 2, _latest_cycle_label(wb) or "")
    ws.cell(2, 3, "Start")
    ws.cell(2, 4, "=$N$3")
    ws.cell(2, 5, "Display")
    ws.cell(2, 6, "PLN")
    ws.cell(2, 14, _RATE_FORMULA)   # N2
    ws.cell(3, 14, _N3_FORMULA)     # N3
    ws.cell(4, 14, _N4_FORMULA)     # N4

    # Formats
    ws.cell(2, 4).number_format = _DATE_FORMAT
    ws.cell(3, 14).number_format = _DATE_FORMAT
    ws.cell(4, 14).number_format = _DATE_FORMAT

    # Summary block
    ws.cell(9, 1, "📅 Selected Cycle")
    ws.cell(10, 1, "Salary")
    ws.cell(10, 2, _sumifs_salary(salary_category))
    ws.cell(11, 1, "Income")
    ws.cell(11, 2, _sumifs_type("Income"))
    ws.cell(12, 1, "Expenses")
    ws.cell(12, 2, _sumifs_type("Expense"))
    ws.cell(13, 1, "Savings")
    ws.cell(13, 2, _sumifs_type("Savings"))
    ws.cell(14, 1, "Unaccounted")
    ws.cell(14, 2, "=B10-B12-B13")
    ws.cell(15, 1, "Cycle Days")
    ws.cell(15, 2, '=IF($N$3="","",$N$4-$N$3)')
    ws.cell(15, 2).number_format = "0"

    # Category table
    ws.cell(9, 8, "🏷 Expenses by Category")
    for i, hdr in enumerate(_CATEGORY_HEADERS):
        ws.cell(10, 8 + i, hdr)
    _write_category_block(ws, _dashboard_categories(wb))

    # B2 cycle dropdown fed by the Cycles ledger
    dv = DataValidation(
        type="list",
        formula1=f"=Cycles!$B$2:$B${_MAX_ROW}",
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("B2")

    log.info("Created %s sheet in workbook", CYCLE_DASHBOARD_SHEET_NAME)
    return ws


def sync_cycle_dashboard_categories(wb) -> int:
    """
    Make the Cycle Dashboard category block match the Dashboard's exactly.
    Returns the number of category rows written (0 = already in sync).
    """
    if CYCLE_DASHBOARD_SHEET_NAME not in wb.sheetnames:
        ws = ensure_cycle_dashboard(wb)
        return len(_read_category_list(ws) or [])

    ws = wb[CYCLE_DASHBOARD_SHEET_NAME]
    dashboard_cats = _dashboard_categories(wb)
    current_cats = _read_category_list(ws)
    if current_cats == dashboard_cats:
        return 0

    # Full rebuild: delete the existing H11..TOTAL block rows' contents, rewrite.
    clear_category_block(ws)
    _write_category_block(ws, dashboard_cats)
    log.info(
        "Synced %s categories: %d row(s) rewritten",
        CYCLE_DASHBOARD_SHEET_NAME, len(dashboard_cats),
    )
    return len(dashboard_cats)
