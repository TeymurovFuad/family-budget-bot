"""
excel_ops.py — async write operations on the Excel file.
All writes go through asyncio.get_running_loop().run_in_executor so they
don't block the Telegram event loop.
"""

import asyncio

from config import log
from log_decorators import log_call
from excel_schema import (
    find_next_data_row,
    lists_currency_range,
    write_transaction_row,
    ensure_monthly_summary_rows_from_masterdata,
)
from audit import audit_span
from storage_backends import ConcurrentModificationError
from file_storage import (
    ExcelFileContext,
    append_transactions_batch,
    atomic_save,
    delete_transaction_row,
    update_currency_rates_in_excel,
    append_to_recovery_queue,
    flush_recovery_queue,
    delete_recovery_queue_file,
    requeue_rows,
    _excel_write_lock,
)
from models import Transaction


def _invalidate_reference_cache() -> None:
    from data import invalidate_reference_cache
    invalidate_reference_cache()


@log_call()
def _do_append_transaction(transaction: Transaction) -> None:
    from openpyxl import load_workbook

    row = transaction.to_row()
    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["MasterData"]
        r = find_next_data_row(ws)
        write_transaction_row(ws, r, row, lists_currency_range(wb))
        ensure_monthly_summary_rows_from_masterdata(wb)
        atomic_save(wb, excel_path)
        log.info("Appended transaction row %d: %s", r, row)


@log_call()
async def append_transaction(transaction: Transaction, user=None) -> None:
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        with audit_span("append", rows=1, user=user):
            try:
                await loop.run_in_executor(None, _do_append_transaction, transaction)
            except ConcurrentModificationError:
                # Conflict: another write won the race — do NOT requeue. The
                # workbook may already contain the competing version of this
                # data, and a replay would blindly re-append a duplicate row.
                # Surface the failure so the user can retry intentionally.
                log.error("Concurrent modification — transaction NOT queued for replay")
                raise
            except Exception as e:
                log.error("Upload failed — saving to recovery queue: %s", e)
                append_to_recovery_queue(transaction.to_row())
                raise
        _invalidate_reference_cache()


async def async_delete_transaction_row(row_idx: int, expected: dict | None = None, user=None) -> None:
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        with audit_span("delete", rows=1, user=user):
            await loop.run_in_executor(None, delete_transaction_row, row_idx, expected)
        _invalidate_reference_cache()


async def async_update_currency_rates(new_rates: dict, user=None) -> None:
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        with audit_span("update_rates", rows=len(new_rates), user=user):
            await loop.run_in_executor(None, update_currency_rates_in_excel, new_rates)
        _invalidate_reference_cache()


async def async_append_batch(transactions: list, user=None) -> None:
    loop = asyncio.get_running_loop()
    async with _excel_write_lock:
        with audit_span("append_batch", rows=len(transactions), user=user):
            await loop.run_in_executor(None, append_transactions_batch, transactions)
        _invalidate_reference_cache()


def replay_recovery_queue() -> None:
    """
    Re-apply transactions persisted after a failed write. One open/save cycle
    for the whole batch; rows that fail are re-queued instead of dropped.
    """
    from openpyxl import load_workbook

    pending = flush_recovery_queue()
    if not pending:
        return
    log.warning("Re-applying %d transactions from recovery queue", len(pending))

    # Queue rows were JSON-roundtripped (json.dumps default=str): dates became
    # strings, numbers may be strings. Rehydrate so replayed rows are typed
    # identically to normally appended ones.
    from datetime import date as _date
    for row in pending:
        if isinstance(row.get("date"), str):
            try:
                row["date"] = _date.fromisoformat(row["date"][:10])
            except ValueError:
                pass
        for numeric_field in ("value", "year"):
            if isinstance(row.get(numeric_field), str):
                try:
                    row[numeric_field] = float(row[numeric_field]) if numeric_field == "value" else int(row[numeric_field])
                except ValueError:
                    pass
        if isinstance(row.get("is_recurring"), str):
            row["is_recurring"] = row["is_recurring"].lower() in {"true", "1", "yes"}
    failed: list[dict] = []
    outcome = "ok"
    import time as _time
    _start = _time.perf_counter()
    try:
        with ExcelFileContext() as excel_path:
            wb = load_workbook(excel_path)
            ws = wb["MasterData"]
            lu_range = lists_currency_range(wb)
            r = find_next_data_row(ws)
            for row in pending:
                try:
                    write_transaction_row(ws, r, row, lu_range)
                    log.info("Recovery queue: re-applied row %d: %s", r, row)
                    r += 1
                except Exception as e:
                    log.error("Recovery queue: failed to re-apply row %s: %s", row, e)
                    failed.append(row)
            added = ensure_monthly_summary_rows_from_masterdata(wb)
            if added:
                log.info("Recovery queue: ensured %d Monthly Summary row(s)", added)
            atomic_save(wb, excel_path)
    except Exception as e:
        # Requeueing here is safe even on ConcurrentModificationError: the
        # replay's upload failed as a whole, so none of these rows were
        # persisted — retrying later cannot produce duplicates.
        log.error("Recovery queue: replay aborted, re-queueing all rows: %s", e)
        failed = pending
        outcome = "error"

    # Replay attempt is fully finished (success or failure) — safe to drop
    # the old journal now and persist only the rows that still failed, in a
    # single atomic write.
    delete_recovery_queue_file()
    requeue_rows(failed)

    from audit import audit_save
    audit_save(source="replay_recovery_queue", rows=len(pending) - len(failed),
               outcome=outcome if not failed else f"{outcome}_requeued_{len(failed)}",
               duration_ms=(_time.perf_counter() - _start) * 1000)
    if len(pending) - len(failed) > 0:
        _invalidate_reference_cache()
