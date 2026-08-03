"""
file_storage.py
===================
Public facade for all access to the Excel file.

The storage layer is split into focused modules; this module re-exports
their public API so existing imports (`from file_storage import ...`) and
test monkeypatching of `file_storage.<CONFIG>` keep working unchanged:

  storage_backends.py   backend selection (local/GCS/S3), atomic saves,
                        download/upload, lost-update protection,
                        ExcelFileContext, get_excel_path_for_reading
  workbook_template.py  creating/repairing workbooks from the repo template
  recovery_queue.py     append-only JSONL journal for failed writes

What remains here is the workbook "repository": reference lists, budgets,
currency rates, user prefs, and transaction CRUD on MasterData.

Supports three storage backends, selected by environment variable:

  STORAGE_BACKEND=local  →  read/write from local disk  (default if nothing set)
  STORAGE_BACKEND=gcs    →  Google Cloud Storage
  STORAGE_BACKEND=s3     →  Any S3-compatible storage:
                             AWS S3, Oracle Object Storage, Cloudflare R2,
                             Backblaze B2, MinIO, DigitalOcean Spaces

An explicitly set STORAGE_BACKEND always wins; bucket-name variables only
select a backend when STORAGE_BACKEND itself was not set.

─────────────────────────────────────────────────────────────────────────────
Environment variables by backend
─────────────────────────────────────────────────────────────────────────────

Local (default — development, phone hosting, Oracle VM with local disk):
  XLSX_PATH          Path to the Excel file. Default: data/Expenses_Improved.xlsx

GCS (Google Cloud Storage — free tier, recommended for reports + phone setup):
  STORAGE_BACKEND    gcs
  GCS_BUCKET_NAME    Bucket name, e.g. your-bucket-name
  GCS_OBJECT_NAME    Object name inside bucket. Default: Expenses_Improved.xlsx
  GCS_KEY_JSON       Full contents of service account JSON key (for GitHub Actions
                     or any host where you can't write a file to disk)
                     Leave empty if GOOGLE_APPLICATION_CREDENTIALS is set instead.

S3-compatible (Oracle Object Storage, Cloudflare R2, Backblaze B2, AWS S3):
  STORAGE_BACKEND    s3
  S3_BUCKET_NAME     Bucket name
  S3_OBJECT_NAME     Object name. Default: Expenses_Improved.xlsx
  S3_ENDPOINT_URL    Full endpoint URL. Examples:
                       Oracle:     https://<namespace>.compat.objectstorage.<region>.oraclecloud.com
                       Cloudflare: https://<account-id>.r2.cloudflarestorage.com
                       Backblaze:  https://s3.<region>.backblazeb2.com
                       AWS:        leave empty (boto3 uses default)
  S3_ACCESS_KEY      Access key ID (AWS_ACCESS_KEY_ID equivalent)
  S3_SECRET_KEY      Secret access key (AWS_SECRET_ACCESS_KEY equivalent)
  S3_REGION          Region name. Default: us-east-1
                     Oracle uses: us-ashburn-1, eu-frankfurt-1, etc.
"""

import asyncio
import json
from pathlib import Path

import settings
from logger import get_logger

log = get_logger(__name__)

from excel_schema import CyclesSchema, ListsSchema, MasterDataSchema, find_col, col_indices, header_of

# ── Configuration (monkeypatch-friendly module globals) ───────────────────────
# Submodules read these lazily through this module, so tests may monkeypatch
# e.g. file_storage.LOCAL_XLSX_PATH or file_storage.RECOVERY_QUEUE_PATH.

STORAGE_BACKEND = settings.STORAGE_BACKEND
LOCAL_XLSX_PATH = settings.XLSX_PATH
USER_PREFS_PATH = settings.USER_PREFS_PATH

GCS_BUCKET_NAME = settings.GCS_BUCKET_NAME
GCS_OBJECT_NAME = settings.GCS_OBJECT_NAME
GCS_KEY_JSON    = settings.GCS_KEY_JSON

S3_BUCKET_NAME  = settings.S3_BUCKET_NAME
S3_OBJECT_NAME  = settings.S3_OBJECT_NAME
S3_ENDPOINT_URL = settings.S3_ENDPOINT_URL
S3_ACCESS_KEY   = settings.S3_ACCESS_KEY
S3_SECRET_KEY   = settings.S3_SECRET_KEY
S3_REGION       = settings.S3_REGION

RECOVERY_QUEUE_PATH = settings.RECOVERY_QUEUE_PATH
TEMPLATE_PATH = settings.DEFAULT_TEMPLATE_PATH

# ── Re-exports from the split storage modules ─────────────────────────────────

from storage_backends import (  # noqa: E402,F401
    ConcurrentModificationError,
    ExcelFileContext,
    _active_backend,
    _download_to_temp_file,
    _gcs_client,
    _replace_with_retry,
    _s3_client,
    _temp_files,
    _upload_from_local_file,
    atomic_save,
    cleanup_temp_files,
    get_excel_path_for_reading,
)
from workbook_template import (  # noqa: E402,F401
    _VALIDATION_LAST_ROW,
    _repair_template_workbook,
    create_blank_excel,
    create_workbook_from_template,
    lists_categories_populated,
)
from recovery_queue import (  # noqa: E402,F401
    append_to_recovery_queue,
    delete_recovery_queue_file,
    flush_recovery_queue,
    requeue_rows,
)

_excel_write_lock = asyncio.Lock()
# Public alias — external modules (scheduled.py etc.) should use this name
# rather than reaching for the underscore-private one.
excel_write_lock = _excel_write_lock


class RowMovedError(Exception):
    """
    Raised when a row picked in /delete or /edit no longer matches what the
    user saw at pick time — another write shifted rows in between, so the
    row index is stale. Callers should abort instead of silently mutating
    the wrong transaction.
    """


def _row_matches_snapshot(ws, headers: dict, row_idx: int, expected: dict) -> bool:
    """
    Best-effort check that MasterData row `row_idx` still holds the same
    date/value/description the user saw when they picked it. Used to guard
    against the row having shifted due to a concurrent delete/edit.
    """
    if row_idx < 2 or row_idx > ws.max_row:
        return False

    for col_name in ("Date", "Value", "Description"):
        if col_name not in expected:
            continue
        col_idx = headers.get(col_name)
        if col_idx is None:
            continue
        current = ws.cell(row_idx, col_idx).value
        target = expected[col_name]

        if col_name == "Date":
            current_cmp = current.date() if hasattr(current, "date") else current
            target_cmp = target.date() if hasattr(target, "date") else target
        elif col_name == "Value":
            try:
                current_cmp = round(float(current), 2)
                target_cmp = round(float(target), 2)
            except (TypeError, ValueError):
                current_cmp, target_cmp = current, target
        else:
            current_cmp = str(current or "")
            target_cmp = str(target or "")

        if current_cmp != target_cmp:
            return False
    return True


def _invalidate_reference_cache() -> None:
    """Drop data.py's reference-data TTL cache after a write to Lists."""
    try:
        from data import invalidate_reference_cache
        invalidate_reference_cache()
    except Exception:
        log.debug("Could not invalidate reference cache", exc_info=True)


# ── User preferences ──────────────────────────────────────────────────────────

def load_user_prefs() -> dict:
    """Load persisted per-user settings (display currency etc.) from JSON."""
    try:
        if USER_PREFS_PATH.exists():
            return json.loads(USER_PREFS_PATH.read_text(encoding="utf-8"))
    except Exception as e:
        log.warning("Could not load user prefs: %s", e)
    return {}


def save_user_prefs(prefs: dict) -> None:
    """Persist per-user settings to JSON file alongside the Excel file."""
    try:
        USER_PREFS_PATH.parent.mkdir(parents=True, exist_ok=True)
        USER_PREFS_PATH.write_text(json.dumps(prefs, indent=2), encoding="utf-8")
    except Exception as e:
        log.warning("Could not save user prefs: %s", e)


# ── Budget loading ────────────────────────────────────────────────────────────

def load_budgets_from_excel(excel_path: Path) -> dict[str, float]:
    """
    Read monthly budget limits from the Lists sheet Budget column.

    Only categories that have a non-zero budget value are returned.
    Categories with no limit set are simply absent from the result dict.

    Returns {category_name: base_amount}.
    Falls back to an empty dict if the sheet cannot be read.
    """
    try:
        lists = load_lists(excel_path)
        budgets = {cat: amt for cat, amt in lists.get("budgets", {}).items() if amt > 0}
        log.info("Loaded %d budget entries from Lists sheet", len(budgets))
        return budgets
    except Exception as error:
        log.warning("Could not load budgets from Lists: %s", error)
        return {}


# ── Reference lists ───────────────────────────────────────────────────────────

def load_lists(excel_path: Path) -> dict[str, list]:
    """
    Read all reference lists from the Lists sheet.

    Returns a dict with keys:
      months      — ['Jan', 'Feb', ..., 'Dec']
      txn_types   — ['Expense', 'Income', 'Savings']
      categories  — ['Groceries', 'Housing', ...]
      persons     — ['<YOUR_NAME>', '<FAMILY_MEMBER_1>', '<FAMILY_MEMBER_2>', '<FAMILY_MEMBER_3>']
      years       — [2024, 2025, 2026, 2027]

    Currency codes and rates are loaded separately by load_rates() since
    they include a rate column and need different handling.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb["Lists"]

        idx = col_indices(ws, ListsSchema)

        def read_col(field_name: str) -> list:
            c = idx.get(field_name)
            if c is None:
                log.warning("Lists sheet: column '%s' not found",
                            header_of(ListsSchema, field_name))
                return []
            values = []
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row, c).value
                if val is None or (isinstance(val, str) and val.startswith("←")):
                    break
                values.append(val)
            return values

        result = {
            "months":     read_col("months"),
            "txn_types":  read_col("txn_types"),
            "categories": read_col("categories"),
            "persons":    read_col("persons"),
            "years":      read_col("years"),
        }

        # Build category→budget mapping
        cat_c = idx.get("categories")
        bud_c = idx.get("budget_base")
        budgets: dict[str, float] = {}
        if cat_c and bud_c:
            for row in range(2, ws.max_row + 1):
                cat = ws.cell(row, cat_c).value
                bud = ws.cell(row, bud_c).value
                if cat is None:
                    break
                cat_str = str(cat).strip()
                if cat_str and bud is not None:
                    try:
                        budgets[cat_str] = float(bud)
                    except (TypeError, ValueError):
                        pass
        result["budgets"] = budgets

        return result

    except Exception as error:
        log.warning("Could not load Lists sheet: %s — using empty lists", error)
        return {
            "months": [], "txn_types": [], "categories": [],
            "persons": [], "years": [], "budgets": {},
        }


# ── Currency rate management ──────────────────────────────────────────────────

def update_currency_rates_in_excel(new_rates: dict[str, float]) -> None:
    """
    Write updated currency rates back to Lists sheet columns I (code) and J (rate).

    Only updates rows where the currency code already exists in the sheet.
    Does not add new currencies or remove existing ones.
    """
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["Lists"]

        idx = col_indices(ws, ListsSchema)
        ccy_col  = idx.get("currency")
        rate_col = idx.get("rate_to_base")

        if ccy_col is None or rate_col is None:
            log.warning("Currency or Rate column not found in Lists sheet — rates not updated")
            return

        for row in range(2, ws.max_row + 1):
            ccy = ws.cell(row, ccy_col).value
            if ccy and str(ccy).strip().upper() in new_rates:
                ws.cell(row, rate_col).value = round(new_rates[str(ccy).strip().upper()], 4)

        atomic_save(wb, excel_path)
        log.info("Updated %d currency rates in Lists sheet", len(new_rates))
    _invalidate_reference_cache()


def update_category_budget_in_excel(category: str, new_budget_base: float) -> None:
    """
    Write a new monthly budget limit (in base currency) for one category into the Lists
    sheet Budget (base) column. Only updates the row whose Categories cell
    already matches `category` — never adds or removes a category row.
    """
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["Lists"]

        idx      = col_indices(ws, ListsSchema)
        cat_col  = idx.get("categories")
        bud_col  = idx.get("budget_base")

        if cat_col is None or bud_col is None:
            log.warning("Categories or Budget (base) column not found in Lists sheet — budget not updated")
            return

        for row in range(2, ws.max_row + 1):
            cat = ws.cell(row, cat_col).value
            if cat and str(cat).strip() == category:
                ws.cell(row, bud_col).value = round(new_budget_base, 2)
                break
        else:
            log.warning("Category '%s' not found in Lists sheet — budget not updated", category)
            return

        atomic_save(wb, excel_path)
        log.info("Updated budget for category '%s' to %.2f (base)", category, new_budget_base)
    _invalidate_reference_cache()


# ── Transaction management ────────────────────────────────────────────────────

def get_recent_transactions(excel_path: Path, n: int = 5) -> list[dict]:
    """
    Return the last N data rows from MasterData with their Excel row indices.

    Each dict includes all column values plus '_row_idx' (1-based Excel row number).
    Used by the /delete command so it knows which row to remove.
    """
    from openpyxl import load_workbook

    wb      = load_workbook(excel_path, data_only=True)
    ws      = wb["MasterData"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows    = []

    for row_idx in range(2, ws.max_row + 1):
        row_data = {headers[c]: ws.cell(row_idx, c + 1).value
                    for c in range(len(headers))}
        if row_data.get("Value") is not None:
            row_data["_row_idx"] = row_idx
            rows.append(row_data)

    return rows[-n:] if len(rows) > n else rows


def append_transactions_batch(transactions: list) -> None:
    """
    Write multiple Transaction rows in a single open/save/upload cycle.

    Use this instead of calling append_transaction_row N times when you have
    multiple rows ready at once (e.g. bulk import from a receipt image).
    """
    from openpyxl import load_workbook

    if not transactions:
        return

    from excel_schema import (
        find_next_data_row, lists_currency_range, write_transaction_row,
        ensure_monthly_summary_rows_from_masterdata,
    )

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["MasterData"]

        lu_range = lists_currency_range(wb)
        r = find_next_data_row(ws)
        for transaction in transactions:
            write_transaction_row(ws, r, transaction.to_row(), lu_range)
            r += 1

        added = ensure_monthly_summary_rows_from_masterdata(wb)
        if added:
            log.info("Batch-appended: ensured %d Monthly Summary row(s)", added)
        atomic_save(wb, excel_path)
        log.info("Batch-appended %d transactions to MasterData", len(transactions))


def delete_transaction_row(row_idx: int, expected: dict | None = None) -> None:
    """
    Delete a single row from MasterData by its 1-based Excel row index.

    Shifts all rows below up by one. Uploads to remote storage on exit.

    If `expected` (a snapshot of Date/Value/Description captured when the
    user picked the row) is given, the row is re-verified under the write
    lock before deleting — protects against a stale row index if another
    delete/edit shifted rows in the meantime. Raises RowMovedError if the
    row no longer matches.
    """
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["MasterData"]
        if expected is not None:
            headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            if not _row_matches_snapshot(ws, headers, row_idx, expected):
                raise RowMovedError(
                    f"Row {row_idx} no longer matches the selected transaction — it may have moved."
                )
        ws.delete_rows(row_idx)
        atomic_save(wb, excel_path)
        log.info("Deleted MasterData row %d", row_idx)


def append_category_to_lists_sheet(name: str) -> None:
    """Append a new category name to the first empty row in the Categories column of the Lists sheet."""
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["Lists"]

        idx = col_indices(ws, ListsSchema)
        cat_col = idx.get("categories")
        if cat_col is None:
            log.warning("Categories column not found in Lists sheet — category not appended")
            return

        # Find first empty cell in the categories column (from row 2 down)
        target_row = 2
        for row in range(2, ws.max_row + 2):
            val = ws.cell(row, cat_col).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                target_row = row
                break

        ws.cell(target_row, cat_col, name)
        atomic_save(wb, excel_path)
        log.info("Appended category '%s' to Lists sheet row %d", name, target_row)
    _invalidate_reference_cache()


def rename_category_in_lists_sheet(old: str, new: str) -> None:
    """Rename a category in the Categories column of the Lists sheet."""
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["Lists"]

        idx = col_indices(ws, ListsSchema)
        cat_col = idx.get("categories")
        if cat_col is None:
            log.warning("Categories column not found in Lists sheet — category not renamed")
            return

        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, cat_col).value
            if val is not None and str(val).strip() == old:
                ws.cell(row, cat_col, new)
                atomic_save(wb, excel_path)
                log.info("Renamed category '%s' → '%s' in Lists sheet", old, new)
                _invalidate_reference_cache()
                return

        log.warning("Category '%s' not found in Lists sheet — not renamed", old)


def append_person_to_lists_sheet(name: str) -> None:
    """Append a new person name to the first empty row in the Persons column of the Lists sheet."""
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["Lists"]

        idx = col_indices(ws, ListsSchema)
        persons_col = idx.get("persons")
        if persons_col is None:
            log.warning("Persons column not found in Lists sheet — person not appended")
            return

        target_row = 2
        for row in range(2, ws.max_row + 2):
            val = ws.cell(row, persons_col).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                target_row = row
                break

        ws.cell(target_row, persons_col, name)
        atomic_save(wb, excel_path)
        log.info("Appended person '%s' to Lists sheet row %d", name, target_row)
    _invalidate_reference_cache()


def update_transaction_field(row_idx: int, field, value=None, expected: dict | None = None) -> None:
    """
    Update one or more fields of a MasterData row in a single save.

    `field` is either a column name (with `value` as the new cell value) or a
    dict of {column name: value} pairs applied atomically in one write. This
    function is purely generic — any domain logic (e.g. keeping Year/Month in
    sync with Date) belongs to the caller.

    If `expected` is given, the row is re-verified under the write lock
    before applying the change (see delete_transaction_row). Raises
    RowMovedError if the row no longer matches.
    """
    from openpyxl import load_workbook

    with ExcelFileContext() as excel_path:
        wb = load_workbook(excel_path)
        ws = wb["MasterData"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        if expected is not None and not _row_matches_snapshot(ws, headers, row_idx, expected):
            raise RowMovedError(
                f"Row {row_idx} no longer matches the selected transaction — it may have moved."
            )
        updates = field if isinstance(field, dict) else {field: value}
        for col_name, col_value in updates.items():
            col_idx = headers.get(col_name)
            if col_idx is None:
                raise ValueError(f"Column '{col_name}' not found")
            ws.cell(row_idx, col_idx, col_value)
        atomic_save(wb, excel_path)
        log.info("Updated MasterData row %d columns %s", row_idx, list(updates))
